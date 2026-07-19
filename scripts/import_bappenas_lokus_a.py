# -*- coding: utf-8 -*-
"""Import lokus pendukung Aspek A "Prioritas & Nilai Strategis" (penilaian
Bappenas) -> MySQL. Sumber: sheet "Kumpulan Data" baris 3-18 di
docs/docs/2_Analisis Prioritas untuk Bappenas dan Teknokratis 15.7.2026.xlsx
-- lihat docs/spec/Draf Penilaian Bappenas.md dan scripts/schema_bappenas_lokus_a.sql.

8 kriteria baru diimpor di sini (masing-masing dari file/sheet berbeda,
struktur kolom tidak seragam jadi tiap kriteria punya fungsi ekstraksi
sendiri): LOKPRI_RPJMN, PKSN, PERBATASAN, SR, SEKOLAH_GARUDA, KNMP, KDMP,
SWASEMBADA_PANGAN_RPJMN.

5 kriteria lain (PKPN, PERKEBUNAN, PERIKANAN, TRANSMIGRASI, KI_PRIORITAS)
SUDAH ada di tabel kawasan_tematik (scripts/import_kawasan_tematik.py) --
TIDAK diduplikasi di sini. BBM_1_HARGA (sheet "Lokus IJD BBM 1 HARGA", cuma
5 baris nasional per 17 Jul 2026, teks bebas + koordinat tanpa kolom
kabupaten/kota bersih) diimpor via regex "Kab./Kec. <nama>" pada gabungan
seluruh sel per baris -- lihat import_bbm_1_harga().

Semua fungsi import_* punya signature seragam (wb, ctx) -> list-of-rows
(ctx = dict kab_idx/kec_idx/kab_by_name/prov_idx dari build_master_index())
supaya bisa dipanggil baik dari CLI (main(), baca file dari docs/docs/) MAUPUN
dari endpoint upload di app.py (POST /api/bappenas-lokus-a/import) yang
menerima file xlsx dari browser -- lihat KRITERIA_SOURCES di bawah.

Usage (venv aktif):
    python scripts/import_bappenas_lokus_a.py
"""

import os
import re
import sys
from pathlib import Path

import openpyxl
import pymysql

BASE_DIR = Path(__file__).resolve().parent.parent
DOCS_DIR = BASE_DIR / "docs" / "docs"
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_bappenas_lokus_a.sql"

DB_HOST = os.environ.get("MYSQL_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("MYSQL_PORT", "3306"))
DB_USER = os.environ.get("MYSQL_USER", "root")
DB_PASS = os.environ.get("MYSQL_PASS", "")
DB_NAME = os.environ.get("MYSQL_DB", "route_gis")


def connect():
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        charset="utf8mb4", autocommit=False, cursorclass=pymysql.cursors.DictCursor,
    )
    with conn.cursor() as cur:
        cur.execute(
            "CREATE DATABASE IF NOT EXISTS %s CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
            % DB_NAME
        )
    conn.select_db(DB_NAME)
    return conn


def run_schema(conn):
    sql_text = SCHEMA_PATH.read_text(encoding="utf-8")
    code = "\n".join(l for l in sql_text.splitlines() if not l.strip().startswith("--"))
    with conn.cursor() as cur:
        for stmt in [s.strip() for s in code.split(";") if s.strip()]:
            cur.execute(stmt)
    conn.commit()


def norm(s):
    return " ".join(t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if t)


def norm_prov(s):
    toks = [t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if t]
    toks = ["DI" if t in ("DI", "D", "DAERAH") else t for t in toks]
    return " ".join(t for t in toks if t not in ("I", "ISTIMEWA"))


def strip_kab_prefix(s):
    n = norm(s)
    for p in ("KABUPATEN ", "KAB ", "KOTA "):
        if n.startswith(p):
            return n[len(p):]
    return n


def clean(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def build_master_index(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT DISTINCT kode_provinsi, provinsi, kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan")
        kab_master = cur.fetchall()
        cur.execute("SELECT kode_kabupaten, kecamatan, kode_kecamatan FROM penduduk_kecamatan")
        kec_master = cur.fetchall()
    kab_idx = {(norm_prov(m["provinsi"]), strip_kab_prefix(m["kabupaten_kota"])): m for m in kab_master}
    kec_idx = {(m["kode_kabupaten"], norm(m["kecamatan"])): m["kode_kecamatan"] for m in kec_master}
    # index tanpa provinsi -- utk sumber yang cuma punya nama kabupaten (mis.
    # Swasembada Pangan RPJMN). penduduk_kecamatan.kabupaten_kota tidak
    # menyimpan prefiks "Kabupaten"/"Kota" sama sekali (mis. keduanya cuma
    # "BANDUNG") -- Kab. Bandung (3204) & Kota Bandung (3273) makanya sama
    # persis namanya di master. Dibedakan pakai konvensi kode BPS: 2 digit
    # terakhir kode_kabupaten >= 71 = Kota (fakta administratif nasional,
    # bukan tebakan) -- key = (is_kota, nama dasar).
    kab_by_name = {}
    for m in kab_master:
        is_kota = (m["kode_kabupaten"] % 100) >= 71
        key = (is_kota, strip_kab_prefix(m["kabupaten_kota"]))
        kab_by_name[key] = None if key in kab_by_name else m
    prov_idx = {norm_prov(m["provinsi"]): m["kode_provinsi"] for m in kab_master}
    return {"kab_idx": kab_idx, "kec_idx": kec_idx, "kab_by_name": kab_by_name, "prov_idx": prov_idx}


def match_kab(kab_idx, provinsi, kabupaten):
    if not provinsi or not kabupaten:
        return None
    return kab_idx.get((norm_prov(provinsi), strip_kab_prefix(kabupaten)))


def match_kab_by_name(kab_by_name, kabupaten):
    if not kabupaten:
        return None
    is_kota = norm(kabupaten).startswith("KOTA ")
    return kab_by_name.get((is_kota, strip_kab_prefix(kabupaten)))


def match_kec(kec_idx, kode_kabupaten, kecamatan):
    if not kode_kabupaten or not kecamatan or "," in kecamatan or kecamatan.strip() in ("-", ""):
        return None
    return kec_idx.get((kode_kabupaten, norm(kecamatan)))


def _row(kriteria, level, provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
         keterangan, sumber_file, sumber_sheet):
    return (kriteria, level, provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
            keterangan, sumber_file, sumber_sheet)


def import_lokpri_rpjmn(wb, ctx, sumber_file="Data Highlight Intervensi Bab 4 RPJMN 2025 - 2029_R1_Konektivitas Jalan_20062025.xlsx"):
    """Sheet1 "Data Highlight Intervensi..." -- cuma level PROVINSI (kolom
    Provinsi terisi sekali per grup baris, kolom lain tidak relevan utk
    matching lokasi)."""
    ws = wb[wb.sheetnames[0]]
    out, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        provinsi = clean(row[1]) if len(row) > 1 else None
        if not provinsi or provinsi in seen:
            continue
        seen.add(provinsi)
        kode_prov = ctx["prov_idx"].get(norm_prov(provinsi))
        out.append(_row("LOKPRI_RPJMN", "PROVINSI", provinsi, None, None, kode_prov, None, None,
                         None, sumber_file, ws.title))
    return out


def import_pksn(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus PKSN" -- header baris 3, PROVINSI (kolom B) & KABUPATEN
    (kolom D) sebagian kosong di baris lanjutan (gaya sel gabung) -- perlu
    forward-fill provinsi. KABUPATEN & KECAMATAN (kolom H) terisi tiap baris."""
    ws = wb["Lokus PKSN"]
    kab_idx, kec_idx = ctx["kab_idx"], ctx["kec_idx"]
    out = []
    last_prov = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 8:
            continue
        provinsi = clean(row[1]) or last_prov
        kabupaten = clean(row[3])
        kecamatan = clean(row[7])
        if provinsi:
            last_prov = provinsi
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        kode_kec = match_kec(kec_idx, kode_kab, kecamatan) if kode_kab else None
        level = "KECAMATAN" if kecamatan else "KABUPATEN"
        out.append(_row("PKSN", level, provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
                         None, sumber_file, ws.title))
    return out


def import_perbatasan(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus Perbatasan" -- header baris 3, PROVINSI/KABUPATEN/
    KECAMATAN terisi tiap baris (tidak ada sel gabung)."""
    ws = wb["Lokus Perbatasan"]
    kab_idx, kec_idx = ctx["kab_idx"], ctx["kec_idx"]
    out = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        provinsi, kabupaten, kecamatan = clean(row[1]), clean(row[2]), clean(row[3])
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        kode_kec = match_kec(kec_idx, kode_kab, kecamatan) if kode_kab else None
        out.append(_row("PERBATASAN", "KECAMATAN" if kecamatan else "KABUPATEN",
                         provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
                         None, sumber_file, ws.title))
    return out


def import_sr(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus SR 2025" -- header baris 1, level Kab/Kota saja."""
    ws = wb["Lokus SR 2025"]
    kab_idx = ctx["kab_idx"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        provinsi, kabupaten = clean(row[1]), clean(row[4])
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        out.append(_row("SR", "KABUPATEN", provinsi, kabupaten, None, kode_prov, kode_kab, None,
                         clean(row[3]), sumber_file, ws.title))
    return out


_SEKOLAH_GARUDA_RX = re.compile(
    r"Kabupaten\s+([^,]+?)\s*,\s*Provinsi\s+(.+)$|Kota\s+([^,]+?)\s*,\s*Provinsi\s+(.+)$", re.I)


def import_sekolah_garuda(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus Sekolah Garuda " -- kolom "Lokus" teks bebas "Desa X,
    Kabupaten Y, Provinsi Z" -- kabupaten/provinsi diambil via regex."""
    ws = wb["Lokus Sekolah Garuda "]
    kab_idx = ctx["kab_idx"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3:
            continue
        lokus = clean(row[2])
        if not lokus:
            continue
        m = _SEKOLAH_GARUDA_RX.search(lokus)
        if not m:
            continue
        if m.group(1):
            kabupaten, provinsi = "Kabupaten " + m.group(1), m.group(2)
        else:
            kabupaten, provinsi = "Kota " + m.group(3), m.group(4)
        mk = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = mk["kode_provinsi"] if mk else None
        kode_kab = mk["kode_kabupaten"] if mk else None
        out.append(_row("SEKOLAH_GARUDA", "KABUPATEN", provinsi, kabupaten, None, kode_prov, kode_kab, None,
                         clean(row[1]), sumber_file, ws.title))
    return out


def import_knmp(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "KNMP 2025" -- header baris 2, Provinsi/Kabupaten/Kecamatan
    terisi tiap baris."""
    ws = wb["KNMP 2025"]
    kab_idx, kec_idx = ctx["kab_idx"], ctx["kec_idx"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 4:
            continue
        provinsi, kabupaten, kecamatan = clean(row[1]), clean(row[2]), clean(row[3])
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        kode_kec = match_kec(kec_idx, kode_kab, kecamatan) if kode_kab else None
        out.append(_row("KNMP", "KECAMATAN" if kecamatan else "KABUPATEN",
                         provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
                         clean(row[8]) if len(row) > 8 else None,
                         sumber_file, ws.title))
    return out


def import_kdmp(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus IJD KDMP" -- data mulai baris 5, baris header
    grup KODIM (kolom PROVINSI kosong) dilewati; level Kab/Kota."""
    ws = wb["Lokus IJD KDMP"]
    kab_idx = ctx["kab_idx"]
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or len(row) < 4:
            continue
        provinsi, kabupaten = clean(row[2]), clean(row[3])
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        out.append(_row("KDMP", "KABUPATEN", provinsi, kabupaten, None, kode_prov, kode_kab, None,
                         clean(row[5]) if len(row) > 5 else None,
                         sumber_file, ws.title))
    return out


_PAREN_SUFFIX_RX = re.compile(r"\s*\([^)]*\)\s*$")


def import_swasembada_pangan_lokus(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus Swasembada Pangan" -- header baris 2 (Provinsi kolom E,
    Kab/kota kolom F -- kadang ada anotasi "(padi)" dsb. yang perlu dibuang
    dulu sblm match, Kecamatan kolom G tapi isinya kosong/NBSP semua di data
    contoh -- tidak dipakai). Level Kab/Kota. BEDA dari SWASEMBADA_PANGAN_RPJMN
    (Lampiran IV RPJMN) yang sumbernya file lain & daftar kabupatennya beda."""
    ws = wb["Lokus Swasembada Pangan"]
    kab_idx = ctx["kab_idx"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 6:
            continue
        provinsi, kabupaten = clean(row[4]), clean(row[5])
        if not provinsi or not kabupaten:
            continue
        kabupaten = _PAREN_SUFFIX_RX.sub("", kabupaten).strip()
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        out.append(_row("SWASEMBADA_PANGAN_LOKUS", "KABUPATEN", provinsi, kabupaten, None,
                         kode_prov, kode_kab, None, clean(row[1]), sumber_file, ws.title))
    return out


_KEC_RX = re.compile(
    r"Kec(?:amatan)?\.?\s+([A-Za-zÀ-ÿ][\w'-]*(?:\s+[A-Za-zÀ-ÿ][\w'-]*)*?)"
    r"(?=\s*(?:,|\.|\n|\||$|Kab\b|Kel\b))", re.I)
_KAB_RX = re.compile(
    r"Kab(?:upaten)?\.?\s+([A-Za-zÀ-ÿ][\w'-]*(?:\s+[A-Za-zÀ-ÿ][\w'-]*)*?)"
    r"(?=\s*(?:,|\.|\n|\||$|Kec\b|Kel\b|Provinsi\b|Jalan\b))", re.I)


def import_bbm_1_harga(wb, ctx, sumber_file="6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"):
    """Sheet "Lokus IJD BBM 1 HARGA" -- HANYA 5 baris nasional (per 17 Jul
    2026), TANPA kolom kabupaten/kecamatan bersih: lokasi ada di teks bebas
    tersebar di beberapa kolom ("Ruas Jalan"/"Identifikasi Ruas"/"Lokasi
    Ruas yang Rusak" Dari-Ke/"Lokasi Yang dituju"), dan header sel gabung
    bikin index kolom antar baris tidak konsisten -- makanya SELURUH sel per
    baris digabung jadi satu teks lalu dicari pola "Kab./Kabupaten X" &
    "Kec./Kecamatan X" dgn regex, bukan baca kolom tetap. Kolom "Provinsi"
    (index 7) konsisten terisi bersih di kelima baris -- dipakai sbg batas
    provinsi supaya nama kab/kec pendek tidak salah cocok ke provinsi lain.

    Kabupaten yang disingkat di sumbernya (mis. "Kab. Lutra" utk "Luwu
    Utara") gagal cocok by name -- fallback: cocokkan kecamatan hasil regex
    ke SEMUA kabupaten di provinsi itu (kec_by_prov), pakai kalau hasilnya
    unik. Baris yang menyebut >1 kabupaten tanpa kecamatan spesifik (mis.
    "wilayah Kab. Dogiyai, Kab. Deiyai, dan Kab. Paniai") menghasilkan
    BEBERAPA baris keluaran, satu per kabupaten yang match, level KABUPATEN
    (bukan lokasi pasti -- SENGAJA beda dari pola kriteria lain yg 1 baris
    sumber = 1 baris keluaran).

    Regex ini diverifikasi manual cocok 5/5 baris sumber (17 Jul 2026) --
    KALAU sumbernya bertambah baris/berubah format, cek ulang manual,
    jangan percaya regex begitu saja."""
    ws = wb["Lokus IJD BBM 1 HARGA"]
    kab_idx, kec_idx = ctx["kab_idx"], ctx["kec_idx"]
    kec_by_prov = ctx.get("_kec_by_prov")
    if kec_by_prov is None:
        kec_by_prov = {}
        for (kode_kab, nama_kec_norm), kode_kec in kec_idx.items():
            kec_by_prov.setdefault((kode_kab // 100, nama_kec_norm), set()).add((kode_kab, kode_kec))
        ctx["_kec_by_prov"] = kec_by_prov

    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 8:
            continue
        provinsi = clean(row[7])
        if not provinsi:
            continue
        blob = " | ".join(clean(v) or "" for v in row)
        kab_candidates = list(dict.fromkeys(m.strip() for m in _KAB_RX.findall(blob)))
        kec_candidates = list(dict.fromkeys(m.strip() for m in _KEC_RX.findall(blob)))
        kode_prov = ctx["prov_idx"].get(norm_prov(provinsi))

        matched_kabs = []
        for kabc in kab_candidates:
            m = match_kab(kab_idx, provinsi, kabc)
            if m:
                matched_kabs.append(m["kode_kabupaten"])
        matched_kabs = list(dict.fromkeys(matched_kabs))

        if not matched_kabs and kec_candidates:
            for kecc in kec_candidates:
                cands = kec_by_prov.get((kode_prov, norm(kecc)))
                if cands and len(cands) == 1:
                    matched_kabs = [next(iter(cands))[0]]
                    break

        if len(matched_kabs) > 1:
            for kode_kab in matched_kabs:
                out.append(_row("BBM_1_HARGA", "KABUPATEN", provinsi, None, None,
                                 kode_prov, kode_kab, None, clean(row[3]), sumber_file, ws.title))
            continue

        kode_kab = matched_kabs[0] if matched_kabs else None
        kode_kec = None
        if kode_kab:
            for kecc in kec_candidates:
                kode_kec = match_kec(kec_idx, kode_kab, kecc)
                if kode_kec:
                    break
        level = "KECAMATAN" if kode_kec else "KABUPATEN"
        out.append(_row("BBM_1_HARGA", level, provinsi,
                         kab_candidates[0] if kab_candidates else None,
                         kec_candidates[0] if kode_kec else None,
                         kode_prov, kode_kab, kode_kec,
                         clean(row[3]), sumber_file, ws.title))
    return out


def import_swasembada_pangan_rpjmn(wb, ctx, sumber_file="Lampiran IV RPJMN 2025-2029.xlsx"):
    """Sheet2 "Lampiran IV RPJMN 2025-2029.xlsx" -- cuma kolom KAB/KOTA
    (tanpa provinsi terpisah) -- dicocokkan by nama kabupaten saja."""
    ws = wb["Sheet2"]
    kab_by_name = ctx["kab_by_name"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        kabupaten = clean(row[1])
        if not kabupaten:
            continue
        m = match_kab_by_name(kab_by_name, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        out.append(_row("SWASEMBADA_PANGAN_RPJMN", "KABUPATEN", None, kabupaten, None,
                         kode_prov, kode_kab, None, clean(row[2]) if len(row) > 2 else None,
                         sumber_file, ws.title))
    return out


def import_kpp_desa(wb, ctx, sumber_file="ALL LOKPRI PDAT (1).xlsx"):
    """Sheet "30 KPP Desa" -- Kawasan Perdesaan Prioritas, 30 kabupaten,
    tabel bersih satu baris per kawasan (Provinsi/Kabupaten/Nomenklatur KPP/
    Jumlah Desa). Kriteria BARU (tidak ada di 12 kriteria sebelumnya, dan
    tidak disebut di sheet Kumpulan Data -- ditemukan langsung dari file
    "ALL LOKPRI PDAT (1).xlsx" yang juga jadi sumber TRANSMIGRASI-nya versi
    lain; 19 Jul 2026). BEDA dari sheet "22 PPKP" di file yang sama --
    itu SENGAJA tidak diimpor krn cuma pecahan level kecamatan dari PKSN
    yang sudah ada (kolom "Nama PKSN" persis sama isinya)."""
    ws = wb["30 KPP Desa"]
    kab_idx = ctx["kab_idx"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3:
            continue
        provinsi, kabupaten = clean(row[1]), clean(row[2])
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        out.append(_row("KPP_DESA", "KABUPATEN", provinsi, kabupaten, None, kode_prov, kode_kab, None,
                         clean(row[3]) if len(row) > 3 else None, sumber_file, ws.title))
    return out


# Registry dipakai CLI (main() di bawah) MAUPUN endpoint upload di app.py
# (POST /api/bappenas-lokus-a/import) -- "file" = nama file default di
# docs/docs/ (dipakai CLI), "label" = teks utk dropdown UI upload.
KRITERIA_SOURCES = {
    "LOKPRI_RPJMN": {
        "label": "Lokpri RPJMN (Data Highlight Intervensi Bab 4 RPJMN)",
        "file": "Data Highlight Intervensi Bab 4 RPJMN 2025 - 2029_R1_Konektivitas Jalan_20062025.xlsx",
        "importer": import_lokpri_rpjmn,
    },
    "PKSN": {
        "label": "PKSN Perbatasan (6_Usulan Lokus IJD 2026, sheet Lokus PKSN)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_pksn,
    },
    "PERBATASAN": {
        "label": "Kecamatan Perbatasan Prioritas (6_Usulan Lokus IJD 2026, sheet Lokus Perbatasan)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_perbatasan,
    },
    "SR": {
        "label": "Sekolah Rakyat (6_Usulan Lokus IJD 2026, sheet Lokus SR 2025)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_sr,
    },
    "SEKOLAH_GARUDA": {
        "label": "Sekolah Unggul Garuda (6_Usulan Lokus IJD 2026, sheet Lokus Sekolah Garuda)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_sekolah_garuda,
    },
    "KNMP": {
        "label": "Kampung Nelayan Merah Putih (6_Usulan Lokus IJD 2026, sheet KNMP 2025)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_knmp,
    },
    "KDMP": {
        "label": "Koperasi Desa Merah Putih (6_Usulan Lokus IJD 2026, sheet Lokus IJD KDMP)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_kdmp,
    },
    "SWASEMBADA_PANGAN_RPJMN": {
        "label": "Swasembada Pangan RPJMN (Lampiran IV RPJMN 2025-2029, Sheet2)",
        "file": "Lampiran IV RPJMN 2025-2029.xlsx",
        "importer": import_swasembada_pangan_rpjmn,
    },
    "SWASEMBADA_PANGAN_LOKUS": {
        "label": "Lokus Swasembada Pangan (6_Usulan Lokus IJD 2026, sheet Lokus Swasembada Pangan)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_swasembada_pangan_lokus,
    },
    "BBM_1_HARGA": {
        "label": "BBM Satu Harga (6_Usulan Lokus IJD 2026, sheet Lokus IJD BBM 1 HARGA — regex teks bebas)",
        "file": "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx",
        "importer": import_bbm_1_harga,
    },
    "KPP_DESA": {
        "label": "Kawasan Perdesaan Prioritas (ALL LOKPRI PDAT, sheet 30 KPP Desa)",
        "file": "ALL LOKPRI PDAT (1).xlsx",
        "importer": import_kpp_desa,
    },
}


def import_kriteria(kriteria, wb, ctx):
    """Dipanggil dari app.py (upload) dan main() (CLI) -- return list of rows
    siap INSERT (lihat _row())."""
    spec = KRITERIA_SOURCES.get(kriteria)
    if not spec:
        raise ValueError(f"Kriteria tidak dikenal: {kriteria}")
    return spec["importer"](wb, ctx)


def main():
    conn = connect()
    try:
        run_schema(conn)
        ctx = build_master_index(conn)

        lokus_path = DOCS_DIR / "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"
        wb_lokus = openpyxl.load_workbook(lokus_path, read_only=True, data_only=True)

        all_rows = []
        for kriteria, spec in KRITERIA_SOURCES.items():
            path = DOCS_DIR / spec["file"]
            if not path.exists():
                print(f"  WARNING: file tidak ditemukan: {spec['file']} -- {kriteria} dilewati", file=sys.stderr)
                continue
            wb = wb_lokus if spec["file"] == lokus_path.name else openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = spec["importer"](wb, ctx)
            n_match = sum(1 for r in rows if r[6] is not None)
            print(f"  {kriteria}: {len(rows)} baris ({n_match} match kabupaten)")
            all_rows.extend(rows)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM bappenas_lokus_a WHERE kriteria IN %s", (tuple(KRITERIA_SOURCES.keys()),))
            cur.executemany(
                "INSERT INTO bappenas_lokus_a (kriteria, level, provinsi_asli, kabupaten_asli, "
                "kecamatan_asli, kode_provinsi, kode_kabupaten, kode_kecamatan, keterangan, "
                "sumber_file, sumber_sheet) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                all_rows,
            )
        conn.commit()
        print(f"\nTotal: {len(all_rows)} baris dimuat ke bappenas_lokus_a")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
