# -*- coding: utf-8 -*-
"""Impor statistik kecelakaan lalu lintas Korlantas POLRI 2020-2025
(docs/New/8. KESELAMATAN/ANEV LAKA LANTAS TAHUN 2020-2025 (BAPENAS
B-17549) 2.xlsx) ke 2 tabel referensi non-spasial -- lihat
scripts/schema_anev_laka_lantas.sql untuk skema & alasan long/tidy vs.
rectangular, dan docs/kajian_data_baru_docs_new.md §9.1 untuk hasil
telaah datanya. TIDAK terkait usulan_inpres/IJD.

Sheet REKAP: laporan hierarkis 25 kategori, tiap kategori (baris header:
NO terisi + URAIAN=nama kategori + SATUAN kosong) diikuti baris data
(NO kosong, URAIAN=label, SATUAN terisi, 6 kolom tahun terisi).
Sheet POLDA: rectangular per (polda, tahun) -- 5 kolom (KEJADIAN/MD/LB/
LR/RUMAT) per blok tahun, 6 blok tahun.

Idempotent: DELETE + INSERT ulang seluruh isi kedua tabel tiap run (bukan
UPSERT per baris) -- sumbernya satu file utuh yang diganti sekaligus tiap
update, bukan data yang di-append per usulan.

Usage (venv aktif):
    python scripts/import_anev_laka_lantas.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

from db import db_cursor as pg_cursor  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = (
    REPO_ROOT / "docs" / "New" / "8. KESELAMATAN-20260820T015309Z-1-001"
    / "8. KESELAMATAN" / "ANEV LAKA LANTAS TAHUN 2020-2025 (BAPENAS B-17549) 2.xlsx"
)
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_anev_laka_lantas.sql"

REKAP_YEAR_COLS = [4, 5, 6, 7, 8, 9]  # idx dlm row tuple -> THN 2020..JAN-30 OKT 2025
REKAP_YEAR_LABELS = ["2020", "2021", "2022", "2023", "2024", "JAN - 30 OKT 2025"]

# POLDA: 5 kolom (KEJADIAN, KORBAN MD, KORBAN LB, KORBAN LR, RUMAT) per blok
# tahun, mulai kolom idx 3, 6 blok tahun berturutan (idx 3-7, 8-12, ...).
POLDA_YEAR_LABELS = ["2020", "2021", "2022", "2023", "2024", "JAN - 30 OKT 2025"]
POLDA_BLOCK_START = 3
POLDA_BLOCK_WIDTH = 5


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _load_rekap():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["REKAP"]
    rows = []
    kategori_no, kategori = None, None
    for row in ws.iter_rows(min_row=11, values_only=True):
        no, uraian, satuan = row[1], row[2], row[3]
        if uraian is None:
            continue
        uraian = str(uraian).strip()
        if not uraian:
            continue
        if satuan is None and isinstance(no, int):
            # baris header kategori baru
            kategori_no, kategori = no, uraian
            continue
        if kategori is None:
            continue  # jaga-jaga baris sebelum kategori pertama terdeteksi
        for col, label in zip(REKAP_YEAR_COLS, REKAP_YEAR_LABELS):
            nilai = _num(row[col])
            if nilai is None:
                continue
            rows.append((kategori_no, kategori, uraian, satuan, label, nilai))
    return rows


def _load_polda():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["POLDA"]
    rows = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        polda = row[2]
        if not polda or not isinstance(polda, str):
            continue
        polda = polda.strip()
        for i, label in enumerate(POLDA_YEAR_LABELS):
            base = POLDA_BLOCK_START + i * POLDA_BLOCK_WIDTH
            kejadian, md, lb, lr, rumat = (row[base + k] for k in range(5))
            if all(v is None for v in (kejadian, md, lb, lr, rumat)):
                continue
            rows.append((
                polda, label,
                int(_num(kejadian)) if _num(kejadian) is not None else None,
                int(_num(md)) if _num(md) is not None else None,
                int(_num(lb)) if _num(lb) is not None else None,
                int(_num(lr)) if _num(lr) is not None else None,
                _num(rumat),
            ))
    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"GAGAL: tidak ditemukan {XLSX_PATH}")
        sys.exit(1)

    with pg_cursor() as cur:
        cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

    print(f"Membaca {XLSX_PATH.name}...")
    rekap_rows = _load_rekap()
    polda_rows = _load_polda()
    print(f"  REKAP: {len(rekap_rows)} baris (kategori x uraian x tahun)")
    print(f"  POLDA: {len(polda_rows)} baris (polda x tahun)")

    with pg_cursor() as cur:
        cur.execute("DELETE FROM anev_laka_lantas_nasional")
        cur.executemany(
            "INSERT INTO anev_laka_lantas_nasional "
            "(kategori_no, kategori, uraian, satuan, tahun, nilai) VALUES (%s, %s, %s, %s, %s, %s)",
            rekap_rows,
        )
        cur.execute("DELETE FROM anev_laka_lantas_polda")
        cur.executemany(
            "INSERT INTO anev_laka_lantas_polda "
            "(polda, tahun, kejadian, korban_md, korban_lb, korban_lr, kerugian_materi) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            polda_rows,
        )

    print("\nSelesai: anev_laka_lantas_nasional & anev_laka_lantas_polda di-refresh.")


if __name__ == "__main__":
    main()
