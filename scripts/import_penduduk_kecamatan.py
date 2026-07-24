# -*- coding: utf-8 -*-
"""Import/export xlsx "ID dan Jumlah Penduduk Indonesia" <-> MySQL (route_gis).

Sumber: docs/docs/3_ID dan Jumlah Penduduk Indonesia Tahun 2025.xlsx —
master penyelarasan kode wilayah BPS (provinsi/kabupaten/kecamatan) + jumlah
penduduk per kecamatan seluruh Indonesia, acuan kerangka CPIT dan dasar skor
C.A1 (Jumlah Penduduk Kecamatan) IJD.

Usage (venv aktif):
    python scripts/import_penduduk_kecamatan.py                 # import file default di docs/docs/
    python scripts/import_penduduk_kecamatan.py path/ke/file.xlsx
    python scripts/import_penduduk_kecamatan.py --export        # export DB -> xlsx
    python scripts/import_penduduk_kecamatan.py --export out.xlsx

Import bersifat upsert per (kode_kecamatan, tahun): baris yang sudah ada
di-UPDATE, yang belum di-INSERT. Header file boleh diawali baris judul —
baris header dikenali dari sel "Kode Provinsi". Export menghasilkan xlsx
dengan tata letak kolom yang sama sehingga bisa di-import balik.
"""

import argparse
import datetime
import os
import sys
from pathlib import Path

import openpyxl
import psycopg

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = BASE_DIR / "docs" / "docs" / "3_ID dan Jumlah Penduduk Indonesia Tahun 2025.xlsx"
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_penduduk_kecamatan.sql"

DB_HOST = os.environ.get("PG_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("PG_PORT", "5432"))
DB_USER = os.environ.get("PG_USER", "postgres")
DB_PASS = os.environ.get("PG_PASS", "")
DB_NAME = os.environ.get("PG_DB", "route_gis")

TAHUN = 2025

HEADERS = [
    "Kode Provinsi", "Provinsi", "Kode Kabupaten/Kota", "Kabupaten/Kota",
    "Kode Kecamatan/Distrik", "Kecamatan/Distrik", "Jumlah Penduduk", "Keterangan",
]
COLS = [
    "kode_provinsi", "provinsi", "kode_kabupaten", "kabupaten_kota",
    "kode_kecamatan", "kecamatan", "jumlah_penduduk", "keterangan",
]


def connect(select_db=True):
    return psycopg.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS, dbname=DB_NAME,
    )


def run_schema(conn):
    """Tabel sudah dibuat via scripts/migrate_pg_01_schema.py (lihat
    docs/migrasi_mysql_ke_postgresql.md) -- schema_penduduk_kecamatan.sql
    aslinya DDL MySQL, tidak bisa dieksekusi langsung ke PostgreSQL. Di
    sini cuma pastikan tabelnya benar-benar ada."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema='public' AND table_name='penduduk_kecamatan'"
        )
        if not cur.fetchone():
            raise RuntimeError(
                "Tabel penduduk_kecamatan belum ada di PostgreSQL -- jalankan "
                "scripts/migrate_pg_01_schema.py dulu."
            )


def _to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def upsert_xlsx(source, conn):
    """Upsert workbook (path atau file-like) ke penduduk_kecamatan. Return
    dict statistik. Raises ValueError bila header tidak dikenali."""
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    # cari baris header (file sumber punya 3 baris judul di atasnya)
    header = None
    for row in rows:
        if row and str(row[0]).strip() == "Kode Provinsi":
            header = [str(c).strip() if c is not None else "" for c in row]
            break
    if header is None:
        raise ValueError('Baris header dengan kolom "Kode Provinsi" tidak ditemukan di file xlsx')
    idx = {}
    for name in HEADERS:
        if name not in header:
            raise ValueError(f'Kolom wajib "{name}" tidak ada di file xlsx')
        idx[name] = header.index(name)

    with conn.cursor() as cur:
        cur.execute("SELECT kode_kecamatan FROM penduduk_kecamatan WHERE tahun = %s", (TAHUN,))
        existing = {r[0] for r in cur.fetchall()}

    sql = (
        f"INSERT INTO penduduk_kecamatan ({', '.join(COLS)}, tahun) "
        f"VALUES ({', '.join(['%s'] * len(COLS))}, {TAHUN}) "
        "ON CONFLICT (kode_kecamatan, tahun) DO UPDATE SET " +
        ", ".join(f"{c}=EXCLUDED.{c}" for c in COLS if c != "kode_kecamatan")
    )

    batch, total, inserted, updated = [], 0, 0, 0
    with conn.cursor() as cur:
        for row in rows:
            if len(row) < len(header):
                row = row + (None,) * (len(header) - len(row))
            kode_kec = _to_int(row[idx["Kode Kecamatan/Distrik"]])
            if kode_kec is None:
                continue  # baris judul/sumber/kosong
            ket = row[idx["Keterangan"]]
            batch.append((
                _to_int(row[idx["Kode Provinsi"]]),
                str(row[idx["Provinsi"]] or "").strip(),
                _to_int(row[idx["Kode Kabupaten/Kota"]]),
                str(row[idx["Kabupaten/Kota"]] or "").strip(),
                kode_kec,
                str(row[idx["Kecamatan/Distrik"]] or "").strip(),
                _to_int(row[idx["Jumlah Penduduk"]]),
                str(ket).strip() if ket is not None else None,
            ))
            if kode_kec in existing:
                updated += 1
            else:
                inserted += 1
            total += 1
            if len(batch) >= 1000:
                cur.executemany(sql, batch)
                batch.clear()
        if batch:
            cur.executemany(sql, batch)
    conn.commit()

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM penduduk_kecamatan")
        total_db = cur.fetchone()[0]
    return {"total": total, "inserted": inserted, "updated": updated, "total_db": total_db}


def export_xlsx(out, conn=None):
    """Export penduduk_kecamatan ke xlsx dengan tata letak sama dengan file
    sumber (bisa di-import balik). `out` boleh path atau file-like."""
    own_conn = conn is None
    if own_conn:
        conn = connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {', '.join(COLS)} FROM penduduk_kecamatan "
                f"WHERE tahun = %s ORDER BY kode_kecamatan", (TAHUN,)
            )
            rows = cur.fetchall()
    finally:
        if own_conn:
            conn.close()

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Lembar1")
    ws.append([f"Jumlah Penduduk Indonesia Tahun {TAHUN} Menurut Kecamatan/Distrik"])
    ws.append(["(Jiwa)"])
    ws.append([])
    ws.append(HEADERS)
    for row in rows:
        ws.append(list(row))
    wb.save(out)
    return len(rows)


def import_xlsx(xlsx_path):
    conn = connect()
    try:
        run_schema(conn)
        stats = upsert_xlsx(xlsx_path, conn)
    finally:
        conn.close()
    print(f"Selesai import {stats['total']} baris dari {xlsx_path.name}: "
          f"{stats['inserted']} baru diinsert, {stats['updated']} sudah ada di-update.")
    print("Total penduduk_kecamatan:", stats["total_db"])


def main():
    ap = argparse.ArgumentParser(
        description="Import (upsert) / export xlsx penduduk per kecamatan <-> MySQL route_gis")
    ap.add_argument("xlsx", nargs="?", default=None,
                    help="path file xlsx (import: default file 3_ID dan Jumlah Penduduk "
                         "di docs/docs/; export: default docs/penduduk_kecamatan_export_<ts>.xlsx)")
    ap.add_argument("--export", action="store_true",
                    help="export isi tabel ke xlsx alih-alih import")
    args = ap.parse_args()

    if args.export:
        out = Path(args.xlsx) if args.xlsx else BASE_DIR / "docs" / (
            "penduduk_kecamatan_export_%s.xlsx"
            % datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        n = export_xlsx(out)
        print(f"Export {n} baris penduduk_kecamatan -> {out}")
    else:
        xlsx = Path(args.xlsx) if args.xlsx else DEFAULT_XLSX
        if not xlsx.exists():
            sys.exit(f"File tidak ditemukan: {xlsx}")
        import_xlsx(xlsx)


if __name__ == "__main__":
    main()
