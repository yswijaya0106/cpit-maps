# -*- coding: utf-8 -*-
"""Import kawasan tematik (dasar parameter A3 "Tematik Tambahan") dari
docs/docs/6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx -> MySQL.

5 sheet dipilih (lihat scripts/schema_kawasan_tematik.sql untuk alasan
pemilihan & pemetaan nilai): Perkebunan, Kelautan-Perikanan, Transmigrasi,
4 sheet KI PSN/Hilirisasi/Prioritas RPJMN/Dirgantara (digabung kategori
KI_PRIORITAS), dan PKPN 3T (9.377 baris level desa, diagregasi ke kecamatan
unik supaya tabel tidak meledak).

kode_kabupaten dicari via pencocokan nama (provinsi+kabupaten) terhadap
master penduduk_kecamatan — teknik sama dengan build_wilayah_mapping.py;
kode_kecamatan hanya diisi bila nama kecamatan tunggal (bukan daftar
dipisah koma) dan match persis/longgar ke master.

Usage (venv aktif):
    python scripts/import_kawasan_tematik.py
"""

import os
import re
import sys
from pathlib import Path

import openpyxl
import psycopg
from psycopg.rows import dict_row

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = BASE_DIR / "docs" / "docs" / "6_Usulan Lokus IJD 2026 Sektor Bappenas.xlsx"
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_kawasan_tematik.sql"

DB_HOST = os.environ.get("PG_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("PG_PORT", "5432"))
DB_USER = os.environ.get("PG_USER", "postgres")
DB_PASS = os.environ.get("PG_PASS", "")
DB_NAME = os.environ.get("PG_DB", "route_gis")


def connect():
    return psycopg.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        dbname=DB_NAME, row_factory=dict_row,
    )


def run_schema(conn):
    """Tabel sudah dibuat via scripts/migrate_pg_01_schema.py -- di sini
    cuma pastikan ada (lihat docs/migrasi_mysql_ke_postgresql.md)."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema='public' AND table_name='kawasan_tematik'"
        )
        if not cur.fetchone():
            raise RuntimeError(
                "Tabel kawasan_tematik belum ada di PostgreSQL -- jalankan "
                "scripts/migrate_pg_01_schema.py dulu."
            )


def norm(s):
    return " ".join(t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if t)


def norm_prov(s):
    toks = [t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if t]
    toks = ["DI" if t in ("DI", "D", "DAERAH") else t for t in toks]
    return " ".join(t for t in toks if t not in ("I", "ISTIMEWA"))


def strip_kab_prefix(s):
    n = norm(s)
    for p in ("KABUPATEN ", "KAB ", "KOTA "):
        if n.startswith(p):
            return n[len(p):]
    return n


def data_rows(ws):
    """Baris data = baris dengan kolom pertama mirip nomor urut '1.0'/'1'."""
    for row in ws.iter_rows(values_only=True):
        if row and row[0] is not None and str(row[0]).strip().replace(".0", "").isdigit():
            yield row


def clean(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


# (sheet, kategori, idx_provinsi, idx_kabupaten, idx_kecamatan|None)
SHEET_SPECS = [
    ("Lokus Perkebunan ", "PERKEBUNAN", 4, 5, 6),
    ("Lokus Kelautan dan Perikanan", "PERIKANAN", 1, 2, 3),
    ("Lokus Transmigrasi", "TRANSMIGRASI", 3, 2, None),
    ("Lokus PKPN KI Dirgantara", "KI_PRIORITAS", 2, 3, 4),
    ("Lokus PKPN KI Hilirisasi", "KI_PRIORITAS", 2, 3, 4),
    ("Lokus PKPN KI Prioritas RPJMN", "KI_PRIORITAS", 2, 3, 4),
    ("Lokus KI PSN IUKI Sudah Terbit", "KI_PRIORITAS", 2, 3, 4),
]


def build_master_index(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT DISTINCT kode_provinsi, provinsi, kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan")
        kab_master = cur.fetchall()
        cur.execute("SELECT kode_kabupaten, kecamatan, kode_kecamatan FROM penduduk_kecamatan")
        kec_master = cur.fetchall()
    kab_idx = {(norm_prov(m["provinsi"]), strip_kab_prefix(m["kabupaten_kota"])): m for m in kab_master}
    kec_idx = {(m["kode_kabupaten"], norm(m["kecamatan"])): m["kode_kecamatan"] for m in kec_master}
    return kab_idx, kec_idx


def match_kab(kab_idx, provinsi, kabupaten):
    if not provinsi or not kabupaten:
        return None
    return kab_idx.get((norm_prov(provinsi), strip_kab_prefix(kabupaten)))


def match_kec(kec_idx, kode_kabupaten, kecamatan):
    if not kode_kabupaten or not kecamatan or "," in kecamatan or kecamatan.strip() in ("-", ""):
        return None
    return kec_idx.get((kode_kabupaten, norm(kecamatan)))


def import_sheet_rows(wb, sheet, kategori, i_prov, i_kab, i_kec, kab_idx, kec_idx):
    ws = wb[sheet]
    out = []
    for row in data_rows(ws):
        provinsi = clean(row[i_prov]) if i_prov < len(row) else None
        kabupaten = clean(row[i_kab]) if i_kab < len(row) else None
        kecamatan = clean(row[i_kec]) if i_kec is not None and i_kec < len(row) else None
        if not provinsi or not kabupaten:
            continue
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        kode_kec = match_kec(kec_idx, kode_kab, kecamatan) if kode_kab else None
        out.append((kategori, provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec, None, sheet))
    return out


def import_pkpn3t(wb, kab_idx, kec_idx):
    """Level desa (9.377 baris) -> agregasi unik (provinsi, kabupaten, kecamatan)."""
    ws = wb["Lokus PKPN 3T"]
    seen = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[3] or not row[4]:
            continue
        provinsi, kabupaten = clean(row[3]), clean(row[4])
        kecamatan = clean(row[6]) if len(row) > 6 else None
        key = (provinsi, kabupaten, kecamatan)
        if key not in seen:
            seen[key] = True
    out = []
    for provinsi, kabupaten, kecamatan in seen:
        m = match_kab(kab_idx, provinsi, kabupaten)
        kode_prov = m["kode_provinsi"] if m else None
        kode_kab = m["kode_kabupaten"] if m else None
        kode_kec = match_kec(kec_idx, kode_kab, kecamatan) if kode_kab else None
        out.append(("PKPN", provinsi, kabupaten, kecamatan, kode_prov, kode_kab, kode_kec,
                    "Diagregasi dari data level desa (3T/tertinggal)", "Lokus PKPN 3T"))
    return out


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.exists():
        sys.exit(f"File tidak ditemukan: {path}")

    conn = connect()
    try:
        run_schema(conn)
        kab_idx, kec_idx = build_master_index(conn)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        all_rows = []
        for sheet, kategori, i_prov, i_kab, i_kec in SHEET_SPECS:
            rows = import_sheet_rows(wb, sheet, kategori, i_prov, i_kab, i_kec, kab_idx, kec_idx)
            print(f"  {sheet.strip()!r} ({kategori}): {len(rows)} baris")
            all_rows.extend(rows)
        pkpn_rows = import_pkpn3t(wb, kab_idx, kec_idx)
        print(f"  'Lokus PKPN 3T' (PKPN, diagregasi ke kecamatan): {len(pkpn_rows)} baris unik")
        all_rows.extend(pkpn_rows)

        with conn.cursor() as cur:
            cur.executemany(
                "INSERT INTO kawasan_tematik (kategori, provinsi_asli, kabupaten_asli, "
                "kecamatan_asli, kode_provinsi, kode_kabupaten, kode_kecamatan, keterangan, "
                "sumber_sheet) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                "ON CONFLICT (kategori, sumber_sheet, provinsi_asli, kabupaten_asli, kecamatan_asli) "
                "DO UPDATE SET kode_provinsi=EXCLUDED.kode_provinsi, "
                "kode_kabupaten=EXCLUDED.kode_kabupaten, kode_kecamatan=EXCLUDED.kode_kecamatan, "
                "keterangan=EXCLUDED.keterangan",
                all_rows,
            )
        conn.commit()

        with conn.cursor() as cur:
            cur.execute("SELECT kategori, COUNT(*) n, SUM(kode_kabupaten IS NOT NULL) match_kab "
                        "FROM kawasan_tematik GROUP BY kategori")
            print("\nRekap per kategori (total | match kabupaten):")
            for r in cur.fetchall():
                print(f"  {r['kategori']}: {r['n']} | {r['match_kab']}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
