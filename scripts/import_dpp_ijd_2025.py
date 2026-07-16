# -*- coding: utf-8 -*-
"""Import DPP IJD TA 2025 (docs/docs/DPP_IJD 2025.xlsx) -> MySQL (route_gis)
+ pencocokan ke usulan Inpres 2026 untuk parameter E (Penuntasan).

Sumber dua sheet:
  - "LAMPIRAN BA"     : 534 kegiatan fisik (daftar berita acara)
  - "RINCIAN DPP VAL" : 112 kegiatan DPP; blok pivot kirinya memuat ID SITIA
                        siklus 2025 (TIDAK beririsan dengan id usulan 2026)

Usage (venv aktif):
    python scripts/import_dpp_ijd_2025.py                 # import + pencocokan
    python scripts/import_dpp_ijd_2025.py path/file.xlsx  # file lain
    python scripts/import_dpp_ijd_2025.py --match-only    # ulang pencocokan saja

Import bersifat upsert per (sumber, no_urut); schema dibuat dulu dari
scripts/schema_dpp_ijd_2025.sql bila belum ada, dan kolom
usulan_inpres.lanjutan_ijd_2025 (TINYINT NULL) ditambahkan bila belum ada.

Pencocokan (deterministik, tanpa fuzzy): nama ruas dinormalisasi (huruf besar,
buang isi tanda kurung, buang kata kegiatan pembuka spt "Preservasi Jalan",
semua tanda baca jadi spasi) lalu dicocokkan per (provinsi, kab/kota, ruas);
fallback (provinsi, ruas) bila unik. Hasil: dpp_ijd_2025.matched_usulan_id
terisi dan usulan_inpres.lanjutan_ijd_2025 = 1 (cocok) / 0 (tidak) — NULL
berarti pencocokan belum pernah dijalankan.
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pymysql

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = BASE_DIR / "docs" / "docs" / "DPP_IJD 2025.xlsx"
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_dpp_ijd_2025.sql"

DB_HOST = os.environ.get("MYSQL_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("MYSQL_PORT", "3306"))
DB_USER = os.environ.get("MYSQL_USER", "root")
DB_PASS = os.environ.get("MYSQL_PASS", "")
DB_NAME = os.environ.get("MYSQL_DB", "route_gis")

COLS = [
    "sumber", "no_urut", "id_sitia_2025", "nama_kegiatan", "jenis_penanganan",
    "status_jalan", "provinsi", "kewenangan", "pjg_jalan_km", "pjg_jbt_m",
    "alokasi_rp", "alokasi_ta2025_rp", "alokasi_ta2026_rp", "keterangan", "tematik",
]

# Kata pembuka nama kegiatan yang bukan bagian nama ruas — dibuang saat
# normalisasi kunci pencocokan.
_ACTIVITY_WORDS = {
    "PRESERVASI", "PEMBANGUNAN", "PENINGKATAN", "PENGGANTIAN", "PERBAIKAN",
    "REKONSTRUKSI", "REHABILITASI", "PEMELIHARAAN", "PELEBARAN", "PENANGANAN",
    "RUAS", "JALAN", "JEMBATAN", "JL", "JLN",
}


def connect():
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        charset="utf8mb4", autocommit=False,
    )
    with conn.cursor() as cur:
        cur.execute(
            "CREATE DATABASE IF NOT EXISTS %s CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
            % DB_NAME
        )
    conn.select_db(DB_NAME)
    return conn


def run_schema(conn):
    sql_text = SCHEMA_PATH.read_text(encoding="utf-8")
    code_lines = [ln for ln in sql_text.splitlines() if not ln.strip().startswith("--")]
    with conn.cursor() as cur:
        for stmt in [s.strip() for s in "\n".join(code_lines).split(";") if s.strip()]:
            cur.execute(stmt)
        # Kolom flag penuntasan di usulan_inpres (MySQL 8 belum mendukung
        # ADD COLUMN IF NOT EXISTS, jadi cek information_schema dulu).
        cur.execute(
            "SELECT COUNT(*) FROM information_schema.columns "
            "WHERE table_schema = %s AND table_name = 'usulan_inpres' "
            "AND column_name = 'lanjutan_ijd_2025'",
            (DB_NAME,),
        )
        if cur.fetchone()[0] == 0:
            cur.execute(
                "ALTER TABLE usulan_inpres ADD COLUMN lanjutan_ijd_2025 TINYINT NULL "
                "COMMENT 'Parameter E: 1=lanjutan/penuntasan IJD TA 2025, 0=usulan baru, NULL=belum dicocokkan'"
            )
    conn.commit()


def _num(v):
    if v is None or str(v).strip() in ("", "-", "#N/A"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _txt(v, maxlen=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "#N/A":
        return None
    return s[:maxlen] if maxlen else s


def _is_no(v):
    return v is not None and str(v).strip().isdigit()


def parse_workbook(path):
    """Return list of dict per kegiatan dari kedua sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []

    ws = wb["LAMPIRAN BA"]
    for r in ws.iter_rows(values_only=True):
        if not _is_no(r[5]) or not _txt(r[6]):
            continue
        out.append({
            "sumber": "BA", "no_urut": int(str(r[5]).strip()), "id_sitia_2025": None,
            "nama_kegiatan": _txt(r[6], 300), "jenis_penanganan": _txt(r[7], 60),
            "status_jalan": _txt(r[9], 4), "provinsi": _txt(r[10], 80),
            "kewenangan": _txt(r[11], 120), "pjg_jalan_km": _num(r[12]),
            "pjg_jbt_m": _num(r[13]), "alokasi_rp": _num(r[14]),
            "alokasi_ta2025_rp": None, "alokasi_ta2026_rp": None,
            "keterangan": _txt(r[15], 120), "tematik": _txt(r[16], 180),
        })

    ws = wb["RINCIAN DPP VAL"]
    rows = list(ws.iter_rows(values_only=True))
    # blok helper pivot (kolom H:ID, I:nama) -> peta nama ternormalisasi ke ID
    # SITIA 2025; dipakai hanya bila nama unik.
    helper = defaultdict(set)
    for r in rows:
        if _is_no(r[7]) and _txt(r[8]):
            helper[norm_ruas(str(r[8]))].add(int(str(r[7]).strip()))
    for r in rows:
        if not _is_no(r[11]) or not _txt(r[12]):
            continue
        ids = helper.get(norm_ruas(str(r[12])), set())
        out.append({
            "sumber": "DPP", "no_urut": int(str(r[11]).strip()),
            "id_sitia_2025": ids.pop() if len(ids) == 1 else None,
            "nama_kegiatan": _txt(r[12], 300), "jenis_penanganan": _txt(r[13], 60),
            "status_jalan": _txt(r[14], 4), "provinsi": _txt(r[15], 80),
            "kewenangan": _txt(r[16], 120), "pjg_jalan_km": _num(r[17]),
            "pjg_jbt_m": _num(r[18]), "alokasi_rp": _num(r[19]),
            "alokasi_ta2025_rp": _num(r[20]), "alokasi_ta2026_rp": _num(r[21]),
            "keterangan": None, "tematik": _txt(r[22], 180),
        })
    return out


def upsert(conn, rows):
    sql = (
        f"INSERT INTO dpp_ijd_2025 ({', '.join(COLS)}) "
        f"VALUES ({', '.join(['%s'] * len(COLS))}) "
        "ON DUPLICATE KEY UPDATE "
        + ", ".join(f"{c} = VALUES({c})" for c in COLS if c not in ("sumber", "no_urut"))
    )
    with conn.cursor() as cur:
        cur.executemany(sql, [[row[c] for c in COLS] for row in rows])
    conn.commit()


# --- Pencocokan usulan 2026 <-> DPP 2025 (parameter E Penuntasan) ---

def _norm_tokens(s):
    s = re.sub(r"\([^)]*\)", " ", str(s).upper())      # buang isi tanda kurung
    return [t for t in re.split(r"[^A-Z0-9]+", s) if t]


def norm_ruas(s):
    toks = _norm_tokens(s)
    while toks and toks[0] in _ACTIVITY_WORDS:
        toks.pop(0)
    return " ".join(toks)


def norm_wilayah(s):
    toks = _norm_tokens(s)
    toks = ["KABUPATEN" if t in ("KAB", "KABUPATEN") else t for t in toks]
    return " ".join(toks)


def norm_prov(s):
    toks = _norm_tokens(s)
    toks = ["DI" if t in ("DI", "D", "DAERAH") else t for t in toks]
    out = " ".join(t for t in toks if t not in ("I", "ISTIMEWA"))
    return out


def run_match(conn):
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute("SELECT id, nama_ruas, nama_kegiatan, provinsi, kabupaten_kota FROM usulan_inpres")
        usulan = cur.fetchall()
        cur.execute("SELECT id, nama_kegiatan, provinsi, kewenangan FROM dpp_ijd_2025")
        dpp = cur.fetchall()

    by_prov_kab = defaultdict(set)   # (prov, kab, ruas) -> {usulan_id}
    by_prov = defaultdict(set)       # (prov, ruas) -> {usulan_id}
    for u in usulan:
        prov = norm_prov(u["provinsi"] or "")
        kab = norm_wilayah(u["kabupaten_kota"] or "")
        for nama in (u["nama_ruas"], u["nama_kegiatan"]):
            ruas = norm_ruas(nama or "")
            if not ruas:
                continue
            by_prov_kab[(prov, kab, ruas)].add(u["id"])
            by_prov[(prov, ruas)].add(u["id"])

    matches, stats = [], defaultdict(int)
    for d in dpp:
        prov = norm_prov(d["provinsi"] or "")
        kab = norm_wilayah(d["kewenangan"] or "")
        ruas = norm_ruas(d["nama_kegiatan"] or "")
        cand = by_prov_kab.get((prov, kab, ruas), set())
        metode = "PROV_KAB_RUAS"
        if len(cand) != 1:
            cand = by_prov.get((prov, ruas), set())
            metode = "PROV_RUAS"
        if len(cand) == 1:
            matches.append((cand.copy().pop(), metode, d["id"]))
            stats[metode] += 1
        else:
            stats["AMBIGU" if len(cand) > 1 else "TIDAK_COCOK"] += 1

    with conn.cursor() as cur:
        cur.execute("UPDATE dpp_ijd_2025 SET matched_usulan_id = NULL, match_metode = NULL")
        cur.executemany(
            "UPDATE dpp_ijd_2025 SET matched_usulan_id = %s, match_metode = %s WHERE id = %s",
            matches,
        )
        # NULL -> 0/1 untuk seluruh usulan: skorer E membaca kolom ini.
        cur.execute("UPDATE usulan_inpres SET lanjutan_ijd_2025 = 0")
        matched_ids = sorted({m[0] for m in matches})
        if matched_ids:
            fmt = ",".join(["%s"] * len(matched_ids))
            cur.execute(
                f"UPDATE usulan_inpres SET lanjutan_ijd_2025 = 1 WHERE id IN ({fmt})",
                matched_ids,
            )
    conn.commit()
    return stats, len(matched_ids)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX))
    ap.add_argument("--match-only", action="store_true", help="jalankan ulang pencocokan tanpa import")
    args = ap.parse_args()

    conn = connect()
    try:
        run_schema(conn)
        if not args.match_only:
            path = Path(args.xlsx)
            if not path.exists():
                sys.exit(f"File tidak ditemukan: {path}")
            rows = parse_workbook(path)
            ba = sum(1 for r in rows if r["sumber"] == "BA")
            upsert(conn, rows)
            print(f"Import OK: {len(rows)} kegiatan (BA={ba}, DPP={len(rows) - ba}) dari {path.name}")

        stats, n_usulan = run_match(conn)
        total = sum(stats.values())
        print(f"Pencocokan: {total} kegiatan DPP 2025 -> {n_usulan} usulan 2026 ditandai lanjutan_ijd_2025=1")
        for k in ("PROV_KAB_RUAS", "PROV_RUAS", "AMBIGU", "TIDAK_COCOK"):
            if stats.get(k):
                print(f"  {k}: {stats[k]}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
