"""
Spatial-join tiap ruas LHR (`map_layers` layer='Jalan Nasional', attrs->>'LINKID')
terhadap poligon BATAS KECAMATAN untuk menentukan provinsi/kabupaten/kecamatan
yang dilalui, lalu tulis hasilnya sebagai kolom baru pada sheet "Per Ruas"
di file xlsx sumber LHR.

Dijalankan sekali (ad-hoc), bukan bagian dari pipeline import_*.py yang
biasa direrun.
"""
import sys
from collections import defaultdict

import openpyxl

from db import db_cursor

XLSX_PATH = (
    r"docs\New\8. KESELAMATAN-20260820T015309Z-1-001"
    r"\8. KESELAMATAN\Data Lalu Lintas Harian Rata-Rata (LHR)"
    r" Jalan Nasional Tahun 2024.xlsx"
)
SHEET_NAME = "Per Ruas"

BATCH_SIZE = 200

QUERY = """
WITH ruas AS (
    SELECT attrs->>'LINKID' AS linkid, geom
    FROM map_layers
    WHERE provinsi = 'JALAN NASIONAL' AND layer = 'Jalan Nasional'
      AND attrs->>'LINKID' = ANY(%s)
)
SELECT
    r.linkid,
    bk.attrs->>'PROVINSI' AS provinsi,
    bk.attrs->>'KABUPATEN_KOTA' AS kabupaten,
    bk.attrs->>'KECAMATAN' AS kecamatan,
    ST_Length(ST_Intersection(r.geom, bk.geom)::geography) AS panjang_m
FROM ruas r
JOIN map_layers bk
    ON bk.provinsi = 'BATAS KECAMATAN'
    AND ST_Intersects(r.geom, bk.geom)
"""

# Fallback untuk ruas yang tidak beririsan langsung (celah kecil di batas
# poligon) - kecamatan terdekat via KNN (operator <->, pakai idx_map_layers_geom
# GiST) lalu baru divalidasi jaraknya <= 200m di sisi Python. ST_DWithin/
# ST_Distance versi ::geography pada 7273 poligon tanpa bantuan KNN index
# ternyata brute-force (~36 ribu perhitungan jarak poligon per linkid,
# beberapa menit untuk cuma 5 baris) -- KNN membatasi ke 5 kandidat terdekat
# dulu baru dihitung jarak presisinya, jauh lebih murah.
FALLBACK_KNN_QUERY = """
SELECT
    bk.attrs->>'PROVINSI' AS provinsi,
    bk.attrs->>'KABUPATEN_KOTA' AS kabupaten,
    bk.attrs->>'KECAMATAN' AS kecamatan,
    ST_Distance(r.geom::geography, bk.geom::geography) AS jarak_m
FROM map_layers r, map_layers bk
WHERE r.provinsi = 'JALAN NASIONAL' AND r.layer = 'Jalan Nasional'
  AND r.attrs->>'LINKID' = %s
  AND bk.provinsi = 'BATAS KECAMATAN'
ORDER BY r.geom <-> bk.geom
LIMIT 5
"""

FALLBACK_RADIUS_M = 200


def _semua_linkid():
    with db_cursor() as cur:
        cur.execute(
            "SELECT attrs->>'LINKID' AS linkid FROM map_layers "
            "WHERE provinsi = 'JALAN NASIONAL' AND layer = 'Jalan Nasional'"
        )
        return [r["linkid"] for r in cur.fetchall()]


def main():
    semua_linkid = _semua_linkid()
    total = len(semua_linkid)
    batches = [semua_linkid[i:i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]

    panjang_per_ruas = defaultdict(lambda: defaultdict(float))
    selesai = 0
    for batch in batches:
        with db_cursor() as cur:
            cur.execute(QUERY, (batch,))
            rows = cur.fetchall()
        for row in rows:
            linkid = row["linkid"]
            key = (row["provinsi"], row["kabupaten"], row["kecamatan"])
            panjang_per_ruas[linkid][key] += float(row["panjang_m"] or 0)
        selesai += len(batch)
        pct = selesai / total * 100
        print(f"Progress spatial join: {selesai}/{total} ruas ({pct:.1f}%)", flush=True)

    linkid_hasil = {}
    for linkid, kecs in panjang_per_ruas.items():
        urut = sorted(kecs.items(), key=lambda kv: -kv[1])
        provinsi_list, kab_list, kec_list = [], [], []
        for (prov, kab, kec), _panjang in urut:
            if prov and prov not in provinsi_list:
                provinsi_list.append(prov)
            if kab and kab not in kab_list:
                kab_list.append(kab)
            if kec and kec not in kec_list:
                kec_list.append(kec)
        linkid_hasil[linkid] = (
            "; ".join(provinsi_list),
            "; ".join(kab_list),
            "; ".join(kec_list),
        )

    print(f"Ruas dengan hasil spatial join (irisan langsung): {len(linkid_hasil)}", flush=True)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    header_row = 1
    col_provinsi = ws.max_column + 1
    col_kabupaten = col_provinsi + 1
    col_kecamatan = col_provinsi + 2
    ws.cell(row=header_row, column=col_provinsi, value="Provinsi")
    ws.cell(row=header_row, column=col_kabupaten, value="Kabupaten/Kota")
    ws.cell(row=header_row, column=col_kecamatan, value="Kecamatan")

    linkid_col = 2  # kolom B = Linkid
    missing_linkids = []
    for r in range(4, ws.max_row + 1):
        linkid_cell = ws.cell(row=r, column=linkid_col).value
        if linkid_cell is None:
            continue
        linkid = str(linkid_cell).strip()
        hasil = linkid_hasil.get(linkid)
        if hasil is None:
            missing_linkids.append((r, linkid))
            continue
        ws.cell(row=r, column=col_provinsi, value=hasil[0])
        ws.cell(row=r, column=col_kabupaten, value=hasil[1])
        ws.cell(row=r, column=col_kecamatan, value=hasil[2])

    print(f"Baris tanpa hasil langsung (butuh fallback jarak): {len(missing_linkids)}", flush=True)

    if missing_linkids:
        fb_map = {}
        with db_cursor() as cur:
            for _r, linkid in missing_linkids:
                cur.execute(FALLBACK_KNN_QUERY, (linkid,))
                kandidat = cur.fetchall()
                terdekat = kandidat[0] if kandidat else None
                if terdekat and terdekat["jarak_m"] <= FALLBACK_RADIUS_M:
                    fb_map[linkid] = (terdekat["provinsi"], terdekat["kabupaten"], terdekat["kecamatan"])
                print(
                    f"Fallback KNN {linkid}: "
                    + (f"{terdekat['kecamatan']} ({terdekat['jarak_m']:.0f}m)" if terdekat else "tidak ada kandidat"),
                    flush=True,
                )
        still_missing = []
        for r, linkid in missing_linkids:
            hasil = fb_map.get(linkid)
            if hasil is None:
                still_missing.append(linkid)
                continue
            ws.cell(row=r, column=col_provinsi, value=hasil[0])
            ws.cell(row=r, column=col_kabupaten, value=hasil[1])
            ws.cell(row=r, column=col_kecamatan, value=hasil[2])
        print(f"Terisi via fallback (radius {FALLBACK_RADIUS_M}m): {len(fb_map)}", flush=True)
        if still_missing:
            print(f"Masih tanpa hasil sama sekali ({len(still_missing)}): {still_missing}", flush=True)

    print("Menyimpan file xlsx...", flush=True)
    wb.save(XLSX_PATH)
    print(f"Tersimpan: {XLSX_PATH}", flush=True)


if __name__ == "__main__":
    main()
