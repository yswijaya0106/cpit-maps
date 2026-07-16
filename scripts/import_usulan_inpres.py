"""Import/export xlsx usulan Inpres (SITIA) <-> MySQL (route_gis).

Usage (venv aktif):
    python scripts/import_usulan_inpres.py                  # import xlsx terbaru di docs/
    python scripts/import_usulan_inpres.py path/ke/file.xlsx  # import file tertentu
    python scripts/import_usulan_inpres.py --export         # export DB -> xlsx
    python scripts/import_usulan_inpres.py --export out.xlsx

Import bersifat upsert per ID usulan: baris yang sudah ada di tabel di-UPDATE,
yang belum ada di-INSERT (kolom geom_geojson/geom_fetched_at hasil fetch KML
tidak disentuh). Skema dibuat dulu dari scripts/schema_usulan_inpres.sql bila
belum ada. Export menghasilkan xlsx dengan tata letak kolom yang sama dengan
file sumber SITIA sehingga hasilnya bisa di-import balik.
"""

import argparse
import datetime
import os
import sys
from pathlib import Path

import openpyxl
import pymysql

BASE_DIR = Path(__file__).resolve().parent.parent
DOCS_DIR = BASE_DIR / "docs"
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_usulan_inpres.sql"

DB_HOST = os.environ.get("MYSQL_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("MYSQL_PORT", "3306"))
DB_USER = os.environ.get("MYSQL_USER", "root")
DB_PASS = os.environ.get("MYSQL_PASS", "")
DB_NAME = os.environ.get("MYSQL_DB", "route_gis")

# Kolom sumber -> kolom tabel usulan_inpres. Kolom yang tidak disebut di sini
# (terutama seluruh "File ..." dan jalur Kompetensi yang nyaris kosong)
# sengaja tidak diimpor sebagai kolom tabel utama.
COLUMN_MAP = {
    "ID": "id",
    "Pemda Simpan/Submit": "pemda_status",
    "Nama Pengusul": "nama_pengusul",
    "Provinsi": "provinsi",
    "Kabupaten/Kota": "kabupaten_kota",
    "Nama Kegiatan": "nama_kegiatan",
    "Prioritas": "prioritas",
    "Surat Dukungan": "surat_dukungan",
    "Pemberi Dukungan": "pemberi_dukungan",
    "Tgl. Pengusulan": "tgl_pengusulan",
    "No. Surat": "no_surat",
    "Tgl. Surat": "tgl_surat",
    "Perihal": "perihal",
    "Kode Koridor": "kode_koridor",
    "Nama Koridor": "nama_koridor",
    "Panjang Koridor": "panjang_koridor_km",
    "Kode Ruas": "kode_ruas",
    "Nama Ruas": "nama_ruas",
    "Panjang Ruas (KM)": "panjang_ruas_km",
    "Status Ruas": "status_ruas",
    "Lebar Jalan (M)": "lebar_jalan_m",
    "Kesesuaian Lebar Jalan (Balai)": "kesesuaian_lebar_jalan_balai",
    "No. Jembatan": "no_jembatan",
    "Nama Jembatan": "nama_jembatan",
    "Panjang Jembatan (M)": "panjang_jembatan_m",
    "Kondisi Jembatan": "kondisi_jembatan",
    "Jenis Penanganan": "jenis_penanganan",
    "Komponen (Balai)": "komponen_balai",
    "Komponen (Kompetensi)": "komponen_kompetensi",
    "Panjang Penanganan (Pemda)": "panjang_penanganan_pemda",
    "Panjang Penanganan (Balai)": "panjang_penanganan_balai",
    "Panjang Penanganan (Kompetensi)": "panjang_penanganan_kompetensi",
    "Satuan": "satuan",
    "Alokasi Usulan (Pemda)": "alokasi_usulan_pemda",
    "Alokasi Usulan (Balai)": "alokasi_usulan_balai",
    "Alokasi Usulan (Kompetensi)": "alokasi_usulan_kompetensi",
    "Kapasitas Fiskal": "kapasitas_fiskal",
    "Tematik Kawasan (Pemda)": "tematik_kawasan_pemda",
    "Tematik Kawasan (Balai)": "tematik_kawasan_balai",
    "Tematik Kawasan (Kompetensi)": "tematik_kawasan_kompetensi",
    "Kondisi Ruas Jalan Baik (KM)": "kondisi_baik_km",
    "Kondisi Ruas Jalan Sedang (KM)": "kondisi_sedang_km",
    "Kondisi Ruas Jalan Ringan (KM)": "kondisi_ringan_km",
    "Kondisi Ruas Jalan Berat (KM)": "kondisi_berat_km",
    "Seleksi Sistem": "seleksi_sistem",
    "Kelengkapan Input Data": "kelengkapan_input_data",
    "Verifikasi Balai": "verifikasi_balai",
    "Verifikasi Balai Oleh": "verifikasi_balai_oleh",
    "Catatan Pembahasan Balai": "catatan_pembahasan_balai",
    "Alasan Tolak/Terima Verifikasi Balai": "alasan_tolak_terima_balai",
    "Verifikasi Kompetensi": "verifikasi_kompetensi",
    "Dir. Pengampu Verifikasi Balai": "dir_pengampu_verifikasi_balai",
    "Prioritas Balai": "prioritas_balai",
    "Prioritas Kompetensi": "prioritas_kompetensi",
    "RC DED Pemda": "rc_ded_pemda",
    "RC DED Balai": "rc_ded_balai",
    "Catatan RC DED Balai": "catatan_rc_ded_balai",
    "RC FS Pemda": "rc_fs_pemda",
    "RC FS Balai": "rc_fs_balai",
    "Catatan RC FS Balai": "catatan_rc_fs_balai",
    "RC Lahan Pemda": "rc_lahan_pemda",
    "RC Lahan Balai": "rc_lahan_balai",
    "Catatan RC Lahan Balai": "catatan_rc_lahan_balai",
    "RC Dokling Pemda": "rc_dokling_pemda",
    "RC Dokling Balai": "rc_dokling_balai",
    "Catatan RC Dokling Balai": "catatan_rc_dokling_balai",
    "RAB Pemda": "rab_pemda",
    "RAB Balai": "rab_balai",
    "Catatan RAB Balai": "catatan_rab_balai",
    "Keterangan": "keterangan",
    "Catatan": "catatan",
    "Map Ruas KML Original": "kml_original_url",
    "Map Ruas KML dengan Data IJD": "kml_ijd_url",
}

# Kolom yang baru muncul/terisi mulai tarikan 15 Juli 2026 (jalur verifikasi
# Kompetensi + kolom prioritisasi). Opsional: diimpor hanya bila ada di file,
# sehingga tarikan lama (6 Juli, 139 kolom) tetap bisa di-import/-export.
# Nilai = (kolom tabel, DDL untuk migrasi ALTER TABLE otomatis di run_schema).
OPTIONAL_COLUMN_MAP = {
    "Prioritas DPR": ("prioritas_dpr", "SMALLINT UNSIGNED"),
    "Indikasi Ruas Prioritas PU": ("indikasi_prioritas_pu", "VARCHAR(10)"),
    # masih kosong di tarikan 15 Juli, tapi jadi komponen 10% Skor Prioritas
    # Nasional (G9) — otomatis terisi saat tarikan berikutnya diimpor
    "Indikasi Ruas Prioritas Bappenas": ("indikasi_prioritas_bappenas", "VARCHAR(10)"),
    "Indikasi Ruas Prioritas Kemenko IPK": ("indikasi_prioritas_kemenko", "VARCHAR(10)"),
    "Verifikasi Kompetensi Oleh": ("verifikasi_kompetensi_oleh", "VARCHAR(255)"),
    "Catatan Pembahasan Kompetensi": ("catatan_pembahasan_kompetensi", "TEXT"),
    "Alasan Tolak/Terima Verifikasi Kompetensi": ("alasan_tolak_terima_kompetensi", "TEXT"),
    "Dir. Pengampu Verifikasi Kompetensi": ("dir_pengampu_verifikasi_kompetensi", "VARCHAR(100)"),
    "RC DED Kompetensi": ("rc_ded_kompetensi", "VARCHAR(20)"),
    "Catatan RC DED Kompetensi": ("catatan_rc_ded_kompetensi", "TEXT"),
    "RC FS Kompetensi": ("rc_fs_kompetensi", "VARCHAR(20)"),
    "Catatan RC FS Kompetensi": ("catatan_rc_fs_kompetensi", "TEXT"),
    "RC Lahan Kompetensi": ("rc_lahan_kompetensi", "VARCHAR(20)"),
    "Catatan RC Lahan Kompetensi": ("catatan_rc_lahan_kompetensi", "TEXT"),
    "RC Dokling Kompetensi": ("rc_dokling_kompetensi", "VARCHAR(20)"),
    "Catatan RC Dokling Kompetensi": ("catatan_rc_dokling_kompetensi", "TEXT"),
    "RAB Kompetensi": ("rab_kompetensi", "VARCHAR(20)"),
    "Catatan RAB Kompetensi": ("catatan_rab_kompetensi", "TEXT"),
    "Jenis RC Dokling Balai": ("jenis_rc_dokling_balai", "VARCHAR(40)"),
    # SESUAI/TIDAK SESUAI — dipakai _ijd_score_koridor() (parameter D 2026)
    "Status Koridor Prioritas Balai": ("status_koridor_balai", "VARCHAR(20)"),
    # KP2B/LP2B dkk — calon sumber sub-parameter A4 (data dukung tematik)
    "Jenis Data Dukung Tematik (Kompetensi)": ("jenis_data_dukung_tematik_kompetensi", "VARCHAR(60)"),
    # YA = flag resmi penuntasan — diprioritaskan _ijd_score_penuntasan()
    "Penuntasan IJD Sebelumnya (Kompetensi)": ("penuntasan_ijd_kompetensi", "VARCHAR(10)"),
}

INT_COLUMNS = {
    "prioritas", "prioritas_balai", "prioritas_kompetensi", "prioritas_dpr",
    "alokasi_usulan_pemda", "alokasi_usulan_balai", "alokasi_usulan_kompetensi",
}
DATE_COLUMNS = {"tgl_pengusulan", "tgl_surat"}
DECIMAL_COLUMNS = {
    "panjang_koridor_km", "panjang_ruas_km", "lebar_jalan_m",
    "panjang_jembatan_m", "panjang_penanganan_pemda", "panjang_penanganan_balai",
    "panjang_penanganan_kompetensi", "kondisi_baik_km", "kondisi_sedang_km",
    "kondisi_ringan_km", "kondisi_berat_km",
}

# Kolom sumber "File ..." -> jenis_dokumen di usulan_dokumen
DOCUMENT_COLUMN_MAP = {
    "File Surat Resmi": "surat_resmi",
    "File Surat Dukungan": "surat_dukungan",
    "File Data Dukung Tematik": "data_dukung_tematik",
    "File Data Dukung Ruas Non-Status": "data_dukung_ruas_non_status",
    "File RC DED Pemda": "rc_ded_pemda",
    "File RC DED Balai": "rc_ded_balai",
    "File RC FS Pemda": "rc_fs_pemda",
    "File RC FS Balai": "rc_fs_balai",
    "File RC Lahan Pemda": "rc_lahan_pemda",
    "File RC Lahan Balai": "rc_lahan_balai",
    "File RC Dokling Pemda": "rc_dokling_pemda",
    "File RC Dokling Balai": "rc_dokling_balai",
    "File RAB Pemda": "rab_pemda",
    "File RAB Balai": "rab_balai",
    "File Project Digest Pemda": "project_digest_pemda",
    "File Project Digest Balai": "project_digest_balai",
    "File Telaah Balai": "telaah_balai",
    "File Berita Acara Balai": "berita_acara_balai",
    "File Surat Pernyataan Kesiapan Menerima Hibah": "pernyataan_hibah",
    "File Surat Pernyataan Terbebas dari Tuntutan Pihak Ketiga": "pernyataan_bebas_tuntutan",
    "File BAST dan/atau Surat Pertanyaan Kesiapan Menerima Aset dan Komitmen Menyediakan Anggaran Pemeliharaan": "bast_aset",
}


def clean_value(col, value):
    if value == "" or value is None:
        return None
    if col in INT_COLUMNS:
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    if col in DECIMAL_COLUMNS:
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    if col in DATE_COLUMNS:
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        try:
            return datetime.datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    if isinstance(value, str):
        return value.strip() or None
    return value


def run_schema(conn):
    sql_text = SCHEMA_PATH.read_text(encoding="utf-8")
    code_lines = [
        line for line in sql_text.splitlines()
        if not line.strip().startswith("--")
    ]
    statements = [s.strip() for s in "\n".join(code_lines).split(";") if s.strip()]
    with conn.cursor() as cur:
        for stmt in statements:
            cur.execute(stmt)
        # Migrasi kolom opsional (tarikan 15 Juli+) untuk DB yang tabelnya
        # sudah terlanjur dibuat dari schema lama (MySQL 8 belum mendukung
        # ADD COLUMN IF NOT EXISTS).
        cur.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_schema = %s AND table_name = 'usulan_inpres'",
            (DB_NAME,),
        )
        existing_cols = {r[0] for r in cur.fetchall()}
        for dest_col, ddl in OPTIONAL_COLUMN_MAP.values():
            if dest_col not in existing_cols:
                cur.execute(f"ALTER TABLE usulan_inpres ADD COLUMN {dest_col} {ddl}")
    conn.commit()


def connect(select_db=True):
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        charset="utf8mb4", autocommit=False,
    )
    if select_db:
        with conn.cursor() as cur:
            cur.execute(
                "CREATE DATABASE IF NOT EXISTS %s CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                % DB_NAME
            )
        conn.select_db(DB_NAME)
    return conn


def latest_xlsx():
    candidates = sorted(p for p in DOCS_DIR.glob("usulan_inpres_*.xlsx")
                        if "export" not in p.name)
    if not candidates:
        sys.exit(f"Tidak ada file usulan_inpres_*.xlsx di {DOCS_DIR}")
    return candidates[-1]


def upsert_xlsx(source, conn):
    """Upsert isi workbook xlsx (path atau file-like) ke usulan_inpres +
    usulan_dokumen lewat koneksi `conn` (database sudah terpilih, skema sudah
    ada). Return dict statistik. Dipakai CLI ini dan endpoint import di app.py.

    Raises ValueError bila kolom wajib tidak ada di file.
    """
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["Worksheet"] if "Worksheet" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_idx = {name: i for i, name in enumerate(header)}

    missing = [c for c in COLUMN_MAP if c not in header_idx]
    if missing:
        raise ValueError(f"Kolom wajib tidak ada di file xlsx: {missing}")
    doc_cols = {s: j for s, j in DOCUMENT_COLUMN_MAP.items() if s in header_idx}
    # kolom opsional: hanya yang benar-benar ada di file ini
    active_map = dict(COLUMN_MAP)
    active_map.update({s: dest for s, (dest, _ddl) in OPTIONAL_COLUMN_MAP.items()
                       if s in header_idx})

    with conn.cursor() as cur:
        cur.execute("SELECT id FROM usulan_inpres")
        existing_ids = {r[0] for r in cur.fetchall()}

    usulan_cols = list(active_map.values())
    insert_sql = (
        f"INSERT INTO usulan_inpres ({', '.join(usulan_cols)}) "
        f"VALUES ({', '.join(['%s'] * len(usulan_cols))}) "
        f"ON DUPLICATE KEY UPDATE " +
        ", ".join(f"{c}=VALUES({c})" for c in usulan_cols if c != "id")
    )
    doc_insert_sql = (
        "INSERT INTO usulan_dokumen (usulan_id, jenis_dokumen, url) VALUES (%s, %s, %s) "
        "ON DUPLICATE KEY UPDATE url=VALUES(url)"
    )

    usulan_batch, doc_batch = [], []
    total = inserted = updated = 0
    with conn.cursor() as cur:
        for row in rows:
            if len(row) < len(header):
                # openpyxl memangkas sel kosong di ujung baris
                row = row + (None,) * (len(header) - len(row))
            record = {}
            for src_name, dest_col in active_map.items():
                raw = row[header_idx[src_name]]
                record[dest_col] = clean_value(dest_col, raw)
            if record["id"] is None:
                continue
            usulan_batch.append(tuple(record[c] for c in usulan_cols))
            if record["id"] in existing_ids:
                updated += 1
            else:
                inserted += 1

            for src_name, jenis in doc_cols.items():
                url = row[header_idx[src_name]]
                if url:
                    doc_batch.append((record["id"], jenis, url))

            total += 1
            if len(usulan_batch) >= 500:
                cur.executemany(insert_sql, usulan_batch)
                usulan_batch.clear()
                if doc_batch:
                    cur.executemany(doc_insert_sql, doc_batch)
                    doc_batch.clear()

        if usulan_batch:
            cur.executemany(insert_sql, usulan_batch)
        if doc_batch:
            cur.executemany(doc_insert_sql, doc_batch)

    conn.commit()

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM usulan_inpres")
        total_usulan = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM usulan_dokumen")
        total_dokumen = cur.fetchone()[0]
    return {
        "total": total, "inserted": inserted, "updated": updated,
        "total_usulan": total_usulan, "total_dokumen": total_dokumen,
    }


def import_xlsx(xlsx_path):
    conn = connect()
    try:
        run_schema(conn)
        stats = upsert_xlsx(xlsx_path, conn)
    finally:
        conn.close()
    print(f"Selesai import {stats['total']} baris dari {xlsx_path.name}: "
          f"{stats['inserted']} baru diinsert, {stats['updated']} sudah ada di-update.")
    print("Total usulan_inpres:", stats["total_usulan"])
    print("Total usulan_dokumen:", stats["total_dokumen"])


def export_xlsx(out, conn=None):
    """Export usulan_inpres (+ URL dokumen) ke xlsx dengan header yang sama
    dengan file sumber SITIA, sehingga hasilnya bisa di-import balik.
    `out` boleh path atau file-like (BytesIO). Return jumlah baris."""
    own_conn = conn is None
    if own_conn:
        conn = connect()
    try:
        # sertakan kolom opsional (tarikan 15 Juli+) supaya roundtrip
        # export -> import tidak menghilangkan data — tapi hanya yang sudah
        # ada di DB (export tidak menjalankan migrasi schema)
        with conn.cursor() as cur:
            cur.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = %s AND table_name = 'usulan_inpres'",
                (DB_NAME,),
            )
            existing_cols = {r[0] for r in cur.fetchall()}
        full_map = dict(COLUMN_MAP)
        full_map.update({s: dest for s, (dest, _ddl) in OPTIONAL_COLUMN_MAP.items()
                         if dest in existing_cols})
        usulan_cols = list(full_map.values())
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {', '.join(usulan_cols)} FROM usulan_inpres ORDER BY id"
            )
            usulan_rows = cur.fetchall()
            cur.execute("SELECT usulan_id, jenis_dokumen, url FROM usulan_dokumen")
            docs = {}
            for usulan_id, jenis, url in cur.fetchall():
                docs.setdefault(usulan_id, {})[jenis] = url
    finally:
        if own_conn:
            conn.close()

    doc_headers = list(DOCUMENT_COLUMN_MAP)  # "File ..." sesuai sumber
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Worksheet")
    ws.append(["No."] + list(full_map) + doc_headers)
    for no, row in enumerate(usulan_rows, start=1):
        rec = dict(zip(usulan_cols, row))
        doc = docs.get(rec["id"], {})
        ws.append(
            [no]
            + [rec[c] for c in usulan_cols]
            + [doc.get(DOCUMENT_COLUMN_MAP[h]) for h in doc_headers]
        )
    wb.save(out)
    return len(usulan_rows)


def main():
    ap = argparse.ArgumentParser(
        description="Import (upsert) / export xlsx usulan Inpres <-> MySQL route_gis")
    ap.add_argument("xlsx", nargs="?", default=None,
                    help="path file xlsx (import: default file usulan_inpres_*.xlsx "
                         "terbaru di docs/; export: default docs/usulan_inpres_export_<ts>.xlsx)")
    ap.add_argument("--export", action="store_true",
                    help="export isi tabel ke xlsx alih-alih import")
    args = ap.parse_args()

    if args.export:
        out = Path(args.xlsx) if args.xlsx else DOCS_DIR / (
            "usulan_inpres_export_%s.xlsx"
            % datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        n = export_xlsx(out)
        print(f"Export {n} baris usulan_inpres -> {out}")
    else:
        xlsx = Path(args.xlsx) if args.xlsx else latest_xlsx()
        if not xlsx.exists():
            sys.exit(f"File tidak ditemukan: {xlsx}")
        import_xlsx(xlsx)


if __name__ == "__main__":
    main()
