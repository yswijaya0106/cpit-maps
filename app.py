import io
import json
import re
import secrets
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import List, Optional

import anthropic
import openpyxl
import requests
from dotenv import load_dotenv
import os

from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import geopandas as gpd
import pandas as pd
from pyproj import Geod
import shapely.wkt
from shapely.geometry import LineString, Point
from shapely.strtree import STRtree

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

# Logika import/export xlsx usulan Inpres (COLUMN_MAP, upsert, export) tinggal
# satu di scripts/import_usulan_inpres.py — dipakai bersama oleh CLI dan
# endpoint /api/usulan-inpres/import|export di bawah.
sys.path.insert(0, str(BASE_DIR / "scripts"))
import import_usulan_inpres as usulan_xlsx  # noqa: E402
import import_penduduk_kecamatan as penduduk_xlsx  # noqa: E402
import import_bappenas_lokus_a as bappenas_lokus_xlsx  # noqa: E402

from db import db_cursor  # noqa: E402
from map_layer_labels import map_layer_label as _map_layer_label  # noqa: E402
import chat_providers  # noqa: E402
# _llm_plain/_plain_* (penilaian Bappenas AI) masih tinggal di app.py dan
# butuh konstanta model/URL yang ikut pindah ke chat_providers saat refactor
# Fase 2 — tanpa import ini semua fitur AI Bappenas NameError.
from chat_providers import (  # noqa: E402
    GROQ_MODEL, GROQ_API_URL, GROK_MODEL, GROK_API_URL,
    OPENAI_MODEL, CLAUDE_MODEL, GEMINI_MODEL,
)

app = FastAPI(title="Route to SHP Converter")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
# Kompresi respons (GeoJSON layer peta bisa ~10MB mentah, mis. Jalan
# Nasional -- teks JSON terstruktur/berulang biasanya kompres 70-90%,
# jauh lebih terasa di intranet/LAN drpd naikkan CPU sedikit di server).
# minimum_size supaya respons kecil (mis. endpoint status) tidak ikut
# kena overhead gzip yang tak sepadan.
app.add_middleware(GZipMiddleware, minimum_size=1000)

APP_USERNAME = os.getenv("APP_USERNAME")
APP_PASSWORD = os.getenv("APP_PASSWORD")


@app.middleware("http")
async def basic_auth_middleware(request: Request, call_next):
    """Gerbang HTTP Basic Auth untuk seluruh aplikasi (statis + /api/*).

    Nonaktif otomatis kalau APP_USERNAME/APP_PASSWORD tidak diset di .env,
    supaya alur dev lokal yang sudah ada tidak berubah tanpa konfigurasi
    eksplisit.
    """
    if not APP_USERNAME or not APP_PASSWORD:
        return await call_next(request)

    auth_header = request.headers.get("authorization", "")
    if auth_header.startswith("Basic "):
        try:
            import base64
            decoded = base64.b64decode(auth_header[6:]).decode("utf-8")
            username, _, password = decoded.partition(":")
        except Exception:
            username, password = "", ""
        if secrets.compare_digest(username, APP_USERNAME) and secrets.compare_digest(
            password, APP_PASSWORD
        ):
            return await call_next(request)

    return Response(
        status_code=401,
        headers={"WWW-Authenticate": 'Basic realm="The Next - SiJalan"'},
    )


class Segment(BaseModel):
    segment_id: int
    road_name: Optional[str] = ""
    road_type: Optional[str] = ""
    distance_km: Optional[float] = 0
    duration_min: Optional[float] = 0
    bearing: Optional[float] = None
    start_lat: float
    start_lng: float
    end_lat: float
    end_lng: float


class RoutePayload(BaseModel):
    route_id: str
    route_name: str
    alternative: int
    transport_mode: str
    distance_km: float
    duration_min: float
    coordinates: List[List[float]]  # [[lat, lng], ...]
    segments: Optional[List[Segment]] = []


class ExportRequest(BaseModel):
    format: str
    routes: List[RoutePayload]


class RoadClassRequest(BaseModel):
    coordinates: List[List[float]]  # [[lat, lng], ...]


class RegionQuery(BaseModel):
    provinsi: Optional[str] = None
    kabupaten_kota: Optional[str] = None


class UsulanNearbyRequest(BaseModel):
    regions: List[RegionQuery]


class ChatMessage(BaseModel):
    role: str  # "user" | "assistant"
    text: str


class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    context: Optional[dict] = None


@app.get("/api/config")
def get_config():
    key = os.getenv("GOOGLE_MAPS_API_KEY", "")
    if not key:
        raise HTTPException(500, "GOOGLE_MAPS_API_KEY belum diset di file .env")
    return {"googleMapsApiKey": key}


def _routes_to_geodataframe(routes: List[RoutePayload]) -> gpd.GeoDataFrame:
    records = []
    for r in routes:
        line = LineString([(lng, lat) for lat, lng in r.coordinates])
        records.append({
            "route_id": r.route_id,
            "route_name": r.route_name[:80],
            "distance_km": round(r.distance_km, 3),
            "duration_min": round(r.duration_min, 2),
            "transport_mode": r.transport_mode,
            "alternative": r.alternative,
            "geometry": line,
        })
    return gpd.GeoDataFrame(records, geometry="geometry", crs="EPSG:4326")


def _build_geojson(routes: List[RoutePayload]) -> dict:
    gdf = _routes_to_geodataframe(routes)
    return json.loads(gdf.to_json())


def _build_vertex_csv(routes: List[RoutePayload]) -> str:
    lines = ["route_id,route_name,alternative,seq,lat,lng"]
    for r in routes:
        for seq, (lat, lng) in enumerate(r.coordinates):
            lines.append(f'{r.route_id},"{r.route_name}",{r.alternative},{seq},{lat},{lng}')
    return "\n".join(lines)


def _build_gpx(routes: List[RoutePayload]) -> str:
    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<gpx version="1.1" creator="analytic-maps" xmlns="http://www.topografix.com/GPX/1/1">',
    ]
    for r in routes:
        parts.append(f'  <trk><name>{r.route_name}</name><trkseg>')
        for lat, lng in r.coordinates:
            parts.append(f'    <trkpt lat="{lat}" lon="{lng}"></trkpt>')
        parts.append('  </trkseg></trk>')
    parts.append('</gpx>')
    return "\n".join(parts)


def _build_wkt(routes: List[RoutePayload]) -> str:
    lines = ["route_id|route_name|alternative|transport_mode|distance_km|duration_min|wkt"]
    for r in routes:
        line = LineString([(lng, lat) for lat, lng in r.coordinates])
        lines.append(
            f"{r.route_id}|{r.route_name}|{r.alternative}|{r.transport_mode}|"
            f"{r.distance_km}|{r.duration_min}|{line.wkt}"
        )
    return "\n".join(lines)


@app.post("/api/export")
def export_route(payload: ExportRequest):
    if not payload.routes:
        raise HTTPException(400, "Tidak ada rute untuk diekspor")

    fmt = payload.format.lower()

    if fmt == "geojson":
        content = json.dumps(_build_geojson(payload.routes), ensure_ascii=False, indent=2)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/geo+json",
            headers={"Content-Disposition": "attachment; filename=route.geojson"},
        )

    if fmt == "csv":
        content = _build_vertex_csv(payload.routes)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=route_vertices.csv"},
        )

    if fmt == "gpx":
        content = _build_gpx(payload.routes)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/gpx+xml",
            headers={"Content-Disposition": "attachment; filename=route.gpx"},
        )

    if fmt == "wkt":
        content = _build_wkt(payload.routes)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=route.wkt"},
        )

    if fmt == "shp":
        gdf = _routes_to_geodataframe(payload.routes)
        with TemporaryDirectory() as tmp:
            shp_path = Path(tmp) / "route.shp"
            gdf.to_file(shp_path, driver="ESRI Shapefile", engine="pyogrio")

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in Path(tmp).glob("route.*"):
                    zf.write(f, arcname=f.name)
            zip_buffer.seek(0)
            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=route_shp.zip"},
            )

    raise HTTPException(400, f"Format tidak dikenal: {fmt}")


OVERPASS_MIRRORS = [
    "https://overpass-api.de/api/interpreter",
    "https://lz4.overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]

# Approximate mapping from OSM `highway` tags to Indonesia's legal road
# hierarchy. OSM doesn't carry the official PUPR classification, so this is
# a best-effort estimate, not authoritative data.
HIGHWAY_CLASSIFICATION = {
    "motorway": "Jalan Nasional / Tol (perkiraan)",
    "motorway_link": "Jalan Nasional / Tol (perkiraan)",
    "trunk": "Jalan Nasional (perkiraan)",
    "trunk_link": "Jalan Nasional (perkiraan)",
    "primary": "Jalan Nasional/Provinsi (perkiraan)",
    "primary_link": "Jalan Nasional/Provinsi (perkiraan)",
    "secondary": "Jalan Provinsi (perkiraan)",
    "secondary_link": "Jalan Provinsi (perkiraan)",
    "tertiary": "Jalan Kabupaten/Kota (perkiraan)",
    "tertiary_link": "Jalan Kabupaten/Kota (perkiraan)",
    "unclassified": "Jalan Kabupaten/Kota (perkiraan)",
    "residential": "Jalan Desa/Lingkungan (perkiraan)",
    "living_street": "Jalan Desa/Lingkungan (perkiraan)",
    "service": "Jalan Lingkungan/Service (perkiraan)",
}

_GEOD = Geod(ellps="WGS84")


def _pick_sample_indices(n: int, max_samples: int) -> list:
    if n <= max_samples:
        return list(range(n))
    step = n / max_samples
    return sorted({int(i * step) for i in range(max_samples)} | {n - 1})


def _classify_highway(tag: Optional[str]) -> str:
    return HIGHWAY_CLASSIFICATION.get(tag, "Tidak diketahui")


@app.post("/api/analyze/road-classification")
def analyze_road_classification(payload: RoadClassRequest):
    coords = payload.coordinates
    if len(coords) < 2:
        raise HTTPException(400, "Rute terlalu pendek untuk dianalisis")

    sample_idx = _pick_sample_indices(len(coords), 60)
    samples = [coords[i] for i in sample_idx]
    coord_str = ",".join(f"{lat},{lng}" for lat, lng in samples)

    query = f"""
    [out:json][timeout:25];
    way(around:30,{coord_str})[highway];
    out tags geom;
    """

    data = None
    last_error = None
    for mirror_url in OVERPASS_MIRRORS:
        try:
            resp = requests.post(
                mirror_url,
                data={"data": query},
                headers={
                    "User-Agent": "analytic-maps/1.0 (RouteGIS road classification)",
                    "Accept": "application/json",
                },
                timeout=25,
            )
            resp.raise_for_status()
            data = resp.json()
            break
        except requests.RequestException as e:
            last_error = e
            continue

    if data is None:
        raise HTTPException(
            502,
            f"Semua server Overpass API (OpenStreetMap) sedang tidak dapat diakses/overload: {last_error}",
        )

    way_lines = []
    for way in data.get("elements", []):
        geom = way.get("geometry")
        if not geom or len(geom) < 2:
            continue
        line = LineString([(pt["lon"], pt["lat"]) for pt in geom])
        tags = way.get("tags", {})
        way_lines.append((line, tags.get("highway"), tags.get("ref"), tags.get("name")))

    tree = STRtree([wl[0] for wl in way_lines]) if way_lines else None
    threshold_deg = 0.0007  # ~75m at the equator

    def classify_point(lat, lng):
        if not tree:
            return "Tidak diketahui", None, None
        pt = Point(lng, lat)
        idx = tree.nearest(pt)
        line, highway_tag, ref_tag, name_tag = way_lines[idx]
        if pt.distance(line) > threshold_deg:
            return "Tidak diketahui", None, None
        return _classify_highway(highway_tag), ref_tag, name_tag

    per_vertex = [classify_point(lat, lng) for lat, lng in coords]

    # Aggregate into contiguous runs with real distance (meters) via geodesic calc.
    runs = []
    total_km = 0.0
    for i in range(len(coords) - 1):
        (lat1, lng1), (lat2, lng2) = coords[i], coords[i + 1]
        _, _, dist_m = _GEOD.inv(lng1, lat1, lng2, lat2)
        dist_km = dist_m / 1000
        total_km += dist_km
        label, ref, name = per_vertex[i]
        if runs and runs[-1]["road_type"] == label:
            runs[-1]["distance_km"] += dist_km
            if name and not runs[-1]["road_name"]:
                runs[-1]["road_name"] = name
        else:
            runs.append({"road_type": label, "road_name": name or "", "ref": ref or "", "distance_km": dist_km})

    summary = {}
    for r in runs:
        summary[r["road_type"]] = summary.get(r["road_type"], 0) + r["distance_km"]

    summary_list = [
        {
            "road_type": k,
            "distance_km": round(v, 3),
            "percentage": round((v / total_km * 100) if total_km else 0, 1),
        }
        for k, v in sorted(summary.items(), key=lambda kv: -kv[1])
    ]

    for r in runs:
        r["distance_km"] = round(r["distance_km"], 3)

    return {"total_km": round(total_km, 3), "segments": runs, "summary": summary_list}


USULAN_LIST_FIELDS = """
    id, provinsi, kabupaten_kota, nama_ruas, nama_kegiatan, kode_ruas,
    jenis_penanganan, status_ruas, panjang_ruas_km, alokasi_usulan_pemda,
    prioritas, seleksi_sistem, verifikasi_balai, kapasitas_fiskal,
    tematik_kawasan_pemda, kondisi_baik_km, kondisi_sedang_km,
    kondisi_ringan_km, kondisi_berat_km, kondisi_jembatan,
    (kml_original_url IS NOT NULL) AS has_geometry
"""


@app.post("/api/usulan-inpres/nearby")
def usulan_inpres_nearby(payload: UsulanNearbyRequest):
    """Cari usulan Inpres Jalan/Jembatan (SITIA Bina Marga) yang berada di
    provinsi/kabupaten-kota yang dilalui rute. Pencocokan berbasis wilayah
    administratif (bukan geometri presisi), karena hanya sebagian usulan yang
    punya data KML dan rute jalan raya bisa melintasi banyak ruas sekaligus."""
    seen_regions = []
    for r in payload.regions:
        if not r.provinsi and not r.kabupaten_kota:
            continue
        key = (r.provinsi or "", r.kabupaten_kota or "")
        if key not in seen_regions:
            seen_regions.append(key)

    if not seen_regions:
        return {"regions_matched": 0, "usulan": []}

    results = []
    seen_ids = set()
    with db_cursor() as cur:
        for provinsi, kabupaten_kota in seen_regions:
            conditions = []
            params = []
            if provinsi:
                conditions.append("provinsi LIKE %s")
                params.append(f"%{provinsi}%")
            if kabupaten_kota:
                conditions.append("kabupaten_kota LIKE %s")
                params.append(f"%{kabupaten_kota}%")
            where_clause = " AND ".join(conditions)
            cur.execute(
                f"""
                SELECT {USULAN_LIST_FIELDS}
                FROM usulan_inpres
                WHERE {where_clause}
                ORDER BY (seleksi_sistem = 'LULUS') DESC, prioritas ASC, alokasi_usulan_pemda DESC
                LIMIT 15
                """,
                params,
            )
            for row in cur.fetchall():
                if row["id"] in seen_ids:
                    continue
                seen_ids.add(row["id"])
                results.append(row)

    return {"regions_matched": len(seen_regions), "usulan": results}


@app.get("/api/usulan-inpres/provinsi")
def usulan_inpres_provinsi_list():
    with db_cursor() as cur:
        cur.execute(
            """
            SELECT provinsi, COUNT(*) AS jumlah
            FROM usulan_inpres
            WHERE provinsi IS NOT NULL
            GROUP BY provinsi
            ORDER BY provinsi ASC
            """
        )
        return cur.fetchall()


@app.get("/api/usulan-inpres/kabupaten")
def usulan_inpres_kabupaten_list(provinsi: Optional[str] = None):
    where_clause = "WHERE kabupaten_kota IS NOT NULL"
    params: list = []
    if provinsi:
        where_clause += " AND provinsi = %s"
        params.append(provinsi)
    with db_cursor() as cur:
        cur.execute(
            f"""
            SELECT kabupaten_kota, COUNT(*) AS jumlah
            FROM usulan_inpres
            {where_clause}
            GROUP BY kabupaten_kota
            ORDER BY kabupaten_kota ASC
            """,
            params,
        )
        return cur.fetchall()


@app.get("/api/usulan-inpres")
def usulan_inpres_list(
    provinsi: Optional[str] = None,
    kabupaten_kota: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    conditions = []
    params: list = []
    if provinsi:
        conditions.append("provinsi = %s")
        params.append(provinsi)
    if kabupaten_kota:
        conditions.append("kabupaten_kota LIKE %s")
        params.append(f"%{kabupaten_kota}%")
    if q:
        like = f"%{q}%"
        conditions.append("(nama_ruas LIKE %s OR nama_kegiatan LIKE %s OR kode_ruas LIKE %s)")
        params.extend([like, like, like])
    where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    with db_cursor() as cur:
        cur.execute(f"SELECT COUNT(*) AS total FROM usulan_inpres {where_clause}", params)
        total = cur.fetchone()["total"]
        cur.execute(
            f"""
            SELECT {USULAN_LIST_FIELDS}
            FROM usulan_inpres
            {where_clause}
            ORDER BY (seleksi_sistem = 'LULUS') DESC, prioritas ASC, alokasi_usulan_pemda DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()

    return {"total": total, "limit": limit, "offset": offset, "usulan": rows}


@app.get("/api/usulan-inpres/{usulan_id}")
def usulan_inpres_detail(usulan_id: int):
    with db_cursor() as cur:
        cur.execute("SELECT * FROM usulan_inpres WHERE id = %s", (usulan_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Usulan tidak ditemukan")
        cur.execute(
            "SELECT jenis_dokumen, url FROM usulan_dokumen WHERE usulan_id = %s",
            (usulan_id,),
        )
        row["dokumen"] = cur.fetchall()
    return row


@app.post("/api/usulan-inpres/import")
def usulan_inpres_import(file: UploadFile = File(...)):
    """Upsert xlsx usulan (format SITIA) ke database: ID yang sudah ada
    di-update, yang belum ada di-insert. Kolom geometri hasil fetch KML
    tidak disentuh."""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "File harus berformat .xlsx")
    conn = usulan_xlsx.connect()
    try:
        usulan_xlsx.run_schema(conn)
        try:
            stats = usulan_xlsx.upsert_xlsx(file.file, conn)
        except ValueError as exc:  # kolom wajib tidak ada
            raise HTTPException(400, str(exc))
    finally:
        conn.close()
    return {"filename": file.filename, **stats}


@app.get("/api/usulan-inpres/export/xlsx")
def usulan_inpres_export():
    """Unduh seluruh isi usulan_inpres sebagai xlsx dengan tata letak kolom
    yang sama dengan file sumber SITIA (bisa di-import balik)."""
    buf = io.BytesIO()
    usulan_xlsx.export_xlsx(buf)
    buf.seek(0)
    fname = f"usulan_inpres_export_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/api/penduduk-kecamatan/import")
def penduduk_kecamatan_import(file: UploadFile = File(...)):
    """Upsert xlsx master ID wilayah + jumlah penduduk per kecamatan (format
    "3_ID dan Jumlah Penduduk Indonesia") ke tabel penduduk_kecamatan."""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "File harus berformat .xlsx")
    conn = penduduk_xlsx.connect()
    try:
        penduduk_xlsx.run_schema(conn)
        try:
            stats = penduduk_xlsx.upsert_xlsx(file.file, conn)
        except ValueError as exc:
            raise HTTPException(400, str(exc))
    finally:
        conn.close()
    return {"filename": file.filename, **stats}


# --- Penampil data tabel database (topbar "Data") ---
# Whitelist nama tabel -> label tampilan; menghindari akses tabel arbitrer.
DATA_TABLES = {
    "usulan_inpres": "Usulan Inpres (SITIA)",
    "usulan_dokumen": "Dokumen Usulan",
    "penduduk_kecamatan": "Penduduk per Kecamatan (Nasional 2025)",
    "bps_kecamatan_demografi": "Demografi Kecamatan (Dalam Angka)",
    "bps_kabupaten_padi": "Padi per Kab/Kota (Dalam Angka)",
    "bps_kabupaten_kendaraan": "Kendaraan per Kab/Kota (Dalam Angka)",
    "bps_kabupaten_jalan": "Jalan per Kab/Kota (Dalam Angka)",
    "bps_kabupaten_indeks_penanaman": "Indeks Penanaman per Kab/Kota (Kertas Kerja)",
    "bps_kecamatan_potensi_tematik": "Potensi Tematik Kecamatan (Dalam Angka)",
    "si_panjang_jalan_provinsi": "Panjang Jalan per Provinsi (SI 2026)",
    "si_kendaraan_provinsi": "Kendaraan per Provinsi (SI 2026)",
    "si_lahan_sawah_provinsi": "Lahan Baku Sawah per Provinsi (SI 2026)",
    "ijd_scoring_rules": "Kaidah Skoring IJD",
    "dpp_ijd_2025": "DPP IJD TA 2025 (BA + DPP)",
    "wilayah_mapping": "Pemetaan Wilayah SITIA ↔ Kode BPS",
    "kecamatan_data_turunan": "Data Turunan Kecamatan (C.A1/C.A3)",
    "penilaian_bappenas_ai": "Draf Penilaian Bappenas (AI)",
    "bappenas_lokus_a": "Lokus Aspek A Bappenas (Prioritas & Nilai Strategis)",
    "kawasan_tematik": "Kawasan Tematik (A3 IJD)",
    "kemantapan_ijd_2026": "Kemantapan Jalan IJD 2026 (G8.A2)",
    "bappenas_koridor": "Daftar Koridor Bappenas Admin",
}
# kolom yang tidak ditampilkan (payload besar)
DATA_TABLE_SKIP_COLS = {"geom_geojson"}

# Tabel yang bisa difilter provinsi/kabupaten di viewer "Data" -> (ekspresi SQL
# kode provinsi, ekspresi SQL kode kabupaten), keduanya disetarakan ke kode BPS
# numerik (kode_provinsi 2 digit, kode_kabupaten 4 digit) supaya nilai filter dari
# dropdown (diisi dari penduduk_kecamatan, master nasional) selalu cocok — walau
# sebagian tabel BPS Dalam Angka menyimpan kode_kab sbg CHAR(4), bukan kolom
# kode_provinsi terpisah. usulan_inpres/dpp_ijd_2025/dll TIDAK di sini karena
# wilayahnya teks bebas SITIA (sudah ada filter provinsi sendiri di panel
# "Jelajahi Usulan Inpres"), bukan kode BPS langsung.
DATA_TABLE_GEO = {
    "penduduk_kecamatan": ("kode_provinsi", "kode_kabupaten"),
    "kecamatan_data_turunan": ("(kode_kabupaten / 100)", "kode_kabupaten"),
    "wilayah_mapping": ("kode_provinsi", "kode_kabupaten"),
    "bps_kecamatan_demografi": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bps_kabupaten_padi": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bps_kabupaten_kendaraan": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bps_kabupaten_jalan": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bps_kabupaten_indeks_penanaman": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bps_kecamatan_potensi_tematik": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    "bappenas_lokus_a": ("kode_provinsi", "kode_kabupaten"),
    "kawasan_tematik": ("kode_provinsi", "kode_kabupaten"),
    "kemantapan_ijd_2026": ("kode_provinsi", "kode_wilayah"),
    "bappenas_koridor": ("CAST(LEFT(kode_kab, 2) AS INTEGER)", "CAST(kode_kab AS INTEGER)"),
    # tabel level provinsi (Statistik Indonesia 2026) -- tidak ada dimensi
    # kabupaten, filter kabupaten sengaja dipetakan ke kolom yg sama supaya
    # aman kalau ter-pilih (tidak match apa pun, bukan error).
    "si_panjang_jalan_provinsi": ("kode_provinsi", "kode_provinsi"),
    "si_kendaraan_provinsi": ("kode_provinsi", "kode_provinsi"),
    "si_lahan_sawah_provinsi": ("kode_provinsi", "kode_provinsi"),
}


def _table_columns(cur, table: str) -> list:
    """Nama kolom tabel sesuai urutan asli -- pengganti `SHOW COLUMNS FROM`
    MySQL (tidak ada di PostgreSQL), lewat information_schema."""
    cur.execute(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_schema = 'public' AND table_name = %s ORDER BY ordinal_position",
        (table,),
    )
    return [r["column_name"] for r in cur.fetchall()]


@app.get("/api/data/tables")
def data_tables():
    out = []
    with db_cursor() as cur:
        # Cek keberadaan tabel LEWAT information_schema (bukan try/except di
        # sekitar query yang bisa gagal) -- kalau satu query di tengah loop
        # gagal (mis. tabel belum dibuat), PostgreSQL meng-abort SISA
        # transaksi itu (beda dari MySQL/PyMySQL yang tidak seketat itu),
        # jadi query berikutnya dlm loop yang sama ikut gagal berantai kalau
        # tidak dihindari dari awal.
        cur.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_name = ANY(%s)",
            (list(DATA_TABLES.keys()),),
        )
        existing = {r["table_name"] for r in cur.fetchall()}
        for name, label in DATA_TABLES.items():
            if name not in existing:
                continue  # tabel belum dibuat — sembunyikan dari daftar
            cur.execute(f'SELECT COUNT(*) AS n FROM "{name}"')
            total = cur.fetchone()["n"]
            out.append({"name": name, "label": label, "total": total, "geo": name in DATA_TABLE_GEO})
    return out


@app.get("/api/data/geo/provinces")
def data_geo_provinces():
    """Master provinsi (dari penduduk_kecamatan, cakupan nasional) — dipakai
    dropdown filter provinsi di viewer "Data" untuk semua tabel di DATA_TABLE_GEO."""
    with db_cursor() as cur:
        cur.execute("SELECT DISTINCT kode_provinsi, provinsi FROM penduduk_kecamatan ORDER BY provinsi")
        return cur.fetchall()


@app.get("/api/data/geo/kabupaten")
def data_geo_kabupaten(provinsi: int):
    with db_cursor() as cur:
        cur.execute(
            "SELECT DISTINCT kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan "
            "WHERE kode_provinsi = %s ORDER BY kabupaten_kota",
            (provinsi,),
        )
        return cur.fetchall()


# --- Import/export lokus Aspek A Bappenas (docs/spec/Draf Penilaian Bappenas.md)
# -- upload xlsx per kriteria dari browser, dipetakan ke fungsi ekstraksi yang
# sudah ada di scripts/import_bappenas_lokus_a.py (KRITERIA_SOURCES) supaya
# CLI dan endpoint ini pakai logika parsing yang SAMA persis, bukan duplikat.
# Export cukup lewat viewer "Data" generik (bappenas_lokus_a sudah masuk
# DATA_TABLES di atas) — tidak perlu endpoint export terpisah. ---

@app.get("/api/bappenas-lokus-a/kriteria")
def bappenas_lokus_a_kriteria():
    """Daftar 8 kriteria yang bisa di-upload ulang + jumlah baris saat ini
    (utk dropdown UI upload)."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT kriteria, COUNT(*) n FROM bappenas_lokus_a WHERE kriteria = ANY(%s) GROUP BY kriteria",
            (list(bappenas_lokus_xlsx.KRITERIA_SOURCES.keys()),),
        )
        counts = {r["kriteria"]: r["n"] for r in cur.fetchall()}
    return [
        {"kriteria": k, "label": spec["label"], "default_file": spec["file"], "total": counts.get(k, 0)}
        for k, spec in bappenas_lokus_xlsx.KRITERIA_SOURCES.items()
    ]


@app.post("/api/bappenas-lokus-a/import")
def bappenas_lokus_a_import(kriteria: str, file: UploadFile = File(...)):
    """Upload ulang xlsx sumber utk satu kriteria (update/tambah data) --
    baris lama utk kriteria itu diganti (DELETE + INSERT), kriteria lain
    tidak tersentuh."""
    if kriteria not in bappenas_lokus_xlsx.KRITERIA_SOURCES:
        raise HTTPException(400, f"Kriteria tidak dikenal: {kriteria}")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "File harus berformat .xlsx")

    conn = bappenas_lokus_xlsx.connect()
    try:
        bappenas_lokus_xlsx.run_schema(conn)
        ctx = bappenas_lokus_xlsx.build_master_index(conn)
        try:
            wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
            rows = bappenas_lokus_xlsx.import_kriteria(kriteria, wb, ctx)
        except KeyError as exc:
            raise HTTPException(400, f"Sheet tidak ditemukan di file ini: {exc}")
        if not rows:
            raise HTTPException(400, "Tidak ada baris yang bisa diekstrak dari file ini — cek formatnya sesuai sheet aslinya.")
        n_match = sum(1 for r in rows if r[6] is not None)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM bappenas_lokus_a WHERE kriteria = %s", (kriteria,))
            cur.executemany(
                "INSERT INTO bappenas_lokus_a (kriteria, level, provinsi_asli, kabupaten_asli, "
                "kecamatan_asli, kode_provinsi, kode_kabupaten, kode_kecamatan, keterangan, "
                "sumber_file, sumber_sheet) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                rows,
            )
        conn.commit()
    finally:
        conn.close()
    _ijd_bulk_cache.clear()  # lokus Aspek A baru -> skor bulk kadaluarsa
    _npr_bulk_cache.clear()  # preview NPR ikut membawa skor_teknokratis_100/total_gabungan dari cache di atas
    return {"kriteria": kriteria, "filename": file.filename, "total": len(rows), "match_kabupaten": n_match}


def _data_table_geo_where(table: str, provinsi: int, kabupaten: int, kriteria: str = ""):
    """WHERE + params dari filter provinsi/kabupaten, kalau tabelnya kebagian
    kode geo (DATA_TABLE_GEO) dan filter diisi. Kabupaten menang kalau
    keduanya diisi (provinsinya sudah tersirat). "kriteria" -- khusus
    bappenas_lokus_a (dialog "Lokus Bappenas" navbar) -- filter tambahan
    ANDed di atas geo, supaya dropdown kriteria bisa memfilter preview
    grid, bukan cuma menentukan target upload xlsx."""
    clauses, params = [], []
    geo = DATA_TABLE_GEO.get(table)
    if geo:
        prov_expr, kab_expr = geo
        if kabupaten:
            clauses.append(f"{kab_expr} = %s")
            params.append(kabupaten)
        elif provinsi:
            clauses.append(f"{prov_expr} = %s")
            params.append(provinsi)
    if kriteria and table == "bappenas_lokus_a":
        clauses.append("kriteria = %s")
        params.append(kriteria)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


@app.get("/api/data/{table}/export/xlsx")
def data_table_export(table: str, provinsi: int = 0, kabupaten: int = 0, kriteria: str = ""):
    if table not in DATA_TABLES:
        raise HTTPException(404, "Tabel tidak dikenal")
    where, params = _data_table_geo_where(table, provinsi, kabupaten, kriteria)
    with db_cursor() as cur:
        columns = [c for c in _table_columns(cur, table) if c not in DATA_TABLE_SKIP_COLS]
        col_sql = ", ".join(f'"{c}"' for c in columns)
        cur.execute(f'SELECT {col_sql} FROM "{table}" {where} ORDER BY 1', params)
        rows = cur.fetchall()

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(table[:31])
    ws.append(columns)
    for r in rows:
        ws.append([r[c] for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{table}_export_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/data/{table}")
def data_table_rows(table: str, limit: int = 50, offset: int = 0, provinsi: int = 0, kabupaten: int = 0, kriteria: str = ""):
    if table not in DATA_TABLES:
        raise HTTPException(404, "Tabel tidak dikenal")
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    where, params = _data_table_geo_where(table, provinsi, kabupaten, kriteria)
    with db_cursor() as cur:
        columns = [c for c in _table_columns(cur, table) if c not in DATA_TABLE_SKIP_COLS]
        col_sql = ", ".join(f'"{c}"' for c in columns)
        cur.execute(f'SELECT COUNT(*) AS n FROM "{table}" {where}', params)
        total = cur.fetchone()["n"]
        cur.execute(
            f'SELECT {col_sql} FROM "{table}" {where} ORDER BY 1 LIMIT %s OFFSET %s',
            params + [limit, offset],
        )
        rows = []
        for r in cur.fetchall():
            row = []
            for c in columns:
                v = r[c]
                if isinstance(v, str) and len(v) > 200:
                    v = v[:200] + "…"
                row.append(v)
            rows.append(row)
    return jsonable_encoder({
        "table": table, "label": DATA_TABLES[table], "columns": columns,
        "rows": rows, "total": total, "limit": limit, "offset": offset,
    })


@app.get("/api/penduduk-kecamatan/export/xlsx")
def penduduk_kecamatan_export():
    """Unduh master penduduk per kecamatan sebagai xlsx (bisa di-import balik)."""
    buf = io.BytesIO()
    penduduk_xlsx.export_xlsx(buf)
    buf.seek(0)
    fname = f"penduduk_kecamatan_export_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# --- Skoring "Prioritisasi Teknokratik" IJD (Inpres Jalan Daerah No. 11/2025) ---
# Kaidah nilai/bobot per parameter (Tabel 1, 3, 6, A1/A2) disimpan di tabel
# ijd_scoring_rules (lihat scripts/schema_ijd_scoring.sql) supaya bisa
# disesuaikan tanpa redeploy saat kebijakan IJD berubah tiap tahun anggaran.
# Hanya parameter A (tematik, sub A1/A2 saja), B (kemantapan), D (koridor,
# pendekatan), dan F (kelengkapan RC; hanya kaidah 2025 — dihapus dari
# penilaian 2026) yang punya sumber data di tabel usulan_inpres — C
# (Kemanfaatan) dan E (Penuntasan IJD sebelumnya) dilaporkan "belum tersedia"
# karena sumber datanya (BPS/LHR untuk C, pencocokan DPP IJD 2025 untuk E)
# belum diimpor ke aplikasi ini. Bobot maksimal parameter pending berbeda per
# tahun kaidah (Tabel 1 dokumen 14072026: C naik 20 -> 25).
IJD_PENDING_PARAMETERS = {
    "C": {
        "parameter_label": "Kemanfaatan (kepadatan penduduk, produktivitas lahan, LHR)",
        "bobot_maks_per_tahun": {2025: 20, 2026: 25},
        "alasan": "Belum ada sumber data kepadatan penduduk per kecamatan, produktivitas lahan, dan volume lalu lintas (LHR) di aplikasi ini.",
    },
    "E": {
        # Terpakai hanya untuk tahun kaidah tanpa rules E di DB (2025) —
        # kaidah 2026 menilai E lewat _ijd_score_penuntasan().
        "parameter_label": "Kegiatan IJD Sebelumnya (Penuntasan)",
        "bobot_maks_per_tahun": {2025: 10, 2026: 10},
        "alasan": "Skoring penuntasan hanya diimplementasikan pada kaidah 2026 (pencocokan DPP IJD TA 2025).",
    },
}

# (label komponen, kolom deklarasi Pemda, kolom hasil verifikasi Balai)
_RC_FIELD_MAP = [
    ("DED", "rc_ded_pemda", "rc_ded_balai"),
    ("RAB", "rab_pemda", "rab_balai"),
    ("DOKLING", "rc_dokling_pemda", "rc_dokling_balai"),
    ("LAHAN", "rc_lahan_pemda", "rc_lahan_balai"),
]


def _load_ijd_rules(tahun: int) -> dict:
    with db_cursor() as cur:
        cur.execute(
            "SELECT parameter_kode, parameter_label, bobot_maks, sub_kode, kondisi_label, nilai "
            "FROM ijd_scoring_rules WHERE tahun_berlaku = %s",
            (tahun,),
        )
        rows = cur.fetchall()
    rules: dict = {}
    for r in rows:
        p = rules.setdefault(r["parameter_kode"], {"label": r["parameter_label"], "bobot_maks": float(r["bobot_maks"]), "subs": {}})
        p["subs"][r["sub_kode"]] = {"label": r["kondisi_label"], "nilai": float(r["nilai"])}
    return rules


def _ijd_score_tematik(row: dict, rules: dict, ctx: dict = None) -> dict:
    rule = rules.get("A")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah tematik belum diset di database."}
    # A1 di dokumen 14072026 berjudul "Tematik SITIA (Kompetensi)" — pakai
    # hasil verifikasi kompetensi bila ada, fallback verifikasi Balai, baru
    # deklarasi Pemda (pola sama dengan RC balai->pemda).
    tematik, sumber = "", "deklarasi Pemda"
    for col, label_sumber in (
        ("tematik_kawasan_kompetensi", "verifikasi kompetensi"),
        ("tematik_kawasan_balai", "verifikasi Balai"),
        ("tematik_kawasan_pemda", "deklarasi Pemda"),
    ):
        tematik = (row.get(col) or "").strip()
        if tematik:
            sumber = label_sumber
            break
    sub = rule["subs"].get(tematik)
    if not sub:
        keterangan = (
            f"Kategori tematik '{tematik}' ({sumber}) tidak dikenali kaidah." if tematik
            else "Usulan tidak mencantumkan tematik kawasan."
        )
        return {"tersedia": False, "keterangan": keterangan}

    nilai = sub["nilai"]
    detail = [f"{sub['label']} — sumber: {sumber}"]
    # A4 data dukung tematik (kaidah 2026: rules ber-sub_kode "A4_<STATUS>",
    # nilai sudah tertimbang x10%) dari kolom hasil verifikasi kompetensi.
    # Kaidah tanpa sub A4 (2025) memakai pesan lama apa adanya.
    if any(k.startswith("A4_") for k in rule["subs"]):
        status_a4 = (row.get("jenis_data_dukung_tematik_kompetensi") or "").strip().upper()
        sub_a4 = rule["subs"].get(f"A4_{status_a4}") if status_a4 else None
        if sub_a4:
            nilai += sub_a4["nilai"]
            detail.append(sub_a4["label"])
        elif status_a4:
            detail.append(f"A4: status data dukung '{status_a4}' tidak dikenali kaidah")
        else:
            detail.append("A4: data dukung belum dinilai verifikasi kompetensi")
        # A3 tematik tambahan: DUA sumber independen digabung, nilai tertinggi
        # dipakai bila cocok >1 kategori dan/atau >1 sumber —
        #  (a) tabel kawasan_tematik (data lokus Bappenas) by kabupaten/kecamatan
        #  (b) kecamatan_data_turunan.potensi_* (BPS Dalam Angka, parsial) by
        #      kecamatan — lihat scripts/extract_dalam_angka.py POTENSI_TABLES
        if any(k.startswith("A3_") for k in rule["subs"]):
            kategori_cocok = []
            kode_kab = (row["kode_kecamatan"] // 1000) if row.get("kode_kecamatan") else None
            if not kode_kab:
                if ctx and "kab_by_wilayah" in ctx:
                    kode_kab = ctx["kab_by_wilayah"].get((row.get("provinsi"), row.get("kabupaten_kota")))
                else:
                    with db_cursor() as cur:
                        cur.execute(
                            "SELECT kode_kabupaten FROM wilayah_mapping "
                            "WHERE provinsi_sitia = %s AND kabupaten_kota_sitia = %s",
                            (row.get("provinsi"), row.get("kabupaten_kota")),
                        )
                        r = cur.fetchone()
                    kode_kab = r["kode_kabupaten"] if r else None
            if kode_kab:
                if ctx and "kawasan_by_kab" in ctx:
                    kategori_cocok += [f"A3_{r['kategori']}" for r in ctx["kawasan_by_kab"].get(kode_kab, [])
                                        if r["kode_kecamatan"] is None or r["kode_kecamatan"] == row.get("kode_kecamatan")]
                else:
                    with db_cursor() as cur:
                        cur.execute(
                            "SELECT DISTINCT kategori FROM kawasan_tematik WHERE kode_kabupaten = %s "
                            "AND (kode_kecamatan IS NULL OR kode_kecamatan = %s)",
                            (kode_kab, row.get("kode_kecamatan")),
                        )
                        kategori_cocok += [f"A3_{r['kategori']}" for r in cur.fetchall()]

            kode_kec = row.get("kode_kecamatan")
            potensi = None
            if kode_kec:
                if ctx and "potensi_by_kec" in ctx:
                    potensi = ctx["potensi_by_kec"].get(kode_kec)
                else:
                    with db_cursor() as cur:
                        cur.execute(
                            "SELECT potensi_pertanian, potensi_perkebunan, potensi_peternakan, "
                            "potensi_perikanan FROM kecamatan_data_turunan WHERE kode_kecamatan = %s "
                            "ORDER BY tahun DESC LIMIT 1",
                            (kode_kec,),
                        )
                        potensi = cur.fetchone()
            if potensi:
                # potensi_pertanian SENGAJA tidak dicocokkan ke A3 -- Tabel 2
                # dokumen 14072026 tidak punya kategori Pertanian di A3 (mulai
                # dari Perkebunan), beda dengan A1 yang punya Pertanian.
                for field, kategori in (("potensi_perkebunan", "PERKEBUNAN"),
                                         ("potensi_peternakan", "PETERNAKAN"),
                                         ("potensi_perikanan", "PERIKANAN")):
                    if potensi.get(field):
                        kategori_cocok.append(f"A3_{kategori}")

            kandidat = [rule["subs"][k] for k in kategori_cocok if k in rule["subs"]]
            sub_a3 = max(kandidat, key=lambda s: s["nilai"]) if kandidat else None
            if sub_a3:
                nilai += sub_a3["nilai"]
                detail.append(sub_a3["label"])
            else:
                detail.append(
                    "A3: tidak ada kawasan tematik tambahan yang cocok (Perkebunan/Peternakan/"
                    "Perikanan/Transmigrasi/Kawasan Industri Prioritas/PKPN)"
                )
        else:
            detail.append("A3 (tematik tambahan) belum tersedia — menunggu data lokus gdrive")
        keterangan = "; ".join(detail) + "."
    else:
        keterangan = (
            f"{sub['label']} (sumber: {sumber}) — hanya mencakup sub-parameter tematik "
            "utama (A1/A2); tematik tambahan & data dukung (A3/A4) belum tersedia."
        )
    return {"tersedia": True, "nilai": nilai, "keterangan": keterangan}


def _ijd_score_kemantapan(row: dict, rules: dict, ctx: dict = None) -> dict:
    rule = rules.get("B")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah kemantapan belum diset di database."}
    if "pembangunan" in (row.get("jenis_penanganan") or "").lower():
        sub = rule["subs"]["PEMBANGUNAN"]
        return {"tersedia": True, "nilai": sub["nilai"], "keterangan": sub["label"]}

    baik = float(row.get("kondisi_baik_km") or 0)
    sedang = float(row.get("kondisi_sedang_km") or 0)
    ringan = float(row.get("kondisi_ringan_km") or 0)
    berat = float(row.get("kondisi_berat_km") or 0)
    total = baik + sedang + ringan + berat
    if total <= 0:
        return {"tersedia": False, "keterangan": "Data kondisi ruas (baik/sedang/ringan/berat) belum diisi."}

    pct_mantap = (baik + sedang) / total * 100
    sub = rule["subs"]["TIDAK_MANTAP" if pct_mantap < 60 else "MANTAP"]
    return {
        "tersedia": True,
        "nilai": sub["nilai"],
        "keterangan": f"{sub['label']} (kemantapan eksisting {pct_mantap:.1f}%)",
    }


def _ijd_score_koridor(row: dict, rules: dict, ctx: dict = None) -> dict:
    rule = rules.get("D")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah koridor belum diset di database."}
    # Hasil verifikasi Balai (kolom Status Koridor Prioritas Balai, terisi
    # mulai tarikan 15 Juli) diprioritaskan; usulan yang belum dinilai Balai
    # jatuh ke proksi kode_koridor terisi/kosong.
    status_balai = (row.get("status_koridor_balai") or "").strip().upper()
    if status_balai == "SESUAI":
        sub, sumber = rule["subs"]["TERIDENTIFIKASI"], "verifikasi Balai"
    elif status_balai:
        sub = rule["subs"].get("LAINNYA_BALAI") or rule["subs"]["LAINNYA"]
        sumber = "verifikasi Balai"
    elif (row.get("kode_koridor") or "").strip():
        sub, sumber = rule["subs"]["TERIDENTIFIKASI"], "perkiraan dari kode koridor"
    else:
        sub, sumber = rule["subs"]["LAINNYA"], "perkiraan dari kode koridor"
    return {
        "tersedia": True,
        "nilai": sub["nilai"],
        "keterangan": f"{sub['label']} ({sumber}).",
    }


def _ijd_score_rc(row: dict, rules: dict, ctx: dict = None) -> dict:
    rule = rules.get("F")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah RC belum diset di database."}

    nilai_list, detail = [], []
    for label, pemda_col, balai_col in _RC_FIELD_MAP:
        status = row.get(balai_col) or row.get(pemda_col)
        if not status:
            detail.append(f"{label}: belum diisi")
            continue
        sumber = "Balai" if row.get(balai_col) else "Pemda"
        sub_key = f"{label}_{status.strip().upper().replace(' ', '_')}"
        sub = rule["subs"].get(sub_key)
        if not sub:
            detail.append(f"{label}: status '{status}' tidak dikenali kaidah")
            continue
        nilai_list.append(sub["nilai"])
        detail.append(f"{label}: {status} ({sumber}) → {sub['nilai']:.0f}")

    if not nilai_list:
        return {"tersedia": False, "keterangan": "Belum ada status RC (DED/RAB/Dokling/Lahan) yang terisi."}

    nilai = sum(nilai_list) / len(nilai_list)
    return {
        "tersedia": True,
        "nilai": round(nilai, 1),
        "keterangan": "; ".join(detail) + f" — rata-rata dari {len(nilai_list)}/4 komponen RC yang terisi.",
    }


# bucket_ip raster (scripts/import_indeks_penanaman_raster.py, cuma 3 nilai
# krn raster sumbernya cuma 3 kelas tanam) -> sub_kode ijd_scoring_rules.
_A2IP_RASTER_BUCKET_TO_SUB = {
    "100-150": "A2IP_100_150", "150-199": "A2IP_150_200", "GT300": "A2IP_GT300",
}


def _ijd_score_kemanfaatan(row: dict, rules: dict, ctx: dict = None) -> dict:
    """Parameter C kaidah 2026 — sub A1 (kepadatan penduduk kecamatan, bobot
    35%) + A2 produktivitas padi kabupaten (proksi "Produktivitas Ton/Ha",
    12% dari 35% A2) + A2 Indeks Penanaman kabupaten (11% dari 35% A2,
    Kertas Kerja.xlsx -- Luas Lahan 12% masih menunggu data yg skalanya
    cocok, lihat schema_kertas_kerja.sql) + A3 lalu lintas (kepemilikan
    kendaraan per km jalan kabupaten, substitusi LHR yang belum ada sumber
    -- dokumen resmi mengizinkan rasio kabupaten sbg fallback). Nilai rules
    sudah tertimbang.

    A1 primer dari kecamatan_data_turunan (kunci kode_kecamatan). Kalau
    usulan.kode_kecamatan NULL atau kecamatannya tak dikenal/kosong
    datanya, A1 fallback ke PROKSI kabupaten (23 Jul 2026): kepadatan
    rata-rata tertimbang (total penduduk / total luas) seluruh kecamatan
    kabupaten dari bps_kecamatan_demografi -- BUKAN penerapan literal
    dokumen resmi (Table 4 A1 eksplisit per-kecamatan, beda dari A3 yang
    eksplisit mengizinkan rasio kabupaten), selalu ditandai "PROKSI
    kabupaten" di keterangan supaya tidak disalahartikan sbg kepadatan
    kecamatan usulan sebenarnya. A1 cuma benar-benar "belum tersedia"
    kalau kode_kab-nya juga tak ter-resolve (wilayah_mapping tak match)
    ATAU bps_kecamatan_demografi kosong utk kabupaten itu. Ini
    diputuskan 22 Jul 2026 setelah audit menemukan kode_kecamatan NULL utk
    SEMUA usulan nasional (lihat docs/MEMORY.md): memblokir C secara total
    krn 1 sub-parameter yang genuinely butuh geometri terlalu konservatif,
    padahal A2 (bps_kabupaten_padi, bps_kabupaten_indeks_penanaman) & A3
    (bps_kabupaten_kendaraan ÷ bps_kabupaten_jalan) murni level kabupaten
    dan sudah lama pakai fallback kab_by_wilayah/wilayah_mapping yang sama
    dgn A3 tematik di _ijd_score_tematik & NPR (_bappenas_kode_kab) --
    sekarang A2/A3 dibuka pakai fallback yang sama, hanya A1 yang tetap bisa
    "belum tersedia" sendirian. `nilai` jadi jumlah sub yang match (bisa
    kurang dari 3 sub kalau sebagian belum tersedia); `tersedia: False`
    HANYA kalau tidak ada satupun sub yang berhasil dihitung (bukan cuma A1
    kosong)."""
    rule = rules.get("C")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah kemanfaatan belum diset di database."}

    kode_kec = row.get("kode_kecamatan")
    if kode_kec:
        kode_kab = kode_kec // 1000
    elif ctx and "kab_by_wilayah" in ctx:
        kode_kab = ctx["kab_by_wilayah"].get((row.get("provinsi"), row.get("kabupaten_kota")))
    else:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kabupaten FROM wilayah_mapping "
                "WHERE provinsi_sitia = %s AND kabupaten_kota_sitia = %s",
                (row.get("provinsi"), row.get("kabupaten_kota")),
            )
            r = cur.fetchone()
        kode_kab = r["kode_kabupaten"] if r else None

    def _a1_sub_kode(kepadatan):
        if kepadatan > 1000:
            return "A1_GT1000"
        if kepadatan >= 500:
            return "A1_500_1000"
        if kepadatan >= 100:
            return "A1_100_500"
        return "A1_LT100"

    nilai = 0.0
    ada_sub = False
    detail = []

    kec = None
    if kode_kec:
        if ctx and "kepadatan_by_kec" in ctx:
            kec = ctx["kepadatan_by_kec"].get(kode_kec)
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT kecamatan, kepadatan_per_km2 FROM kecamatan_data_turunan "
                    "WHERE kode_kecamatan = %s ORDER BY tahun DESC LIMIT 1",
                    (kode_kec,),
                )
                kec = cur.fetchone()

    if kec and kec["kepadatan_per_km2"] is not None:
        kepadatan = float(kec["kepadatan_per_km2"])
        sub = rule["subs"][_a1_sub_kode(kepadatan)]
        nilai += sub["nilai"]
        ada_sub = True
        detail.append(f"Kec. {kec['kecamatan']}: {sub['label']} (kepadatan {kepadatan:,.0f} jiwa/km2)")
    else:
        # Fallback PROKSI kabupaten (23 Jul 2026) -- kepadatan rata-rata
        # tertimbang (total penduduk / total luas) seluruh kecamatan
        # kabupaten dari bps_kecamatan_demografi, dipakai HANYA kalau
        # kode_kecamatan usulan tidak diketahui atau kecamatannya tidak
        # dikenal/kosong datanya. Ini BUKAN penerapan literal dokumen resmi
        # (Table 4 A1 eksplisit per-kecamatan, beda dari A3 yang eksplisit
        # mengizinkan rasio kabupaten) -- ditandai jelas "proksi kabupaten"
        # di keterangan supaya tidak disalahpahami sbg kepadatan kecamatan
        # usulan yang sebenarnya.
        if not kode_kec:
            alasan_kec = "usulan belum dihubungkan ke kecamatan (kolom kode_kecamatan — interim manual, spatial-join menunggu SHP batas kecamatan)"
        elif not kec:
            alasan_kec = f"kode kecamatan {kode_kec} tidak dikenal di master"
        else:
            alasan_kec = f"kepadatan penduduk Kec. {kec['kecamatan']} belum tersedia (buku Dalam Angka provinsi ybs. belum diimpor)"

        kepadatan_kab = (ctx["kepadatan_kab_by_kab"].get(kode_kab) if ctx and "kepadatan_kab_by_kab" in ctx
                          else None)
        if kepadatan_kab is None and not (ctx and "kepadatan_kab_by_kab" in ctx) and kode_kab:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT SUM(jumlah_penduduk) AS total_penduduk, SUM(luas_km2_derived) AS total_luas "
                    "FROM bps_kecamatan_demografi WHERE kode_kab = %s AND jumlah_penduduk IS NOT NULL "
                    "AND luas_km2_derived IS NOT NULL AND luas_km2_derived > 0",
                    (f"{kode_kab:04d}",),
                )
                r = cur.fetchone()
                if r and r["total_luas"]:
                    kepadatan_kab = float(r["total_penduduk"]) / float(r["total_luas"])

        if kepadatan_kab is not None:
            sub = rule["subs"][_a1_sub_kode(kepadatan_kab)]
            nilai += sub["nilai"]
            ada_sub = True
            detail.append(
                f"A1 ({alasan_kec}): {sub['label']} — PROKSI kabupaten "
                f"(kepadatan rata-rata kab. {kepadatan_kab:,.0f} jiwa/km2, bukan kecamatan usulan sebenarnya)"
            )
        else:
            detail.append(f"A1: {alasan_kec}, dan kepadatan rata-rata kabupaten juga tidak tersedia")

    if not kode_kab:
        detail.append("A2: kabupaten usulan tidak diketahui (wilayah_mapping tidak match, lihat build_wilayah_mapping.py)")
    elif any(k.startswith("A2_") for k in rule["subs"]):
        padi = (ctx["produktivitas_padi_by_kab"].get(kode_kab) if ctx and "produktivitas_padi_by_kab" in ctx
                else None)
        if padi is None and not (ctx and "produktivitas_padi_by_kab" in ctx):
            with db_cursor() as cur:
                cur.execute(
                    "SELECT nama_kab, produktivitas_ku_ha FROM bps_kabupaten_padi "
                    "WHERE kode_kab = %s ORDER BY tahun DESC LIMIT 1",
                    (f"{kode_kab:04d}",),
                )
                padi = cur.fetchone()
        if padi and padi.get("produktivitas_ku_ha") is not None:
            ku_ha = float(padi["produktivitas_ku_ha"])
            if ku_ha > 60:
                sub_a2 = rule["subs"]["A2_GT6"]
            elif ku_ha >= 50:
                sub_a2 = rule["subs"]["A2_5_6"]
            elif ku_ha >= 40:
                sub_a2 = rule["subs"]["A2_4_5"]
            elif ku_ha >= 30:
                sub_a2 = rule["subs"]["A2_3_4"]
            else:
                sub_a2 = rule["subs"]["A2_LT3"]
            nilai += sub_a2["nilai"]
            ada_sub = True
            detail.append(f"{sub_a2['label']} (produktivitas padi kab. {ku_ha/10:.1f} ton/ha, proksi kabupaten)")
        else:
            detail.append("A2: produktivitas padi kabupaten belum tersedia (buku Dalam Angka belum diimpor)")
    else:
        detail.append("Produktivitas (A2) belum tersedia.")

    if not kode_kab:
        detail.append("A2 IP: kabupaten usulan tidak diketahui (wilayah_mapping tidak match)")
    elif any(k.startswith("A2IP_") for k in rule["subs"]):
        # Sumber PRIMER (21 Jul 2026): raster resmi Dit. SDA (Maps/IP2019-2024,
        # zonal statistics per kabupaten, scripts/import_indeks_penanaman_raster.py)
        # -- diutamakan drpd sumber SEKUNDER (Kertas Kerja.xlsx, "DATA SEKUNDER"
        # per label sumbernya sendiri). Raster cuma py 3 kelas kasar (bucket_ip
        # 100-150/150-199/GT300, lihat KELAS_TO_BUCKET di importer), fallback
        # ke Kertas Kerja (skala % kontinu, 5 bucket) kalau kabupaten ybs tak
        # tercakup raster.
        raster_bucket = (ctx["ip_raster_by_kab"].get(kode_kab) if ctx and "ip_raster_by_kab" in ctx
                          else None)
        if raster_bucket is None and not (ctx and "ip_raster_by_kab" in ctx):
            with db_cursor() as cur:
                cur.execute(
                    "SELECT bucket_ip FROM bps_kabupaten_indeks_penanaman_raster "
                    "WHERE kode_kab = %s ORDER BY tahun DESC LIMIT 1",
                    (f"{kode_kab:04d}",),
                )
                r = cur.fetchone()
                raster_bucket = r["bucket_ip"] if r else None
        sub_ip = None
        if raster_bucket:
            sub_ip = rule["subs"].get(_A2IP_RASTER_BUCKET_TO_SUB.get(raster_bucket, ""))
            if sub_ip:
                detail.append(f"{sub_ip['label']} (raster resmi Dit. SDA, kelas tanam kabupaten dominan)")

        if not sub_ip:
            ip = (ctx["indeks_penanaman_by_kab"].get(kode_kab) if ctx and "indeks_penanaman_by_kab" in ctx
                  else None)
            if ip is None and not (ctx and "indeks_penanaman_by_kab" in ctx):
                with db_cursor() as cur:
                    cur.execute(
                        "SELECT indeks_penanaman_pct FROM bps_kabupaten_indeks_penanaman "
                        "WHERE kode_kab = %s ORDER BY tahun DESC LIMIT 1",
                        (f"{kode_kab:04d}",),
                    )
                    r = cur.fetchone()
                    ip = float(r["indeks_penanaman_pct"]) if r and r["indeks_penanaman_pct"] is not None else None
            if ip is not None:
                if ip > 300:
                    sub_ip = rule["subs"]["A2IP_GT300"]
                elif ip >= 200:
                    sub_ip = rule["subs"]["A2IP_200_300"]
                elif ip >= 150:
                    sub_ip = rule["subs"]["A2IP_150_200"]
                elif ip >= 100:
                    sub_ip = rule["subs"]["A2IP_100_150"]
                else:
                    sub_ip = rule["subs"]["A2IP_LT100"]
                detail.append(f"{sub_ip['label']} (indeks penanaman kab. {ip:.0f}%, Kertas Kerja.xlsx — sumber sekunder)")

        if sub_ip:
            nilai += sub_ip["nilai"]
            ada_sub = True
        else:
            detail.append("A2: indeks penanaman kabupaten belum tersedia (raster & Kertas Kerja.xlsx tak mencakup kab. ini)")
        detail.append("Luas Lahan (sub A2) belum tersedia.")
    else:
        detail.append("Indeks Penanaman (A2) belum tersedia.")

    if not kode_kab:
        detail.append("A3: kabupaten usulan tidak diketahui (wilayah_mapping tidak match)")
    elif any(k.startswith("A3_") for k in rule["subs"]):
        rasio = (ctx["kendaraan_per_km_by_kab"].get(kode_kab) if ctx and "kendaraan_per_km_by_kab" in ctx
                 else None)
        if rasio is None and not (ctx and "kendaraan_per_km_by_kab" in ctx):
            with db_cursor() as cur:
                cur.execute(
                    "SELECT k.jumlah / j.panjang_total_km AS per_km FROM bps_kabupaten_kendaraan k "
                    "JOIN bps_kabupaten_jalan j ON j.kode_kab = k.kode_kab "
                    "WHERE k.kode_kab = %s AND j.panjang_total_km > 0 AND k.jumlah IS NOT NULL "
                    "ORDER BY k.tahun DESC LIMIT 1",
                    (f"{kode_kab:04d}",),
                )
                r = cur.fetchone()
                rasio = float(r["per_km"]) if r else None
        if rasio is not None:
            if rasio > 1000:
                sub_a3 = rule["subs"]["A3_GT1000"]
            elif rasio >= 600:
                sub_a3 = rule["subs"]["A3_600_1000"]
            elif rasio >= 300:
                sub_a3 = rule["subs"]["A3_300_600"]
            elif rasio >= 100:
                sub_a3 = rule["subs"]["A3_100_300"]
            else:
                sub_a3 = rule["subs"]["A3_LT100"]
            nilai += sub_a3["nilai"]
            ada_sub = True
            detail.append(f"{sub_a3['label']} (kepemilikan kendaraan kab. {rasio:,.0f}/km, proksi kabupaten)")
        else:
            detail.append("A3: rasio kendaraan/km kabupaten belum tersedia (data kendaraan atau panjang jalan belum lengkap)")
    else:
        detail.append("Lalu lintas (A3) belum tersedia.")

    if not ada_sub:
        return {"tersedia": False, "keterangan": "; ".join(detail) + "."}
    return {"tersedia": True, "nilai": nilai, "keterangan": "; ".join(detail) + "."}


def _ijd_score_penuntasan(row: dict, rules: dict, ctx: dict = None) -> dict:
    """Parameter E kaidah 2026: lanjutan/penuntasan IJD TA 2025 vs usulan baru.
    Flag usulan_inpres.lanjutan_ijd_2025 diisi scripts/import_dpp_ijd_2025.py
    (pencocokan nama ruas + wilayah terhadap tabel dpp_ijd_2025); NULL berarti
    pencocokan belum pernah dijalankan."""
    rule = rules.get("E")
    if not rule:
        return {"tersedia": False, "keterangan": "Kaidah penuntasan belum diset di database."}
    # Flag resmi hasil verifikasi kompetensi (kolom Penuntasan IJD Sebelumnya,
    # terisi mulai tarikan 15 Juli) diprioritaskan di atas hasil pencocokan
    # nama ruas kita sendiri.
    if (row.get("penuntasan_ijd_kompetensi") or "").strip().upper() == "YA":
        sub = rule["subs"]["LANJUTAN"]
        return {
            "tersedia": True,
            "nilai": sub["nilai"],
            "keterangan": sub["label"] + " (flag resmi verifikasi kompetensi SITIA).",
        }
    flag = row.get("lanjutan_ijd_2025")
    if flag is None:
        return {
            "tersedia": False,
            "keterangan": "Pencocokan DPP IJD 2025 belum dijalankan (scripts/import_dpp_ijd_2025.py).",
        }
    sub = rule["subs"]["LANJUTAN" if flag else "BARU"]
    return {
        "tersedia": True,
        "nilai": sub["nilai"],
        "keterangan": sub["label"] + " (pencocokan nama ruas + wilayah terhadap DPP IJD TA 2025, bukan penetapan resmi).",
    }


_IJD_SCORERS = {
    "A": _ijd_score_tematik, "B": _ijd_score_kemantapan, "C": _ijd_score_kemanfaatan,
    "D": _ijd_score_koridor, "E": _ijd_score_penuntasan, "F": _ijd_score_rc,
}


def _compute_ijd_score(row: dict, tahun: int = 2026, rules: dict = None, ctx: dict = None) -> dict:
    if rules is None:
        rules = _load_ijd_rules(tahun)

    komponen = []
    skor_tertimbang = 0.0
    bobot_tersedia = 0.0

    # Daftar parameter mengikuti kaidah tahun terpilih: gabungan kode yang
    # punya rules di DB dan yang masih pending — F ada di kaidah 2025 tapi
    # dihapus dari penilaian 2026, jadi tidak boleh muncul (bahkan sebagai
    # "belum tersedia") saat tahun=2026.
    for kode in sorted(set(rules) | set(IJD_PENDING_PARAMETERS)):
        if kode in rules and kode in _IJD_SCORERS:
            rule = rules.get(kode, {})
            hasil = _IJD_SCORERS[kode](row, rules, ctx)
            bobot_maks = rule.get("bobot_maks", 0)
            entry = {
                "kode": kode,
                "label": rule.get("label", kode),
                "bobot_maks": bobot_maks,
                "tersedia": hasil["tersedia"],
                "keterangan": hasil["keterangan"],
            }
            if hasil["tersedia"]:
                kontribusi = hasil["nilai"] / 100 * bobot_maks
                entry["nilai"] = hasil["nilai"]
                entry["kontribusi"] = round(kontribusi, 2)
                skor_tertimbang += kontribusi
                bobot_tersedia += bobot_maks
            komponen.append(entry)
        else:
            pending = IJD_PENDING_PARAMETERS.get(kode, {})
            bobot_maks = rules.get(kode, {}).get("bobot_maks") or \
                pending.get("bobot_maks_per_tahun", {}).get(tahun, 0)
            komponen.append({
                "kode": kode,
                "label": pending.get("parameter_label") or rules.get(kode, {}).get("label", kode),
                "bobot_maks": bobot_maks,
                "tersedia": False,
                "keterangan": pending.get("alasan", "Parameter belum diimplementasikan."),
            })

    skor_ternormalisasi = round(skor_tertimbang / bobot_tersedia * 100, 1) if bobot_tersedia else None

    return {
        "tahun_berlaku": tahun,
        "komponen": komponen,
        "skor_tertimbang": round(skor_tertimbang, 2),
        "bobot_tersedia": bobot_tersedia,
        "skor_ternormalisasi_100": skor_ternormalisasi,
        "catatan": (
            "Skor 'Prioritisasi Teknokratik' perkiraan berdasarkan Tabel 1 dokumen Penentuan Parameter "
            f"Penilaian IJD No. 11/2025 (kaidah tahun {tahun}). skor_tertimbang dihitung hanya dari parameter yang datanya "
            f"tersedia (total bobot tersedia: {bobot_tersedia:.0f} dari 100); skor_ternormalisasi_100 "
            "menyetarakannya ke skala 0-100 seolah hanya parameter tersedia yang dinilai. Bukan skor "
            "resmi Bina Marga/Bappenas."
        ),
    }


def _ijd_ranking_sort_key(npr_skor: dict) -> tuple:
    """Kunci sort bersama utk ranking lintas-usulan (No. urutan nasional &
    peringkat per provinsi di _ijd_score_bulk_rows). SEJAK 22 Jul 2026 (request
    eksplisit user) basis ranking DIGANTI dari skor teknokratis A-E ke skor NPR
    (Nilai Prioritas Ruas -- metodologi ALTERNATIF/eksperimental, BELUM policy
    resmi, lihat _compute_npr) -- BUKAN lagi skor_ternormalisasi_100 dari
    _compute_ijd_score. Usulan yang NPR-nya belum bisa dihitung (SI/SC sama
    sekali tak tersedia) ditaruh paling bawah; dalam grup yang punya NPR, nilai
    NPR tertinggi dulu (sama pola sort dgn _npr_bulk_rows). skor_ternormalisasi_100
    (teknokratis A-E) tetap disertakan apa adanya sbg kolom export terpisah,
    jadi basis ranking lama masih bisa dibaca user, cuma tidak lagi dipakai
    utk mengurutkan baris/peringkat provinsi."""
    npr = npr_skor["npr"]
    return (npr is None, -(npr or 0))


@app.get("/api/usulan-inpres/{usulan_id}/ijd-score")
def usulan_inpres_ijd_score(usulan_id: int, tahun: int = 2026):
    with db_cursor() as cur:
        cur.execute("SELECT * FROM usulan_inpres WHERE id = %s", (usulan_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Usulan tidak ditemukan")
    return _compute_ijd_score(row, tahun)


IJD_EXPORT_IDENTITAS_COLS = [
    ("Nama Pengusul", "nama_pengusul"), ("Provinsi", "provinsi"),
    ("Kabupaten/Kota", "kabupaten_kota"), ("Nama Kegiatan", "nama_kegiatan"),
    ("Kode Koridor", "kode_koridor"), ("Nama Koridor", "nama_koridor"),
    ("Panjang Koridor", "panjang_koridor_km"), ("Kode Ruas", "kode_ruas"),
    ("Nama Ruas", "nama_ruas"), ("Panjang Ruas (KM)", "panjang_ruas_km"),
    ("Status Ruas", "status_ruas"), ("No. Jembatan", "no_jembatan"),
    ("Nama Jembatan", "nama_jembatan"), ("Panjang Jembatan (M)", "panjang_jembatan_m"),
    ("Komponen (Kompetensi)", "komponen_kompetensi"),
    ("Panjang Penanganan (Kompetensi)", "panjang_penanganan_kompetensi"),
    ("Satuan", "satuan"), ("Alokasi Usulan (Kompetensi)", "alokasi_usulan_kompetensi"),
    ("Tematik Kawasan (Kompetensi)", "tematik_kawasan_kompetensi"),
    ("Keterangan", "keterangan"), ("Seleksi Sistem (dipake yang lulus)", "seleksi_sistem"),
    ("Verifikasi Kompetensi", "verifikasi_kompetensi"),
    ("Catatan Pembahasan Kompetensi", "catatan_pembahasan_kompetensi"),
    ("Lebar Jalan (M)", "lebar_jalan_m"),
    ("Kondisi Ruas Jalan Baik (KM)", "kondisi_baik_km"),
    ("Kondisi Ruas Jalan Sedang (KM)", "kondisi_sedang_km"),
    ("Kondisi Ruas Jalan Ringan (KM)", "kondisi_ringan_km"),
    ("Kondisi Ruas Jalan Berat (KM)", "kondisi_berat_km"),
    ("Kondisi Jembatan", "kondisi_jembatan"), ("Kapasitas Fiskal", "kapasitas_fiskal"),
]
# Header lengkap (multi-baris, dipakai xlsx -- ikuti persis wording template
# Bappenas) vs pendek (satu baris, dipakai preview JSON di UI -- tabel data-
# viewer generik tidak dirancang utk header berparagraf).
IJD_EXPORT_BAPPENAS_HEADERS = [
    "ASPEK PRIORITAS DAN NILAI STRATEGIS\n"
    "(Menilai sejauh mana usulan berada di lokasi yang menjadi prioritas nasional "
    "dan membutuhkan afirmasi khusus dari pemerintah pusat.)\n\n"
    "Data referensi penilaian adalah di sheet Kumpulan Data, Row Penilaian Lokpri.",
    "TOTAL CHECKLIST ASPEK A\n"
    "(Jumlah kriteria yang tercentang [v] pada kolom sebelumnya -- angka mentah, "
    "BEDA dari \"Poin\" yang ditampilkan di dalam kolom itu sendiri, yang dikelompokkan "
    "3 kelas 0/1/2 dgn ambang 0 / 1-2 / >=3 kriteria.)",
    "DAYA UNGKIT EKONOMI & KINERJA SEKTORAL\n"
    "(Menilai kontribusi ruas terhadap ketahanan pangan nasional, kelancaran logistik, "
    "dan stimulasi pertumbuhan ekonomi lokal).\n\n"
    "Data referensi penilaian adalah di sheet Kumpulan Data, Row Penilaian Ruas.",
    "DAYA UNGKIT EKONOMI & KINERJA SEKTORAL - NARASI AI\n"
    "(Pelengkap kolom sebelumnya, narasi naratif hasil AI. Kosong bila usulan belum pernah "
    "digenerate lewat tombol \"Proses Narasi AI\" (bulk per provinsi) di preview atau fitur "
    "\"Draf Penilaian Bappenas (AI)\" per-usulan.)",
    "RANGKING DALAM PROVINSI\n"
    "(Urutan prioritas ruas dalam masing-masing provinsi merupakan kesimpulan integratif "
    "dari 2 aspek yaitu nilai strategis dan dampak ekonomi dan sektoral)",
    "KESIMPULAN PENILAIAN BAPPENAS",
    "TOTAL",
]
IJD_EXPORT_BAPPENAS_HEADERS_SHORT = [
    "Aspek A: Prioritas & Nilai Strategis",
    "Total Checklist Aspek A",
    "Aspek B: Daya Ungkit Ekonomi & Sektoral",
    "Aspek B: Narasi AI",
    "Ranking Bappenas (per Provinsi)",
    "Kesimpulan Penilaian Bappenas",
    "Total (Bappenas)",
]
IJD_EXPORT_TEKNOKRATIS_KODE = ["A", "B", "C", "D", "E"]
IJD_EXPORT_TEKNOKRATIS_HEADERS = {
    "A": "TEMATIK & DATA DUKUNG (A)", "B": "KONDISI KEMANTAPAN EKSISTING RUAS (B)",
    "C": "KEMANFAATAN (C)", "D": "KORIDOR (D)", "E": "KEBERLANJUTAN KEGIATAN INPRES SEBELUMNYA (E)",
}
IJD_EXPORT_RANKING_LABEL = "PENILAIAN PRIORITASI USULAN PER PROVINSI"
# SEJAK 22 Jul 2026 (request eksplisit user) kolom ranking ini diurutkan
# turun berdasar SKOR NPR (Nilai Prioritas Ruas, eksperimental/belum policy
# resmi -- lihat IJD_EXPORT_NPR_LABEL & _ijd_ranking_sort_key), BUKAN lagi
# skor_ternormalisasi_100 teknokratis A-E spt sebelumnya. Skor teknokratis
# tetap dihitung & ditampilkan apa adanya lewat IJD_EXPORT_SKOR_TEKNOKRATIS_100_LABEL
# (kolom terpisah), cuma tidak lagi jadi basis urutan baris/peringkat provinsi.
# Sebelum perubahan ini, ranking dibasiskan skor_ternormalisasi_100 dgn
# prioritas kelompok "data lengkap (bobot_tersedia=100) dulu" per keputusan
# 21 Jul 2026 (lihat docs/verifikasi_ijd_ciparay_cikumpay.md) -- aturan
# kelompok itu TIDAK dipertahankan di basis NPR krn _compute_npr sendiri
# sudah menormalisasi SI/SC masing2 thd bobot parameter yang tersedia
# (pola sama "belum tersedia, bukan 0"), jadi kolom Kelengkapan Data di
# bawah tetap merujuk kelengkapan skor TEKNOKRATIS, bukan NPR.
IJD_EXPORT_KELENGKAPAN_LABEL = "Kelengkapan Data Skor Teknokratis"
# Skor Prioritas Nasional (_skor_prioritas_nasional -- 70% Teknokratis + 10%
# Kementerian PU + 10% Bappenas + 10% Kemenko Infra, dokumen 14072026 bagian
# C) BEDA dari IJD_EXPORT_RANKING_LABEL di atas (itu ranking teknokratis A-E
# SAJA per provinsi) -- kolom ini rangkuman skor+peringkat NASIONAL lintas
# provinsi memakai formula gabungan 4-komponen itu. Ditambahkan 21 Jul 2026.
IJD_EXPORT_PRIORITAS_NASIONAL_LABEL = "PENILAIAN PRIORITASI USULAN NASIONAL"
# Kolom angka murah-sortir terpisah dari IJD_EXPORT_PRIORITAS_NASIONAL_LABEL
# (yang isinya teks deskriptif skor+peringkat+formula) -- pola sama dgn
# IJD_EXPORT_RANKING_LABEL (angka polos, bukan teks) supaya user bisa
# sort/filter Excel langsung tanpa parse teks. Ditambahkan 21 Jul 2026.
IJD_EXPORT_RANKING_NASIONAL_LABEL = "RANKING NASIONAL"
# Skor teknokratis A-E ternormalisasi (skor_ternormalisasi_100 dari
# _compute_ijd_score) sbg kolom ANGKA tersendiri -- sebelumnya cuma tersirat
# lewat kontribusi tiap komponen A-E, tidak ada satu kolom totalnya. Ditambah
# 22 Jul 2026 supaya "basis ranking lama" tetap kebaca stlh ranking
# (IJD_EXPORT_RANKING_LABEL) dipindah ke basis NPR (lihat _ijd_ranking_sort_key).
IJD_EXPORT_SKOR_TEKNOKRATIS_100_LABEL = "Skor Teknokratis A-E (Ternormalisasi 0-100)"
# NPR (Nilai Prioritas Ruas, _compute_npr) digabung ke export ini per request
# eksplisit user 22 Jul 2026 -- metodologi ALTERNATIF/eksperimental, BELUM
# policy resmi (lihat docs/kajian_metodologi_skala_prioritas_ruas.md), TETAP
# dilabeli begitu di header supaya tidak disalahartikan sbg skor resmi
# Bina Marga/Bappenas. Rincian 27-kolom per-komponen SI/SC tetap HANYA di
# export NPR sendiri (/api/usulan-inpres/npr/export/xlsx) -- di sini cuma
# total NPR + kategori, sesuai permintaan ("masukkan skor NPR").
IJD_EXPORT_NPR_LABEL = "Skor NPR (Eksperimental, Belum Policy Resmi)"
IJD_EXPORT_NPR_KATEGORI_LABEL = "Kategori NPR (Eksperimental)"

# Ambang heuristik "tidak masuk akal" per kolom produksi kecamatan
# (bps_kecamatan_potensi_tematik) -- BUKAN batas resmi apa pun, cuma jauh di
# atas produksi realistis 1 kecamatan (lihat docs/verifikasi_npr_ciparay_cikumpay.md
# §3 "Temuan Data Quality": Sanggau perkebunan_produksi_ton sampai 30,5 juta
# ton/kecamatan, Lombok Utara peternakan_produksi_telur_kg sampai 2,2 miliar
# kg, Sumba Timur pertanian_produksi_ton sampai 200 juta ton -- smua salah
# parse skala/kolom PDF BPS di extract_dalam_angka.py, BELUM diperbaiki per
# 21 Jul 2026). Dipakai utk nandai per-baris di export skor IJD supaya user
# tahu usulan mana yang kecamatan/kabupatennya bersumber dari data produksi
# yang diduga rusak -- bukan cuma 3 kabupaten yang sudah dikonfirmasi di
# atas, krn re-validasi menemukan kasus baru tiap kali dicek ulang.
IJD_OUTLIER_PRODUKSI_AMBANG = {
    "pertanian_produksi_ton": ("Pertanian (Padi/Jagung)", 1_000_000),
    "perkebunan_produksi_ton": ("Perkebunan", 1_000_000),
    "peternakan_produksi_daging_kg": ("Peternakan (daging)", 10_000_000),
    "peternakan_produksi_telur_kg": ("Peternakan (telur)", 10_000_000),
    "perikanan_produksi_ton": ("Perikanan", 1_000_000),
}
IJD_EXPORT_OUTLIER_LABEL = "Temuan Data Quality — Outlier Produksi Kecamatan"


def _outlier_produksi_keterangan(produksi: dict) -> str:
    """Bandingkan tiap kolom produksi kecamatan (bps_kecamatan_potensi_tematik)
    thd IJD_OUTLIER_PRODUKSI_AMBANG. Return keterangan temuan kalau ada yang
    kelewat ambang, atau None kalau tidak ada indikasi (termasuk kalau
    kecamatan ini tidak punya baris data sama sekali)."""
    if not produksi:
        return None
    temuan = []
    for kolom, (label, ambang) in IJD_OUTLIER_PRODUKSI_AMBANG.items():
        nilai = produksi.get(kolom)
        if nilai is not None and nilai > ambang:
            temuan.append(f"{label}: {nilai:,.0f} (ambang wajar ~{ambang:,.0f})".replace(",", "."))
    if not temuan:
        return None
    return ("⚠ Produksi kecamatan tidak masuk akal, kemungkinan salah parse skala/kolom PDF BPS "
            "(extract_dalam_angka.py, belum diperbaiki) — " + "; ".join(temuan))

# Cache hasil komputasi bulk (kunci: provinsi+tahun) -- skoring ±3.000 usulan
# makan waktu beberapa detik (query batch + loop scorer), jadi TIDAK dihitung
# ulang tiap kali user membuka/navigasi halaman preview. Di-invalidasi manual
# di titik-titik yang mengubah datanya (draf AI Bappenas baru, import lokus
# Aspek A) -- restart server juga membersihkannya (sama pola dgn cache Maps
# overlay di _map_layer_geojson_cache).
_ijd_bulk_cache: dict = {}


def _normalisasi_provinsi_multi(provinsi) -> tuple:
    """str tunggal (kompatibilitas lama) / list / tuple nama provinsi ->
    tuple TERURUT tanpa string kosong (urutan disamakan supaya cache key
    _ijd_bulk_cache konsisten terlepas urutan pilihan user di UI). Tuple
    kosong = nasional (tanpa filter)."""
    if isinstance(provinsi, str):
        provinsi = [provinsi] if provinsi else []
    return tuple(sorted({p for p in provinsi if p}))


def _ijd_score_bulk_rows(provinsi, tahun: int):
    """Skoring IJD massal (opsional difilter per provinsi, BISA lebih dari
    satu -- provinsi: tuple/list nama provinsi, tuple/list kosong = nasional,
    lihat _normalisasi_provinsi_multi): identitas usulan + penilaian
    Bappenas (Aspek A/B rule-based) + penilaian Teknokratis (A-E) + ranking,
    satu baris per usulan. Dipakai bareng oleh endpoint preview JSON (paged)
    dan export xlsx supaya logikanya cuma ditulis sekali.

    _ijd_score_tematik (A3) dan _ijd_score_kemanfaatan (C) biasanya query DB
    per-usulan — untuk cakupan nasional (±3.000 usulan) itu berarti ribuan
    koneksi kecil. Di sini semua lookup dibatch di muka (kepadatan per
    kecamatan, kawasan tematik per kabupaten, fallback kode kabupaten dari
    wilayah_mapping) lalu dioper lewat `ctx` supaya scorer tidak query per baris.

    Return (header_row_full, header_row_short, data_rows) — data_rows berupa
    list of list, urutan kolom SAMA dgn header_row_*.
    """
    provinsi = _normalisasi_provinsi_multi(provinsi)
    key = (provinsi, tahun)
    if key in _ijd_bulk_cache:
        return _ijd_bulk_cache[key]

    with db_cursor() as cur:
        if provinsi:
            cur.execute("SELECT * FROM usulan_inpres WHERE provinsi = ANY(%s)", (list(provinsi),))
        else:
            cur.execute("SELECT * FROM usulan_inpres")
        rows = cur.fetchall()
    if not rows:
        raise HTTPException(404, "Tidak ada usulan yang cocok filter provinsi tsb.")

    rules = _load_ijd_rules(tahun)

    with db_cursor() as cur:
        cur.execute("SELECT provinsi_sitia, kabupaten_kota_sitia, kode_kabupaten FROM wilayah_mapping")
        kab_by_wilayah = {(r["provinsi_sitia"], r["kabupaten_kota_sitia"]): r["kode_kabupaten"]
                           for r in cur.fetchall()}

    kode_kab_set = set()
    for r in rows:
        if r.get("kode_kecamatan"):
            kode_kab_set.add(r["kode_kecamatan"] // 1000)
        else:
            kab = kab_by_wilayah.get((r.get("provinsi"), r.get("kabupaten_kota")))
            if kab:
                kode_kab_set.add(kab)
    # String version -- sejumlah tabel bps_kabupaten_*/bps_kecamatan_demografi
    # menyimpan kode_kab sbg CHAR(4) (lihat scripts/schema_bps_kemanfaatan.sql
    # dkk.), BEDA dari kode_kabupaten INT di tabel lain (kawasan_tematik,
    # bappenas_lokus_a, dst.) -- MySQL membiarkan int vs varchar dibandingkan
    # diam-diam, PostgreSQL tidak (butuh tipe persis sama utk operator "=").
    kode_kab_str = [str(k) for k in kode_kab_set]
    kawasan_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kabupaten, kode_kecamatan, kategori FROM kawasan_tematik "
                "WHERE kode_kabupaten = ANY(%s)", (list(kode_kab_set),),
            )
            for r in cur.fetchall():
                kawasan_by_kab.setdefault(r["kode_kabupaten"], []).append(r)

    kode_kec_set = {r["kode_kecamatan"] for r in rows if r.get("kode_kecamatan")}
    kepadatan_by_kec = {}
    if kode_kec_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kecamatan, kecamatan, kepadatan_per_km2, jumlah_penduduk, "
                "kendaraan_total, potensi_pertanian, potensi_perkebunan, potensi_peternakan, "
                "potensi_perikanan FROM kecamatan_data_turunan WHERE kode_kecamatan = ANY(%s) ORDER BY tahun DESC",
                (list(kode_kec_set),),
            )
            for r in cur.fetchall():
                kepadatan_by_kec.setdefault(r["kode_kecamatan"], r)  # tahun terbaru menang (ORDER BY di atas)

    # Angka produksi riil (bukan sekadar flag ada/tidak) utk narasi Aspek B --
    # sumbernya bps_kecamatan_potensi_tematik (kolom *_produksi_*), BEDA dari
    # kepadatan_by_kec di atas (yang cuma simpan boolean potensi_* dari
    # kecamatan_data_turunan). Lihat _bappenas_aspek_b_ekonomi.
    potensi_produksi_by_kec = {}
    if kode_kec_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kecamatan, pertanian_produksi_ton, perkebunan_produksi_ton, "
                "peternakan_produksi_daging_kg, peternakan_produksi_telur_kg, perikanan_produksi_ton "
                "FROM bps_kecamatan_potensi_tematik WHERE kode_kecamatan = ANY(%s) ORDER BY tahun DESC",
                (list(kode_kec_set),),
            )
            for r in cur.fetchall():
                potensi_produksi_by_kec.setdefault(r["kode_kecamatan"], r)  # tahun terbaru menang

    # Batch utk Aspek A/B Bappenas (rule-based, lihat _bappenas_aspek_a_lokus/
    # _bappenas_aspek_b_ekonomi) — sama alasan: hindari query per baris.
    bappenas_lokus_by_kab, bappenas_lokus_by_prov = {}, {}
    kode_prov_set = {k // 100 for k in kode_kab_set}
    if kode_kab_set or kode_prov_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kriteria, level, kode_provinsi, kode_kabupaten, kode_kecamatan "
                "FROM bappenas_lokus_a WHERE kode_kabupaten = ANY(%s) OR kode_provinsi = ANY(%s)",
                (list(kode_kab_set) or [0], list(kode_prov_set) or [0]),
            )
            for r in cur.fetchall():
                if r["level"] == "PROVINSI" and r["kode_provinsi"]:
                    bappenas_lokus_by_prov.setdefault(r["kode_provinsi"], []).append(r["kriteria"])
                elif r["kode_kabupaten"]:
                    bappenas_lokus_by_kab.setdefault(r["kode_kabupaten"], []).append(r)
    kemantapan_kab_set = set()
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT DISTINCT kode_wilayah FROM kemantapan_ijd_2026 WHERE kode_wilayah = ANY(%s) "
                "AND jenis_adm IN ('Kab.','Kota')",
                (list(kode_kab_set),),
            )
            kemantapan_kab_set = {r["kode_wilayah"] for r in cur.fetchall()}

    # Produktivitas padi kabupaten (C.A2, proksi "Produktivitas Ton/Ha" Tabel 4
    # — hanya komoditas padi & level kabupaten, lihat _ijd_score_kemanfaatan).
    produktivitas_padi_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kab, nama_kab, produktivitas_ku_ha FROM bps_kabupaten_padi "
                "WHERE kode_kab = ANY(%s) ORDER BY tahun DESC",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                produktivitas_padi_by_kab.setdefault(int(r["kode_kab"]), r)  # tahun terbaru menang

    # Kepadatan kabupaten (C.A1 fallback PROKSI saat kode_kecamatan NULL —
    # rata-rata tertimbang penduduk/luas seluruh kecamatan kabupaten, BUKAN
    # kepadatan kecamatan usulan sebenarnya; lihat _ijd_score_kemanfaatan).
    kepadatan_kab_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kab, SUM(jumlah_penduduk) AS total_penduduk, "
                "SUM(luas_km2_derived) AS total_luas FROM bps_kecamatan_demografi "
                "WHERE kode_kab = ANY(%s) AND jumlah_penduduk IS NOT NULL "
                "AND luas_km2_derived IS NOT NULL AND luas_km2_derived > 0 "
                "GROUP BY kode_kab",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                if r["total_luas"]:
                    kepadatan_kab_by_kab[int(r["kode_kab"])] = float(r["total_penduduk"]) / float(r["total_luas"])

    # Kepemilikan kendaraan per km jalan kabupaten (C.A3, substitusi LHR —
    # lihat _ijd_score_kemanfaatan).
    kendaraan_per_km_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT k.kode_kab, k.jumlah / j.panjang_total_km AS per_km FROM bps_kabupaten_kendaraan k "
                "JOIN bps_kabupaten_jalan j ON j.kode_kab = k.kode_kab "
                "WHERE k.kode_kab = ANY(%s) AND j.panjang_total_km > 0 AND k.jumlah IS NOT NULL "
                "ORDER BY k.tahun DESC",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                kendaraan_per_km_by_kab.setdefault(int(r["kode_kab"]), float(r["per_km"]))  # tahun terbaru menang

    # Indeks Penanaman kabupaten (C.A2 IP, lihat _ijd_score_kemanfaatan).
    indeks_penanaman_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kab, indeks_penanaman_pct FROM bps_kabupaten_indeks_penanaman "
                "WHERE kode_kab = ANY(%s) AND indeks_penanaman_pct IS NOT NULL ORDER BY tahun DESC",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                indeks_penanaman_by_kab.setdefault(int(r["kode_kab"]), float(r["indeks_penanaman_pct"]))  # tahun terbaru menang

    # Luas Lahan Baku Sawah (LBS) 2024 kabupaten -- tabel SAMA dgn
    # indeks_penanaman_by_kab di atas (bps_kabupaten_indeks_penanaman), kolom
    # beda; dipakai checklist Aspek B "LBS" (_bappenas_aspek_b_ekonomi).
    lbs_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kab, lahan_baku_sawah_ha FROM bps_kabupaten_indeks_penanaman "
                "WHERE kode_kab = ANY(%s) AND lahan_baku_sawah_ha IS NOT NULL AND lahan_baku_sawah_ha > 0 "
                "ORDER BY tahun DESC",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                lbs_by_kab.setdefault(int(r["kode_kab"]), float(r["lahan_baku_sawah_ha"]))

    # Indeks Penanaman kabupaten — sumber PRIMER (raster resmi Dit. SDA,
    # lihat _A2IP_RASTER_BUCKET_TO_SUB); indeks_penanaman_by_kab di atas
    # (Kertas Kerja.xlsx) jadi fallback SEKUNDER kalau kabupaten tak tercakup.
    ip_raster_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kab, bucket_ip FROM bps_kabupaten_indeks_penanaman_raster "
                "WHERE kode_kab = ANY(%s) ORDER BY tahun DESC",
                (kode_kab_str,),
            )
            for r in cur.fetchall():
                ip_raster_by_kab.setdefault(int(r["kode_kab"]), r["bucket_ip"])

    # kepadatan_by_kec juga menyimpan kolom potensi_* (satu query, satu tabel
    # sumber) — dipakai ulang sebagai ctx["potensi_by_kec"] utk A3.
    ctx = {"kab_by_wilayah": kab_by_wilayah, "kawasan_by_kab": kawasan_by_kab,
           "kepadatan_by_kec": kepadatan_by_kec, "potensi_by_kec": kepadatan_by_kec,
           "potensi_produksi_by_kec": potensi_produksi_by_kec,
           "bappenas_lokus_by_kab": bappenas_lokus_by_kab, "bappenas_lokus_by_prov": bappenas_lokus_by_prov,
           "kemantapan_kab_set": kemantapan_kab_set,
           "produktivitas_padi_by_kab": produktivitas_padi_by_kab,
           "kepadatan_kab_by_kab": kepadatan_kab_by_kab,
           "ip_raster_by_kab": ip_raster_by_kab,
           "kendaraan_per_km_by_kab": kendaraan_per_km_by_kab,
           "indeks_penanaman_by_kab": indeks_penanaman_by_kab, "lbs_by_kab": lbs_by_kab}
    # NPR (Nilai Prioritas Ruas) per usulan -- dibatch sama pola dgn ctx
    # teknokratis di atas (lihat _npr_bulk_ctx), dipakai SEJAK 22 Jul 2026
    # (request eksplisit user) sbg basis ranking export ini (gantikan
    # skor_ternormalisasi_100 teknokratis, lihat _ijd_ranking_sort_key) dan
    # jadi kolom export tersendiri (Skor NPR/Kategori NPR). Key
    # "jaringan_jalan_by_kab"/"simpul_by_kab"/"spasial_terhubung_by_usulan"
    # dari ctx ini di-reuse (bukan query ulang) oleh checklist Aspek B
    # JARINGAN_JALAN/SIMPUL_TRANSPORTASI (_bappenas_aspek_b_ekonomi),
    # ditambah 23 Jul 2026 -- lihat merge ke `ctx` di bawah.
    npr_ctx = _npr_bulk_ctx(rows)
    npr_by_id = {row["id"]: _compute_npr(row, npr_ctx) for row in rows}
    ctx["jaringan_jalan_by_kab"] = npr_ctx["jaringan_jalan_by_kab"]
    ctx["simpul_by_kab"] = npr_ctx["simpul_by_kab"]
    ctx["spasial_terhubung_by_usulan"] = npr_ctx["spasial_terhubung_by_usulan"]

    hasil = [(row, _compute_ijd_score(row, tahun, rules, ctx)) for row in rows]
    hasil.sort(key=lambda x: _ijd_ranking_sort_key(npr_by_id[x[0]["id"]]))

    # Peringkat per provinsi (kolom 43 template) dihitung terpisah dari urutan
    # baris utama (yang dalam cakupan filter, bisa nasional) supaya tetap
    # benar walau ekspor mencakup >1 provinsi sekaligus.
    by_provinsi: dict = {}
    for row, skor in hasil:
        by_provinsi.setdefault(row.get("provinsi"), []).append((row, skor))
    rank_in_provinsi = {}
    for items in by_provinsi.values():
        items_sorted = sorted(items, key=lambda x: _ijd_ranking_sort_key(npr_by_id[x[0]["id"]]))
        for i, (row, _) in enumerate(items_sorted, start=1):
            rank_in_provinsi[row["id"]] = i

    # Aspek A/B Bappenas (rule-based, lihat _bappenas_aspek_a_lokus/
    # _bappenas_aspek_b_ekonomi) — dihitung skr karena ctx sudah dibatch di
    # atas, jadi murah (tanpa query per baris). "Kesimpulan" TETAP kosong
    # kecuali usulan itu sudah pernah digenerate satu-per-satu lewat
    # POST /api/usulan-inpres/{id}/penilaian-bappenas (AI, di-cache) — bulk
    # export tidak memanggil LLM ribuan kali.
    usulan_ids = [row["id"] for row in rows]
    kesimpulan_cache = {}
    aspek_b_narasi_ai_cache = {}
    if usulan_ids:
        with db_cursor() as cur:
            cur.execute(
                "SELECT usulan_id, kesimpulan, aspek_b_narasi_ai FROM penilaian_bappenas_ai "
                "WHERE usulan_id = ANY(%s)",
                (list(usulan_ids),),
            )
            for r in cur.fetchall():
                kesimpulan_cache[r["usulan_id"]] = r["kesimpulan"]
                aspek_b_narasi_ai_cache[r["usulan_id"]] = r["aspek_b_narasi_ai"]

    bappenas_hasil = {}
    for row in rows:
        aspek_a = _bappenas_aspek_a_lokus(row, ctx)
        aspek_b = _bappenas_aspek_b_ekonomi(row, ctx)
        poin_a = _bappenas_poin_from_total(aspek_a["total_kriteria"])
        poin_b = _bappenas_poin_b_from_total(aspek_b["total_indikator"])
        bappenas_hasil[row["id"]] = {"aspek_a": aspek_a, "aspek_b": aspek_b,
                                      "poin_a": poin_a, "poin_b": poin_b, "total": poin_a + poin_b}

    # Peringkat Bappenas per provinsi (kolom "RANGKING DALAM PROVINSI
    # (Bappenas)") — total poin A+B, terpisah dari peringkat teknokratis.
    rank_bappenas_in_provinsi = {}
    for items in by_provinsi.values():
        items_sorted = sorted(items, key=lambda x: -bappenas_hasil[x[0]["id"]]["total"])
        for i, (row, _) in enumerate(items_sorted, start=1):
            rank_bappenas_in_provinsi[row["id"]] = i

    # Skor & peringkat PRIORITAS NASIONAL (_skor_prioritas_nasional, 70%
    # teknokratis + 10% PU + 10% Bappenas + 10% Kemenko Infra) — BEDA dari
    # rank_in_provinsi di atas (itu ranking teknokratis A-E SAJA, per
    # provinsi). Peringkat dihitung thd SEMUA usulan nasional yang sudah
    # py urutan_prioritas_kompetensi (basis komponen 70%), TERLEPAS dari
    # filter provinsi export ini, sama semantik dgn endpoint tunggal
    # /api/usulan-inpres/{id}/skor-prioritas-nasional — dibatch sekali di
    # sini (bukan query per baris) krn dipanggil utk ekspor massal.
    with db_cursor() as cur:
        cur.execute(f"SELECT {_SPN_FIELDS} FROM usulan_inpres WHERE prioritas_kompetensi IS NOT NULL")
        spn_semua_rows = cur.fetchall()
    spn_by_id = {r["id"]: _skor_prioritas_nasional(r) for r in spn_semua_rows}
    spn_sorted_scores = sorted((s["skor_total"] for s in spn_by_id.values() if s["skor_total"] is not None),
                                reverse=True)
    # peringkat = posisi kemunculan PERTAMA nilai itu di daftar terurut —
    # usulan dgn skor sama dapat peringkat yang sama, persis semantik
    # `semua.index(skor) + 1` di endpoint tunggal, tapi dihitung sekali
    # (bukan O(n) per baris) supaya tetap murah utk ekspor ribuan usulan.
    spn_rank_by_score: dict = {}
    for i, sk in enumerate(spn_sorted_scores, start=1):
        spn_rank_by_score.setdefault(sk, i)
    spn_jumlah_ternilai = len(spn_sorted_scores)

    # Struktur & label kolom mengikuti sheet "Output Penilaian" di
    # docs/docs/2_Analisis Prioritas untuk Bappenas dan Teknokratis
    # 15.7.2026.xlsx (template kosong dari Bappenas) — kolom 1-32 identitas/
    # administratif dari usulan_inpres, 33-38 penilaian Bappenas (Aspek A/B
    # rule-based + narasi AI Aspek B + ranking + total; kesimpulan cuma
    # terisi kalau sudah pernah digenerate AI per-usulan), 39-44 penilaian
    # Teknokratis (A-E dari _compute_ijd_score + peringkat per provinsi).
    header_row = (["No.", "ID"] + [label for label, _ in IJD_EXPORT_IDENTITAS_COLS]
                  + IJD_EXPORT_BAPPENAS_HEADERS
                  + [IJD_EXPORT_TEKNOKRATIS_HEADERS[k] for k in IJD_EXPORT_TEKNOKRATIS_KODE]
                  + [IJD_EXPORT_RANKING_LABEL, IJD_EXPORT_KELENGKAPAN_LABEL, IJD_EXPORT_OUTLIER_LABEL,
                     IJD_EXPORT_PRIORITAS_NASIONAL_LABEL, IJD_EXPORT_RANKING_NASIONAL_LABEL,
                     IJD_EXPORT_SKOR_TEKNOKRATIS_100_LABEL, IJD_EXPORT_NPR_LABEL, IJD_EXPORT_NPR_KATEGORI_LABEL])
    header_row_short = (["No.", "ID"] + [label for label, _ in IJD_EXPORT_IDENTITAS_COLS]
                         + IJD_EXPORT_BAPPENAS_HEADERS_SHORT
                         + [IJD_EXPORT_TEKNOKRATIS_HEADERS[k] for k in IJD_EXPORT_TEKNOKRATIS_KODE]
                         + [IJD_EXPORT_RANKING_LABEL, IJD_EXPORT_KELENGKAPAN_LABEL, IJD_EXPORT_OUTLIER_LABEL,
                            IJD_EXPORT_PRIORITAS_NASIONAL_LABEL, IJD_EXPORT_RANKING_NASIONAL_LABEL,
                            IJD_EXPORT_SKOR_TEKNOKRATIS_100_LABEL, IJD_EXPORT_NPR_LABEL, IJD_EXPORT_NPR_KATEGORI_LABEL])

    data_rows = []
    for i, (row, skor) in enumerate(hasil, start=1):
        komponen_by_kode = {k["kode"]: k for k in skor["komponen"]}
        bh = bappenas_hasil[row["id"]]
        data_row = [i, row["id"]] + [row.get(field) for _, field in IJD_EXPORT_IDENTITAS_COLS]
        data_row += [
            # Kolom "ASPEK PRIORITAS.../DAYA UNGKIT..." -- HANYA baris yang
            # TERCENTANG ([v]), baris [ ] (tidak cocok) DIBUANG -- request
            # eksplisit user 21 Jul 2026, supaya kolom export ringkas &
            # gampang di-scan, bukan daftar 11-14 kriteria penuh per sel.
            # Angka "[Poin N]" DALAM SEL ini = jumlah checklist mentah
            # (total_kriteria, PERSIS sama dgn kolom "Total Checklist Aspek
            # A" di sebelahnya) -- BUKAN skala 0/1/2 (bh['poin_a']) yang
            # dipakai kolom terpisah "Total (Bappenas)"/"Ranking Bappenas
            # per Provinsi" di bawah; user 21 Jul 2026 menandai angka di
            # dalam sel checklist ini seharusnya cocok jumlah tercentang,
            # bukan skala poin itu -- skala 0/1/2 TETAP dipakai apa adanya
            # utk ranking/total (tidak diubah, beda tujuan/kolom).
            f"[Poin {bh['aspek_a']['total_kriteria']}] "
            + "\n".join(l for l in bh['aspek_a']['checklist_lines'] if l.startswith("[v]")),
            bh['aspek_a']['total_kriteria'],
            # Badge "[Poin N]" DIHILANGKAN dari sel Aspek B (beda dari Aspek
            # A di atas) -- request eksplisit user 21 Jul 2026: skala 0/1/2
            # (bh['poin_b']) selalu tampil sama ("POIN 2") utk usulan dgn
            # jumlah checklist berbeda-beda (3, 4, 5 kriteria dst.), jadi
            # membingungkan drpd informatif di dalam sel checklist ini --
            # skala poin TETAP dipakai apa adanya di kolom "Total
            # (Bappenas)"/"Ranking Bappenas per Provinsi" terpisah, cuma
            # tidak lagi ditampilkan sbg prefix di sini.
            "\n".join(l for l in bh['aspek_b']['checklist_lines'] if l.startswith("[v]")),
            aspek_b_narasi_ai_cache.get(row["id"]),
            rank_bappenas_in_provinsi[row["id"]],
            kesimpulan_cache.get(row["id"]),
            bh["total"],
        ]
        data_row += [komponen_by_kode.get(k, {}).get("kontribusi") for k in IJD_EXPORT_TEKNOKRATIS_KODE]
        kode_hilang = [k for k in IJD_EXPORT_TEKNOKRATIS_KODE if not komponen_by_kode.get(k, {}).get("tersedia")]
        kelengkapan = (
            f"Lengkap ({skor['bobot_tersedia']:.0f}/100)" if not kode_hilang
            else f"PARSIAL ({skor['bobot_tersedia']:.0f}/100) — {'/'.join(kode_hilang)} belum tersedia; "
                 "skor dinormalisasi cuma dari parameter tersedia, bisa terlihat lebih tinggi drpd usulan berdata lengkap"
        )
        produksi_kec = ctx["potensi_produksi_by_kec"].get(row.get("kode_kecamatan"))
        outlier_keterangan = _outlier_produksi_keterangan(produksi_kec) or "Tidak ada indikasi outlier"

        spn = spn_by_id.get(row["id"])
        if spn and spn["skor_total"] is not None:
            peringkat_nasional = spn_rank_by_score[spn["skor_total"]]
            prioritas_nasional_ket = (
                f"{spn['skor_total']:.2f} — Peringkat {peringkat_nasional}/{spn_jumlah_ternilai} nasional "
                "(70% Prioritisasi Teknokratis + 10% Prioritisasi Kementerian PU + "
                "10% Indikasi Prioritas Bappenas + 10% Indikasi Prioritas Kemenko Infra)"
            )
        else:
            peringkat_nasional = None
            prioritas_nasional_ket = ("Belum tersedia — usulan belum punya urutan prioritas kompetensi "
                                       "(basis komponen 70% Prioritisasi Teknokratis).")

        npr_skor = npr_by_id[row["id"]]
        data_row += [rank_in_provinsi[row["id"]], kelengkapan, outlier_keterangan,
                     prioritas_nasional_ket, peringkat_nasional,
                     skor["skor_ternormalisasi_100"], npr_skor["npr"], npr_skor["kategori"]]
        data_rows.append(data_row)

    result = (header_row, header_row_short, data_rows)
    _ijd_bulk_cache[key] = result
    return result


@app.get("/api/usulan-inpres/ijd-score/preview")
def usulan_inpres_ijd_score_preview(provinsi: list[str] = Query(default=[]), tahun: int = 2026,
                                     limit: int = 50, offset: int = 0):
    """Versi JSON, dipaging, dari _ijd_score_bulk_rows — dipakai modal preview
    di UI (tabel gaya "Data") supaya user bisa cek isi sebelum benar-benar
    unduh xlsx. Kontrak responsnya sengaja mengikuti GET /api/data/{table}
    supaya bisa pakai styling/komponen tabel yang sama di frontend.
    provinsi BISA berulang (?provinsi=A&provinsi=B, multi-select combo di
    UI, 21 Jul 2026) -- kosong = nasional, satu = provinsi itu saja."""
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    provinsi = _normalisasi_provinsi_multi(provinsi)
    _header_full, header_short, data_rows = _ijd_score_bulk_rows(provinsi, tahun)
    page = data_rows[offset:offset + limit]
    # Kolom prosa AI: kirim utuh (batas longgar) — frontend menampilkannya
    # terpotong tapi memberi tooltip berisi teks lengkap saat hover.
    prose_cols = {i for i, h in enumerate(header_short)
                  if h in ("Aspek B: Narasi AI", "Kesimpulan Penilaian Bappenas")}

    def _trim(v, i):
        if not isinstance(v, str):
            return v
        # Sel checklist Aspek A/B: JANGAN potong di 200 char — butir tercentang
        # ([v]) sering di urutan bawah daftar (mis. Kemantapan baris 39) dan
        # ikut terpotong, membuat poinnya tampak tidak cocok dgn centangnya.
        # Tinggi sel dikendalikan CSS (.checklist-cell), bukan pemotongan.
        limit_chars = 1500 if (i in prose_cols or "[v]" in v or "[ ]" in v) else 200
        return (v[:limit_chars] + "…") if len(v) > limit_chars else v

    trimmed = [[_trim(v, i) for i, v in enumerate(r)] for r in page]
    if not provinsi:
        scope = "Nasional"
    elif len(provinsi) == 1:
        scope = provinsi[0]
    elif len(provinsi) <= 3:
        scope = ", ".join(provinsi)
    else:
        scope = f"{len(provinsi)} Provinsi"
    return jsonable_encoder({
        "table": "ijd_skor_preview", "label": f"Preview Export Skor IJD — {scope} ({tahun})",
        "columns": header_short, "rows": trimmed, "total": len(data_rows),
        "limit": limit, "offset": offset,
    })


@app.get("/api/usulan-inpres/ijd-score/export/xlsx")
def usulan_inpres_ijd_score_bulk_export(provinsi: list[str] = Query(default=[]), tahun: int = 2026):
    provinsi = _normalisasi_provinsi_multi(provinsi)
    header_row, _header_short, data_rows = _ijd_score_bulk_rows(provinsi, tahun)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Output Penilaian")
    group_row = [None] * (2 + len(IJD_EXPORT_IDENTITAS_COLS))
    group_row += ["FORMAT PENILAIAN BAPPENAS"] + [None] * (len(IJD_EXPORT_BAPPENAS_HEADERS) - 1)
    # Sisa kolom (komponen A-E + ranking per-provinsi + Kelengkapan Data +
    # Temuan Data Quality Outlier + Prioritisasi Nasional) semua di bawah
    # payung grup "TEKNOKRATIS" walau sebagian bukan komponen A-E sendiri --
    # dihitung DINAMIS dari total len(header_row) (bukan hardcode jumlah
    # kolom trailing) supaya tidak drift lagi tiap kali kolom baru
    # ditambahkan di akhir (pernah kelewat off-by-one sebelum 21 Jul 2026
    # saat kolom Prioritas Nasional ditambahkan -- ditemukan & diperbaiki
    # sekalian di sini).
    n_teknokratis_group = len(header_row) - len(group_row) - 1
    group_row += ["FORMAT PENILAIAN TEKNOKRATIS"] + [None] * n_teknokratis_group
    ws.append(group_row)
    ws.append(header_row)
    for data_row in data_rows:
        ws.append(data_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if not provinsi:
        scope = "Nasional"
    elif len(provinsi) <= 3:
        scope = re.sub(r"[^\w]+", "_", "_".join(provinsi))
    else:
        scope = f"{len(provinsi)}Provinsi"
    fname = f"ijd_skor_{scope}_{tahun}_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _ijd_score_dashboard(provinsi: str = "", tahun: int = 2026) -> dict:
    """Ringkasan siap-pakai skor IJD Prioritisasi Teknokratik A-E (KPI,
    peringkat, cakupan tiap komponen, komposisi kategori) -- pola SAMA dgn
    _laporan_prioritas_dashboard (Laporan Prioritas), reuse
    _ijd_score_bulk_rows yang sudah ada & di-cache (_ijd_bulk_cache) drpd
    query/skoring baru. Ditambahkan 21 Jul 2026 (request eksplisit user).

    PAKAI skor_tertimbang MENTAH (sum kontribusi A-E yang tersedia), BUKAN
    skor_ternormalisasi_100 spt /ijd-score -- utk avg_total/komposisi/cakupan
    komponen di bawah, angka mentah lebih adil (renormalisasi thd
    bobot_tersedia bisa menaikkan usulan yg datanya sedikit-tapi-kebetulan-
    bagus, lihat docs/analisis_keadilan_kemanfaatan_skoring.md §2). Label
    jelas "skor_tertimbang" di respons supaya tidak tertukar dgn endpoint
    lain.

    "top10" (peringkat) SEJAK 22 Jul 2026 (request eksplisit user) diurutkan
    berdasar skor NPR -- BUKAN lagi skor_tertimbang A-E -- supaya konsisten
    dgn ranking di export Skor IJD (IJD_EXPORT_RANKING_LABEL, lihat
    _ijd_ranking_sort_key) yang sudah dipindah ke basis NPR lebih dulu. NPR
    tetap metodologi eksperimental/belum policy resmi (lihat _compute_npr) --
    avg_total/komposisi/cakupan_komponen TIDAK ikut berubah, tetap basis
    skor_tertimbang teknokratis A-E apa adanya."""
    header_full, header_short, data_rows = _ijd_score_bulk_rows(provinsi, tahun)
    npr_by_id = {r["id"]: r["npr"] for r in _npr_bulk_rows(provinsi)}
    idx_prov = header_short.index("Provinsi")
    idx_kab = header_short.index("Kabupaten/Kota")
    idx_ruas = header_short.index("Nama Ruas")
    # Dicari lewat label (BUKAN offset tetap dari akhir header_short) --
    # offset tetap ("N kolom dari belakang") jadi salah tiap kali kolom
    # trailing baru ditambahkan (Prioritas Nasional/Ranking Nasional dkk.,
    # 21-22 Jul 2026) krn asumsi "cuma 1 kolom di belakang komponen E" jadi
    # basi; ditemukan lewat crash TypeError float+str (idx_komponen
    # kebablasan ikut kolom teks) saat verifikasi combo provinsi multi-
    # select, diperbaiki sekalian di sini.
    idx_komponen = [header_short.index(IJD_EXPORT_TEKNOKRATIS_HEADERS[k]) for k in IJD_EXPORT_TEKNOKRATIS_KODE]

    items = []
    for r in data_rows:
        nilai_komponen = [r[i] for i in idx_komponen]
        n_tersedia = sum(1 for v in nilai_komponen if v is not None)
        total = round(sum(v for v in nilai_komponen if v is not None), 2)
        items.append({"id": r[1], "nama_ruas": r[idx_ruas], "provinsi": r[idx_prov],
                       "kabupaten_kota": r[idx_kab], "total": total, "n_komponen_tersedia": n_tersedia,
                       "npr": npr_by_id.get(r[1])})

    n = len(items)
    n_tanpa_data = sum(1 for it in items if it["n_komponen_tersedia"] == 0)
    avg_total = round(sum(it["total"] for it in items) / n, 2) if n else 0

    ranked = sorted(items, key=lambda it: (it["npr"] is None, -(it["npr"] or 0)))
    top10 = ranked[:10]

    def _kategori(it):
        if it["n_komponen_tersedia"] == 0:
            return "Tidak ada data"
        if it["total"] <= 25:
            return "Rendah"
        if it["total"] <= 50:
            return "Sedang"
        return "Tinggi"
    komposisi_count = {"Tinggi": 0, "Sedang": 0, "Rendah": 0, "Tidak ada data": 0}
    for it in items:
        komposisi_count[_kategori(it)] += 1
    komposisi = [{"label": k, "count": komposisi_count[k]} for k in ("Tinggi", "Sedang", "Rendah", "Tidak ada data")]

    cakupan_komponen = []
    for kode, idx in zip(IJD_EXPORT_TEKNOKRATIS_KODE, idx_komponen):
        n_match = sum(1 for r in data_rows if r[idx] is not None)
        cakupan_komponen.append({"label": IJD_EXPORT_TEKNOKRATIS_HEADERS[kode], "count": n_match,
                                  "pct": round(n_match / n * 100, 1) if n else 0})

    # KPI & komposisi kategori NPR -- ditambah 22 Jul 2026 (request user,
    # "apakah ada masukan agar lebih informatif") supaya dashboard tidak
    # cuma nampilin top10 berbasis NPR tanpa konteks sebaran/cakupan NPR
    # nasionalnya. "Belum Tersedia" (NPR None, SI & SC sama sekali tak bisa
    # dihitung) DIPISAH dari "Belum Prioritas" (NPR < 50 tapi terhitung) --
    # pola "belum tersedia, bukan 0" yang sama dipakai di seluruh app ini.
    n_dengan_npr = sum(1 for it in items if it["npr"] is not None)
    avg_npr = round(sum(it["npr"] for it in items if it["npr"] is not None) / n_dengan_npr, 2) if n_dengan_npr else None

    def _kategori_npr(it):
        if it["npr"] is None:
            return "Belum Tersedia"
        for ambang, label in _NPR_KATEGORI:
            if it["npr"] >= ambang:
                return label
        return "Belum Prioritas"
    urutan_kategori_npr = [label for _, label in _NPR_KATEGORI] + ["Belum Prioritas", "Belum Tersedia"]
    komposisi_npr_count = {label: 0 for label in urutan_kategori_npr}
    for it in items:
        komposisi_npr_count[_kategori_npr(it)] += 1
    komposisi_npr = [{"label": k, "count": komposisi_npr_count[k],
                       "pct": round(komposisi_npr_count[k] / n * 100, 1) if n else 0}
                      for k in urutan_kategori_npr]

    return {
        "n_usulan": n, "avg_total": avg_total, "maks_total": 100, "n_tanpa_data": n_tanpa_data,
        "top10": top10, "komposisi": komposisi, "cakupan_komponen": cakupan_komponen,
        "avg_npr": avg_npr, "n_dengan_npr": n_dengan_npr, "komposisi_npr": komposisi_npr,
        "provinsi": provinsi or None, "tahun": tahun,
        "catatan": (
            "avg_total/komposisi/cakupan komponen = jumlah kontribusi komponen A-E (skor_tertimbang) yang "
            "datanya tersedia -- BUKAN skor_ternormalisasi_100 (yang dipakai /ijd-score per-usulan). "
            "Renormalisasi sengaja tidak dipakai di sini krn bisa menaikkan usulan yang datanya sedikit "
            "tapi kebetulan bagus, kurang adil utk ranking lintas-usulan skala nasional. Peringkat 10 "
            "usulan (top10), avg_npr, & komposisi_npr TERPISAH -- semuanya berbasis skor NPR (Nilai "
            "Prioritas Ruas, metodologi eksperimental/belum policy resmi), konsisten dgn ranking di "
            "export Skor IJD, BUKAN skor_tertimbang A-E yang dipakai avg_total/komposisi/cakupan_komponen."
        ),
    }


@app.get("/api/usulan-inpres/ijd-score/dashboard")
def usulan_inpres_ijd_score_dashboard(provinsi: str = "", tahun: int = 2026):
    return jsonable_encoder(_ijd_score_dashboard(provinsi, tahun))


# --- Skor Prioritas Nasional (dokumen 14072026 bagian C: 70% teknokratis +
# 10% Kementerian PU + 10% Bappenas + 10% Kemenko Infra) ---

_SPN_FIELDS = (
    "id, nama_kegiatan, nama_ruas, provinsi, kabupaten_kota, "
    "prioritas_kompetensi, prioritas_balai, indikasi_prioritas_bappenas, "
    "indikasi_prioritas_kemenko, alokasi_usulan_kompetensi, alokasi_usulan_pemda"
)


def _skor_prioritas_nasional(row: dict) -> dict:
    """Skor kuantitatif per usulan. Hanya usulan yang sudah punya urutan
    prioritas kompetensi (lolos verifikasi) yang bisa dihitung — komponen
    teknokratis 70% adalah basisnya."""
    pk = row.get("prioritas_kompetensi")
    pb = row.get("prioritas_balai")
    bappenas = (row.get("indikasi_prioritas_bappenas") or "").strip().upper() == "YA"
    kemenko = (row.get("indikasi_prioritas_kemenko") or "").strip().upper() == "YA"

    skor_kompetensi = max(101 - int(pk), 1) if pk is not None else None
    skor_pu = 100 if pb == 1 else (50 if pb is not None else 0)

    komponen = [
        {
            "kode": "A", "label": "Prioritisasi Teknokratis", "bobot_pct": 70,
            "nilai": skor_kompetensi,
            "keterangan": (
                f"101 − urutan prioritas kompetensi ({pk}), minimal 1." if pk is not None
                else "Belum ada urutan prioritas kompetensi (usulan belum selesai dinilai)."
            ),
        },
        {
            "kode": "B", "label": "Prioritisasi Kementerian PU", "bobot_pct": 10,
            "nilai": skor_pu,
            "keterangan": (
                f"Prioritas Balai {pb} ({'peringkat 1 = 100' if pb == 1 else 'peringkat 2+ = 50'})."
                if pb is not None else "Belum ada urutan prioritas Balai — dinilai 0."
            ),
        },
        {
            "kode": "C", "label": "Indikasi Prioritas Bappenas", "bobot_pct": 10,
            "nilai": 100 if bappenas else 0,
            "keterangan": (
                "Ditandai prioritas Bappenas." if bappenas else
                "Belum/tidak ditandai prioritas Bappenas (kolom indikasi SITIA masih kosong per tarikan 15 Juli)."
            ),
        },
        {
            "kode": "D", "label": "Indikasi Prioritas Kemenko Infra", "bobot_pct": 10,
            "nilai": 100 if kemenko else 0,
            "keterangan": (
                "Ditandai prioritas Kemenko Infra." if kemenko else
                "Belum/tidak ditandai prioritas Kemenko Infra (kolom indikasi SITIA masih kosong per tarikan 15 Juli)."
            ),
        },
    ]
    total = None
    if skor_kompetensi is not None:
        total = round(sum(k["nilai"] * k["bobot_pct"] / 100 for k in komponen), 2)
    for k in komponen:
        k["kontribusi"] = round((k["nilai"] or 0) * k["bobot_pct"] / 100, 2)

    return {
        "skor_total": total,
        "komponen": komponen,
        "catatan": (
            "Skor Prioritas Nasional perkiraan per dokumen Penentuan Parameter Penilaian IJD "
            "14 Juli 2026 (70% teknokratis + 10% PU + 10% Bappenas + 10% Kemenko Infra). "
            "Bukan penetapan resmi."
        ),
    }


@app.get("/api/usulan-inpres/{usulan_id}/skor-prioritas-nasional")
def usulan_inpres_skor_prioritas(usulan_id: int):
    with db_cursor() as cur:
        cur.execute(f"SELECT {_SPN_FIELDS} FROM usulan_inpres WHERE id = %s", (usulan_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Usulan tidak ditemukan")
    hasil = _skor_prioritas_nasional(row)
    if hasil["skor_total"] is not None:
        # peringkat nasional = posisi skor usulan ini di antara semua yang ternilai
        with db_cursor() as cur:
            cur.execute(
                f"SELECT {_SPN_FIELDS} FROM usulan_inpres WHERE prioritas_kompetensi IS NOT NULL"
            )
            semua = sorted((_skor_prioritas_nasional(r)["skor_total"] for r in cur.fetchall()),
                           reverse=True)
        hasil["peringkat_nasional"] = semua.index(hasil["skor_total"]) + 1
        hasil["jumlah_ternilai"] = len(semua)
    return hasil


@app.get("/api/prioritas-nasional")
def prioritas_nasional_list(provinsi: str = "", limit: int = 50, offset: int = 0):
    """Peringkat Prioritas Nasional: semua usulan yang punya urutan prioritas
    kompetensi, diurutkan skor tertinggi -> terendah (basis alokasi 2 lapis)."""
    limit = max(1, min(limit, 200))
    with db_cursor() as cur:
        cur.execute(f"SELECT {_SPN_FIELDS} FROM usulan_inpres WHERE prioritas_kompetensi IS NOT NULL")
        rows = cur.fetchall()
    dinilai = []
    for r in rows:
        hasil = _skor_prioritas_nasional(r)
        dinilai.append({
            "id": r["id"], "nama_kegiatan": r["nama_kegiatan"], "nama_ruas": r["nama_ruas"],
            "provinsi": r["provinsi"], "kabupaten_kota": r["kabupaten_kota"],
            "prioritas_kompetensi": r["prioritas_kompetensi"], "prioritas_balai": r["prioritas_balai"],
            "alokasi": r["alokasi_usulan_kompetensi"] or r["alokasi_usulan_pemda"],
            "skor_total": hasil["skor_total"],
        })
    dinilai.sort(key=lambda x: (-x["skor_total"], x["id"]))
    for i, d in enumerate(dinilai, start=1):
        d["peringkat_nasional"] = i
    if provinsi:
        dinilai = [d for d in dinilai if d["provinsi"] == provinsi]
    total = len(dinilai)
    return {"total": total, "usulan": dinilai[offset:offset + limit]}


# --- Pagu Indikatif Provinsi (dokumen 14072026 bagian B) + Alokasi 2 Lapis
# (bagian C akhir). Komponen pagu yang datanya ada: A1 (panjang jalan
# daerah, SI 2026 Tabel 10.1.1), A2 kemantapan (kemantapan_ijd_2026), A3
# kawasan pangan (proksi lahan sawah ATR/BPN, lihat catatan di bawah), dan
# A4 (kapasitas fiskal, kolom SITIA usulan gubernur) — hanya A5 IKK
# (publikasi BPS terpisah) belum ada sumbernya. Tiap komponen dinyatakan
# sebagai pangsa nasional (jumlah antar provinsi = 100%) supaya total pagu
# tepat mendistribusikan alokasi nasional; skor akhir dinormalisasi ulang ke
# bobot komponen yang tersedia (85 dari 100 tanpa A5).

_FISKAL_SKOR = {"SANGAT TINGGI": 10, "TINGGI": 15, "SEDANG": 20, "RENDAH": 25, "SANGAT RENDAH": 30}
_PAGU_KOMPONEN_PENDING = {
    "A5": "Indeks Kemahalan Konstruksi (bobot 15) — publikasi IKK BPS belum diimpor.",
}


def _skor_indeks_ketidakmantapan(pct: float) -> float:
    """Tabel 7 dokumen 14072026: skor indeks ketidakmantapan jalan daerah."""
    if pct > 65:
        return 100
    if pct >= 40:
        return 75
    if pct >= 25:
        return 50
    return 25


def _norm_prov_nama(s: str) -> str:
    toks = [t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if t]
    toks = ["DI" if t in ("DI", "D", "DAERAH") else t for t in toks]
    return " ".join(t for t in toks if t not in ("I", "ISTIMEWA"))


def _pagu_provinsi(alokasi_nasional: Optional[float] = None) -> dict:
    with db_cursor() as cur:
        cur.execute(
            "SELECT provinsi, NULLIF(COALESCE(provinsi_km, 0) + COALESCE(kabkota_km, 0), 0) AS daerah_km "
            "FROM si_panjang_jalan_provinsi WHERE kode_provinsi > 0 "
            "AND tahun = (SELECT MAX(tahun) FROM si_panjang_jalan_provinsi)"
        )
        jalan = {_norm_prov_nama(r["provinsi"]): float(r["daerah_km"])
                 for r in cur.fetchall() if r["daerah_km"] is not None}
        cur.execute(
            "SELECT provinsi, MAX(kapasitas_fiskal) AS fiskal FROM usulan_inpres "
            "WHERE nama_pengusul LIKE 'Gubernur%%' AND kapasitas_fiskal IS NOT NULL "
            "GROUP BY provinsi"
        )
        fiskal = {_norm_prov_nama(r["provinsi"]): r["fiskal"] for r in cur.fetchall()}
        cur.execute(
            "SELECT provinsi, SUM(panjang_km) AS total_km, SUM(tidak_mantap_km) AS total_tm "
            "FROM kemantapan_ijd_2026 GROUP BY provinsi"
        )
        kemantapan_raw = {_norm_prov_nama(r["provinsi"]): r for r in cur.fetchall()
                          if r["total_km"]}
        cur.execute(
            "SELECT provinsi, lahan_sawah_2024_km2 FROM si_lahan_sawah_provinsi "
            "WHERE kode_provinsi > 0 AND lahan_sawah_2024_km2 IS NOT NULL"
        )
        sawah = {_norm_prov_nama(r["provinsi"]): float(r["lahan_sawah_2024_km2"]) for r in cur.fetchall()}
        cur.execute("SELECT DISTINCT provinsi FROM usulan_inpres ORDER BY provinsi")
        frame = [r["provinsi"] for r in cur.fetchall()]

    # A2: pct ketidakmantapan per provinsi -> skor indeks Tabel 7 -> pangsa nasional
    indeks_ketidakmantapan = {}
    for key, r in kemantapan_raw.items():
        pct = float(r["total_tm"] or 0) / float(r["total_km"]) * 100
        indeks_ketidakmantapan[key] = _skor_indeks_ketidakmantapan(pct)

    total_jalan = sum(jalan.get(_norm_prov_nama(p), 0) for p in frame)
    total_fiskal = sum(_FISKAL_SKOR.get((fiskal.get(_norm_prov_nama(p)) or "").upper(), 0) for p in frame)
    total_indeks = sum(indeks_ketidakmantapan.get(_norm_prov_nama(p), 0) for p in frame)
    total_sawah = sum(sawah.get(_norm_prov_nama(p), 0) for p in frame)

    provinsi_rows, total_skor = [], 0.0
    for p in frame:
        key = _norm_prov_nama(p)
        km = jalan.get(key)
        f_label = fiskal.get(key)
        f_skor = _FISKAL_SKOR.get((f_label or "").upper())
        idx = indeks_ketidakmantapan.get(key)
        sawah_km2 = sawah.get(key)
        a1 = km / total_jalan * 100 if km and total_jalan else None
        a2 = idx / total_indeks * 100 if idx and total_indeks else None
        a3 = sawah_km2 / total_sawah * 100 if sawah_km2 and total_sawah else None
        a4 = f_skor / total_fiskal * 100 if f_skor and total_fiskal else None
        bobot, nilai = 0.0, 0.0
        if a1 is not None:
            bobot += 20
            nilai += a1 * 20
        if a2 is not None:
            bobot += 30
            nilai += a2 * 30
        if a3 is not None:
            bobot += 20
            nilai += a3 * 20
        if a4 is not None:
            bobot += 15
            nilai += a4 * 15
        skor = nilai / bobot if bobot else None
        pct_tm = (float(kemantapan_raw[key]["total_tm"] or 0) / float(kemantapan_raw[key]["total_km"]) * 100
                  if key in kemantapan_raw else None)
        provinsi_rows.append({
            "provinsi": p,
            "jalan_daerah_km": km,
            "a1_pangsa_pct": round(a1, 3) if a1 is not None else None,
            "tidak_mantap_pct": round(pct_tm, 2) if pct_tm is not None else None,
            "a2_pangsa_pct": round(a2, 3) if a2 is not None else None,
            "lahan_sawah_km2": sawah_km2,
            "a3_pangsa_pct": round(a3, 3) if a3 is not None else None,
            "kapasitas_fiskal": f_label,
            "a4_pangsa_pct": round(a4, 3) if a4 is not None else None,
            "bobot_tersedia": bobot,
            "skor_pct": round(skor, 3) if skor is not None else None,
        })
        total_skor += skor or 0

    # normalisasi ulang supaya jumlah pangsa = 100% (provinsi tanpa data salah
    # satu komponen tetap proporsional terhadap yang lain)
    for r in provinsi_rows:
        if r["skor_pct"] is not None and total_skor:
            r["pangsa_final_pct"] = round(r["skor_pct"] / total_skor * 100, 3)
            if alokasi_nasional:
                r["pagu_rp"] = round(alokasi_nasional * r["pangsa_final_pct"] / 100)
        else:
            r["pangsa_final_pct"] = None
            if alokasi_nasional:
                r["pagu_rp"] = 0

    provinsi_rows.sort(key=lambda r: -(r["pangsa_final_pct"] or 0))
    return {
        "tahun_data_jalan": "SI 2026 (tahun terbaru di tabel)",
        "alokasi_nasional_rp": alokasi_nasional,
        "provinsi": provinsi_rows,
        "komponen_belum_tersedia": _PAGU_KOMPONEN_PENDING,
        "catatan": (
            "Skor pagu provinsi PARSIAL: komponen A1 panjang jalan daerah (bobot 20), A2 panjang "
            "jalan tidak mantap (bobot 30, dari docs/docs/5_IJD 2026 - DATA.xlsx), A3 kawasan pangan "
            "(bobot 20, proksi lahan sawah ATR/BPN SK 446.1/2024 dari si_lahan_sawah_provinsi — "
            "sawah saja, bukan kawasan pangan multi-komoditas resmi PDF 14072026, karena itu satu-"
            "satunya sumber ATR/BPN nasional level provinsi yang tersedia) dan A4 kapasitas fiskal "
            "(bobot 15) dari 5 komponen dokumen 14072026 — hanya A5 Indeks Kemahalan Konstruksi yang "
            "masih kosong; pangsa dinormalisasi ulang ke bobot yang tersedia (85/100). Bukan penetapan resmi."
        ),
    }


@app.get("/api/pagu-provinsi")
def pagu_provinsi(alokasi_nasional: Optional[float] = None):
    return _pagu_provinsi(alokasi_nasional)


@app.get("/api/alokasi-2-lapis")
def alokasi_2_lapis(alokasi_nasional: float):
    """Simulasi Metode Alokasi 2 Lapis (dokumen 14072026): Lapis 1 = usulan
    peringkat 1-2 per pemda mengikuti Peringkat Nasional, kumulatif <= pagu
    provinsi; Lapis 2 = sisa pagu untuk pemda yang belum kebagian, lalu
    pengisi celah."""
    pagu = {r["provinsi"]: r.get("pagu_rp") or 0 for r in _pagu_provinsi(alokasi_nasional)["provinsi"]}

    with db_cursor() as cur:
        cur.execute(f"SELECT {_SPN_FIELDS} FROM usulan_inpres WHERE prioritas_kompetensi IS NOT NULL")
        rows = cur.fetchall()
    usulan = []
    for r in rows:
        skor = _skor_prioritas_nasional(r)["skor_total"]
        nilai = r["alokasi_usulan_kompetensi"] or r["alokasi_usulan_pemda"] or 0
        usulan.append({
            "id": r["id"], "provinsi": r["provinsi"], "pemda": r["kabupaten_kota"],
            "nama_ruas": r["nama_ruas"], "skor_total": skor, "nilai_rp": int(nilai),
        })
    usulan.sort(key=lambda x: (-x["skor_total"], x["id"]))
    for i, u in enumerate(usulan, start=1):
        u["peringkat_nasional"] = i

    # peringkat dalam pemda (kunci aturan kualifikasi Lapis 1)
    rank_pemda: dict = {}
    for u in usulan:
        key = (u["provinsi"], u["pemda"])
        rank_pemda[key] = rank_pemda.get(key, 0) + 1
        u["peringkat_pemda"] = rank_pemda[key]

    terpakai = {p: 0 for p in pagu}
    pemda_dapat = set()
    hasil = []

    # Lapis 1
    for u in usulan:
        if u["peringkat_pemda"] > 2 or not u["nilai_rp"]:
            continue
        prov = u["provinsi"]
        if terpakai.get(prov, 0) + u["nilai_rp"] <= pagu.get(prov, 0):
            terpakai[prov] = terpakai.get(prov, 0) + u["nilai_rp"]
            pemda_dapat.add((prov, u["pemda"]))
            u["lapis"] = 1
            hasil.append(u)

    # Lapis 2
    teralokasi_ids = {u["id"] for u in hasil}
    for prov in pagu:
        while True:
            sisa = pagu[prov] - terpakai.get(prov, 0)
            if sisa <= 0:
                break
            kandidat = [u for u in usulan
                        if u["provinsi"] == prov and u["id"] not in teralokasi_ids
                        and 0 < u["nilai_rp"] <= sisa]
            if not kandidat:
                break
            # prioritas pemerataan: pemda yang belum dapat alokasi dulu
            belum = [u for u in kandidat if (prov, u["pemda"]) not in pemda_dapat]
            pilih = min(belum or kandidat, key=lambda x: x["peringkat_nasional"])
            terpakai[prov] += pilih["nilai_rp"]
            pemda_dapat.add((prov, pilih["pemda"]))
            teralokasi_ids.add(pilih["id"])
            pilih["lapis"] = 2
            hasil.append(pilih)

    rekap = []
    for prov in sorted(pagu, key=lambda p: -pagu[p]):
        alloc = [u for u in hasil if u["provinsi"] == prov]
        rekap.append({
            "provinsi": prov,
            "pagu_rp": pagu[prov],
            "terpakai_rp": terpakai.get(prov, 0),
            "sisa_rp": pagu[prov] - terpakai.get(prov, 0),
            "jumlah_usulan": len(alloc),
            "lapis1": sum(1 for u in alloc if u["lapis"] == 1),
            "lapis2": sum(1 for u in alloc if u["lapis"] == 2),
        })

    return {
        "alokasi_nasional_rp": alokasi_nasional,
        "total_teralokasi_rp": sum(terpakai.values()),
        "jumlah_usulan_teralokasi": len(hasil),
        "rekap_provinsi": rekap,
        "alokasi": sorted(hasil, key=lambda x: x["peringkat_nasional"]),
        "catatan": (
            "Simulasi alokasi 2 lapis di atas pagu provinsi PARSIAL (komponen A1+A4 saja) dan "
            "Skor Prioritas Nasional perkiraan — hasil berubah saat komponen pagu/flag prioritas "
            "lain terisi. Bukan penetapan resmi."
        ),
    }


# --- Draf narasi "Output Penilaian" Bappenas per usulan, dihasilkan AI (gap
# G11). Format mengikuti sheet Output Penilaian file 2 (aspek A Prioritas-
# Strategis + aspek B Daya Ungkit, poin 0/1/2 + narasi + kesimpulan); hasil
# di-cache di tabel penilaian_bappenas_ai. SELALU berlabel draf AI. ---

# --- Aspek A "Prioritas & Nilai Strategis" (Bappenas) -- RULE-BASED, bukan
# AI lagi (docs/spec/Draf Penilaian Bappenas.md). Cocokkan usulan ke 11
# kriteria lokus yg sudah punya sumber data: 8 dari bappenas_lokus_a
# (scripts/import_bappenas_lokus_a.py) + 3 dari kawasan_tematik (PKPN/
# TRANSMIGRASI/KI_PRIORITAS, sudah diimpor utk IJD A3 — dipakai ulang di
# sini krn sheet "Kumpulan Data" baris 4/7/14-17 memang bagian Aspek A).
# BBM_1_HARGA belum ada sumber data bersih (lihat catatan di
# import_bappenas_lokus_a.py), tidak dicek. ---

BAPPENAS_KRITERIA_LABEL = {
    "LOKPRI_RPJMN": "Lokasi Prioritas (Lokpri) RPJMN 2025-2029",
    "PKPN": "Kawasan Mendukung PKPN (3T/Tertinggal)",
    "PKSN": "Pusat Kegiatan Strategis Nasional (PKSN) Perbatasan",
    "PERBATASAN": "Kecamatan Perbatasan Prioritas",
    "TRANSMIGRASI": "Kawasan Transmigrasi Prioritas",
    "SR": "Lokasi Sekolah Rakyat (SR)",
    "SEKOLAH_GARUDA": "Lokasi Sekolah Unggul Garuda (SUGB)",
    "KNMP": "Kampung Nelayan Merah Putih (KNMP)",
    "KDMP": "Koperasi Desa Merah Putih (KDMP)",
    "KI_PRIORITAS": "Kawasan Industri Prioritas (PSN/Hilirisasi/RPJMN/Dirgantara)",
    "SWASEMBADA_PANGAN_RPJMN": "Kawasan Komoditas Unggulan Swasembada Pangan RPJMN",
    "BBM_1_HARGA": "Lokasi Mendukung Distribusi BBM Satu Harga",
    "KPP_DESA": "Kawasan Perdesaan Prioritas (KPP)",
    # 3 kriteria ditambahkan 21 Jul 2026 -- disinkronkan dgn LAPORAN_ASPEK_A
    # (baris 13-15 sheet "Kumpulan Data", ditambahkan sebelumnya ke Laporan
    # Prioritas tapi kelewat disinkronkan ke sini). Label PERSIS
    # LAPORAN_ASPEK_A supaya kedua checklist (per-usulan di export Skor IJD
    # ini vs agregat per-kabupaten di Laporan Prioritas) konsisten. Sebelum
    # fix ini, SWASEMBADA_PANGAN_LOKUS SUDAH ikut kehitung di total_kriteria
    # (kriteria_cocok tidak difilter whitelist) tapi TIDAK PERNAH muncul di
    # checklist_lines (yang iterasi dict ini) -- bug tersembunyi, kriteria
    # nyumbang skor tapi tak kelihatan di centangan.
    "PERKEBUNAN": "Lokasi Kawasan Perkebunan (RPJMN)",
    "PERIKANAN": "Lokus Kelautan & Perikanan (RPJMN)",
    "SWASEMBADA_PANGAN_LOKUS": "Lokus Swasembada Pangan (RPJMN)",
}
_BAPPENAS_KAWASAN_TEMATIK_KATEGORI = ("PKPN", "TRANSMIGRASI", "KI_PRIORITAS", "PERKEBUNAN", "PERIKANAN")


def _bappenas_kode_kab(row: dict, ctx: dict = None) -> int:
    """kode_kabupaten usulan -- dipakai Aspek A & B. Prioritas wilayah_mapping
    (kabupaten_kota SITIA yg dideklarasikan pemda, otoritatif, "100%
    coverage" per build_wilayah_mapping.py) DI ATAS kode_kecamatan//1000
    (fallback) -- dibalik 23 Jul 2026 dari urutan semula. Alasan: semua 15+
    kriteria Aspek A itu granularitas Kab/Kota (lihat docstring
    _bappenas_aspek_a_lokus), jadi kode_kecamatan usulan tidak dibutuhkan
    utk resolusi kabupaten di sini -- dan JUSTRU berisiko: kalau
    kode_kecamatan usulan salah (hasil spatial-join yg keliru, mis. rute
    ke-assign ke kecamatan tetangga di kabupaten lain krn poligon BATAS
    KECAMATAN bermasalah di area itu), Aspek A/B ikut menghitung kabupaten
    yang SALAH, sampai kelihatan tercentang kriteria yg bukan milik
    kabupaten usulan itu (kasus nyata: usulan Kab. Nduga ID 241602 ter-
    assign kode_kecamatan Kec. Sawa Erma/Kab. Asmat, bikin "TOTAL CHECKLIST
    ASPEK A" di export Skor IJD keliru 6, padahal Nduga cuma 1 kriteria
    sesuai dashboard Laporan Daerah Prioritas yg SELALU pakai kode_kab asli,
    bukan lewat kode_kecamatan). ctx["kab_by_wilayah"] (sama dgn ctx IJD)
    dipakai kalau ada, supaya bulk export tidak query per baris."""
    if ctx and "kab_by_wilayah" in ctx:
        kode_kab = ctx["kab_by_wilayah"].get((row.get("provinsi"), row.get("kabupaten_kota")))
    else:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kabupaten FROM wilayah_mapping "
                "WHERE provinsi_sitia = %s AND kabupaten_kota_sitia = %s",
                (row.get("provinsi"), row.get("kabupaten_kota")),
            )
            r = cur.fetchone()
        kode_kab = r["kode_kabupaten"] if r else None
    if kode_kab:
        return kode_kab
    kode_kec = row.get("kode_kecamatan")
    return kode_kec // 1000 if kode_kec else None


def _bappenas_aspek_a_lokus(row: dict, ctx: dict = None) -> dict:
    """Cocokkan usulan ke kriteria lokus Aspek A. Return {"checklist": bool,
    "total_kriteria": int, "kriteria_cocok": [kode,...], "narasi": str}.
    ctx (opsional, dipakai bulk export) — "bappenas_lokus_by_kab"/
    "_by_prov": dict hasil batch dari bappenas_lokus_a; "kawasan_by_kab":
    dict hasil batch dari kawasan_tematik (sama dgn ctx IJD A3).

    Matching KABUPATEN/KOTA-level utk SEMUA kriteria Aspek A (bukan cuma yang
    levelnya='KABUPATEN' di tabel) -- diperbaiki 23 Jul 2026 setelah dicek
    ulang thd docs/docs/2_Analisis Prioritas untuk Bappenas dan Teknokratis
    15.7.2026.xlsx sheet "Kumpulan Data" baris 3-21 ("PENILAIAN LOKPRI"):
    KOLOM "Level Data" utk SELURUH kriteria Aspek A di situ berisi "Kab/Kota"
    / "Kab" / "Kabupaten" -- TIDAK SATU PUN "Kecamatan" -- dan
    docs/spec/Draf Penilaian Bappenas.md baris 5 eksplisit: "kalau Kab/Kota
    berarti itu kabupaten ... berdasarkan nama, jika ketemu skor 1". Sebelum
    perbaikan ini, baris bappenas_lokus_a yang level='KECAMATAN' (PERBATASAN,
    PKSN, KNMP, sebagian BBM_1_HARGA -- lihat import_bappenas_lokus_a.py,
    kode_kecamatan terisi krn resolusi nama lokasi sampai ke kecamatan) dan
    baris kawasan_tematik dgn kode_kecamatan terisi HANYA dihitung cocok
    kalau kode_kecamatan usulan PERSIS sama -- keliru, krn granularitas
    RESMI kriteria ini Kab/Kota, bukan kecamatan: usulan lain di kabupaten
    yang sama (kecamatan beda) semestinya tetap tercentang. Ini jugalah yang
    bikin Aspek A endpoint per-usulan (dipakai "TOTAL CHECKLIST ASPEK A" di
    export Skor IJD) sebelumnya BISA beda hasil dari "Laporan Daerah
    Prioritas" (LAPORAN_ASPEK_A) yang dari awal sudah kabupaten-lebar --
    sekarang keduanya konsisten satu semantik."""
    kode_kab = _bappenas_kode_kab(row, ctx)
    kode_prov = kode_kab // 100 if kode_kab else None

    kriteria_cocok = []
    if ctx and "bappenas_lokus_by_kab" in ctx:
        if kode_kab:
            for r in ctx["bappenas_lokus_by_kab"].get(kode_kab, []):
                if r["level"] in ("KABUPATEN", "KECAMATAN"):
                    kriteria_cocok.append(r["kriteria"])
        if kode_prov:
            kriteria_cocok += ctx["bappenas_lokus_by_prov"].get(kode_prov, [])
    elif kode_kab or kode_prov:
        with db_cursor() as cur:
            cur.execute(
                "SELECT DISTINCT kriteria FROM bappenas_lokus_a WHERE "
                "(level IN ('KABUPATEN','KECAMATAN') AND kode_kabupaten=%s) OR "
                "(level='PROVINSI' AND kode_provinsi=%s)",
                (kode_kab, kode_prov),
            )
            kriteria_cocok += [r["kriteria"] for r in cur.fetchall()]

    if kode_kab:
        if ctx and "kawasan_by_kab" in ctx:
            kriteria_cocok += [r["kategori"] for r in ctx["kawasan_by_kab"].get(kode_kab, [])
                                if r["kategori"] in _BAPPENAS_KAWASAN_TEMATIK_KATEGORI]
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT kategori FROM kawasan_tematik WHERE kategori = ANY(%s) "
                    "AND kode_kabupaten=%s",
                    (list(_BAPPENAS_KAWASAN_TEMATIK_KATEGORI), kode_kab),
                )
                kriteria_cocok += [r["kategori"] for r in cur.fetchall()]
    kriteria_cocok = sorted(set(kriteria_cocok))

    # Narasi Aspek A HANYA soal keanggotaan lokus prioritas nasional --
    # potensi produksi (bps_kecamatan_potensi_tematik) itu indikator EKONOMI,
    # sengaja cuma dipakai di Aspek B (_bappenas_aspek_b_ekonomi), bukan di sini.
    kriteria_label = [BAPPENAS_KRITERIA_LABEL.get(k, k) for k in kriteria_cocok]
    # Checklist eksplisit atas SELURUH kriteria di BAPPENAS_KRITERIA_LABEL
    # (bukan cuma yang cocok) -- "kolom AG jika salah satu ada maka check
    # list" (docs/spec/Draf Penilaian Bappenas.md) diminta ditampilkan
    # sebagai daftar centang/silang, plus keterangan naratif terpisah di
    # bawahnya.
    checklist_lines = [
        f"{'[v]' if kode in kriteria_cocok else '[ ]'} {label}"
        for kode, label in BAPPENAS_KRITERIA_LABEL.items()
    ]
    if kriteria_label:
        keterangan = (
            f"Usulan berada di lokasi yang termasuk {len(kriteria_label)} kriteria prioritas nasional: "
            + "; ".join(kriteria_label) + "."
        )
    else:
        keterangan = "Usulan tidak terindikasi berada di lokasi kriteria prioritas nasional manapun dari data yang tersedia."
    narasi = "\n".join(checklist_lines) + "\n\nKeterangan: " + keterangan

    return {
        "checklist": len(kriteria_cocok) > 0,
        "total_kriteria": len(kriteria_cocok),
        "kriteria_cocok": kriteria_cocok,
        "checklist_lines": checklist_lines,
        "keterangan": keterangan,
        "narasi": narasi,
    }


# --- Aspek B "Daya Ungkit Ekonomi & Kinerja Sektoral" (Bappenas) -- RULE-BASED
# juga, sama seperti Aspek A (docs/spec/Draf Penilaian Bappenas.md, minta baca
# baris 20-43 sheet Kumpulan Data, bagian "PENILAIAN RUAS"). 9 indikator yang
# sudah punya sumber data di aplikasi ini dicek; sisanya (Tata Guna Lahan,
# Indeks Penanaman, SHP Jaringan Jalan/Simpul Transportasi, KP2B/LP2B,
# Penuntasan Koridor OneDrive) belum ada sumber bersih -- tidak dicek. ---

BAPPENAS_ASPEK_B_INDIKATOR_LABEL = {
    "PENDUDUK": "Jumlah penduduk kecamatan (baris 20)",
    "SWASEMBADA_PANGAN_LOKUS": "Lokus Swasembada Pangan (baris 23)",
    "PERKEBUNAN": "Kawasan Perkebunan (baris 24)",
    "PERIKANAN": "Kawasan Kelautan & Perikanan (baris 22)",
    "PRODUKSI_PERTANIAN": "Produksi Padi/Jagung (baris 25)",
    "PRODUKSI_PERKEBUNAN": "Produksi Kelapa Sawit/Kelapa/Tebu/Karet (baris 26)",
    "PRODUKSI_PETERNAKAN": "Produktivitas Peternakan (baris 27)",
    "PRODUKSI_PERIKANAN": "Produktivitas Perikanan tangkap (baris 28)",
    "LBS": "Luas Lahan Baku Sawah (LBS) 2024 (baris 31)",
    "INDEKS_PENANAMAN": "Indeks Penanaman (IP) (baris 32)",
    "KI_PRIORITAS": "Konektivitas Kawasan Industri/KEK (baris 33-36)",
    "JARINGAN_JALAN": "Konektivitas Jaringan Jalan (baris 38)",
    "SIMPUL_TRANSPORTASI": "Konektivitas Simpul Transportasi (baris 39)",
    "KEMANTAPAN_JALAN": "Kemantapan Jalan IJD (baris 40)",
    "KENDARAAN": "Jumlah Kendaraan kecamatan (baris 41)",
    "KEBERLANJUTAN_IJD": "Keberlanjutan Kegiatan IJD (baris 42)",
    "PENUNTASAN_KORIDOR": "Penuntasan Koridor (baris 43)",
}


def _bappenas_aspek_b_ekonomi(row: dict, ctx: dict = None) -> dict:
    """12 indikator Daya Ungkit Ekonomi & Kinerja Sektoral yg sudah punya
    sumber data. Return {"checklist": bool, "total_indikator": int,
    "indikator_ada": [kode,...], "narasi": str}. ctx (opsional, bulk export)
    — reuse "kawasan_by_kab"/"bappenas_lokus_by_kab" dari Aspek A,
    "kepadatan_by_kec" (sama dgn ctx IJD C/A3 — kalau sudah di-extend
    dgn jumlah_penduduk/kendaraan_total/kolom bps_* di caller),
    "potensi_produksi_by_kec" (angka produksi riil dari
    bps_kecamatan_potensi_tematik, dipakai memperkonkret keterangan --
    checklist indikator PRODUKSI_* tetap dari flag ada/tidak di
    kepadatan_by_kec, bukan dari sini) dan "kemantapan_kab_set" (set
    kode_kab yg punya baris Kab./Kota)."""
    kode_kec = row.get("kode_kecamatan")
    kode_kab = _bappenas_kode_kab(row, ctx)

    indikator = []
    penduduk_n = kendaraan_n = None

    # baris 22/24/33-36: kawasan_tematik
    if kode_kab:
        if ctx and "kawasan_by_kab" in ctx:
            indikator += [r["kategori"] for r in ctx["kawasan_by_kab"].get(kode_kab, [])
                          if r["kategori"] in ("PERKEBUNAN", "PERIKANAN", "KI_PRIORITAS")
                          and (r["kode_kecamatan"] is None or r["kode_kecamatan"] == kode_kec)]
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT kategori FROM kawasan_tematik WHERE kategori IN "
                    "('PERKEBUNAN','PERIKANAN','KI_PRIORITAS') AND kode_kabupaten=%s "
                    "AND (kode_kecamatan IS NULL OR kode_kecamatan=%s)",
                    (kode_kab, kode_kec),
                )
                indikator += [r["kategori"] for r in cur.fetchall()]

    # baris 23: bappenas_lokus_a SWASEMBADA_PANGAN_LOKUS
    if kode_kab:
        if ctx and "bappenas_lokus_by_kab" in ctx:
            if any(r["kriteria"] == "SWASEMBADA_PANGAN_LOKUS" for r in ctx["bappenas_lokus_by_kab"].get(kode_kab, [])):
                indikator.append("SWASEMBADA_PANGAN_LOKUS")
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT 1 FROM bappenas_lokus_a WHERE kriteria='SWASEMBADA_PANGAN_LOKUS' "
                    "AND kode_kabupaten=%s LIMIT 1",
                    (kode_kab,),
                )
                if cur.fetchone():
                    indikator.append("SWASEMBADA_PANGAN_LOKUS")

    # baris 20/25-28/40: kecamatan_data_turunan + bps_kecamatan_potensi_tematik
    if kode_kec:
        if ctx and "kepadatan_by_kec" in ctx:
            kdt = ctx["kepadatan_by_kec"].get(kode_kec)
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT jumlah_penduduk, kendaraan_total, potensi_pertanian, potensi_perkebunan, "
                    "potensi_peternakan, potensi_perikanan FROM kecamatan_data_turunan "
                    "WHERE kode_kecamatan=%s ORDER BY tahun DESC LIMIT 1",
                    (kode_kec,),
                )
                kdt = cur.fetchone()
        if kdt:
            penduduk_n = kdt.get("jumlah_penduduk")
            kendaraan_n = kdt.get("kendaraan_total")
            if penduduk_n:
                indikator.append("PENDUDUK")
            if kendaraan_n:
                indikator.append("KENDARAAN")
            if kdt.get("potensi_pertanian"):
                indikator.append("PRODUKSI_PERTANIAN")
            if kdt.get("potensi_perkebunan"):
                indikator.append("PRODUKSI_PERKEBUNAN")
            if kdt.get("potensi_peternakan"):
                indikator.append("PRODUKSI_PETERNAKAN")
            if kdt.get("potensi_perikanan"):
                indikator.append("PRODUKSI_PERIKANAN")

    # Angka produksi riil (bukan sekadar flag ada/tidak di atas) --
    # bps_kecamatan_potensi_tematik, dipakai utk memperkonkret keterangan di
    # bawah (bukan komponen checklist tambahan; checklist tetap dari flag
    # potensi_* kecamatan_data_turunan supaya tidak dobel-hitung).
    produksi = None
    if kode_kec:
        if ctx and "potensi_produksi_by_kec" in ctx:
            produksi = ctx["potensi_produksi_by_kec"].get(kode_kec)
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT pertanian_produksi_ton, perkebunan_produksi_ton, "
                    "peternakan_produksi_daging_kg, peternakan_produksi_telur_kg, "
                    "perikanan_produksi_ton FROM bps_kecamatan_potensi_tematik "
                    "WHERE kode_kecamatan=%s ORDER BY tahun DESC LIMIT 1",
                    (kode_kec,),
                )
                produksi = cur.fetchone()

    # baris 39: kemantapan_ijd_2026 (level kab/kota)
    if kode_kab:
        if ctx and "kemantapan_kab_set" in ctx:
            if kode_kab in ctx["kemantapan_kab_set"]:
                indikator.append("KEMANTAPAN_JALAN")
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT mantap_pct FROM kemantapan_ijd_2026 WHERE kode_wilayah=%s "
                    "AND jenis_adm IN ('Kab.','Kota') LIMIT 1",
                    (kode_kab,),
                )
                if cur.fetchone():
                    indikator.append("KEMANTAPAN_JALAN")

    # baris 31/32: bps_kabupaten_indeks_penanaman (LBS & IP, level kabupaten)
    # -- sumber SAMA dgn _npr_build_lbs/_npr_build_ip (NPR Skor Intensitas),
    # ditambah 23 Jul 2026 setelah dicek Kumpulan Data baris 24-43 ("PENILAIAN
    # RUAS") ternyata belum semua ke-checklist di sini.
    if kode_kab:
        if ctx and "indeks_penanaman_by_kab" in ctx:
            if kode_kab in ctx["indeks_penanaman_by_kab"]:
                indikator.append("INDEKS_PENANAMAN")
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT 1 FROM bps_kabupaten_indeks_penanaman WHERE kode_kab=%s "
                    "AND indeks_penanaman_pct IS NOT NULL LIMIT 1", (f"{kode_kab:04d}",),
                )
                if cur.fetchone():
                    indikator.append("INDEKS_PENANAMAN")
        if ctx and "lbs_by_kab" in ctx:
            if kode_kab in ctx["lbs_by_kab"]:
                indikator.append("LBS")
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT 1 FROM bps_kabupaten_indeks_penanaman WHERE kode_kab=%s "
                    "AND lahan_baku_sawah_ha IS NOT NULL AND lahan_baku_sawah_ha > 0 LIMIT 1", (f"{kode_kab:04d}",),
                )
                if cur.fetchone():
                    indikator.append("LBS")

    # baris 38/39: konektivitas jaringan jalan & simpul transportasi -- sumber
    # SAMA dgn NPR Skor Cakupan (_npr_skor_cakupan), ctx-nya di sini reuse
    # dari _npr_bulk_ctx (dioper caller lewat ctx["jaringan_jalan_by_kab"]/
    # "simpul_by_kab"/"spasial_terhubung_by_usulan"), ditambah 23 Jul 2026.
    if kode_kab:
        if ctx and "jaringan_jalan_by_kab" in ctx:
            jj = ctx["jaringan_jalan_by_kab"].get(kode_kab)
            jalan_daerah = bool(jj and jj.get("ada_jalan_daerah"))
            jalan_nasional = bool(jj and jj.get("ada_jalan_nasional"))
            jalan_spasial = bool(ctx.get("spasial_terhubung_by_usulan", {}).get(row.get("id"), False))
        else:
            with db_cursor() as cur:
                cur.execute(
                    "SELECT ada_jalan_daerah, ada_jalan_nasional FROM konektivitas_jaringan_jalan "
                    "WHERE kode_kabupaten=%s", (kode_kab,),
                )
                jj = cur.fetchone()
            jalan_daerah = bool(jj and jj["ada_jalan_daerah"])
            jalan_nasional = bool(jj and jj["ada_jalan_nasional"])
            jalan_spasial = False
            if row.get("id"):
                with db_cursor() as cur:
                    cur.execute("SELECT terhubung FROM usulan_konektivitas_jalan WHERE usulan_id=%s", (row["id"],))
                    r = cur.fetchone()
                    jalan_spasial = bool(r["terhubung"]) if r else False
        if jalan_daerah or jalan_nasional or jalan_spasial:
            indikator.append("JARINGAN_JALAN")

        if ctx and "simpul_by_kab" in ctx:
            simpul_jenis = ctx["simpul_by_kab"].get(kode_kab, set())
        else:
            with db_cursor() as cur:
                cur.execute("SELECT DISTINCT jenis FROM simpul_transportasi WHERE kode_kabupaten=%s", (kode_kab,))
                simpul_jenis = {r["jenis"] for r in cur.fetchall()}
        if simpul_jenis & {"PELABUHAN_NASIONAL", "PELABUHAN_PENYEBERANGAN"}:
            indikator.append("SIMPUL_TRANSPORTASI")

    # baris 42: keberlanjutan IJD (flag resmi kompetensi atau pencocokan DPP 2025)
    if (row.get("penuntasan_ijd_kompetensi") or "").strip().upper() == "YA" or row.get("lanjutan_ijd_2025"):
        indikator.append("KEBERLANJUTAN_IJD")

    # baris 43: penuntasan koridor -- row-level, sama sumber dgn NPR Skor
    # Cakupan PENUNTASAN_KORIDOR (biner murni dari kode_koridor terisi).
    if (row.get("kode_koridor") or "").strip():
        indikator.append("PENUNTASAN_KORIDOR")

    indikator = sorted(set(indikator))
    indikator_label = [BAPPENAS_ASPEK_B_INDIKATOR_LABEL.get(k, k) for k in indikator]
    # Checklist eksplisit atas SELURUH 12 indikator (bukan cuma yang ada),
    # sama pola dengan Aspek A -- lihat _bappenas_aspek_a_lokus.
    checklist_lines = [
        f"{'[v]' if kode in indikator else '[ ]'} {label}"
        for kode, label in BAPPENAS_ASPEK_B_INDIKATOR_LABEL.items()
    ]

    keterangan_parts = []
    if indikator_label:
        keterangan_parts.append(
            f"Ditemukan {len(indikator_label)} indikator daya ungkit ekonomi/sektoral yang didukung data: "
            + "; ".join(indikator_label) + "."
        )
    else:
        keterangan_parts.append("Belum ada indikator daya ungkit ekonomi/sektoral yang didukung data untuk lokasi usulan ini.")
    if penduduk_n:
        keterangan_parts.append(f"Jumlah penduduk kecamatan: {penduduk_n:,} jiwa.".replace(",", "."))
    if kendaraan_n:
        keterangan_parts.append(f"Jumlah kendaraan kecamatan: {kendaraan_n:,} unit.".replace(",", "."))
    if produksi:
        def _fmt(n):
            return f"{n:,.0f}".replace(",", ".")
        if produksi.get("pertanian_produksi_ton"):
            keterangan_parts.append(f"Produksi padi & jagung kecamatan: {_fmt(produksi['pertanian_produksi_ton'])} ton.")
        if produksi.get("perkebunan_produksi_ton"):
            keterangan_parts.append(
                f"Produksi perkebunan kecamatan (agregat seluruh jenis tanaman perkebunan rakyat): "
                f"{_fmt(produksi['perkebunan_produksi_ton'])} ton."
            )
        if produksi.get("peternakan_produksi_daging_kg"):
            keterangan_parts.append(f"Produksi daging ternak kecamatan: {_fmt(produksi['peternakan_produksi_daging_kg'])} kg.")
        if produksi.get("peternakan_produksi_telur_kg"):
            keterangan_parts.append(f"Produksi telur kecamatan: {_fmt(produksi['peternakan_produksi_telur_kg'])} kg.")
        if produksi.get("perikanan_produksi_ton"):
            keterangan_parts.append(f"Produksi perikanan tangkap kecamatan: {_fmt(produksi['perikanan_produksi_ton'])} ton.")
    keterangan = " ".join(keterangan_parts)
    narasi = "\n".join(checklist_lines) + "\n\nKeterangan: " + keterangan

    return {
        "checklist": len(indikator) > 0,
        "total_indikator": len(indikator),
        "indikator_ada": indikator,
        "checklist_lines": checklist_lines,
        "keterangan": keterangan,
        "narasi": narasi,
    }


# --- NPR (Nilai Prioritas Ruas) -- metodologi ALTERNATIF, sumber
# UTAMA sheet "Pembobotan Ruas" di
# `docs/docs/2_Analisis Prioritas untuk Bappenas dan Teknokratis
# 15.7.2026.xlsx` (bobot & daftar item RESMI, ditemukan 21 Jul 2026 --
# sheet ini baru muncul di file, sebelumnya cuma ada contoh ilustratif di
# `Metodologi Penilaian skala prioritas Ruas.docx` yang jadi basis draf
# implementasi PERTAMA -- SUDAH DIROMBAK mengikuti sheet resmi ini, lihat
# docs/kajian_metodologi_skala_prioritas_ruas.md). BELUM policy final --
# endpoint TERPISAH dari /ijd-score (A-E), tidak menggantikannya. Bobot
# disimpan sbg konstanta Python (bukan tabel rules-as-data spt
# ijd_scoring_rules) krn metodologi ini belum diadopsi -- migrasi ke tabel
# kalau/ketika resmi.
#
# SI (Skor Intensitas, bobot 70% NPR): 9 parameter, skala 0-100 per
# parameter via KELAS NASIONAL DINAMIS 4-tingkat (100/75/50/25 -- sheet
# resmi mencatat "Dibagi menjadi 4 kelas" utk item pertama tiap seksi,
# diterapkan seragam ke semua parameter numerik SI, BEDA dari draf awal yg
# pakai 5-kuintil ilustratif docx). SEMUA 9 parameter kini AKTIF (Padi&
# Jagung, Peternakan, Perikanan yg sebelumnya "belum tersedia" di draf
# pertama, ternyata memang bagian resmi & datanya sudah ada di
# bps_kecamatan_potensi_tematik).
#
# SC (Skor Cakupan, bobot 30% NPR): 7 kategori (2 tunggal + 4 sub KI/KEK +
# 2 dari eksistensi row-level usulan) -- SEMUA reuse data yang SUDAH ada
# (kawasan_tematik/bappenas_lokus_a/konektivitas_jaringan_jalan/
# simpul_transportasi/status_koridor_balai/lanjutan_ijd_2025), berbeda
# total dari draf pertama yang pakai kategori ilustratif (Bandara/KEK/
# Pariwisata/PSN-umum) yang sebagian besar genuinely tak ada sumbernya.

NPR_BOBOT_SI_SC = {"SI": 0.7, "SC": 0.3}

_npr_kelas_cache: dict = {}  # {cache_key: (value_map, (vmin,vmax))} -- sekali per proses, restart server utk refresh (pola sama dgn _map_layer_geojson_cache)


def _npr_kelas_dinamis(cache_key: str, kode, builder) -> tuple:
    """4-kelas (100/75/50/25) via RENTANG MIN-MAX linier -- BUKAN kuantil
    distribusi (metode lama, diganti 21 Jul 2026 setelah user tunjukkan
    contoh konkret di kolom J sheet "Pembobotan Ruas": lebar_kelas =
    (max-min)/4, lalu 4 rentang non-overlapping [min, min+lebar],
    (min+lebar, min+2*lebar], (min+2*lebar, min+3*lebar], (min+3*lebar,
    max] -> skor 25/50/75/100. Contoh sheet: min=10.000, max=1.500.000
    (asumsi dari tabel usulan_inpres) -> lebar=372.500; nilai 250.000 jatuh
    di rentang pertama -> skor 25. BEDA dari kuantil (yang bikin ~jumlah
    kabupaten sama tiap kelas, tapi lebar rentang nilai bisa timpang) --
    min-max bikin lebar rentang SAMA tiap kelas, tapi jumlah kabupaten per
    kelas bisa timpang kalau distribusinya skewed (wajar utk data begini,
    apa adanya sesuai kerangka sheet, bukan bug).

    `builder` (callable tanpa argumen) dipanggil SEKALI saja per cache_key
    lalu di-cache modul-level. Return (nilai_mentah, skor) atau (None,
    None) kalau `kode` tidak py data atau titik data nasionalnya <2 (min==
    max, rentang tidak bisa dibagi)."""
    if cache_key not in _npr_kelas_cache:
        value_map = builder()
        if len(value_map) >= 2:
            vmin, vmax = min(value_map.values()), max(value_map.values())
            if vmax <= vmin:
                vmin = vmax = None
        else:
            vmin = vmax = None
        _npr_kelas_cache[cache_key] = (value_map, (vmin, vmax))
    value_map, (vmin, vmax) = _npr_kelas_cache[cache_key]
    nilai = value_map.get(kode)
    if nilai is None or vmin is None:
        return None, None
    lebar = (vmax - vmin) / 4
    if nilai <= vmin + lebar:
        skor = 25
    elif nilai <= vmin + 2 * lebar:
        skor = 50
    elif nilai <= vmin + 3 * lebar:
        skor = 75
    else:
        skor = 100
    return nilai, skor


def _npr_build_penduduk_kecamatan():
    with db_cursor() as cur:
        cur.execute("SELECT kode_kecamatan, jumlah_penduduk FROM penduduk_kecamatan WHERE jumlah_penduduk IS NOT NULL")
        return {r["kode_kecamatan"]: float(r["jumlah_penduduk"]) for r in cur.fetchall()}


def _npr_build_penduduk_kabupaten():
    # Varian KABUPATEN (total penduduk seluruh kecamatan) dari
    # _npr_build_penduduk_kecamatan -- dipakai Aspek B Laporan Prioritas
    # (agregat per kabupaten, beda basis dari NPR yang per-usulan/
    # kecamatan) supaya "Jumlah Penduduk" kabupaten py makna yg jelas
    # (total, bukan penduduk 1 kecamatan usulan tertentu).
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_kabupaten, SUM(jumlah_penduduk) v FROM penduduk_kecamatan "
            "WHERE jumlah_penduduk IS NOT NULL GROUP BY kode_kabupaten"
        )
        return {int(r["kode_kabupaten"]): float(r["v"]) for r in cur.fetchall()}


def _npr_build_kab_sum(kolom):
    """Factory: sum kolom produksi (bps_kecamatan_potensi_tematik) per
    kabupaten -- dipakai Padi&Jagung (pertanian_produksi_ton) dan
    Perkebunan (perkebunan_produksi_ton)."""
    def builder():
        with db_cursor() as cur:
            cur.execute(
                f"SELECT kode_kab, SUM({kolom}) v FROM bps_kecamatan_potensi_tematik "
                f"WHERE {kolom} IS NOT NULL GROUP BY kode_kab HAVING SUM({kolom}) > 0"
            )
            return {int(r["kode_kab"]): float(r["v"]) for r in cur.fetchall()}
    return builder


def _npr_build_peternakan():
    # Produktivitas Peternakan (Sapi/unggas/kambing/telur) -- daging+telur
    # digabung jadi satu angka kg (sumbernya sendiri sudah gabungan per
    # jenis produk, bukan per jenis hewan, lihat schema_bps_potensi_tematik.sql).
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_kab, SUM(COALESCE(peternakan_produksi_daging_kg,0) + "
            "COALESCE(peternakan_produksi_telur_kg,0)) v FROM bps_kecamatan_potensi_tematik "
            "WHERE peternakan_produksi_daging_kg IS NOT NULL OR peternakan_produksi_telur_kg IS NOT NULL "
            "GROUP BY kode_kab HAVING SUM(COALESCE(peternakan_produksi_daging_kg,0) + "
            "COALESCE(peternakan_produksi_telur_kg,0)) > 0"
        )
        return {int(r["kode_kab"]): float(r["v"]) for r in cur.fetchall()}


def _npr_build_padi_jagung_provinsi():
    """Fallback PROVINSI utk NPR PADI_JAGUNG (bps_kecamatan_potensi_tematik
    tidak py baris utk kabupaten usulan) -- si_padi_jagung_provinsi, dari
    BPS "Statistik Indonesia 2026" Tabel 5.1.2 (Padi) + 5.1.4 (Jagung),
    diisi scripts/extract_statistik_indonesia.py. Kode 0 ("Indonesia" total)
    dibuang -- bukan wilayah asli."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_provinsi, total_ton FROM si_padi_jagung_provinsi "
            "WHERE total_ton IS NOT NULL AND kode_provinsi > 0"
        )
        return {int(r["kode_provinsi"]): float(r["total_ton"]) for r in cur.fetchall()}


def _npr_build_perikanan_provinsi():
    """Fallback PROVINSI utk NPR PERIKANAN -- si_perikanan_tangkap_provinsi
    (Statistik Indonesia 2026 Tabel 5.6.1, "Tangkap Darat & Laut" -- sama
    cakupan dgn label parameter ini)."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_provinsi, volume_ton FROM si_perikanan_tangkap_provinsi "
            "WHERE volume_ton IS NOT NULL AND kode_provinsi > 0"
        )
        return {int(r["kode_provinsi"]): float(r["volume_ton"]) for r in cur.fetchall()}


# Fallback PROVINSI dipakai _npr_skor_intensitas() saat builder kabupaten
# tidak punya nilai utk kode_kab usulan -- lihat docs/MEMORY.md soal kenapa
# fallback ini TIDAK ikut dicampur ke dict builder kabupaten (min-max
# nasionalnya beda skala, provinsi vs kabupaten). PETERNAKAN sengaja tidak
# ada entri di sini (belum diekstrak, lihat schema_si_produksi_provinsi.sql).
_NPR_SI_PROVINSI_FALLBACK = {
    "PADI_JAGUNG": _npr_build_padi_jagung_provinsi,
    "PERIKANAN": _npr_build_perikanan_provinsi,
}


def _npr_build_ip():
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_kab, indeks_penanaman_pct FROM bps_kabupaten_indeks_penanaman "
            "WHERE indeks_penanaman_pct IS NOT NULL ORDER BY tahun DESC"
        )
        out = {}
        for r in cur.fetchall():
            out.setdefault(int(r["kode_kab"]), float(r["indeks_penanaman_pct"]))  # tahun terbaru menang
        return out


def _npr_build_lbs():
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_kab, lahan_baku_sawah_ha FROM bps_kabupaten_indeks_penanaman "
            "WHERE lahan_baku_sawah_ha IS NOT NULL AND lahan_baku_sawah_ha > 0 ORDER BY tahun DESC"
        )
        out = {}
        for r in cur.fetchall():
            out.setdefault(int(r["kode_kab"]), float(r["lahan_baku_sawah_ha"]))
        return out


def _npr_build_kemantapan():
    with db_cursor() as cur:
        cur.execute(
            "SELECT kode_wilayah, mantap_pct FROM kemantapan_ijd_2026 "
            "WHERE jenis_adm IN ('Kab.','Kota') AND mantap_pct IS NOT NULL"
        )
        return {int(r["kode_wilayah"]): float(r["mantap_pct"]) for r in cur.fetchall()}


def _npr_build_kendaraan():
    with db_cursor() as cur:
        cur.execute("SELECT kode_kab, jumlah FROM bps_kabupaten_kendaraan WHERE jumlah IS NOT NULL ORDER BY tahun DESC")
        out = {}
        for r in cur.fetchall():
            out.setdefault(int(r["kode_kab"]), float(r["jumlah"]))
        return out


# (kode, label, bobot_maks, basis "kecamatan"/"kabupaten", builder, satuan)
# -- bobot PERSIS sheet "Pembobotan Ruas" (jumlah = 100, skala internal SI).
_NPR_SI_DEF = [
    ("PENDUDUK", "Jumlah Penduduk", 5, "kecamatan", _npr_build_penduduk_kecamatan, "jiwa"),
    ("PADI_JAGUNG", "Produksi Padi & Jagung", 15, "kabupaten", _npr_build_kab_sum("pertanian_produksi_ton"), "ton"),
    ("PERKEBUNAN", "Produksi Kelapa Sawit/Kelapa/Tebu/Karet", 15, "kabupaten", _npr_build_kab_sum("perkebunan_produksi_ton"), "ton"),
    ("PETERNAKAN", "Produktivitas Peternakan (Sapi/Unggas/Kambing, dst.)", 15, "kabupaten", _npr_build_peternakan, "kg"),
    ("PERIKANAN", "Produktivitas Perikanan (Tangkap Darat & Laut)", 15, "kabupaten", _npr_build_kab_sum("perikanan_produksi_ton"), "ton"),
    ("IP", "Indeks Penanaman (IP)", 5, "kabupaten", _npr_build_ip, "%"),
    ("KEMANTAPAN_JALAN", "Kemantapan Jalan (IJD)", 10, "kabupaten", _npr_build_kemantapan, "%"),
    ("JUMLAH_KENDARAAN", "Jumlah Kendaraan", 10, "kabupaten", _npr_build_kendaraan, "unit"),
    ("LBS", "Luas Lahan Baku Sawah (LBS) 2024", 10, "kabupaten", _npr_build_lbs, "ha"),
]


def _npr_skor_intensitas(row: dict, kode_kab: int) -> dict:
    komponen = []
    skor_tertimbang = 0.0
    bobot_tersedia = 0.0
    kode_kec = row.get("kode_kecamatan")
    for kode, label, bobot, basis, builder, satuan in _NPR_SI_DEF:
        target = kode_kec if basis == "kecamatan" else kode_kab
        nilai_raw, skor = (_npr_kelas_dinamis(kode, target, builder) if target else (None, None))
        proksi_provinsi = False
        if skor is None and kode in _NPR_SI_PROVINSI_FALLBACK and kode_kab:
            # Fallback PROVINSI (Statistik Indonesia 2026) saat kabupaten
            # usulan tidak punya baris di bps_kecamatan_potensi_tematik --
            # cache_key & min-max TERPISAH dari basis kabupaten di atas
            # (skala provinsi jauh lebih besar, tidak boleh dicampur ke
            # rentang min-max nasional kabupaten, lihat docs/MEMORY.md).
            nilai_raw, skor = _npr_kelas_dinamis(
                f"{kode}_PROV", kode_kab // 100, _NPR_SI_PROVINSI_FALLBACK[kode])
            proksi_provinsi = skor is not None
        if skor is not None:
            keterangan = (f"{label} {basis} ini: {nilai_raw:,.1f} {satuan} "
                          "(kelas nasional dinamis, 4 kelas: 100/75/50/25).").replace(",", ".")
            if proksi_provinsi:
                keterangan = (f"{label} PROKSI PROVINSI (kabupaten ini tidak punya data): "
                              f"{nilai_raw:,.1f} {satuan} (kelas nasional provinsi, "
                              "4 kelas: 100/75/50/25).").replace(",", ".")
            komponen.append({
                "kode": kode, "label": label, "bobot_maks": bobot, "tersedia": True,
                "nilai": skor, "kontribusi": round(skor / 100 * bobot, 2),
                "keterangan": keterangan,
            })
            skor_tertimbang += skor / 100 * bobot
            bobot_tersedia += bobot
        else:
            komponen.append({"kode": kode, "label": label, "bobot_maks": bobot, "tersedia": False,
                              "keterangan": f"Data {label.lower()} tidak ditemukan untuk {basis} ini."})
    # skor_ternormalisasi_100 = skor_tertimbang APA ADANYA (bukan direnormalisasi
    # thd bobot_tersedia lagi -- perubahan 22 Jul 2026 atas permintaan eksplisit
    # user). Karena total bobot_maks seluruh 9 indikator sudah = 100, indikator
    # yang datanya tidak tersedia kini efektif MENYUMBANG 0 ke skor, bukan
    # dikecualikan dari pembagi -- BEDA dari pola "normalisasi thd bobot
    # tersedia" yang masih dipakai skor IJD A-E (_ijd_score_bulk_rows dkk.,
    # lihat CLAUDE.md) dan sebelumnya juga dipakai di sini.
    skor_100 = round(skor_tertimbang, 1) if bobot_tersedia else None
    return {"komponen": komponen, "skor_tertimbang": round(skor_tertimbang, 2),
            "bobot_tersedia": bobot_tersedia, "skor_ternormalisasi_100": skor_100}


# sumber_sheet asli (kawasan_tematik) -> sub-kategori KI/KEK -- sama pola
# dgn _KI_SHEET_TO_KODE di seksi Laporan Prioritas (duplikasi sengaja,
# scorer ini self-contained -- lihat konvensi di CLAUDE.md/skill
# ijd-scoring-parameter).
_NPR_KI_SHEET_TO_KODE = {
    "Lokus KI PSN IUKI Sudah Terbit": "KI_PSN_IUKI",
    "Lokus PKPN KI Prioritas RPJMN": "KI_PRIO_RPJMN",
    "Lokus PKPN KI Hilirisasi": "KI_HILIRISASI",
    "Lokus PKPN KI Dirgantara": "KI_DIRGANTARA",
}

# (kode, label, bobot_maks) -- bobot PERSIS sheet "Pembobotan Ruas" (jumlah
# = 100, skala internal SC). KEBERLANJUTAN_IJD/PENUNTASAN_KORIDOR ditangani
# khusus di _npr_skor_cakupan (row-level, bukan lookup kabupaten).
_NPR_SC_DEF = [
    ("KNMP_PERIKANAN", "Lokus Kelautan & Perikanan (KNMP)", 20),
    ("SWASEMBADA_PANGAN", "Lokus Swasembada Pangan", 20),
    ("KI_PSN_IUKI", "Konektivitas KI/KEK - KI PSN IUKI", 3),
    ("KI_PRIO_RPJMN", "Konektivitas KI/KEK - KI Prio RPJMN", 3),
    ("KI_HILIRISASI", "Konektivitas KI/KEK - KI Hilirisasi", 2),
    ("KI_DIRGANTARA", "Konektivitas KI/KEK - KI Dirgantara", 2),
    ("JARINGAN_JALAN", "Konektivitas Jaringan Jalan", 15),
    ("SIMPUL_TRANSPORTASI", "Konektivitas Simpul Transportasi", 10),
    ("KEBERLANJUTAN_IJD", "Keberlanjutan IJD", 10),
    ("PENUNTASAN_KORIDOR", "Penuntasan Koridor", 15),
]


def _npr_kelas_dari_hitung(matched: int, total: int) -> int:
    """4-kelas (100/75/50/25) dari rasio "jumlah sinyal cocok / total sinyal
    yg didefinisikan utk kategori itu" -- request user 21 Jul 2026: catatan
    sel E10/E22 sheet "Pembobotan Ruas" ("Dibagi menjadi 4 kelas") adalah
    MERGED CELL yg mencakup SELURUH baris di kedua tabel (SI & SC), bukan
    cuma baris pertama -- jadi Skor Cakupan JUGA harus 4-kelas, bukan biner
    ada/tidak spt draf sebelumnya. Kategori dgn total=1 (tak py sub-sinyal
    terpisah) otomatis cuma 2 nilai efektif (100/0), krn ratio cuma bisa 0
    atau 1 -- itu bawaan strukturnya, bukan bug."""
    if total <= 0:
        return 0
    ratio = matched / total
    if ratio >= 1:
        return 100
    if ratio >= 0.5:
        return 75
    if ratio > 0:
        return 50
    return 0


def _npr_skor_cakupan(row: dict, kode_kab: int, ctx: dict = None) -> dict:
    # Sinyal MENTAH (bukan hasil OR/union langsung) per kategori, supaya
    # bisa dihitung jumlah cocok utk _npr_kelas_dari_hitung -- BEDA dari
    # draf sebelumnya yang langsung OR-kan jadi satu boolean "ada".
    kawasan_kategori, ki_sheet_set, lokus_kriteria = set(), set(), set()
    jalan_daerah = jalan_nasional = jalan_spasial = False
    simpul_jenis = set()
    if kode_kab:
        if ctx and "kawasan_by_kab" in ctx:
            kawasan_rows = ctx["kawasan_by_kab"].get(kode_kab, [])
            lokus_kriteria = set(ctx["lokus_by_kab"].get(kode_kab, ())) | set(ctx["lokus_by_prov"].get(kode_kab // 100, ()))
            jj = ctx["jaringan_jalan_by_kab"].get(kode_kab)
            if jj:
                jalan_daerah, jalan_nasional = bool(jj.get("ada_jalan_daerah")), bool(jj.get("ada_jalan_nasional"))
            simpul_jenis = ctx["simpul_by_kab"].get(kode_kab, set())
            jalan_spasial = ctx["spasial_terhubung_by_usulan"].get(row.get("id"), False)
        else:
            with db_cursor() as cur:
                cur.execute("SELECT DISTINCT kategori, sumber_sheet FROM kawasan_tematik WHERE kode_kabupaten = %s", (kode_kab,))
                kawasan_rows = cur.fetchall()
                # level='PROVINSI' saja yang boleh match by kode_provinsi --
                # baris KECAMATAN/KABUPATEN kadang ikut menyimpan kode_provinsi
                # sbg field denormalisasi (bukan penanda "berlaku se-provinsi"),
                # OR polos di sini sempat salah cocok ke kabupaten lain dlm
                # provinsi yang sama (bug ditemukan 21 Jul 2026 lewat crosscheck
                # thd hasil ctx-batched _npr_bulk_ctx yang levelnya sudah benar).
                cur.execute(
                    "SELECT DISTINCT kriteria FROM bappenas_lokus_a WHERE kode_kabupaten = %s "
                    "OR (kode_provinsi = %s AND level = 'PROVINSI')", (kode_kab, kode_kab // 100),
                )
                lokus_kriteria = {r["kriteria"] for r in cur.fetchall()}
                cur.execute(
                    "SELECT ada_jalan_daerah, ada_jalan_nasional FROM konektivitas_jaringan_jalan "
                    "WHERE kode_kabupaten = %s", (kode_kab,),
                )
                jj = cur.fetchone()
                if jj:
                    jalan_daerah, jalan_nasional = bool(jj["ada_jalan_daerah"]), bool(jj["ada_jalan_nasional"])
                cur.execute("SELECT DISTINCT jenis FROM simpul_transportasi WHERE kode_kabupaten = %s", (kode_kab,))
                simpul_jenis = {r["jenis"] for r in cur.fetchall()}
                if row.get("id"):
                    cur.execute("SELECT terhubung FROM usulan_konektivitas_jalan WHERE usulan_id = %s", (row["id"],))
                    r = cur.fetchone()
                    jalan_spasial = bool(r["terhubung"]) if r else False
        kawasan_kategori = {r["kategori"] for r in kawasan_rows}
        ki_sheet_set = {_NPR_KI_SHEET_TO_KODE[r["sumber_sheet"]] for r in kawasan_rows
                        if r["sumber_sheet"] in _NPR_KI_SHEET_TO_KODE}

    komponen = []
    skor_tertimbang = 0.0
    bobot_tersedia = 0.0

    def _tambah(kode, label, bobot, tersedia, nilai=None, keterangan=""):
        nonlocal skor_tertimbang, bobot_tersedia
        entry = {"kode": kode, "label": label, "bobot_maks": bobot, "tersedia": tersedia, "keterangan": keterangan}
        if tersedia:
            entry["nilai"] = nilai
            entry["kontribusi"] = round(nilai / 100 * bobot, 2)
            skor_tertimbang += nilai / 100 * bobot
            bobot_tersedia += bobot
        komponen.append(entry)

    for kode, label, bobot in _NPR_SC_DEF:
        if kode in ("KI_PSN_IUKI", "KI_PRIO_RPJMN", "KI_HILIRISASI", "KI_DIRGANTARA"):
            # Sub KI/KEK: masing2 SUDAH jadi baris bobot tersendiri (bukan
            # sub-sinyal 1 kategori gabungan) -- tetap biner per baris, "4
            # kelas" tidak relevan di level ini (satu lokus flag, tak py
            # gradasi lagi di bawahnya).
            cocok = kode in ki_sheet_set
            _tambah(kode, label, bobot, True, 100 if cocok else 0,
                    f"{'Ada' if cocok else 'Tidak ada'} data {label.lower()} di kabupaten/kota ini.")
            continue
        if kode == "KNMP_PERIKANAN":
            # 2 sinyal: lokus KNMP (bappenas_lokus_a) & kawasan PERIKANAN (kawasan_tematik).
            matched = int("KNMP" in lokus_kriteria) + int("PERIKANAN" in kawasan_kategori)
            nilai = _npr_kelas_dari_hitung(matched, 2)
            _tambah(kode, label, bobot, True, nilai,
                    f"{matched}/2 sinyal cocok (Lokus KNMP, Kawasan Perikanan) -> kelas {nilai}.")
            continue
        if kode == "SWASEMBADA_PANGAN":
            matched = int("SWASEMBADA_PANGAN_LOKUS" in lokus_kriteria) + int("SWASEMBADA_PANGAN_RPJMN" in lokus_kriteria)
            nilai = _npr_kelas_dari_hitung(matched, 2)
            _tambah(kode, label, bobot, True, nilai,
                    f"{matched}/2 sinyal cocok (Lokus Swasembada Pangan, RPJMN) -> kelas {nilai}.")
            continue
        if kode == "JARINGAN_JALAN":
            # 3 sinyal: jalan daerah terpetakan, dilewati jalan nasional
            # (keduanya administratif/kabupaten), validasi spasial KML
            # usulan INI sendiri (terkuat, per-usulan bukan per-kabupaten).
            matched = int(jalan_daerah) + int(jalan_nasional) + int(jalan_spasial)
            nilai = _npr_kelas_dari_hitung(matched, 3)
            _tambah(kode, label, bobot, True, nilai,
                    f"{matched}/3 sinyal cocok (jalan daerah, jalan nasional, validasi spasial KML) -> kelas {nilai}.")
            continue
        if kode == "SIMPUL_TRANSPORTASI":
            matched = len(simpul_jenis & {"PELABUHAN_NASIONAL", "PELABUHAN_PENYEBERANGAN"})
            nilai = _npr_kelas_dari_hitung(matched, 2)
            _tambah(kode, label, bobot, True, nilai,
                    f"{matched}/2 sinyal cocok (Pelabuhan Nasional, Pelabuhan Penyeberangan) -> kelas {nilai}.")
            continue
        if kode == "KEBERLANJUTAN_IJD":
            # Sumber SAMA dgn parameter E skor IJD A-E (_ijd_score_penuntasan)
            # -- row-level, 2 sinyal: flag resmi verifikasi kompetensi &
            # pencocokan DPP IJD 2025.
            flag_kompetensi = (row.get("penuntasan_ijd_kompetensi") or "").strip().upper() == "YA"
            flag_dpp = bool(row.get("lanjutan_ijd_2025"))
            if row.get("penuntasan_ijd_kompetensi") is None and row.get("lanjutan_ijd_2025") is None:
                _tambah(kode, label, bobot, False, keterangan=(
                    "Belum ada data verifikasi kompetensi atau pencocokan DPP IJD 2025 utk usulan ini."))
                continue
            matched = int(flag_kompetensi) + int(flag_dpp)
            nilai = _npr_kelas_dari_hitung(matched, 2)
            _tambah(kode, label, bobot, True, nilai,
                    f"{matched}/2 sinyal cocok (verifikasi kompetensi, pencocokan DPP 2025) -> kelas {nilai}.")
            continue
        if kode == "PENUNTASAN_KORIDOR":
            # Disederhanakan 21 Jul 2026 (request eksplisit user, koreksi dari
            # draf 2-sinyal sebelumnya): checklist biner murni dari
            # `kode_koridor` -- terisi (bukan null/kosong) = koridor
            # teridentifikasi. status_koridor_balai TIDAK dipakai lagi di
            # sini (itu tetap dipakai parameter D skor IJD A-E,
            # _ijd_score_koridor, yang beda skema/tujuan).
            kode_koridor = (row.get("kode_koridor") or "").strip()
            nilai = 100 if kode_koridor else 0
            _tambah(kode, label, bobot, True, nilai,
                    f"Kode koridor {'teridentifikasi (' + kode_koridor + ')' if kode_koridor else 'tidak ada'}.")
            continue

    # skor_ternormalisasi_100 = skor_tertimbang APA ADANYA, sama alasan/tanggal
    # perubahan dgn _npr_skor_intensitas di atas (total bobot_maks 10 kategori
    # SC juga = 100).
    skor_100 = round(skor_tertimbang, 1) if bobot_tersedia else None
    return {"komponen": komponen, "skor_tertimbang": round(skor_tertimbang, 2),
            "bobot_tersedia": bobot_tersedia, "skor_ternormalisasi_100": skor_100}


_NPR_KATEGORI = [(80, "Prioritas Sangat Tinggi"), (70, "Prioritas Tinggi"), (60, "Prioritas Menengah"),
                  (50, "Prioritas Rendah")]


def _compute_npr(row: dict, ctx: dict = None) -> dict:
    kode_kab = _bappenas_kode_kab(row, ctx)
    si = _npr_skor_intensitas(row, kode_kab)
    sc = _npr_skor_cakupan(row, kode_kab, ctx)

    npr = None
    kategori = None
    if si["skor_ternormalisasi_100"] is not None and sc["skor_ternormalisasi_100"] is not None:
        npr = round(NPR_BOBOT_SI_SC["SI"] * si["skor_ternormalisasi_100"] +
                     NPR_BOBOT_SI_SC["SC"] * sc["skor_ternormalisasi_100"], 1)
        kategori = "Belum Prioritas"
        for ambang, label in _NPR_KATEGORI:
            if npr >= ambang:
                kategori = label
                break

    return {
        "skor_intensitas": si,
        "skor_cakupan": sc,
        "bobot_si_sc": NPR_BOBOT_SI_SC,
        "npr": npr,
        "kategori": kategori,
        "catatan": (
            "NPR (Nilai Prioritas Ruas) -- metodologi ALTERNATIF, bobot & daftar item dari sheet "
            "\"Pembobotan Ruas\" (2_Analisis Prioritas untuk Bappenas dan Teknokratis "
            "15.7.2026.xlsx), BELUM policy final, TERPISAH dari skor IJD Prioritisasi Teknokratik "
            "A-E (/ijd-score). Kategori (Sangat Tinggi/Tinggi/dst.) masih pakai ambang dari contoh "
            "ilustratif dokumen metodologi awal, BELUM dikonfirmasi resmi. skor_intensitas & "
            "skor_cakupan = skor_tertimbang APA ADANYA (bukan direnormalisasi thd bobot parameter "
            "yang tersedia lagi -- diubah 22 Jul 2026 atas permintaan eksplisit user): parameter "
            "yang datanya belum ada sumbernya kini efektif menyumbang 0, bukan dikecualikan dari "
            "pembagi -- lihat docs/kajian_metodologi_skala_prioritas_ruas.md utk cakupan data & "
            "rencana bertahap."
        ),
    }


@app.get("/api/usulan-inpres/{usulan_id}/npr")
def usulan_inpres_npr(usulan_id: int):
    with db_cursor() as cur:
        cur.execute("SELECT * FROM usulan_inpres WHERE id = %s", (usulan_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Usulan tidak ditemukan")
    return _compute_npr(row)


def _npr_bulk_ctx(rows: list) -> dict:
    """Batch semua lookup KABUPATEN yang dipakai _npr_skor_cakupan supaya
    preview massal (ratusan/ribuan usulan) tidak query per baris -- pola
    sama dgn ctx di _ijd_score_bulk_rows. SI TIDAK butuh ctx sama sekali --
    semua builder-nya sudah cache modul-level nasional (_npr_kelas_cache)."""
    # kab_by_wilayah -- fallback kode_kab utk usulan TANPA kode_kecamatan
    # (dipakai _bappenas_kode_kab). Bug ditemukan 21 Jul 2026: kode_kab_set
    # sebelumnya cuma dari kode_kecamatan//1000, usulan yg kode_kecamatan-nya
    # NULL (resolvable lewat wilayah_mapping) jadi ke-skip dari SEMUA lookup
    # kabupaten di bawah -- hasil bulk beda dari single (mis. usulan 242153,
    # SIMPUL_TRANSPORTASI 100 di single vs 0 di bulk).
    with db_cursor() as cur:
        cur.execute("SELECT provinsi_sitia, kabupaten_kota_sitia, kode_kabupaten FROM wilayah_mapping")
        kab_by_wilayah = {(r["provinsi_sitia"], r["kabupaten_kota_sitia"]): r["kode_kabupaten"] for r in cur.fetchall()}

    kode_kab_set = set()
    for r in rows:
        kab = _bappenas_kode_kab(r, {"kab_by_wilayah": kab_by_wilayah})
        if kab:
            kode_kab_set.add(kab)

    kawasan_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kabupaten, kategori, sumber_sheet FROM kawasan_tematik WHERE kode_kabupaten = ANY(%s)",
                (list(kode_kab_set),),
            )
            for r in cur.fetchall():
                kawasan_by_kab.setdefault(r["kode_kabupaten"], []).append(r)

    kode_prov_set = {k // 100 for k in kode_kab_set}
    lokus_by_kab, lokus_by_prov = {}, {}
    if kode_kab_set or kode_prov_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kriteria, level, kode_provinsi, kode_kabupaten FROM bappenas_lokus_a "
                "WHERE kode_kabupaten = ANY(%s) OR kode_provinsi = ANY(%s)",
                (list(kode_kab_set) or [0], list(kode_prov_set) or [0]),
            )
            for r in cur.fetchall():
                if r["level"] == "PROVINSI" and r["kode_provinsi"]:
                    lokus_by_prov.setdefault(r["kode_provinsi"], set()).add(r["kriteria"])
                elif r["kode_kabupaten"]:
                    lokus_by_kab.setdefault(r["kode_kabupaten"], set()).add(r["kriteria"])

    # Rincian per-sinyal (bukan cuma "ada row"), dipakai _npr_kelas_dari_hitung
    # di _npr_skor_cakupan -- ditambah 21 Jul 2026 saat SC dirombak dari
    # biner jadi 4-kelas berbasis jumlah sinyal cocok.
    jaringan_jalan_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kode_kabupaten, ada_jalan_daerah, ada_jalan_nasional FROM konektivitas_jaringan_jalan "
                "WHERE kode_kabupaten = ANY(%s)", (list(kode_kab_set),),
            )
            jaringan_jalan_by_kab = {r["kode_kabupaten"]: r for r in cur.fetchall()}

    simpul_by_kab = {}
    if kode_kab_set:
        with db_cursor() as cur:
            cur.execute(
                "SELECT DISTINCT kode_kabupaten, jenis FROM simpul_transportasi WHERE kode_kabupaten = ANY(%s)",
                (list(kode_kab_set),),
            )
            for r in cur.fetchall():
                simpul_by_kab.setdefault(r["kode_kabupaten"], set()).add(r["jenis"])

    usulan_ids = [r["id"] for r in rows if r.get("id")]
    spasial_terhubung_by_usulan = {}
    if usulan_ids:
        with db_cursor() as cur:
            cur.execute(
                "SELECT usulan_id, terhubung FROM usulan_konektivitas_jalan WHERE usulan_id = ANY(%s)",
                (usulan_ids,),
            )
            spasial_terhubung_by_usulan = {r["usulan_id"]: bool(r["terhubung"]) for r in cur.fetchall()}

    return {"kawasan_by_kab": kawasan_by_kab, "lokus_by_kab": lokus_by_kab, "lokus_by_prov": lokus_by_prov,
            "jaringan_jalan_by_kab": jaringan_jalan_by_kab, "simpul_by_kab": simpul_by_kab,
            "spasial_terhubung_by_usulan": spasial_terhubung_by_usulan, "kab_by_wilayah": kab_by_wilayah}


_npr_bulk_cache: dict = {}


def _npr_bulk_rows(provinsi: str) -> list:
    """Versi massal _compute_npr, dipakai preview JSON. Cache in-process per
    provinsi (pola sama dgn _ijd_bulk_cache) -- restart server utk refresh.

    Sejak 22 Jul 2026 (request eksplisit user) tiap baris juga membawa
    "skor_teknokratis_100" (skor_ternormalisasi_100 A-E, REUSE dari
    _ijd_score_bulk_rows tahun 2026 -- bukan hitung ulang ctx teknokratis di
    sini) -- kolom "Total Gabungan" (rata-rata NPR & skor teknokratis) yang
    sempat ditambahkan bareng itu DIHAPUS lagi (request eksplisit user,
    sama tanggal) krn dianggap membingungkan tanpa makna kebijakan yang
    jelas. Urutan baris di sini TETAP murni skor NPR (ranking export Skor
    IJD yang sudah dipindah ke basis NPR ada di _ijd_ranking_sort_key,
    terpisah dari sort di sini)."""
    if provinsi in _npr_bulk_cache:
        return _npr_bulk_cache[provinsi]
    with db_cursor() as cur:
        if provinsi:
            cur.execute("SELECT * FROM usulan_inpres WHERE provinsi = %s", (provinsi,))
        else:
            cur.execute("SELECT * FROM usulan_inpres")
        rows = cur.fetchall()
    if not rows:
        raise HTTPException(404, "Tidak ada usulan yang cocok filter provinsi tsb.")
    ctx = _npr_bulk_ctx(rows)

    _ijd_header_full, ijd_header_short, ijd_data_rows = _ijd_score_bulk_rows(provinsi, 2026)
    idx_id = ijd_header_short.index("ID")
    idx_teknokratis_100 = ijd_header_short.index(IJD_EXPORT_SKOR_TEKNOKRATIS_100_LABEL)
    teknokratis_100_by_id = {r[idx_id]: r[idx_teknokratis_100] for r in ijd_data_rows}

    hasil = []
    for row in rows:
        npr = _compute_npr(row, ctx)
        skor_teknokratis_100 = teknokratis_100_by_id.get(row["id"])
        hasil.append({"id": row["id"], "nama_ruas": row.get("nama_ruas"),
                      "provinsi": row.get("provinsi"), "kabupaten_kota": row.get("kabupaten_kota"),
                      "skor_teknokratis_100": skor_teknokratis_100, **npr})
    hasil.sort(key=lambda r: (r["npr"] is None, -(r["npr"] or 0)))
    _npr_bulk_cache[provinsi] = hasil
    return hasil


@app.get("/api/usulan-inpres/npr/preview")
def usulan_inpres_npr_preview(provinsi: str = "", limit: int = 50, offset: int = 0):
    """Versi JSON, dipaging, dari _npr_bulk_rows -- pola sama dgn
    /ijd-score/preview (kontrak respons GET /api/data/{table}-compatible)."""
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    hasil = _npr_bulk_rows(provinsi)
    page = hasil[offset:offset + limit]
    scope = provinsi or "Nasional"
    rows = [[i + offset + 1, r["id"], r.get("nama_ruas"), r.get("kabupaten_kota"),
             r["npr"], r["kategori"], r["skor_intensitas"]["skor_ternormalisasi_100"],
             r["skor_cakupan"]["skor_ternormalisasi_100"], r["skor_teknokratis_100"]]
            for i, r in enumerate(page)]
    return jsonable_encoder({
        "table": "npr_preview", "label": f"Preview NPR — {scope}",
        "columns": ["No.", "ID", "Nama Ruas", "Kabupaten/Kota", "NPR", "Kategori", "Skor Intensitas", "Skor Cakupan",
                    "Skor Teknokratis A-E (0-100)"],
        "rows": rows, "total": len(hasil), "limit": limit, "offset": offset,
    })


@app.get("/api/usulan-inpres/npr/export/xlsx")
def usulan_inpres_npr_export(provinsi: str = ""):
    """Export xlsx rinci per-komponen (bukan cuma total NPR) -- kolom
    kontribusi tiap parameter SI/SC, pola sama dgn ijd-score/export/xlsx."""
    hasil = _npr_bulk_rows(provinsi)
    si_kode = [kode for kode, *_ in _NPR_SI_DEF]
    sc_kode = [kode for kode, *_ in _NPR_SC_DEF]
    si_label = {kode: label for kode, label, *_ in _NPR_SI_DEF}
    sc_label = {kode: label for kode, label, _ in _NPR_SC_DEF}

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("NPR (eksperimental)")
    group_row = [None, None, None, None] + ["SKOR INTENSITAS (70%)"] + [None] * len(si_kode) \
        + ["SKOR CAKUPAN (30%)"] + [None] * len(sc_kode) + [None, None, None]
    ws.append(group_row)
    header_row = (["No.", "ID", "Nama Ruas", "Kabupaten/Kota"]
                  + [si_label[k] for k in si_kode] + ["SI (0-100)"]
                  + [sc_label[k] for k in sc_kode] + ["SC (0-100)"]
                  + ["NPR", "Kategori", "Skor Teknokratis A-E (0-100)"])
    ws.append(header_row)
    for i, r in enumerate(hasil, start=1):
        si_by_kode = {k["kode"]: k for k in r["skor_intensitas"]["komponen"]}
        sc_by_kode = {k["kode"]: k for k in r["skor_cakupan"]["komponen"]}
        data_row = [i, r["id"], r.get("nama_ruas"), r.get("kabupaten_kota")]
        data_row += [si_by_kode.get(k, {}).get("kontribusi") for k in si_kode]
        data_row += [r["skor_intensitas"]["skor_ternormalisasi_100"]]
        data_row += [sc_by_kode.get(k, {}).get("kontribusi") for k in sc_kode]
        data_row += [r["skor_cakupan"]["skor_ternormalisasi_100"]]
        data_row += [r["npr"], r["kategori"], r["skor_teknokratis_100"]]
        ws.append(data_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    scope = re.sub(r"[^\w]+", "_", provinsi) if provinsi else "Nasional"
    fname = f"npr_{scope}_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# --- Laporan Daerah Prioritas per Kabupaten/Kota (docs/docs/laporan-validator.md)
# -- checklist agregat per kabupaten/kota dlm SATU provinsi terpilih, dua
# aspek (A "Prioritas & Nilai Strategis", B "Daya Ungkit Ekonomi & Kinerja
# Sektoral") persis susunan sheet "Kumpulan Data". BEDA dari
# _bappenas_aspek_a_lokus/_bappenas_aspek_b_ekonomi (yang menjawab "usulan
# INI kena kriteria apa saja", match ke kode_kecamatan spesifik usulan):
# laporan ini menjawab "kabupaten INI py minimal 1 lokus/data utk kriteria
# X dimana saja di wilayahnya" -- makanya query fresh per kabupaten (bukan
# reuse fungsi per-usulan itu), supaya kriteria level KECAMATAN tetap
# tercentang walau kecamatan spesifiknya tidak diketahui (semantik "ada di
# kabupaten ini", bukan "ada di kecamatan usulan ini").
#
# Kolom yg TIDAK disertakan (belum ada sumber data bersih, lihat
# docs/docs/laporan-validator.md): Tata Guna Lahan, KP2B/LP2B, Keberlanjutan
# IJD (dpp_ijd_2025 blm py kode_kabupaten bersih), Penuntasan Koridor (E43,
# OneDrive manual). Per audit 21 Jul 2026 thd sheet "Kumpulan Data" (file
# `2_Analisis Prioritas untuk Bappenas dan Teknokratis 15.7.2026.xlsx`),
# baris Tata Guna Lahan & KP2B/LP2B TIDAK LAGI muncul di sheet itu (dulu
# berstatus "Konfirm"/"tanya", kini dihapus dari kerangka) -- kemungkinan
# sudah keluar dari cakupan resmi, bukan sekadar belum diproses; tidak
# dihapus dari daftar exclusion di sini krn belum ada konfirmasi eksplisit
# dari user soal maknanya.

LAPORAN_ASPEK_A = [
    ("LOKPRI_RPJMN", "LOKPRI RPJMN"),
    ("PKPN", "Lokus PKPN 3T"),
    ("PKSN", "Lokus PKSN/Perbatasan"),
    ("PERBATASAN", "Lokus Perbatasan"),
    ("TRANSMIGRASI", "Lokus Transmigrasi"),
    ("SR", "Lokus SR"),
    ("SEKOLAH_GARUDA", "Lokus Sekolah Garuda"),
    ("KNMP", "Lokus KNMP (Kampung Nelayan Merah Putih)"),
    ("KDMP", "Lokus KDMP (Koperasi Desa Merah Putih)"),
    ("BBM_1_HARGA", "Lokasi BBM 1 Harga (Ruas)"),
    ("KI_PSN_IUKI", "Konektivitas KI/KEK - KI PSN IUKI"),
    ("KI_PRIO_RPJMN", "Konektivitas KI/KEK - KI Prio RPJMN"),
    ("KI_HILIRISASI", "Konektivitas KI/KEK - KI Hilirisasi"),
    ("KI_DIRGANTARA", "Konektivitas KI/KEK - KI Dirgantara"),
    ("SWASEMBADA_PANGAN_RPJMN", "Lokus Swasembada Pangan RPJMN"),
    # Sub-lokus RPJMN (baris 13-15 sheet Kumpulan Data) -- BEDA dari sub KI/KEK
    # di atas: di sumber, ketiganya py nomor item TERSENDIRI (13/14/15), bukan
    # digabung ke nomor 12 -- jadi masing2 kolom checklist + poin _LAPORAN_GROUP_A
    # sendiri, WALAU kode kriteria sama persis dgn yg dipakai Aspek B (dokumen
    # sengaja mendaftarkan ketiganya di kedua aspek). Ditambahkan 21 Jul 2026.
    ("PERIKANAN", "Lokus Kelautan & Perikanan (RPJMN)"),
    ("SWASEMBADA_PANGAN_LOKUS", "Lokus Swasembada Pangan (RPJMN)"),
    ("PERKEBUNAN", "Lokasi Kawasan Perkebunan (RPJMN)"),
]

# Aspek B DIGANTI TOTAL 21 Jul 2026 -- sumbernya semula sheet "Kumpulan
# Data" (Aspek Daya Ungkit Ekonomi dan Kinerja Sektoral), kini pakai sheet
# "Pembobotan Ruas" (`2_Analisis Prioritas untuk Bappenas dan Teknokratis
# 15.7.2026.xlsx` -- BUKAN `1_ KERANGKA PENGGUNAAN DATA UNTUK APLIKASI
# CPIT.docx`, yang dicek 21 Jul 2026 tidak punya tabel "Pembobotan Ruas"
# sama sekali, cuma teks framework umum). Awalnya hanya bagian **Skor
# Cakupan** (7 kategori, "Skor Cakupan saja") -- diperluas 21 Jul 2026
# sesudahnya menambahkan **Skor Intensitas** (9 parameter) krn user
# menunjukkan Jumlah Penduduk & parameter SI lain seharusnya ikut tampil.
# Checklist "ada/tidak" utk item SI berarti "kabupaten ini py datanya"
# (bukan skor magnitude-nya spt di NPR) -- kode kriteria dipetakan ke kode
# `_ada()` yang SUDAH ADA lebih dulu (dipakai Aspek B versi lama sebelum
# dirombak): PADI_JAGUNG->PRODUKSI_PADI (sumbernya SAMA dgn NPR --
# `bps_kecamatan_potensi_tematik.pertanian_produksi_ton`, diperbaiki 21 Jul
# 2026 saat audit ulang kecocokan tabel-per-item -- SEBELUMNYA salah pakai
# `bps_kabupaten_padi.produktivitas_ku_ha`, tabel produktivitas ku/ha utk
# IJD C.A2, bukan tonase produksi), PERKEBUNAN->PRODUKSI_PERKEBUNAN, PETERNAKAN->PRODUKSI_PETERNAKAN,
# PERIKANAN->PRODUKSI_PERIKANAN_TANGKAP (kode "PERKEBUNAN"/"PERIKANAN" polos
# SENGAJA tidak dipakai lagi di sini -- itu kode kawasan LOKUS di Aspek A,
# beda makna dari produksi/tonase). PENDUDUK kode baru (_ada() return True
# apa adanya -- setiap kabupaten di kab_list sudah pasti py data
# penduduk_kecamatan, jadi checklist ini pada praktiknya akan selalu ✓,
# TETAP ditampilkan sesuai kerangka sheet resmi bukan disembunyikan).
# Kode kriteria & BOBOT reuse persis _NPR_SI_DEF/_NPR_SC_DEF (dekat
# _npr_skor_intensitas/_npr_skor_cakupan) supaya NPR & Laporan Prioritas
# konsisten 1 sumber -- label kolom menampilkan "(bobot N)" apa adanya
# dari kolom D sheet resmi.
_NPR_SI_KODE_TO_ADA = {
    "PENDUDUK": "PENDUDUK", "PADI_JAGUNG": "PRODUKSI_PADI", "PERKEBUNAN": "PRODUKSI_PERKEBUNAN",
    "PETERNAKAN": "PRODUKSI_PETERNAKAN", "PERIKANAN": "PRODUKSI_PERIKANAN_TANGKAP", "IP": "IP",
    "KEMANTAPAN_JALAN": "KEMANTAPAN_JALAN", "JUMLAH_KENDARAAN": "JUMLAH_KENDARAAN", "LBS": "LBS",
}
LAPORAN_ASPEK_B = (
    [(_NPR_SI_KODE_TO_ADA[kode], f"{label} (bobot {bobot})") for kode, label, bobot, *_ in _NPR_SI_DEF]
    + [(kode, f"{label} (bobot {bobot})") for kode, label, bobot in _NPR_SC_DEF]
)

# Laporan-kode (Aspek B) -> (cache_key NPR, builder) utk 9 item SI --
# dipakai _skor_b() (dalam _laporan_prioritas_kabupaten) supaya kelas
# min-max Laporan SAMA & SHARE CACHE dgn NPR (_npr_kelas_cache). PENDUDUK
# pakai builder KABUPATEN (_npr_build_penduduk_kabupaten, total penduduk
# kabupaten), BEDA dari NPR yang basis KECAMATAN (per-usulan) -- laporan
# ini agregat per-kabupaten, tidak ada "kecamatan usulan" tunggal yg
# representatif, jadi dipakai total kabupaten dgn cache_key sendiri
# ("PENDUDUK_KAB") supaya tidak tercampur dgn cache NPR yang basisnya beda.
_LAPORAN_ASPEK_B_SI_SOURCE = {
    "PENDUDUK": ("PENDUDUK_KAB", _npr_build_penduduk_kabupaten),
    "PRODUKSI_PADI": ("PADI_JAGUNG", _npr_build_kab_sum("pertanian_produksi_ton")),
    "PRODUKSI_PERKEBUNAN": ("PERKEBUNAN", _npr_build_kab_sum("perkebunan_produksi_ton")),
    "PRODUKSI_PETERNAKAN": ("PETERNAKAN", _npr_build_peternakan),
    "PRODUKSI_PERIKANAN_TANGKAP": ("PERIKANAN", _npr_build_kab_sum("perikanan_produksi_ton")),
    "IP": ("IP", _npr_build_ip),
    "KEMANTAPAN_JALAN": ("KEMANTAPAN_JALAN", _npr_build_kemantapan),
    "JUMLAH_KENDARAAN": ("JUMLAH_KENDARAAN", _npr_build_kendaraan),
    "LBS": ("LBS", _npr_build_lbs),
}

# sumber_sheet asli (kawasan_tematik) -> sub-kategori KI/KEK -- 4 sheet
# digabung 1 kategori KI_PRIORITAS saat impor (import_kawasan_tematik.py),
# dipecah lagi di sini pakai sumber_sheet supaya laporan bisa 4 kolom
# terpisah persis kerangka dokumen, bukan 1 kolom gabungan.
_KI_SHEET_TO_KODE = {
    "Lokus KI PSN IUKI Sudah Terbit": "KI_PSN_IUKI",
    "Lokus PKPN KI Prioritas RPJMN": "KI_PRIO_RPJMN",
    "Lokus PKPN KI Hilirisasi": "KI_HILIRISASI",
    "Lokus PKPN KI Dirgantara": "KI_DIRGANTARA",
}


def _laporan_prioritas_kabupaten(kode_provinsi: int = 0, kode_kab_only: int = None) -> dict:
    """kode_provinsi=0 (atau falsy) -> "Seluruh Indonesia", tanpa filter
    provinsi -- semua query di bawah IN-kan seluruh kabupaten/kota nasional
    (~514), bukan cuma satu provinsi. kode_kab_only -- batasi ke SATU
    kabupaten/kota (dipakai drill-down tab Distribusi Skor, lihat
    _laporan_prioritas_kabupaten_detail); mengabaikan kode_provinsi kalau
    diisi, reuse SELURUH pipeline ctx-batch di bawah apa adanya (murah krn
    query IN-nya otomatis cuma 1 elemen)."""
    with db_cursor() as cur:
        if kode_kab_only:
            cur.execute(
                "SELECT DISTINCT kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan "
                "WHERE kode_kabupaten = %s", (kode_kab_only,),
            )
        elif kode_provinsi:
            cur.execute(
                "SELECT DISTINCT kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan "
                "WHERE kode_provinsi = %s ORDER BY kode_kabupaten",
                (kode_provinsi,),
            )
        else:
            cur.execute(
                "SELECT DISTINCT kode_kabupaten, kabupaten_kota FROM penduduk_kecamatan "
                "ORDER BY kode_kabupaten"
            )
        kab_list = cur.fetchall()
    if not kab_list:
        raise HTTPException(404, "Provinsi/kabupaten tidak ditemukan atau tidak punya data.")
    kode_kab_set = [r["kode_kabupaten"] for r in kab_list]
    kode_prov_set = sorted({k // 100 for k in kode_kab_set})

    with db_cursor() as cur:
        cur.execute(
            "SELECT kriteria, level, kode_provinsi, kode_kabupaten FROM bappenas_lokus_a "
            "WHERE kode_kabupaten = ANY(%s) OR kode_provinsi = ANY(%s)",
            (kode_kab_set, kode_prov_set),
        )
        lokus_rows = cur.fetchall()
    lokus_by_kab, lokus_prov_wide = {}, {}
    for r in lokus_rows:
        if r["level"] == "PROVINSI" and r["kode_provinsi"]:
            lokus_prov_wide.setdefault(r["kode_provinsi"], set()).add(r["kriteria"])
        elif r["kode_kabupaten"]:
            lokus_by_kab.setdefault(r["kode_kabupaten"], set()).add(r["kriteria"])

    with db_cursor() as cur:
        cur.execute(
            "SELECT kategori, kode_kabupaten, sumber_sheet FROM kawasan_tematik WHERE kode_kabupaten = ANY(%s)",
            (kode_kab_set,),
        )
        kawasan_rows = cur.fetchall()
    kawasan_by_kab = {}
    for r in kawasan_rows:
        kode = _KI_SHEET_TO_KODE.get(r["sumber_sheet"]) if r["kategori"] == "KI_PRIORITAS" else r["kategori"]
        if kode:
            kawasan_by_kab.setdefault(r["kode_kabupaten"], set()).add(kode)

    kode_kab_str = [str(k) for k in kode_kab_set]
    with db_cursor() as cur:
        # "Produksi Padi & Jagung" (SI item #2 sheet "Pembobotan Ruas") --
        # tonase gabungan padi+jagung, SAMA tabel/kolom yg dipakai NPR
        # (_npr_build_kab_sum("pertanian_produksi_ton")). BUKAN
        # bps_kabupaten_padi.produktivitas_ku_ha (itu rasio ku/ha, tabel
        # BEDA, dipakai IJD C.A2 -- ketemu salah pakai saat audit ulang
        # kecocokan tabel-per-item 21 Jul 2026, diperbaiki di sini).
        cur.execute(
            "SELECT DISTINCT kode_kab FROM bps_kecamatan_potensi_tematik "
            "WHERE kode_kab = ANY(%s) AND pertanian_produksi_ton IS NOT NULL", (kode_kab_str,),
        )
        produksi_padi_jagung_set = {int(r["kode_kab"]) for r in cur.fetchall()}
        cur.execute(
            "SELECT kode_kab, lahan_baku_sawah_ha, indeks_penanaman_pct FROM bps_kabupaten_indeks_penanaman "
            "WHERE kode_kab = ANY(%s)", (kode_kab_str,),
        )
        ip_rows = cur.fetchall()
        cur.execute(
            "SELECT DISTINCT kode_wilayah FROM kemantapan_ijd_2026 WHERE kode_wilayah = ANY(%s) "
            "AND jenis_adm IN ('Kab.','Kota')", (kode_kab_set,),
        )
        kemantapan_set = {r["kode_wilayah"] for r in cur.fetchall()}
        cur.execute(
            "SELECT DISTINCT kode_kab FROM bps_kabupaten_kendaraan WHERE kode_kab = ANY(%s) AND jumlah IS NOT NULL",
            (kode_kab_str,),
        )
        kendaraan_set = {int(r["kode_kab"]) for r in cur.fetchall()}
        # Produksi perkebunan/peternakan/perikanan level KECAMATAN (BPS Dalam
        # Angka Bab 5) -- diagregasi ke "ada di kabupaten ini" (any kecamatan
        # py angka produksi > NULL). Tabel sama yg dipakai narasi Aspek B
        # per-usulan (_bappenas_aspek_b_ekonomi ctx potensi_produksi_by_kec)
        # tapi belum pernah disambungkan ke laporan agregat ini sampai 21 Jul.
        cur.execute(
            "SELECT DISTINCT kode_kab FROM bps_kecamatan_potensi_tematik "
            "WHERE kode_kab = ANY(%s) AND perkebunan_produksi_ton IS NOT NULL", (kode_kab_str,),
        )
        produksi_perkebunan_set = {int(r["kode_kab"]) for r in cur.fetchall()}
        cur.execute(
            "SELECT DISTINCT kode_kab FROM bps_kecamatan_potensi_tematik WHERE kode_kab = ANY(%s) "
            "AND (peternakan_produksi_daging_kg IS NOT NULL OR peternakan_produksi_telur_kg IS NOT NULL)",
            (kode_kab_str,),
        )
        produksi_peternakan_set = {int(r["kode_kab"]) for r in cur.fetchall()}
        cur.execute(
            "SELECT DISTINCT kode_kab FROM bps_kecamatan_potensi_tematik "
            "WHERE kode_kab = ANY(%s) AND perikanan_produksi_ton IS NOT NULL", (kode_kab_str,),
        )
        produksi_perikanan_set = {int(r["kode_kab"]) for r in cur.fetchall()}
        # Konektivitas Simpul Transportasi -- Pelabuhan Nasional + Pelabuhan
        # Penyeberangan dicocokkan ke kabupaten via text-matching
        # (scripts/import_simpul_transportasi.py); Bandara & Pelabuhan Laut
        # (PELABUHAN_PT.shp) via spatial join titik->poligon kecamatan
        # (scripts/spatial_join_simpul_transportasi.py, 21 Jul 2026 --
        # lihat docs/kajian_overlay_kecamatan_simpul_jalan.md). Jenis
        # DIPISAH (bukan cuma "ada row") -- dipakai skor 4-kelas Aspek B
        # (_laporan_skor_b_sc, MASIH cuma 2 sinyal Nasional & Penyeberangan
        # yg dihitung -- BANDARA/PELABUHAN_LAUT baru masuk
        # simpul_transportasi_set checklist "ada/tidak" polos di bawah,
        # BELUM ikut skor 4-kelas ini, keputusan disengaja/belum diminta).
        cur.execute(
            "SELECT DISTINCT kode_kabupaten, jenis FROM simpul_transportasi "
            "WHERE kode_kabupaten = ANY(%s)", (kode_kab_set,),
        )
        simpul_jenis_by_kab = {}
        for r in cur.fetchall():
            simpul_jenis_by_kab.setdefault(r["kode_kabupaten"], set()).add(r["jenis"])
        simpul_transportasi_set = set(simpul_jenis_by_kab.keys())
        # Konektivitas Jaringan Jalan -- 2 sumber digabung (kabupaten "ada"
        # kalau SALAH SATU true; baris cuma pernah di-INSERT dgn minimal 1
        # flag=1, jadi row existence saja sudah cukup): jalan daerah
        # terpetakan (scripts/import_konektivitas_jalan.py) ATAU dilewati
        # jalan nasional (scripts/import_jalan_nasional.py, 21 Jul 2026).
        # ada_jalan_daerah/ada_jalan_nasional DIPISAH -- dipakai skor
        # 4-kelas Aspek B (3 sinyal: daerah, nasional, spasial).
        cur.execute(
            "SELECT kode_kabupaten, ada_jalan_daerah, ada_jalan_nasional "
            "FROM konektivitas_jaringan_jalan WHERE kode_kabupaten = ANY(%s)",
            (kode_kab_set,),
        )
        jaringan_jalan_detail_by_kab = {r["kode_kabupaten"]: r for r in cur.fetchall()}
        jaringan_jalan_set = set(jaringan_jalan_detail_by_kab.keys())
        # Sinyal KETIGA (terkuat) -- validasi spasial NYATA: usulan yg
        # jalurnya (KML ter-cache) beneran berada dlm 100m dari ruas Jalan
        # Nasional/Provinsi/Tol (scripts/spatial_konektivitas_jalan.py, 21
        # Jul 2026, request eksplisit user), BEDA dari 2 sinyal administratif
        # di atas ("kabupaten punya data jalan terpetakan/dilewati" -- itu
        # cuma soal ADA-TIDAKNYA data, ini soal jalur usulan itu SENDIRI
        # nyata berdekatan). Kabupaten "ada" kalau SALAH SATU dari 3 sinyal
        # ini true (union, sama pola dgn 2 sinyal sebelumnya).
        cur.execute(
            "SELECT DISTINCT (kk.kode_kecamatan / 1000) AS kode_kabupaten "
            "FROM usulan_konektivitas_jalan kk WHERE kk.terhubung = TRUE "
            "AND kk.kode_kecamatan IS NOT NULL",
        )
        spasial_jalan_set = {r["kode_kabupaten"] for r in cur.fetchall() if r["kode_kabupaten"] in kode_kab_set}
        jaringan_jalan_set |= spasial_jalan_set
        # Keberlanjutan IJD / Penuntasan Koridor (Aspek B sheet "Pembobotan
        # Ruas") -- sumber row-level SAMA dgn NPR (_npr_skor_cakupan)/skor
        # IJD A-E param D-E, tapi di sini diagregasi ke "kabupaten ini py
        # MINIMAL 1 usulan yg ..." (semantik laporan per-kabupaten, beda dari
        # NPR yang per-usulan). Ditambahkan 21 Jul 2026 sekalian rombak
        # Aspek B -- SEBELUMNYA dicap "belum ada sumber", ternyata cukup via
        # agregasi ini, bukan genuinely gap. 2 sinyal DIPISAH (bukan cuma
        # OR-gabung) -- dipakai skor 4-kelas (_laporan_skor_b_sc).
        cur.execute(
            "SELECT DISTINCT (kode_kecamatan / 1000) AS kode_kabupaten FROM usulan_inpres "
            "WHERE kode_kecamatan IS NOT NULL AND UPPER(TRIM(penuntasan_ijd_kompetensi)) = 'YA'"
        )
        keberlanjutan_kompetensi_set = {r["kode_kabupaten"] for r in cur.fetchall() if r["kode_kabupaten"] in kode_kab_set}
        cur.execute(
            "SELECT DISTINCT (kode_kecamatan / 1000) AS kode_kabupaten FROM usulan_inpres "
            "WHERE kode_kecamatan IS NOT NULL AND lanjutan_ijd_2025 = 1"
        )
        keberlanjutan_dpp_set = {r["kode_kabupaten"] for r in cur.fetchall() if r["kode_kabupaten"] in kode_kab_set}
        keberlanjutan_set = keberlanjutan_kompetensi_set | keberlanjutan_dpp_set
        cur.execute(
            "SELECT DISTINCT (kode_kecamatan / 1000) AS kode_kabupaten FROM usulan_inpres "
            "WHERE kode_kecamatan IS NOT NULL AND UPPER(TRIM(status_koridor_balai)) = 'SESUAI'"
        )
        penuntasan_balai_set = {r["kode_kabupaten"] for r in cur.fetchall() if r["kode_kabupaten"] in kode_kab_set}
        cur.execute(
            "SELECT DISTINCT (kode_kecamatan / 1000) AS kode_kabupaten FROM usulan_inpres "
            "WHERE kode_kecamatan IS NOT NULL AND TRIM(COALESCE(kode_koridor,'')) != ''"
        )
        penuntasan_kode_set = {r["kode_kabupaten"] for r in cur.fetchall() if r["kode_kabupaten"] in kode_kab_set}
        penuntasan_koridor_set = penuntasan_balai_set | penuntasan_kode_set
    lbs_set = {int(r["kode_kab"]) for r in ip_rows if r["lahan_baku_sawah_ha"] is not None}
    ip_set = {int(r["kode_kab"]) for r in ip_rows if r["indeks_penanaman_pct"] is not None}

    def _ada(kode_kab, kode_kriteria, aspek):
        if kode_kriteria == "PENDUDUK":
            # SI "Pembobotan Ruas" -- setiap kabupaten di kab_list SUDAH
            # PASTI py baris penduduk_kecamatan (itu sumber kab_list itu
            # sendiri), jadi checklist ini pada praktiknya selalu ✓. Tetap
            # ditampilkan apa adanya sesuai kerangka sheet resmi.
            return True
        if aspek == "A" and kode_kriteria in ("KI_PSN_IUKI", "KI_PRIO_RPJMN", "KI_HILIRISASI", "KI_DIRGANTARA"):
            return kode_kriteria in kawasan_by_kab.get(kode_kab, set())
        if kode_kriteria in ("PKPN", "TRANSMIGRASI", "PERIKANAN", "PERKEBUNAN"):
            return kode_kriteria in kawasan_by_kab.get(kode_kab, set())
        if kode_kriteria in ("KI_PSN_IUKI", "KI_PRIO_RPJMN", "KI_HILIRISASI", "KI_DIRGANTARA"):
            return kode_kriteria in kawasan_by_kab.get(kode_kab, set())
        if kode_kriteria == "SWASEMBADA_PANGAN_LOKUS":
            return kode_kriteria in lokus_by_kab.get(kode_kab, set())
        if kode_kriteria == "KNMP_PERIKANAN":
            # Aspek B "Pembobotan Ruas" -- union lokus KNMP (bappenas_lokus_a)
            # dgn kategori PERIKANAN (kawasan_tematik), sama pola dgn
            # _npr_skor_cakupan.
            s = lokus_by_kab.get(kode_kab, set()) | lokus_prov_wide.get(kode_kab // 100, set())
            return "KNMP" in s or "PERIKANAN" in kawasan_by_kab.get(kode_kab, set())
        if kode_kriteria == "SWASEMBADA_PANGAN":
            s = lokus_by_kab.get(kode_kab, set()) | lokus_prov_wide.get(kode_kab // 100, set())
            return bool(s & {"SWASEMBADA_PANGAN_LOKUS", "SWASEMBADA_PANGAN_RPJMN"})
        if kode_kriteria == "KEBERLANJUTAN_IJD":
            return kode_kab in keberlanjutan_set
        if kode_kriteria == "PENUNTASAN_KORIDOR":
            return kode_kab in penuntasan_koridor_set
        if kode_kriteria == "PRODUKSI_PADI":
            return kode_kab in produksi_padi_jagung_set
        if kode_kriteria == "PRODUKSI_PERKEBUNAN":
            return kode_kab in produksi_perkebunan_set
        if kode_kriteria == "PRODUKSI_PETERNAKAN":
            return kode_kab in produksi_peternakan_set
        if kode_kriteria == "PRODUKSI_PERIKANAN_TANGKAP":
            return kode_kab in produksi_perikanan_set
        if kode_kriteria == "SIMPUL_TRANSPORTASI":
            return kode_kab in simpul_transportasi_set
        if kode_kriteria == "JARINGAN_JALAN":
            return kode_kab in jaringan_jalan_set
        if kode_kriteria == "LBS":
            return kode_kab in lbs_set
        if kode_kriteria == "IP":
            return kode_kab in ip_set
        if kode_kriteria == "KEMANTAPAN_JALAN":
            return kode_kab in kemantapan_set
        if kode_kriteria == "JUMLAH_KENDARAAN":
            return kode_kab in kendaraan_set
        return (kode_kriteria in lokus_by_kab.get(kode_kab, set())
                or kode_kriteria in lokus_prov_wide.get(kode_kab // 100, set()))

    def _skor_b(kode_kab):
        """Nilai 4-kelas (0/25/50/75/100) per kolom Aspek B -- SI via
        _npr_kelas_dinamis (rentang min-max, SAMA cache modul-level dgn
        NPR -- lihat _LAPORAN_ASPEK_B_SI_SOURCE), SC via
        _npr_kelas_dari_hitung (jumlah sinyal cocok, level KABUPATEN:
        "ada MINIMAL 1 usulan dgn sinyal ini" -- beda dari NPR yang
        per-usulan). KI/KEK tetap biner (masing2 sudah 1 baris bobot
        sendiri). Return {kode: (tersedia, nilai)}. Ditambahkan 21 Jul
        2026 -- checklist Aspek B sebelumnya cuma ✓/kosong (ada/tidak),
        sekarang tampilkan Nilai per kriteria + skor gabungan berbobot
        (request eksplisit user, "kajian" NPR diterapkan jg ke laporan
        agregat kabupaten ini)."""
        out = {}
        for kode, (cache_key, builder) in _LAPORAN_ASPEK_B_SI_SOURCE.items():
            _, nilai = _npr_kelas_dinamis(cache_key, kode_kab, builder)
            out[kode] = (nilai is not None, nilai)
        for kode in ("KI_PSN_IUKI", "KI_PRIO_RPJMN", "KI_HILIRISASI", "KI_DIRGANTARA"):
            cocok = kode in kawasan_by_kab.get(kode_kab, set())
            out[kode] = (True, 100 if cocok else 0)
        lokus = lokus_by_kab.get(kode_kab, set()) | lokus_prov_wide.get(kode_kab // 100, set())
        kawasan = kawasan_by_kab.get(kode_kab, set())
        matched = int("KNMP" in lokus) + int("PERIKANAN" in kawasan)
        out["KNMP_PERIKANAN"] = (True, _npr_kelas_dari_hitung(matched, 2))
        matched = int("SWASEMBADA_PANGAN_LOKUS" in lokus) + int("SWASEMBADA_PANGAN_RPJMN" in lokus)
        out["SWASEMBADA_PANGAN"] = (True, _npr_kelas_dari_hitung(matched, 2))
        jj = jaringan_jalan_detail_by_kab.get(kode_kab)
        matched = ((int(bool(jj["ada_jalan_daerah"])) + int(bool(jj["ada_jalan_nasional"]))) if jj else 0) \
            + int(kode_kab in spasial_jalan_set)
        out["JARINGAN_JALAN"] = (True, _npr_kelas_dari_hitung(matched, 3))
        simpul = simpul_jenis_by_kab.get(kode_kab, set())
        matched = len(simpul & {"PELABUHAN_NASIONAL", "PELABUHAN_PENYEBERANGAN"})
        out["SIMPUL_TRANSPORTASI"] = (True, _npr_kelas_dari_hitung(matched, 2))
        matched = int(kode_kab in keberlanjutan_kompetensi_set) + int(kode_kab in keberlanjutan_dpp_set)
        out["KEBERLANJUTAN_IJD"] = (True, _npr_kelas_dari_hitung(matched, 2))
        # Disederhanakan 21 Jul 2026 -- checklist biner murni dari kode_koridor
        # terisi (>=1 usulan di kabupaten ini py kode_koridor), sama pola
        # dgn _npr_skor_cakupan (status_koridor_balai tidak dipakai lagi di sini).
        out["PENUNTASAN_KORIDOR"] = (True, 100 if kode_kab in penuntasan_kode_set else 0)
        return out

    def _total_b(skor_b):
        """Skor gabungan Aspek B (0-100) -- PERSIS formula NPR (0,7×SI +
        0,3×SC), dihitung representatif utk kabupaten (bukan 1 usulan).
        SI/SC = skor_tertimbang APA ADANYA (bukan direnormalisasi thd bobot
        tersedia) -- diubah 22 Jul 2026 bareng _npr_skor_intensitas/
        _npr_skor_cakupan supaya kedua fitur tetap konsisten dari satu
        sumber, sesuai konvensi LAPORAN_ASPEK_B/_LAPORAN_ASPEK_B_SI_SOURCE."""
        si_tertimbang = si_bobot_tersedia = 0.0
        for kode, _, bobot, *_ in _NPR_SI_DEF:
            tersedia, nilai = skor_b[_NPR_SI_KODE_TO_ADA[kode]]
            if tersedia:
                si_tertimbang += nilai / 100 * bobot
                si_bobot_tersedia += bobot
        sc_tertimbang = sc_bobot_tersedia = 0.0
        for kode, _, bobot in _NPR_SC_DEF:
            tersedia, nilai = skor_b[kode]
            if tersedia:
                sc_tertimbang += nilai / 100 * bobot
                sc_bobot_tersedia += bobot
        si_100 = si_tertimbang if si_bobot_tersedia else 0.0
        sc_100 = sc_tertimbang if sc_bobot_tersedia else 0.0
        return round(NPR_BOBOT_SI_SC["SI"] * si_100 + NPR_BOBOT_SI_SC["SC"] * sc_100, 1)

    rows = []
    kab_totals = []  # {kode_kab, nama, total_a, total_b} -- dipakai _laporan_prioritas_distribusi
    for kab in kab_list:
        kode_kab = kab["kode_kabupaten"]
        # Konvensi kode BPS: 2 digit terakhir kode_kabupaten >=71 -> Kota,
        # selain itu Kab. -- sama pola dgn dataViewerOpenLokusBappenas/
        # kab_label di /api/kecamatan (baris ~4254), dipakai di sini krn
        # nama polos ("SERANG"/"TANGERANG") ambigu antara kab & kota
        # kembar dlm satu provinsi.
        jenis = "Kota" if kode_kab % 100 >= 71 else "Kab."
        nama_lengkap = f"{jenis} {kab['kabupaten_kota']}"
        cells_a = ["✓" if _ada(kode_kab, kode, "A") else "" for kode, _ in LAPORAN_ASPEK_A]
        skor_b = _skor_b(kode_kab)
        cells_b = [str(nilai) if tersedia else "" for kode, _ in LAPORAN_ASPEK_B for tersedia, nilai in [skor_b[kode]]]
        rows.append([kode_kab, nama_lengkap] + cells_a + cells_b)
        # Aspek A tetap COUNT kriteria tercentang (0-15, checklist murni --
        # tidak diminta ikut dirombak). Aspek B kini SKOR BERBOBOT 0-100
        # (bukan count lagi) -- formula NPR (0,7 SI + 0,3 SC) dihitung
        # representatif per kabupaten (_total_b). Skala keduanya BEDA
        # (0-15 vs 0-100) -- "skor gabungan" (dashboard) jadi jumlah 2
        # skala berbeda apa adanya, lihat _LAPORAN_MAKS_TOTAL.
        total_a = sum(1 for group in _LAPORAN_GROUP_A if any(_ada(kode_kab, k, "A") for k in group))
        total_b = _total_b(skor_b)
        kab_totals.append({"kode_kab": kode_kab, "kode_prov": kode_kab // 100, "nama": nama_lengkap,
                            "total_a": total_a, "total_b": total_b})

    header = ["Kode Kab/Kota", "Kabupaten/Kota"] + [lbl for _, lbl in LAPORAN_ASPEK_A] + [lbl for _, lbl in LAPORAN_ASPEK_B]
    return {"header": header, "rows": rows, "n_aspek_a": len(LAPORAN_ASPEK_A), "n_aspek_b": len(LAPORAN_ASPEK_B),
            "kab_totals": kab_totals}


# 15 item Aspek A (sumber sheet "Kumpulan Data", tak berubah) / 7 item
# Aspek B (sumber sheet "Pembobotan Ruas" bagian Skor Cakupan, DIGANTI
# TOTAL 21 Jul 2026 -- lihat komentar LAPORAN_ASPEK_B di atas). Urutan
# Aspek A PERSIS sheet "Kumpulan Data" (sub-item KI/KEK-style bertanda "-"
# TANPA nomor sendiri digabung 1 induk; tapi baris 13-15 "Lokus Kelautan &
# Perikanan"/"Lokus Swasembada Pangan"/"Lokasi Kawasan Perkebunan" PUNYA
# nomor item sendiri di sheet -- jadi 3 poin terpisah, bukan digabung ke
# item 12 "Swasembada Pangan RPJMN").
_LAPORAN_GROUP_A = [
    ["LOKPRI_RPJMN"], ["PKPN"], ["PKSN"], ["PERBATASAN"], ["TRANSMIGRASI"],
    ["SR"], ["SEKOLAH_GARUDA"], ["KNMP"], ["KDMP"], ["BBM_1_HARGA"],
    ["KI_PSN_IUKI", "KI_PRIO_RPJMN", "KI_HILIRISASI", "KI_DIRGANTARA"],
    ["SWASEMBADA_PANGAN_RPJMN"],
    # Baris 13-15 sheet Kumpulan Data -- item TERSENDIRI (bukan sub KI/KEK-style
    # yg digabung 1 poin), naikkan skala Aspek A 12 -> 15. Ditambahkan 21 Jul 2026.
    ["PERIKANAN"], ["SWASEMBADA_PANGAN_LOKUS"], ["PERKEBUNAN"],
]
# Aspek B TIDAK LAGI dihitung sbg count induk kriteria (_LAPORAN_GROUP_B
# lama, dihapus 21 Jul 2026) -- diganti skor berbobot 0-100 formula NPR
# (0,7 SI + 0,3 SC, lihat _skor_b/_total_b dalam _laporan_prioritas_kabupaten).
# Skala tetap konstan 100 (bukan len(list) spt Aspek A), disimpan di sini
# supaya tempat rujukan maks_b/_LAPORAN_MAKS_TOTAL cuma satu.
_LAPORAN_MAKS_B = 100


def _laporan_prioritas_distribusi(kode_provinsi: int, step_override: int = None) -> dict:
    """Distribusi jumlah + NAMA (+kode) kabupaten/kota per rentang skor
    INTEGER NON-OVERLAPPING, dari kab_totals (total_a 0-maks_a /
    total_b 0-maks_b) -- dipakai tab "Distribusi Skor" (klik batang ->
    daftar kabupaten -> klik kabupaten -> drill-down kriteria yang match,
    lihat _laporan_prioritas_kabupaten_detail).

    Lebar bucket (step) DEFAULT dinamis mengikuti skala maks (bukan
    hardcode) -- ~6 bucket target, `step = ceil(maks/6)`, supaya skala
    berubah (12->15, 14->12 dst., riwayat sudah terjadi 2x sesi ini) tidak
    perlu ubah kode. `step_override` (dari input "Lebar Rentang" di UI,
    default tombol 2) HANYA dipakai utk Aspek A -- Aspek B SELALU dinamis
    (21 Jul 2026: Aspek B berubah dari count kriteria 0-16 jadi skor
    berbobot 0-100 kontinu, input manual kecil spt "2" tidak lagi cocok,
    bikin 50 bucket kecil2 kalau dipaksakan; auto ceil(100/6)~=17 lebih
    masuk akal & tidak perlu kontrol terpisah di UI). Batas bucket
    non-overlapping ("0-2","3-4","5-6",... utk step=2, bukan "0-2","2-4"
    yang ambigu di titik 2) -- skor 0 SENGAJA digabung ke bucket pertama
    (bukan bucket sendiri), jadi bucket pertama lebih lebar 1 poin drpd
    bucket sesudahnya."""
    data = _laporan_prioritas_kabupaten(kode_provinsi)
    kab_totals = data["kab_totals"]

    def _bucket(field, maks, allow_override=True):
        step = (step_override if (step_override and allow_override) else None) or max(1, -(-maks // 6))
        edges = [0, step]
        while edges[-1] < maks:
            edges.append(min(edges[-1] + step, maks))
        out = []
        for i in range(len(edges) - 1):
            lo = 0 if i == 0 else edges[i] + 1
            hi = edges[i + 1]
            if lo > hi:
                continue
            members = sorted(
                ({"kode_kab": k["kode_kab"], "nama": k["nama"]} for k in kab_totals if lo <= k[field] <= hi),
                key=lambda m: m["nama"],
            )
            # bucket pertama (lo=0, digabung dgn skor 0) ditampilkan cuma "hi"
            # -- "0-1" salah kesan seolah rentangnya persis 0 s.d. 1 poin,
            # padahal cakupannya sama dgn bucket lain (lebar `step`), cuma
            # skor 0 numpang gabung di dalamnya.
            label = str(hi) if i == 0 else f"{lo}-{hi}"
            out.append({"label": label, "min": lo, "max": hi, "count": len(members),
                        "kabupaten": members})
        return out

    return {
        "aspek_a": _bucket("total_a", len(_LAPORAN_GROUP_A)),
        "aspek_b": _bucket("total_b", _LAPORAN_MAKS_B, allow_override=False),
        "n_kabupaten": len(kab_totals),
        "maks_a": len(_LAPORAN_GROUP_A), "maks_b": _LAPORAN_MAKS_B,
        "step": step_override or None,
    }


def _laporan_prioritas_kabupaten_detail(kode_kab: int) -> dict:
    """Drill-down 1 kabupaten/kota: daftar kriteria Aspek A/B -- dipakai
    saat klik nama kabupaten di tab Distribusi Skor. Aspek A tetap
    checklist ada/tidak ("cocok" boolean, sel "✓"). Aspek B kini SKOR
    4-kelas (25/50/75/100), bukan lagi ada/tidak -- sel isinya angka
    (string) atau kosong ("" = belum tersedia) sejak 21 Jul 2026, jadi
    dikembalikan sbg "nilai" (int atau None), BUKAN "cocok" boolean lagi."""
    data = _laporan_prioritas_kabupaten(kode_kab_only=kode_kab)
    row = data["rows"][0]
    n_a = data["n_aspek_a"]
    aspek_a = [{"label": data["header"][2 + i], "cocok": row[2 + i] == "✓"} for i in range(n_a)]
    aspek_b = [{"label": data["header"][2 + n_a + i],
                "nilai": int(row[2 + n_a + i]) if row[2 + n_a + i] != "" else None}
               for i in range(data["n_aspek_b"])]
    return {"kode_kab": row[0], "nama": row[1], "aspek_a": aspek_a, "aspek_b": aspek_b}


# 115 = 15 (Aspek A, count kriteria) + 100 (Aspek B, skor berbobot NPR) --
# DUA SKALA BEDA dijumlahkan apa adanya (21 Jul 2026, Aspek B dirombak jadi
# skor 0-100 spt NPR, Aspek A tetap count checklist) -- bukan skala tunggal
# yg "bersih", tapi tetap konstan & bermakna sbg approx kasar "besaran
# gabungan", dipakai cuma utk kategori Tinggi/Sedang/Rendah Dashboard.
_LAPORAN_MAKS_TOTAL = len(_LAPORAN_GROUP_A) + _LAPORAN_MAKS_B


def _laporan_prioritas_dashboard(kode_provinsi: int = 0) -> dict:
    """Ringkasan siap-pakai utk pengambil kebijakan (tab "Dashboard"):
    KPI ringkas, peringkat 10 kabupaten/kota skor tertinggi, cakupan tiap
    kriteria (dari kolom checklist flat -- lebih rinci dari 12/14 item
    induk yg dipakai skor gabungan), komposisi kategori prioritas, dan
    (khusus mode "Seluruh Indonesia") perbandingan rata-rata skor antar
    provinsi. Semua dihitung dari _laporan_prioritas_kabupaten() yang sudah
    ada -- tidak ada query/skoring baru, cuma agregasi ulang."""
    data = _laporan_prioritas_kabupaten(kode_provinsi)
    kab_totals = data["kab_totals"]
    n = len(kab_totals)

    totals = [k["total_a"] + k["total_b"] for k in kab_totals]
    n_tanpa_data = sum(1 for t in totals if t == 0)
    avg_total = round(sum(totals) / n, 2) if n else 0
    avg_a = round(sum(k["total_a"] for k in kab_totals) / n, 2) if n else 0
    avg_b = round(sum(k["total_b"] for k in kab_totals) / n, 2) if n else 0

    ranked = sorted(kab_totals, key=lambda k: (-(k["total_a"] + k["total_b"]), k["nama"]))
    top10 = [{"nama": k["nama"], "kode_kab": k["kode_kab"], "total_a": k["total_a"],
              "total_b": k["total_b"], "total": k["total_a"] + k["total_b"]} for k in ranked[:10]]

    # Komposisi kategori prioritas -- ambang relatif thd skala gabungan 26
    # (bukan ambang resmi dokumen manapun, murni pembagian 4-kuadran linier
    # utk visualisasi ringkas; "Tidak ada data" dipisah dari "Rendah" krn
    # beda makna: rendah = ada data tp sedikit kriteria cocok, tidak ada
    # data = 0/0, kemungkinan kode_kabupaten/nama belum match ke sumber).
    def _kategori(t):
        if t == 0:
            return "Tidak ada data"
        if t <= round(_LAPORAN_MAKS_TOTAL * 0.25):
            return "Rendah"
        if t <= round(_LAPORAN_MAKS_TOTAL * 0.5):
            return "Sedang"
        return "Tinggi"
    komposisi_count = {"Tinggi": 0, "Sedang": 0, "Rendah": 0, "Tidak ada data": 0}
    for t in totals:
        komposisi_count[_kategori(t)] += 1
    komposisi = [{"label": k, "count": komposisi_count[k]} for k in ("Tinggi", "Sedang", "Rendah", "Tidak ada data")]

    # Cakupan tiap kriteria (dari kolom flat checklist, bukan item induk --
    # lebih rinci, mis. KI PSN IUKI terpisah dari KI Hilirisasi).
    n_id = 2
    def _coverage(start, count):
        out = []
        for i in range(start, start + count):
            label = data["header"][i]
            n_match = sum(1 for row in data["rows"] if row[i] == "✓")
            out.append({"label": label, "count": n_match, "pct": round(n_match / n * 100, 1) if n else 0})
        out.sort(key=lambda x: -x["count"])
        return out
    cakupan_a = _coverage(n_id, data["n_aspek_a"])
    cakupan_b = _coverage(n_id + data["n_aspek_a"], data["n_aspek_b"])

    result = {
        "n_kabupaten": n, "avg_total": avg_total, "avg_a": avg_a, "avg_b": avg_b,
        "maks_total": _LAPORAN_MAKS_TOTAL, "maks_a": len(_LAPORAN_GROUP_A), "maks_b": _LAPORAN_MAKS_B,
        "n_tanpa_data": n_tanpa_data,
        "top10": top10, "komposisi": komposisi,
        "cakupan_a": cakupan_a, "cakupan_b": cakupan_b,
        "provinsi": None,
    }

    if not kode_provinsi:
        # Mode nasional -- tambahan perbandingan rata-rata skor antar
        # provinsi, utk lihat provinsi mana yg secara agregat py banyak
        # kabupaten/kota berprioritas tinggi vs yg datanya masih kosong.
        by_prov: dict = {}
        for k in kab_totals:
            by_prov.setdefault(k["kode_prov"], []).append(k["total_a"] + k["total_b"])
        with db_cursor() as cur:
            cur.execute(
                "SELECT DISTINCT kode_provinsi, provinsi FROM penduduk_kecamatan "
                "WHERE kode_provinsi = ANY(%s)", (list(by_prov.keys()),),
            )
            nama_by_prov = {r["kode_provinsi"]: r["provinsi"] for r in cur.fetchall()}
        provinsi_rows = [
            {"kode_provinsi": kp, "nama": nama_by_prov.get(kp, str(kp)),
             "n_kabupaten": len(vals), "avg_total": round(sum(vals) / len(vals), 2) if vals else 0}
            for kp, vals in by_prov.items()
        ]
        provinsi_rows.sort(key=lambda r: -r["avg_total"])
        result["provinsi"] = provinsi_rows

    return result


@app.get("/api/laporan-daerah-prioritas/dashboard")
def laporan_daerah_prioritas_dashboard(provinsi: int = 0):
    return jsonable_encoder(_laporan_prioritas_dashboard(provinsi))


def _nama_provinsi_laporan(kode_provinsi: int) -> str:
    if not kode_provinsi:
        return "Seluruh Indonesia"
    with db_cursor() as cur:
        cur.execute("SELECT provinsi FROM penduduk_kecamatan WHERE kode_provinsi = %s LIMIT 1", (kode_provinsi,))
        r = cur.fetchone()
    return r["provinsi"] if r else str(kode_provinsi)


@app.get("/api/laporan-daerah-prioritas/distribusi")
def laporan_daerah_prioritas_distribusi(provinsi: int = 0, step: int = None):
    if step is not None and step < 1:
        raise HTTPException(400, "Lebar rentang (step) harus >= 1.")
    return jsonable_encoder(_laporan_prioritas_distribusi(provinsi, step))


@app.get("/api/laporan-daerah-prioritas/detail/{kode_kab}")
def laporan_daerah_prioritas_detail(kode_kab: int):
    """Drill-down 1 kabupaten/kota (klik nama di tab Distribusi Skor) --
    daftar kriteria checklist Aspek A/B yang match utk kabupaten itu."""
    return jsonable_encoder(_laporan_prioritas_kabupaten_detail(kode_kab))


@app.get("/api/laporan-daerah-prioritas/preview")
def laporan_daerah_prioritas_preview(provinsi: int = 0, limit: int = 50, offset: int = 0):
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    data = _laporan_prioritas_kabupaten(provinsi)
    page = data["rows"][offset:offset + limit]
    nama_provinsi = _nama_provinsi_laporan(provinsi)
    return jsonable_encoder({
        "table": "laporan_prioritas_kabupaten",
        "label": f"Laporan Daerah Prioritas — {nama_provinsi} (agregat per kabupaten/kota)",
        "columns": data["header"], "rows": page, "total": len(data["rows"]),
        "limit": limit, "offset": offset,
        "n_aspek_a": data["n_aspek_a"], "n_aspek_b": data["n_aspek_b"],
    })


# Layout export xlsx (kolom & label) PERSIS docs/docs/Laporan Prioritas.xlsx
# -- template resmi yang diminta 21 Jul 2026 menggantikan layout generik
# lama (grup Aspek A + Aspek B berdampingan). Aspek B (Daya Ungkit Ekonomi)
# SENGAJA tidak diikutkan lagi di export ini -- template resminya cuma
# Aspek A (dikonfirmasi user; Aspek B tetap ada di tab Checklist/Dashboard
# UI, cuma export file-nya yang berubah). Indeks None = kolom header GRUP
# yang di template SELALU KOSONG ("Konektivitas KI/KEK" -- item 11 di
# Kumpulan Data TIDAK punya status/level/sumber sendiri, murni label
# pengelompokan utk 4 sub-kriteria KI/KEK di bawahnya).
#
# "Lokus Swasembada Pangan RPJMN" BEDA -- ditemukan 23 Jul 2026 saat
# validasi view vs export: walau di layout xlsx resmi kolom ini SECARA
# VISUAL juga tampak seperti grup header kosong (sama gaya dgn Konektivitas
# KI/KEK), di Kumpulan Data item 12 ("Lokus Swasembada Pangan RPJMN")
# punya status='ada'/level='Kabupaten'/sumber sendiri (Lampiran IV RPJMN
# 2025-2029.xlsx) yang BERBEDA dari sumber 3 sub-itemnya (item 13-15, Dit.
# KP). Artinya kriteria ini py data independen, BUKAN cuma label
# pengelompokan -- kalau dikosongkan begitu saja (spt sebelumnya, None),
# 56/514 kabupaten yang MEMENUHI kriteria 12 ini SENDIRI (tapi tak satupun
# dari 3 sub-itemnya) jadi tak kelihatan tercentang sama sekali di export,
# padahal tercentang di view Checklist -- per keputusan eksplisit user,
# kolom ini SEKARANG diisi nilai asli (idx["SWASEMBADA_PANGAN_RPJMN"]),
# menyimpang dari layout kosong template resmi demi akurasi/konsistensi
# dgn view. Indeks diambil dinamis dari LAPORAN_ASPEK_A (bukan hardcode
# angka) supaya tetap benar kalau list itu berubah lagi.
def _laporan_export_template_cols() -> list:
    idx = {kode: i for i, (kode, _) in enumerate(LAPORAN_ASPEK_A)}
    return [
        (idx["LOKPRI_RPJMN"], "LOKPRI RPJMN"),
        (idx["PKPN"], "Lokus PKPN 3T"),
        (idx["PKSN"], "Lokus PKSN/Perbatasan"),
        (idx["PERBATASAN"], "Lokus Perbatasan"),
        (idx["TRANSMIGRASI"], "Lokus Transmigrasi"),
        (idx["SR"], "Lokus SR"),
        (idx["SEKOLAH_GARUDA"], "Lokus Sekolah Garuda"),
        (idx["KNMP"], "Lokus KNMP (Kampung Nelayan Merah Putih)"),
        (idx["KDMP"], "Lokus KDMP (Koperasi Desa Merah Putih)"),
        (idx["BBM_1_HARGA"], "Lokasi BBM 1 Harga (Ruas)"),
        (None, "Konektivitas KI/KEK"),
        (idx["KI_PSN_IUKI"], "- KI PSN IUKI"),
        (idx["KI_PRIO_RPJMN"], "- KI Prio RPJMN"),
        (idx["KI_HILIRISASI"], "- KI Hilirisasi"),
        (idx["KI_DIRGANTARA"], "- KI Dirgantara"),
        (idx["SWASEMBADA_PANGAN_RPJMN"], "Lokus Swasembada Pangan RPJMN"),
        (idx["PERIKANAN"], "- Lokus Kelautan & Perikanan"),
        (idx["SWASEMBADA_PANGAN_LOKUS"], "- Lokus Swasembada Pangan"),
        (idx["PERKEBUNAN"], "- Lokasi Kawasan Perkebunan"),
    ]


@app.get("/api/laporan-daerah-prioritas/export/xlsx")
def laporan_daerah_prioritas_export(provinsi: int = 0):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    data = _laporan_prioritas_kabupaten(provinsi)
    nama_provinsi = _nama_provinsi_laporan(provinsi)
    template_cols = _laporan_export_template_cols()

    with db_cursor() as cur:
        cur.execute(
            "SELECT DISTINCT kode_kabupaten, kabupaten_kota, provinsi FROM penduduk_kecamatan "
            "WHERE kode_kabupaten = ANY(%s)",
            (list(row[0] for row in data["rows"]),),
        )
        info_by_kab = {r["kode_kabupaten"]: r for r in cur.fetchall()}

    n_id = 4  # No, PROVINSI, KABUPATEN, TOTAL
    n_cols = n_id + len(template_cols)

    # write_only=False -- butuh styling, tidak tersedia di mode write_only
    # (dipakai endpoint export lain yang tidak butuh styling).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Prioritas"

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.append(["No", "PROVINSI", "KABUPATEN", "TOTAL"] + [lbl for _, lbl in template_cols])

    for i, row in enumerate(data["rows"], start=1):
        kode_kab = row[0]
        cells_a = row[2:2 + data["n_aspek_a"]]
        info = info_by_kab.get(kode_kab, {})
        jenis = "KOTA" if kode_kab % 100 >= 71 else "KABUPATEN"
        marks = [("v" if idx is not None and cells_a[idx] == "✓" else None) for idx, _ in template_cols]
        total = sum(1 for m in marks if m)
        ws.append([i, info.get("provinsi", ""), f"{jenis} {info.get('kabupaten_kota', '')}".strip(), total] + marks)

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=10)
        cell.alignment = center
        cell.fill = PatternFill("solid", fgColor="F2F2F2")

    n_rows = 1 + len(data["rows"])
    for r_idx in range(1, n_rows + 1):
        for c_idx in range(1, n_cols + 1):
            ws.cell(row=r_idx, column=c_idx).border = border
    for r_idx in range(2, n_rows + 1):
        for c_idx in range(1, 4):
            ws.cell(row=r_idx, column=c_idx).alignment = left
        for c_idx in range(4, n_cols + 1):
            ws.cell(row=r_idx, column=c_idx).alignment = center

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 8
    for col in range(5, n_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
    ws.row_dimensions[1].height = 60
    ws.freeze_panes = "E2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    scope = re.sub(r"[^\w]+", "_", nama_provinsi)
    fname = f"laporan_prioritas_{scope}_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


PENILAIAN_SYSTEM_PROMPT = """Anda membantu analis Bappenas menyusun DRAF penilaian kualitatif usulan \
Inpres Jalan Daerah. Aspek A (Prioritas & Nilai Strategis) dan Aspek B (Daya Ungkit Ekonomi & Kinerja \
Sektoral) SUDAH DIHITUNG SISTEM (rule-based, bukan tugas Anda menilai ulang) — masing-masing berupa daftar \
kriteria/indikator yang cocok/ada berdasarkan data lokasi usulan. Field "aspek_a_hasil" dan "aspek_b_hasil" \
pada data yang diberikan berisi hasil itu.

Tugas Anda menyusun DUA teks:
1. "kesimpulan": integrasi naratif 2-3 kalimat dari aspek_a_hasil DAN aspek_b_hasil APA ADANYA.
2. "aspek_b_narasi_ai": narasi PELENGKAP khusus Aspek B (Daya Ungkit Ekonomi & Kinerja Sektoral) saja —
   gaya laporan kebijakan yang mengalir dan meyakinkan, BUKAN daftar/enumerasi kaku bergaya "Ditemukan N
   indikator: ...; ...". WAJIB menyinggung SETIAP indikator di aspek_b_hasil["indikator_ada"] — jangan
   diam-diam melewati sebagian demi keringkasan; kalau indikatornya banyak (>4), gunakan lebih banyak
   kalimat (boleh sampai 8-10) supaya semua kebagian tempat, dikelompokkan per tema (mis. produksi pangan
   dijadikan satu kalimat, konektivitas/lalu lintas kalimat lain) alih-alih satu kalimat generik per
   indikator. Rangkai jadi cerita dampak ekonomi/sektoral ruas ini (ketahanan pangan, kelancaran logistik,
   pertumbuhan ekonomi lokal — sesuai definisi Aspek B), sertakan angka konkret dari data bila ada — dan
   SETIAP kali menyebut angka, sertakan sumbernya secara ringkas dalam kalimat (mis. "...jiwa (BPS Dalam
   Angka)", "...unit kendaraan (BPS Dalam Angka)"), jangan sebut angka telanjang tanpa atribusi. Bila
   indikator_ada kosong, nyatakan dengan jujur belum ada indikator yang didukung data, jangan dibuat seolah ada.

CEK SATU-PER-SATU — kalau field-nya ADA pada data, FAKTA di bawah ini WAJIB masuk ke "aspek_b_narasi_ai"
(jangan diam-diam dilewati; kalau TIDAK ada field-nya, jangan menyinggung topik itu sama sekali). Yang WAJIB
adalah ISI/ANGKA-nya, BUKAN kalimat contoh di bawah — kalimat contoh cuma ilustrasi supaya paham fakta mana
yang dimaksud; rangkai fakta itu dengan kalimat Anda sendiri, jangan disalin verbatim:
- "kemantapan_ruas" ada → sebutkan kondisi kemantapan ruas pakai PERSIS field "pct_tidak_mantap" (jangan
  hitung ulang) — persentase itu = kondisi rusak ringan+berat dibanding panjang penanganan kompetensi.
  Contoh ilustrasi (JANGAN disalin persis): "Dari aspek teknis, kondisi ruas ini tergolong <status> dengan
  persentase sekitar <pct_tidak_mantap>%..." — cari cara lain menyampaikan fakta yang sama.
- "konektivitas_jalan" ada → WAJIB disebut, TERMASUK kalau ketiga "terhubung_..." semuanya false (jangan
  dilewati hanya krn tidak ada yang terhubung — ini tetap fakta yang harus dilaporkan): kalau ada yang true,
  sebut jaringan mana ("terhubung_jalan_nasional"/"terhubung_jalan_provinsi"/"terhubung_tol") beserta jarak
  meter dari "jarak_ke_..._m" berpasangan; kalau semuanya false, sampaikan bahwa ruas belum terhubung
  langsung ke jaringan jalan nasional/provinsi/tol dalam radius "ambang_m" meter (boleh dengan kalimat apa
  saja, bukan rumus tetap).
- "simpul_transportasi" ada → sebutkan simpul transportasi (bandara/pelabuhan) terdekat dari
  "simpul_terdekat" — jenis, "nama_simpul", "jarak_km"-nya, dalam radius "radius_km" km.
- "kecamatan_dilalui" ada → sebutkan SEMUA nama kecamatan pada daftar "kecamatan_dilalui" (jangan cuma
  sebagian kalau lebih dari 1) beserta PERSIS field "total_penduduk_dilalui" (jangan hitung ulang/jumlah
  manual dari daftar) — kalau "total_penduduk_dilalui" null (data penduduk sebagian/semua kecamatan tidak
  tersedia), sebut nama kecamatannya saja tanpa angka total, jangan mengarang angka. Kalau field
  "total_kendaraan_dilalui" pada "kecamatan_dilalui" TIDAK null, WAJIB pakai angka itu (jumlah kendaraan
  GABUNGAN semua kecamatan yang dilalui) setiap kali menyebut jumlah kendaraan ruas ini — JANGAN pakai angka
  kendaraan satu kecamatan saja (mis. dari kalimat ringkasan indikator) kalau usulan ini melintasi >1
  kecamatan; kalau "total_kendaraan_dilalui" null, boleh sebut angka kendaraan level kabupaten dari data
  lain yang tersedia (jangan mengarang).
- "komoditas_perkebunan_kecamatan" ada → boleh sebut komoditas perkebunan UTAMA kecamatan ini persis nama
  pada field "komoditas" (urut "produksi_ton" terbesar, maksimal 3 komoditas) beserta angka produksi_ton-nya;
  JANGAN sebut nama komoditas perkebunan LAIN yang tidak ada di daftar ini (mis. jangan menyebut kelapa
  sawit/karet/tebu kalau tidak ada di daftar "komoditas_perkebunan_kecamatan" — daftar ini sudah cek data
  riil per komoditas, bukan cuma tebakan). Kalau field ini TIDAK ada padahal aspek_b_hasil menyebut
  PRODUKSI_PERKEBUNAN, boleh sebut produksi perkebunan secara GENERIK tanpa menyebut nama komoditas spesifik
  apa pun.

Kedua teks WAJIB berbasis fakta di aspek_a_hasil/aspek_b_hasil/field tambahan di atas APA ADANYA — jangan \
mengarang fakta di luar itu, jangan menilai ulang atau mengubah checklist/poin yang sudah diberikan. Sebut \
secara ringkas kriteria/indikator utama yang mendukung (atau ketiadaannya bila kosong).

GAYA BAHASA & VARIASI — WAJIB, bukan saran opsional. Ini narasi untuk SATU usulan spesifik, bukan formulir \
isian: jangan susun dari kerangka kalimat yang sama tiap kali dipanggil. Secara khusus:
- Kalimat PEMBUKA jangan selalu mulai dari kondisi teknis/kemantapan ruas — pilih fakta yang paling menonjol \
untuk usulan INI sebagai pembuka (bisa konektivitas, kepadatan penduduk yang dilintasi, potensi ekonomi/\
produksi kawasan, atau kondisi teknis — sesuai mana yang paling relevan/mencolok dari datanya), jangan \
selalu format "<Nama Kegiatan> di <Kabupaten>, <Provinsi>, memiliki panjang ... km" atau "Dari aspek \
teknis, ...".
- Kalimat PENUTUP jangan selalu berumus "Peningkatan/Kegiatan ... jalan ini diharapkan dapat mendukung/\
meningkatkan ... pertumbuhan ekonomi lokal" — kalau mau menutup dengan proyeksi dampak, ganti-ganti sudut \
pandang dan pilihan katanya.
- Setiap kalimat wajib membawa informasi baru (data/indikator/angka); jangan mengulang frasa pembuka atau \
kesimpulan umum ("hal ini menunjukkan...", "secara keseluruhan...", dsb.) yang tidak menambah fakta. Hindari \
basa-basi/pengantar yang tidak perlu — langsung ke fakta relevan sejak kalimat pertama. Jangan menyebut \
ulang satu indikator/angka yang sama di lebih dari satu kalimat.
Variasi ini HANYA soal gaya bahasa/struktur kalimat/urutan penyampaian fakta — kelengkapan fakta, checklist \
wajib di atas, dan larangan mengarang tetap berlaku sama ketatnya.

Balas HANYA JSON valid tanpa teks lain, format:
{"kesimpulan": "...", "aspek_b_narasi_ai": "..."}
Bahasa Indonesia formal, sebut angka/kriteria dari data bila relevan."""


def _plain_openai_compatible(url: str, key: str, model: str, system: str, user: str, max_tokens: int = 2048) -> str:
    resp = requests.post(
        url, headers={"Authorization": f"Bearer {key}"},
        json={"model": model, "temperature": 0.4, "max_tokens": max_tokens,
              "messages": [{"role": "system", "content": system},
                           {"role": "user", "content": user}]},
        timeout=180,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def _plain_claude(key: str, system: str, user: str, max_tokens: int = 2048) -> str:
    client = anthropic.Anthropic(api_key=key)
    resp = client.messages.create(
        model=CLAUDE_MODEL, max_tokens=max_tokens, system=system,
        messages=[{"role": "user", "content": user}],
    )
    return "".join(b.text for b in resp.content if b.type == "text")


def _plain_gemini(key: str, system: str, user: str, max_tokens: int = 2048) -> str:
    resp = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={key}",
        json={"system_instruction": {"parts": [{"text": system}]},
              "contents": [{"parts": [{"text": user}]}],
              "generationConfig": {"maxOutputTokens": max_tokens}},
        timeout=180,
    )
    resp.raise_for_status()
    return resp.json()["candidates"][0]["content"]["parts"][0]["text"]


def _llm_plain(system: str, user: str, max_tokens: int = 2048) -> tuple:
    """Satu completion polos (tanpa tools) lewat provider pertama yang
    tersedia — urutan sama dengan _chat_providers. Return (provider, model, teks).
    max_tokens dinaikkan oleh pemanggil yang outputnya panjang (narasi bulk)."""
    attempts = []
    if os.getenv("GROQ_API_KEY"):
        attempts.append(("Groq", GROQ_MODEL, lambda: _plain_openai_compatible(GROQ_API_URL, os.getenv("GROQ_API_KEY"), GROQ_MODEL, system, user, max_tokens)))
    if os.getenv("GROK_API_KEY"):
        attempts.append(("Grok", GROK_MODEL, lambda: _plain_openai_compatible(GROK_API_URL, os.getenv("GROK_API_KEY"), GROK_MODEL, system, user, max_tokens)))
    if os.getenv("OPEN_AI_API_KEY"):
        attempts.append(("OpenAI", OPENAI_MODEL, lambda: _plain_openai_compatible("https://api.openai.com/v1/chat/completions", os.getenv("OPEN_AI_API_KEY"), OPENAI_MODEL, system, user, max_tokens)))
    if os.getenv("CLOUDE_API_KEY"):
        attempts.append(("Claude", CLAUDE_MODEL, lambda: _plain_claude(os.getenv("CLOUDE_API_KEY"), system, user, max_tokens)))
    if os.getenv("GEMINI_API_KEY"):
        attempts.append(("Gemini", GEMINI_MODEL, lambda: _plain_gemini(os.getenv("GEMINI_API_KEY"), system, user, max_tokens)))
    if not attempts:
        raise HTTPException(500, "Tidak ada API key LLM di .env (GROQ/GROK/OPEN_AI/CLOUDE/GEMINI_API_KEY).")
    errors = []
    for provider, model, fn in attempts:
        try:
            return provider, model, fn()
        except Exception as e:  # coba provider berikutnya
            errors.append(f"{provider}: {e}")
    raise HTTPException(502, "Semua provider LLM gagal — " + "; ".join(errors))


_PENILAIAN_SCHEMA_PATH = BASE_DIR / "scripts" / "schema_penilaian_bappenas.sql"
_penilaian_table_ready = False


def _ensure_penilaian_table():
    """PostgreSQL (sejak migrasi 24 Jul 2026, lihat
    docs/migrasi_mysql_ke_postgresql.md) -- schema_penilaian_bappenas.sql
    ditulis dgn sintaks DDL MySQL (USE, TINYINT(1), ENGINE=, dst.), TIDAK
    bisa dieksekusi langsung ke Postgres. Tabelnya sendiri sudah dibuat
    scripts/migrate_pg_01_schema.py; kolom2 yang dulu ditambah belakangan
    lewat cek manual (MySQL 8 tidak dukung ADD COLUMN IF NOT EXISTS) di
    sini disederhanakan pakai sintaks native Postgres."""
    global _penilaian_table_ready
    if _penilaian_table_ready:
        return
    with db_cursor() as cur:
        cur.execute(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema='public' AND table_name='penilaian_bappenas_ai'"
        )
        if not cur.fetchone():
            raise RuntimeError(
                "Tabel penilaian_bappenas_ai belum ada di PostgreSQL -- jalankan "
                "scripts/migrate_pg_01_schema.py dulu."
            )
        cur.execute(
            "ALTER TABLE penilaian_bappenas_ai "
            "ADD COLUMN IF NOT EXISTS aspek_a_checklist BOOLEAN, "
            "ADD COLUMN IF NOT EXISTS aspek_a_total_kriteria SMALLINT, "
            "ADD COLUMN IF NOT EXISTS aspek_b_checklist BOOLEAN, "
            "ADD COLUMN IF NOT EXISTS aspek_b_total_indikator SMALLINT, "
            "ADD COLUMN IF NOT EXISTS aspek_b_narasi_ai TEXT"
        )
    _penilaian_table_ready = True


def _kemantapan_ruas_fakta(row: dict) -> dict | None:
    """Persentase TIDAK MANTAP ruas usulan -- fakta pendukung narasi AI Aspek
    B. Rumus (diperbarui 21 Jul 2026 atas permintaan user, dipakai jg utk
    perbaikan massal data lama di penilaian_bappenas_ai):
    (kondisi_ringan_km + kondisi_berat_km) / panjang_ruas_km x 100 -- BUKAN
    lagi panjang_penanganan_kompetensi (versi sebelumnya, docs/docs/
    Metodologi Penilaian skala prioritas Ruas.docx §F): penyebut lama itu
    ternyata sering tak konsisten dgn kondisi_ringan_km/kondisi_berat_km
    (51% dari 1.030 usulan >100% mentah, ekstrem sampai 1.522%), sedangkan
    panjang_ruas_km jauh lebih konsisten (cuma 2,9% dari 2.090 usulan
    >100%, dicek 21 Jul 2026). BEDA dari _ijd_score_kemantapan (parameter B
    resmi IJD A-E, penyebutnya total kondisi baik+sedang+ringan+berat) --
    narasi AI ini tetap pakai denominator terpisah sesuai permintaan user.
    None kalau data kondisi/panjang ruas belum diisi (supaya prompt tidak
    mengarang)."""
    ringan = row.get("kondisi_ringan_km")
    berat = row.get("kondisi_berat_km")
    panjang_ruas = row.get("panjang_ruas_km")
    if ringan is None or berat is None or not panjang_ruas or float(panjang_ruas) <= 0:
        return None
    # Dibatasi maks 100% -- sisa 2,9% kasus yg masih >100% mentah (data
    # kondisi_ringan_km/kondisi_berat_km & panjang_ruas_km sumbernya tetap
    # tidak 100% konsisten satu sama lain) di-clamp drpd dibuang sepenuhnya,
    # supaya cakupan fakta tetap ada (status Tidak Mantap tetap benar walau
    # >100% raw, krn ambang keputusannya cuma >50%).
    pct_tidak_mantap = round(min(100.0, (float(ringan) + float(berat)) / float(panjang_ruas) * 100), 2)
    return {
        "pct_tidak_mantap": pct_tidak_mantap,
        "pct_mantap": round(max(0.0, 100 - pct_tidak_mantap), 2),
        "status": "Tidak Mantap" if pct_tidak_mantap > 50 else "Mantap",
        "kondisi_ringan_km": float(ringan),
        "kondisi_berat_km": float(berat),
        "panjang_ruas_km": float(panjang_ruas),
    }


def _konektivitas_jalan_fakta(row: dict) -> dict | None:
    """Fakta konektivitas jaringan jalan usulan -- fakta pendukung narasi AI
    Aspek B (sama pola dgn _kemantapan_ruas_fakta), sumber
    usulan_konektivitas_jalan (spatial-join geometri usulan terhadap
    Maps/JALAN NASIONAL & JALAN PROVINSI & JALAN TOL, ambang_m — tabel yang
    sama dipakai NPR "Konektivitas Jaringan Jalan", lihat
    docs/kajian_metodologi_skala_prioritas_ruas.md). None kalau usulan belum
    ada baris di tabel itu (spatial-join belum menjangkau usulan ini)."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT terhubung_nasional, jarak_nasional_m, terhubung_provinsi, "
            "jarak_provinsi_m, terhubung_tol, jarak_tol_m, terhubung, ambang_m "
            "FROM usulan_konektivitas_jalan WHERE usulan_id = %s",
            (row.get("id"),),
        )
        r = cur.fetchone()
    if not r:
        return None
    return {
        "terhubung_jalan_nasional": bool(r["terhubung_nasional"]),
        "jarak_ke_jalan_nasional_m": float(r["jarak_nasional_m"]) if r["jarak_nasional_m"] is not None else None,
        "terhubung_jalan_provinsi": bool(r["terhubung_provinsi"]),
        "jarak_ke_jalan_provinsi_m": float(r["jarak_provinsi_m"]) if r["jarak_provinsi_m"] is not None else None,
        "terhubung_tol": bool(r["terhubung_tol"]),
        "jarak_ke_tol_m": float(r["jarak_tol_m"]) if r["jarak_tol_m"] is not None else None,
        "ambang_m": float(r["ambang_m"]),
    }


def _simpul_transportasi_fakta(row: dict) -> dict | None:
    """Fakta simpul transportasi (bandara/pelabuhan) terdekat dari kecamatan
    usulan -- fakta pendukung narasi AI Aspek B (sama pola dgn
    _kemantapan_ruas_fakta/_konektivitas_jalan_fakta), sumber
    simpul_transportasi_kecamatan_radius (radius tetap 30km per baris,
    scripts/spatial_join_simpul_transportasi.py) -- BEDA dari
    simpul_transportasi/NPR "Konektivitas Simpul Transportasi" yang cuma
    level kabupaten ada/tidak; tabel ini py jarak_km aktual per simpul
    per kecamatan, jadi bisa disebut nama & jaraknya, bukan cuma ada/tidak.
    None kalau usulan.kode_kecamatan NULL (spatial-join usulan→kecamatan
    belum jalan) ATAU tidak ada simpul dalam radius 30km dari kecamatan itu."""
    kode_kec = row.get("kode_kecamatan")
    if not kode_kec:
        return None
    with db_cursor() as cur:
        cur.execute(
            "SELECT jenis_simpul, nama_simpul, jarak_km, radius_km "
            "FROM simpul_transportasi_kecamatan_radius "
            "WHERE kode_kecamatan = %s ORDER BY jarak_km",
            (kode_kec,),
        )
        rows = cur.fetchall()
    if not rows:
        return None
    terdekat = {}
    for r in rows:
        terdekat.setdefault(r["jenis_simpul"], {  # jarak_km ASC -- yg pertama per jenis = terdekat
            "nama_simpul": r["nama_simpul"], "jarak_km": float(r["jarak_km"]),
        })
    return {"radius_km": rows[0]["radius_km"], "simpul_terdekat": terdekat}


def _kecamatan_dilalui_fakta(row: dict) -> dict | None:
    """Kecamatan yang dilalui SELURUH rute usulan (usulan_kecamatan_dilalui,
    hasil spatial_join_kecamatan_multi.py, 21 Jul 2026 -- BEDA dari
    usulan_inpres.kode_kecamatan yang cuma kecamatan DOMINAN tunggal) +
    jumlah penduduk tiap kecamatan itu (bps_kecamatan_demografi, tahun
    terbaru, sama tabel dgn "demografi_kecamatan_bps" di
    _bappenas_fakta_pendukung tapi utk SEMUA kecamatan yang dilalui, bukan
    cuma satu). "total_penduduk_dilalui" = jumlah penduduk seluruh
    kecamatan yang datanya ketemu (kecamatan yang datanya tidak ketemu
    dilewati dari total, BUKAN dihitung 0 -- "belum tersedia", bukan
    dikarang). "total_kendaraan_dilalui" (ditambah 23 Jul 2026, request
    eksplisit user) -- sama semantik & pola dgn total_penduduk_dilalui, tapi
    dari kendaraan_total (kecamatan_data_turunan, JOIN langsung by
    kode_kecamatan -- beda dari bps_kecamatan_demografi yang perlu
    name-matching) -- supaya narasi AI tidak lagi cuma sebut angka
    kendaraan SATU kecamatan dominan (dari aspek_b["keterangan"]) padahal
    kalimat "kecamatan dilalui" di sampingnya membahas SEMUA kecamatan yang
    dilintasi rute. None kalau usulan ini sama sekali tidak punya baris di
    usulan_kecamatan_dilalui (geometri blm ter-proses/tak ditemukan)."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT ukd.kode_kecamatan, pk.kecamatan, pk.kode_kabupaten, pk.kabupaten_kota "
            "FROM usulan_kecamatan_dilalui ukd "
            "JOIN penduduk_kecamatan pk ON pk.kode_kecamatan = ukd.kode_kecamatan "
            "WHERE ukd.usulan_id = %s ORDER BY ukd.n_titik_sampel DESC",
            (row.get("id"),),
        )
        rows = cur.fetchall()
    if not rows:
        return None
    daftar = []
    total_penduduk = 0
    ada_penduduk = False
    total_kendaraan = 0
    ada_kendaraan = False
    # bps_kecamatan_demografi tidak punya kode_kecamatan -- dicocokkan lewat
    # kode_kab + nama kecamatan master, sama pola dgn
    # _bappenas_fakta_pendukung; per-kecamatan (bukan 1 query IN gabungan)
    # krn jumlah kecamatan per usulan kecil (rata-rata 1,7, maks ~23).
    with db_cursor() as cur:
        for r in rows:
            cur.execute(
                "SELECT jumlah_penduduk FROM bps_kecamatan_demografi "
                "WHERE kode_kab=%s AND UPPER(kecamatan)=UPPER(%s) ORDER BY tahun DESC LIMIT 1",
                (str(r["kode_kabupaten"]), r["kecamatan"]),
            )
            demo = cur.fetchone()
            penduduk = demo["jumlah_penduduk"] if demo else None
            if penduduk is not None:
                total_penduduk += penduduk
                ada_penduduk = True
            cur.execute(
                "SELECT kendaraan_total FROM kecamatan_data_turunan "
                "WHERE kode_kecamatan=%s ORDER BY tahun DESC LIMIT 1",
                (r["kode_kecamatan"],),
            )
            kdt = cur.fetchone()
            kendaraan = kdt["kendaraan_total"] if kdt else None
            if kendaraan is not None:
                total_kendaraan += kendaraan
                ada_kendaraan = True
            daftar.append({
                "kecamatan": r["kecamatan"], "kabupaten_kota": r["kabupaten_kota"],
                "jumlah_penduduk": penduduk, "jumlah_kendaraan": kendaraan,
            })
    return {
        "kecamatan_dilalui": daftar,
        "jumlah_kecamatan_dilalui": len(daftar),
        "total_penduduk_dilalui": total_penduduk if ada_penduduk else None,
        "total_kendaraan_dilalui": total_kendaraan if ada_kendaraan else None,
    }


def _penilaian_context(row: dict, aspek_a_hasil: dict, aspek_b_hasil: dict) -> str:
    """Ringkasan data usulan (JSON) yang jadi satu-satunya bahan penilaian AI
    (skr cuma "kesimpulan" — aspek_a & aspek_b dihitung rule-based, lihat
    _bappenas_aspek_a_lokus/_bappenas_aspek_b_ekonomi, disisipkan APA ADANYA
    sbg fakta given)."""
    ijd = _compute_ijd_score(row, 2026)
    spn = _skor_prioritas_nasional(row)
    data = {
        "aspek_a_hasil": aspek_a_hasil,
        "aspek_b_hasil": aspek_b_hasil,
        "usulan": {k: row.get(k) for k in (
            "id", "nama_kegiatan", "nama_ruas", "provinsi", "kabupaten_kota",
            "jenis_penanganan", "status_ruas", "panjang_ruas_km", "panjang_penanganan_pemda",
            "panjang_penanganan_kompetensi", "alokasi_usulan_pemda", "alokasi_usulan_kompetensi",
            "kapasitas_fiskal", "tematik_kawasan_pemda", "kondisi_baik_km", "kondisi_sedang_km",
            "kondisi_ringan_km", "kondisi_berat_km", "verifikasi_balai",
            "verifikasi_kompetensi", "prioritas_balai", "prioritas_kompetensi",
            "status_koridor_balai", "penuntasan_ijd_kompetensi", "lanjutan_ijd_2025",
            "jenis_data_dukung_tematik_kompetensi", "indikasi_prioritas_pu",
        )},
        "skor_teknokratik_ijd_2026": {
            "skor_ternormalisasi_100": ijd["skor_ternormalisasi_100"],
            "komponen": [{k: c.get(k) for k in ("kode", "label", "tersedia", "nilai", "keterangan")}
                         for c in ijd["komponen"]],
        },
        "skor_prioritas_nasional": {
            "total": spn["skor_total"],
            "komponen": [{k: c.get(k) for k in ("kode", "label", "nilai", "keterangan")}
                         for c in spn["komponen"]],
        },
    }
    kemantapan = _kemantapan_ruas_fakta(row)
    if kemantapan:
        data["kemantapan_ruas"] = kemantapan
    konektivitas = _konektivitas_jalan_fakta(row)
    if konektivitas:
        data["konektivitas_jalan"] = konektivitas
    simpul = _simpul_transportasi_fakta(row)
    if simpul:
        data["simpul_transportasi"] = simpul
    kecamatan_dilalui = _kecamatan_dilalui_fakta(row)
    if kecamatan_dilalui:
        data["kecamatan_dilalui"] = kecamatan_dilalui
    komoditas = _komoditas_perkebunan_fakta(row)
    if komoditas:
        data["komoditas_perkebunan_kecamatan"] = komoditas
    return json.dumps(jsonable_encoder(data), ensure_ascii=False)


@app.get("/api/usulan-inpres/{usulan_id}/penilaian-bappenas")
def penilaian_bappenas_get(usulan_id: int):
    _ensure_penilaian_table()
    with db_cursor() as cur:
        cur.execute("SELECT * FROM penilaian_bappenas_ai WHERE usulan_id = %s", (usulan_id,))
        row = cur.fetchone()
    if not row:
        return {"tersedia": False}
    row["tersedia"] = True
    return jsonable_encoder(row)


def _bappenas_poin_from_total(total: int) -> int:
    """0/1/2 dari jumlah kriteria Aspek A yang cocok -- 0=tidak ada, 1=1-2
    cocok, 2=>=3 cocok. Distribusi nasional (23 Jul 2026, 3.072 usulan)
    masih tersebar wajar dgn ambang ini (Poin 0/1/2 = 5,2%/53,1%/41,7%),
    jadi TIDAK diubah -- lihat _bappenas_poin_b_from_total utk kenapa Aspek
    B butuh ambang terpisah."""
    if total <= 0:
        return 0
    return 2 if total >= 3 else 1


def _bappenas_poin_b_from_total(total: int) -> int:
    """0/1/2 dari jumlah indikator Aspek B yang ada -- ambang TERPISAH dari
    Aspek A (_bappenas_poin_from_total), ditambah 23 Jul 2026 setelah
    ditemukan kolom "TOTAL" (Bappenas, AM di export) jenuh di ujung atas:
    ambang lama "0/1-2/>=3" dikalibrasi wkt BAPPENAS_ASPEK_B_INDIKATOR_LABEL
    baru py 12 indikator -- setelah diperluas ke 17 (LBS, Indeks Penanaman,
    Konektivitas Jaringan Jalan, Konektivitas Simpul Transportasi,
    Penuntasan Koridor ditambahkan 23 Jul 2026, hampir selalu tersedia
    scr nasional), distribusi riil 3.072 usulan JADI 6-14 indikator/usulan
    (minimum jauh di atas ambang lama 3) -- Poin 2 ("≥3") ke-hit 99,6% usulan,
    bikin kolom TOTAL Bappenas (Poin A + Poin B) & RANGKING DALAM PROVINSI
    yang berbasis situ kehilangan daya beda.

    Ambang baru: >=50% dari 17 indikator (>=9) = Poin 2, sisanya (1-8) =
    Poin 1, sama pola "mayoritas tersedia" dgn _npr_kelas_dari_hitung dkk di
    tempat lain -- hasil pada data yg sama: Poin 0/1/2 = 0%/66,7%/33,3%,
    jauh lebih menyebar drpd 0%/0,4%/99,6% sebelumnya. Poin 0 (total<=0)
    scr praktik nyaris tak pernah ke-hit nasional (minimum indikator
    ketemu 2) -- itu bawaan data, bukan bug ambang ini."""
    if total <= 0:
        return 0
    return 2 if total >= 9 else 1


@app.post("/api/usulan-inpres/{usulan_id}/penilaian-bappenas")
def penilaian_bappenas_generate(usulan_id: int):
    _ensure_penilaian_table()
    with db_cursor() as cur:
        cur.execute("SELECT * FROM usulan_inpres WHERE id = %s", (usulan_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Usulan tidak ditemukan")

    aspek_a_hasil = _bappenas_aspek_a_lokus(row)
    aspek_b_hasil = _bappenas_aspek_b_ekonomi(row)
    poin_a = _bappenas_poin_from_total(aspek_a_hasil["total_kriteria"])
    poin_b = _bappenas_poin_b_from_total(aspek_b_hasil["total_indikator"])

    provider, model, teks = _llm_plain(
        PENILAIAN_SYSTEM_PROMPT, _penilaian_context(row, aspek_a_hasil, aspek_b_hasil)
    )
    teks = re.sub(r"^```(?:json)?\s*|\s*```$", "", teks.strip())
    try:
        hasil = json.loads(teks)
        kesimpulan = hasil["kesimpulan"]
        aspek_b_narasi_ai = hasil.get("aspek_b_narasi_ai")
    except (ValueError, KeyError, TypeError) as e:
        raise HTTPException(502, f"Jawaban {provider} tidak sesuai format JSON penilaian: {e}")

    with db_cursor() as cur:
        cur.execute(
            "INSERT INTO penilaian_bappenas_ai (usulan_id, aspek_a_poin, aspek_a_checklist, "
            "aspek_a_total_kriteria, aspek_a_narasi, aspek_b_poin, aspek_b_checklist, "
            "aspek_b_total_indikator, aspek_b_narasi, aspek_b_narasi_ai, total_poin, kesimpulan, "
            "provider, model) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
            "ON CONFLICT (usulan_id) DO UPDATE SET aspek_a_poin=EXCLUDED.aspek_a_poin, "
            "aspek_a_checklist=EXCLUDED.aspek_a_checklist, aspek_a_total_kriteria=EXCLUDED.aspek_a_total_kriteria, "
            "aspek_a_narasi=EXCLUDED.aspek_a_narasi, aspek_b_poin=EXCLUDED.aspek_b_poin, "
            "aspek_b_checklist=EXCLUDED.aspek_b_checklist, aspek_b_total_indikator=EXCLUDED.aspek_b_total_indikator, "
            "aspek_b_narasi=EXCLUDED.aspek_b_narasi, aspek_b_narasi_ai=EXCLUDED.aspek_b_narasi_ai, "
            "total_poin=EXCLUDED.total_poin, "
            # generated_at TIDAK ikut di-EXCLUDED -- MySQL asalnya auto-refresh
            # lewat "ON UPDATE CURRENT_TIMESTAMP" (tidak ada padanan clause-level
            # di PostgreSQL, butuh trigger), diganti set eksplisit di sini spy
            # perilaku "waktu generate ulang" tetap sama.
            "kesimpulan=EXCLUDED.kesimpulan, provider=EXCLUDED.provider, model=EXCLUDED.model, "
            "generated_at=CURRENT_TIMESTAMP",
            (usulan_id, poin_a, aspek_a_hasil["checklist"], aspek_a_hasil["total_kriteria"],
             aspek_a_hasil["narasi"], poin_b, aspek_b_hasil["checklist"], aspek_b_hasil["total_indikator"],
             aspek_b_hasil["narasi"], aspek_b_narasi_ai, poin_a + poin_b, kesimpulan, provider, model),
        )
    _ijd_bulk_cache.clear()  # kesimpulan/narasi AI baru -> preview & export xlsx kadaluarsa
    return penilaian_bappenas_get(usulan_id)


PENILAIAN_BULK_SYSTEM_PROMPT = """Anda membantu analis Bappenas menyusun DRAF narasi Aspek B (Daya Ungkit \
Ekonomi & Kinerja Sektoral) untuk BEBERAPA usulan Inpres Jalan Daerah sekaligus. Untuk SETIAP usulan pada \
data JSON yang diberikan, susun narasi gaya laporan kebijakan yang mengalir dan meyakinkan — BUKAN \
enumerasi kaku bergaya "Ditemukan N indikator: ...; ...". Narasi WAJIB menyinggung SETIAP indikator di \
"indikator_ada" — jangan diam-diam melewati sebagian demi keringkasan; kalau indikatornya banyak (>4), \
pakai lebih banyak kalimat (boleh sampai 8-10) supaya semua kebagian tempat, dikelompokkan per tema (mis. \
produksi pangan jadi satu kalimat, konektivitas/lalu lintas kalimat lain) alih-alih satu kalimat generik \
per indikator. Rangkai indikator beserta angka konkret pada "fakta" menjadi cerita dampak ekonomi/sektoral \
ruas tersebut (ketahanan pangan, kelancaran logistik, pertumbuhan ekonomi lokal). "fakta" berisi \
"ringkasan_indikator" plus data BPS pendukung: "demografi_kecamatan_bps" (level kecamatan) serta \
"padi_kabupaten_bps" dan "kendaraan_kabupaten_bps" (level KABUPATEN/KOTA — bila disebut, WAJIB \
diatribusikan ke kabupaten/kota, jangan ditulis seolah angka kecamatan atau ruas). Sebut nama kegiatan/\
koridor, wilayah, dan angka dari data bila relevan — dan SETIAP kali menyebut angka dari \
demografi_kecamatan_bps/padi_kabupaten_bps/kendaraan_kabupaten_bps, sertakan sumbernya secara ringkas \
dalam kalimat (mis. "...jiwa (BPS Dalam Angka)", "...ton/ha (BPS Dalam Angka)"), jangan sebut angka \
telanjang tanpa atribusi.

CEK SATU-PER-SATU untuk SETIAP usulan — kalau field-nya ADA di "fakta" usulan itu, kalimat terkait WAJIB \
masuk ke narasi usulan itu (jangan diam-diam dilewati; kalau TIDAK ada field-nya utk usulan itu, jangan \
menyinggung topik itu sama sekali utk usulan itu):
- "kemantapan_ruas" ada → satu kalimat pakai PERSIS field "pct_tidak_mantap" (jangan hitung ulang), gaya \
"Dari aspek teknis, kondisi ruas ini tergolong <status> dengan persentase sekitar \
<pct_tidak_mantap>% (kondisi rusak ringan+berat dibanding panjang penanganan kompetensi)."
- "konektivitas_jalan" ada → WAJIB satu kalimat, TERMASUK kalau ketiga "terhubung_..." semuanya false \
(jangan dilewati hanya krn tidak ada yang terhubung — ini tetap fakta yang harus dilaporkan): kalau ada \
yang true, sebut jaringan mana ("terhubung_jalan_nasional"/"terhubung_jalan_provinsi"/"terhubung_tol") \
beserta jarak meter dari "jarak_ke_..._m" berpasangan; kalau semuanya false, tulis persis gaya "Ruas ini \
belum terhubung langsung ke jaringan jalan nasional/provinsi/tol dalam radius <ambang_m> meter."
- "simpul_transportasi" ada → satu kalimat sebut simpul transportasi (bandara/pelabuhan) terdekat dari \
"simpul_terdekat" — jenis, "nama_simpul", "jarak_km"-nya, dalam radius "radius_km" km.
- "kecamatan_dilalui" ada → satu kalimat sebut SEMUA nama kecamatan pada daftar "kecamatan_dilalui" \
(jangan cuma sebagian kalau lebih dari 1) beserta PERSIS field "total_penduduk_dilalui" (jangan hitung \
ulang/jumlah manual dari daftar), gaya "Ruas ini melintasi Kecamatan A, Kecamatan B, dan Kecamatan C \
dengan total penduduk terdampak sekitar <total_penduduk_dilalui> jiwa (BPS Dalam Angka)." — kalau \
"total_penduduk_dilalui" null, sebut nama kecamatannya saja tanpa angka total, jangan mengarang angka. \
Kalau field "total_kendaraan_dilalui" pada "kecamatan_dilalui" TIDAK null, WAJIB pakai angka itu (jumlah \
kendaraan GABUNGAN semua kecamatan yang dilalui) setiap kali menyebut jumlah kendaraan ruas ini — JANGAN \
pakai angka kendaraan satu kecamatan saja (mis. dari "ringkasan_indikator") kalau usulan ini melintasi >1 \
kecamatan; kalau "total_kendaraan_dilalui" null, boleh sebut angka kendaraan level kabupaten dari \
"kendaraan_kabupaten_bps" kalau ada (jangan mengarang).
- "komoditas_perkebunan_kecamatan" ada → boleh satu kalimat sebut komoditas perkebunan UTAMA kecamatan ini \
persis nama pada field "komoditas" (urut "produksi_ton" terbesar, maksimal 3 komoditas) beserta angka \
produksi_ton-nya; JANGAN sebut nama komoditas perkebunan LAIN yang tidak ada di daftar ini (mis. jangan \
menyebut kelapa sawit/karet/tebu kalau tidak ada di daftar "komoditas_perkebunan_kecamatan" — daftar ini \
sudah cek data riil per komoditas, bukan cuma tebakan). Kalau field ini TIDAK ada padahal \
"ringkasan_indikator" menyebut PRODUKSI_PERKEBUNAN, boleh sebut produksi perkebunan secara GENERIK (pakai \
angka ton dari "ringkasan_indikator") TANPA menyebut nama komoditas spesifik apa pun.
SEBELUM membalas, jalankan KELIMA cek di atas SATU USULAN PADA SATU WAKTU (bukan sekali generik utk semua) \
— usulan dalam satu batch ini "fakta"-nya BEDA-BEDA, jangan asumsikan checklist yang berlaku di satu \
usulan otomatis berlaku sama utk usulan lain di batch yang sama.

WAJIB berbasis fakta yang diberikan APA ADANYA — jangan mengarang angka atau fakta di luar data. Bila \
"indikator_ada" sebuah usulan kosong, nyatakan jujur belum ada indikator daya ungkit yang didukung data \
untuk lokasi itu, jangan dibuat seolah ada.

GAYA BAHASA: formal dan padat — setiap kalimat wajib membawa informasi baru (data/indikator/angka), jangan \
mengulang frasa pembuka atau kesimpulan umum ("hal ini menunjukkan...", "secara keseluruhan...", dsb.) yang \
tidak menambah fakta, dan jangan menyebut ulang satu indikator/angka yang sama di lebih dari satu kalimat \
dalam narasi usulan yang sama.

VARIASI ANTAR USULAN — WAJIB, ini bukan saran opsional. Narasi tiap usulan dalam batch ini akan berdampingan \
di satu file export, jadi terasa ditulis khusus utk usulan itu, bukan template yang di-copy-paste dgn nama/\
angka diganti. DILARANG KERAS memakai kalimat pembuka dgn kerangka yang SAMA PERSIS utk >1 usulan dalam \
batch ini, contoh pola yang HARUS dihindari kalau sudah dipakai di usulan sebelumnya pada batch yang sama: \
"<Nama Kegiatan> di <Kabupaten/Kota>, <Provinsi>, memiliki panjang penanganan sekitar <angka> km." — begitu \
juga kalimat PENUTUP, DILARANG selalu berakhir dgn variasi "Peningkatan/Kegiatan ... jalan ini diharapkan \
dapat mendukung/meningkatkan ... pertumbuhan ekonomi lokal" di semua usulan. Untuk tiap usulan pilih SATU dari \
strategi pembuka berikut secara BERGANTIAN dalam batch ini (jangan pakai strategi yang sama 2x berturut-turut): \
(a) mulai dari kondisi kemantapan/teknis ruas, (b) mulai dari lokasi/koridor/kabupaten tanpa menyebut panjang \
km di kalimat pertama, (c) mulai dari isu konektivitas jaringan jalan, (d) mulai dari potensi ekonomi/produksi \
kawasan yang dilalui. Kalimat penutup juga wajib bervariasi kalimat & sudut pandang, bukan rumus tetap \
"diharapkan dapat mendukung X". Variasi ini HANYA soal gaya bahasa/struktur kalimat — kelengkapan fakta, \
checklist wajib di atas, dan larangan mengarang tetap berlaku sama ketatnya utk setiap usulan.

Balas HANYA JSON valid tanpa teks lain, format:
[{"id": <id usulan>, "narasi": "..."}, ...]
— satu objek per usulan, "id" disalin apa adanya dari data. Bahasa Indonesia formal."""

# Usulan per panggilan LLM — model diminta membalas array JSON satu objek per
# usulan, lalu di-upsert per baris. Narasi kini wajib menyinggung SEMUA
# indikator_ada (bisa sampai 8-10 kalimat utk usulan dgn banyak indikator,
# bukan 3-5 lagi) -- batch diperkecil dari 30 ke 20 supaya total token
# output per panggilan LLM tetap wajar di bawah max_tokens (lihat
# _PENILAIAN_BULK_MAX_TOKENS) tanpa mengubah max_tokens setinggi mungkin
# risiko melewati batas keras sebagian provider. Diperkecil lagi 20->12
# (22 Jul 2026) -- audit menemukan model (gpt-4o-mini via OpenAI, provider
# pertama yang tersedia di sesi ini) SERING melewatkan checklist
# konektivitas_jalan/simpul_transportasi (§ instruksi CEK SATU-PER-SATU) di
# batch 20, walau fakta-nya ADA di payload (diverifikasi manual lewat
# _bappenas_fakta_pendukung) -- makin banyak usulan per panggilan, makin
# besar kemungkinan checklist per-usulan "terlewat" krn model memprioritaskan
# panjang/kelengkapan sebagian usulan drpd kepatuhan merata ke semuanya.
# Diperkecil lagi 12->5 (22 Jul 2026, keputusan eksplisit user setelah batch
# 12 masih kadang melewatkan checklist) -- pilihan drpd alternatif "pakai
# Claude khusus utk endpoint ini" (lebih taat instruksi tapi lebih mahal per
# token) supaya tetap gratis/konsisten dgn urutan provider default
# (_llm_plain: Groq->Grok->OpenAI->Claude->Gemini). Konsekuensi: proses
# narasi AI bulk makin banyak panggilan LLM per provinsi (makin lambat/makin
# banyak klik "Proses Narasi AI" di UI) -- trade-off yang disengaja demi
# kepatuhan checklist per usulan, bukan bug performa.
_PENILAIAN_BULK_BATCH = 5
_PENILAIAN_BULK_MAX_TOKENS = 8192


_KOMODITAS_PERKEBUNAN_CANONICAL = {
    "Coconut": "Kelapa", "Coffee": "Kopi", "Cocoa": "Kakao", "Betel": "Sirih",
    "Hazelnut": "Kemiri", "Tobacco": "Tembakau", "Cashew Nut": "Jambu Mete",
    "Areca Nut": "Pinang",
}


def _komoditas_perkebunan_fakta(row: dict) -> list | None:
    """Rincian PER-KOMODITAS produksi perkebunan rakyat kecamatan usulan
    (bps_kecamatan_produksi_komoditas, extract_dalam_angka.py), dipakai
    narasi AI Aspek B supaya TIDAK mengarang nama komoditas generik (bug
    23 Jul 2026: keterangan lama menyebut "kelapa sawit/kelapa/karet/tebu"
    untuk SEMUA usulan tanpa cek data riil -- ditemukan lewat verifikasi
    usulan Sumba Timur yang sama sekali tidak membudidayakan sawit/tebu/
    karet, cuma kelapa). Tabel ini TIDAK punya kode_kecamatan terisi
    (selalu NULL) -- dicocokkan lewat kode_kab + nama kecamatan master,
    sama pola dgn _bappenas_fakta_pendukung. Baris bilingual duplikat
    (mis. Kelapa/Coconut, Kopi/Coffee, hasil header tabel PDF dwibahasa)
    di-dedupe ke nama Indonesia saja lewat _KOMODITAS_PERKEBUNAN_CANONICAL.
    Hanya kategori PERKEBUNAN & produksi_ton>0 -- None kalau tidak ada
    baris cocok (bukan berarti kecamatan itu tidak berkebun sama sekali,
    cuma dalam_angka/ belum ter-ekstrak/tidak ada baris utk kecamatan itu)."""
    kode_kec = row.get("kode_kecamatan")
    if not kode_kec:
        return None
    kode_kab = _bappenas_kode_kab(row)
    if not kode_kab:
        return None
    with db_cursor() as cur:
        cur.execute(
            "SELECT kecamatan FROM penduduk_kecamatan WHERE kode_kecamatan=%s "
            "ORDER BY tahun DESC LIMIT 1", (kode_kec,),
        )
        pk = cur.fetchone()
    if not pk:
        return None
    with db_cursor() as cur:
        cur.execute(
            "SELECT jenis_tanaman, produksi_ton FROM bps_kecamatan_produksi_komoditas "
            "WHERE kode_kab=%s AND UPPER(kecamatan)=UPPER(%s) AND kategori='PERKEBUNAN' "
            "AND produksi_ton IS NOT NULL AND produksi_ton > 0",
            (str(kode_kab), pk["kecamatan"]),
        )
        rows = cur.fetchall()
    if not rows:
        return None
    by_nama = {}
    for r in rows:
        nama = _KOMODITAS_PERKEBUNAN_CANONICAL.get(r["jenis_tanaman"], r["jenis_tanaman"])
        by_nama[nama] = max(by_nama.get(nama, 0), float(r["produksi_ton"]))
    return [
        {"komoditas": nama, "produksi_ton": ton}
        for nama, ton in sorted(by_nama.items(), key=lambda kv: kv[1], reverse=True)
    ]


def _bappenas_fakta_pendukung(row: dict, aspek_b: dict) -> dict:
    """Data pendukung BPS untuk narasi AI Aspek B (bulk) — melengkapi
    aspek_b["keterangan"] (yang sudah memuat angka kecamatan_data_turunan +
    bps_kecamatan_potensi_tematik) dengan angka level kecamatan
    (penduduk_kecamatan, bps_kecamatan_demografi) dan level kabupaten
    (bps_kabupaten_padi, bps_kabupaten_kendaraan), plus
    "konektivitas_jalan" (usulan_konektivitas_jalan, lihat
    _konektivitas_jalan_fakta) — status terhubung/tidak & jarak ke jaringan
    jalan nasional/provinsi/tol, "simpul_transportasi"
    (simpul_transportasi_kecamatan_radius, lihat
    _simpul_transportasi_fakta) — bandara/pelabuhan terdekat & jaraknya,
    dan "kecamatan_dilalui" (usulan_kecamatan_dilalui, lihat
    _kecamatan_dilalui_fakta) — SEMUA kecamatan yang dilintasi rute
    (bukan cuma kode_kecamatan dominan) + total penduduknya gabungan.
    Nilai yang tidak tersedia dihilangkan dari dict supaya prompt tetap
    ringkas. Cakupan bps_* mengikuti isi dalam_angka/ (parsial per
    provinsi) — tidak apa-apa kosong."""
    kode_kec = row.get("kode_kecamatan")
    kode_kab = _bappenas_kode_kab(row)
    fakta = {"ringkasan_indikator": aspek_b["keterangan"]}

    nama_kec = None
    if kode_kec:
        with db_cursor() as cur:
            cur.execute(
                "SELECT kecamatan FROM penduduk_kecamatan "
                "WHERE kode_kecamatan=%s ORDER BY tahun DESC LIMIT 1",
                (kode_kec,),
            )
            pk = cur.fetchone()
        if pk:
            nama_kec = pk["kecamatan"]
            fakta["kecamatan"] = nama_kec
    # bps_kecamatan_demografi tidak punya kode_kecamatan — dicocokkan lewat
    # kode_kab + nama kecamatan master; gagal cocok = lewati saja (best-effort).
    if kode_kab and nama_kec:
        with db_cursor() as cur:
            cur.execute(
                "SELECT tahun, jumlah_penduduk, laju_pertumbuhan_pct, kepadatan_per_km2, "
                "rasio_jenis_kelamin FROM bps_kecamatan_demografi "
                "WHERE kode_kab=%s AND UPPER(kecamatan)=UPPER(%s) ORDER BY tahun DESC LIMIT 1",
                (str(kode_kab), nama_kec),
            )
            demo = cur.fetchone()
        if demo:
            fakta["demografi_kecamatan_bps"] = {k: v for k, v in demo.items() if v is not None}
    if kode_kab:
        with db_cursor() as cur:
            cur.execute(
                "SELECT tahun, luas_panen_ha, produktivitas_ku_ha, produksi_ton "
                "FROM bps_kabupaten_padi WHERE kode_kab=%s ORDER BY tahun DESC LIMIT 1",
                (str(kode_kab),),
            )
            padi = cur.fetchone()
            cur.execute(
                "SELECT tahun, mobil_penumpang, bus, mobil_barang, sepeda_motor, jumlah "
                "FROM bps_kabupaten_kendaraan WHERE kode_kab=%s ORDER BY tahun DESC LIMIT 1",
                (str(kode_kab),),
            )
            kendaraan = cur.fetchone()
        if padi:
            fakta["padi_kabupaten_bps"] = {k: v for k, v in padi.items() if v is not None}
        if kendaraan:
            fakta["kendaraan_kabupaten_bps"] = {k: v for k, v in kendaraan.items() if v is not None}
    kemantapan = _kemantapan_ruas_fakta(row)
    if kemantapan:
        fakta["kemantapan_ruas"] = kemantapan
    konektivitas = _konektivitas_jalan_fakta(row)
    if konektivitas:
        fakta["konektivitas_jalan"] = konektivitas
    simpul = _simpul_transportasi_fakta(row)
    if simpul:
        fakta["simpul_transportasi"] = simpul
    kecamatan_dilalui = _kecamatan_dilalui_fakta(row)
    if kecamatan_dilalui:
        fakta["kecamatan_dilalui"] = kecamatan_dilalui
    komoditas = _komoditas_perkebunan_fakta(row)
    if komoditas:
        fakta["komoditas_perkebunan_kecamatan"] = komoditas
    return fakta


@app.post("/api/usulan-inpres/penilaian-bappenas/bulk")
def penilaian_bappenas_bulk(provinsi: str = "", force: bool = False, after_id: int = 0):
    """Generate narasi AI Aspek B massal — SATU batch (<= _PENILAIAN_BULK_BATCH
    usulan, satu panggilan LLM) per request; frontend memanggil berulang sampai
    sisa=0 supaya ada progres dan request tidak kena timeout. SENGAJA hanya
    per provinsi (bukan nasional) supaya pemakaian kuota LLM terkendali.
    Default resume-able: usulan yang sudah punya aspek_b_narasi_ai dilewati,
    TIDAK ditimpa. force=true membalik itu — proses ULANG semua usulan
    provinsi ini walau sudah punya narasi (dipakai setelah prompt/kaidah
    narasi berubah, mis. narasi lama dari sebelum prompt diperluas supaya
    menyinggung SEMUA indikator_ada, bukan cuma 3-5 kalimat pilihan model).
    Aspek A/B rule-based ikut di-upsert supaya panel detail tetap utuh;
    "kesimpulan" TETAP hanya digenerate per-usulan lewat
    POST /api/usulan-inpres/{id}/penilaian-bappenas."""
    provinsi = (provinsi or "").strip()
    if not provinsi:
        raise HTTPException(400, "Proses bulk narasi AI hanya bisa per provinsi — pilih filter provinsi dulu.")
    _ensure_penilaian_table()
    # force=True mengabaikan filter "narasi belum ada" (jadi TIDAK menyusut
    # otomatis antar panggilan) — pakai kursor id (after_id) supaya batch
    # berikutnya maju, bukan mengulang batch pertama selamanya.
    with db_cursor() as cur:
        if force:
            cur.execute(
                "SELECT u.* FROM usulan_inpres u "
                "WHERE u.provinsi = %s AND u.id > %s ORDER BY u.id",
                (provinsi, after_id),
            )
        else:
            cur.execute(
                "SELECT u.* FROM usulan_inpres u "
                "LEFT JOIN penilaian_bappenas_ai p ON p.usulan_id = u.id "
                "WHERE u.provinsi = %s AND (p.aspek_b_narasi_ai IS NULL OR p.aspek_b_narasi_ai = '') "
                "ORDER BY u.id",
                (provinsi,),
            )
        pending = cur.fetchall()
    if not pending:
        return {"diproses": 0, "sisa": 0, "provinsi": provinsi, "next_after_id": after_id}

    rows = pending[:_PENILAIAN_BULK_BATCH]
    hasil = {}
    payload = []
    for row in rows:
        aspek_a = _bappenas_aspek_a_lokus(row)
        aspek_b = _bappenas_aspek_b_ekonomi(row)
        hasil[row["id"]] = (aspek_a, aspek_b)
        payload.append({
            "id": row["id"],
            "nama_kegiatan": row.get("nama_kegiatan"),
            "nama_koridor": row.get("nama_koridor"),
            "provinsi": row.get("provinsi"),
            "kabupaten_kota": row.get("kabupaten_kota"),
            "jenis_penanganan": row.get("jenis_penanganan"),
            "panjang_penanganan_km": row.get("panjang_penanganan_pemda"),
            "indikator_ada": [BAPPENAS_ASPEK_B_INDIKATOR_LABEL.get(k, k) for k in aspek_b["indikator_ada"]],
            "fakta": _bappenas_fakta_pendukung(row, aspek_b),
        })

    provider, model, teks = _llm_plain(
        PENILAIAN_BULK_SYSTEM_PROMPT, json.dumps(jsonable_encoder(payload), ensure_ascii=False),
        max_tokens=_PENILAIAN_BULK_MAX_TOKENS,
    )
    teks = re.sub(r"^```(?:json)?\s*|\s*```$", "", teks.strip())
    try:
        narasi_by_id = {int(item["id"]): (item.get("narasi") or "").strip()
                        for item in json.loads(teks)}
    except (ValueError, KeyError, TypeError) as e:
        raise HTTPException(502, f"Jawaban {provider} tidak sesuai format JSON narasi bulk: {e}")

    diproses = 0
    with db_cursor() as cur:
        for row in rows:
            narasi = narasi_by_id.get(row["id"])
            if not narasi:
                continue  # id tidak dijawab model — dicoba ulang di batch berikutnya
            aspek_a, aspek_b = hasil[row["id"]]
            poin_a = _bappenas_poin_from_total(aspek_a["total_kriteria"])
            poin_b = _bappenas_poin_b_from_total(aspek_b["total_indikator"])
            # kesimpulan sengaja TIDAK disentuh (kolom per-usulan, lihat docstring)
            cur.execute(
                "INSERT INTO penilaian_bappenas_ai (usulan_id, aspek_a_poin, aspek_a_checklist, "
                "aspek_a_total_kriteria, aspek_a_narasi, aspek_b_poin, aspek_b_checklist, "
                "aspek_b_total_indikator, aspek_b_narasi, aspek_b_narasi_ai, total_poin, provider, model) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                "ON CONFLICT (usulan_id) DO UPDATE SET aspek_a_poin=EXCLUDED.aspek_a_poin, "
                "aspek_a_checklist=EXCLUDED.aspek_a_checklist, aspek_a_total_kriteria=EXCLUDED.aspek_a_total_kriteria, "
                "aspek_a_narasi=EXCLUDED.aspek_a_narasi, aspek_b_poin=EXCLUDED.aspek_b_poin, "
                "aspek_b_checklist=EXCLUDED.aspek_b_checklist, aspek_b_total_indikator=EXCLUDED.aspek_b_total_indikator, "
                "aspek_b_narasi=EXCLUDED.aspek_b_narasi, aspek_b_narasi_ai=EXCLUDED.aspek_b_narasi_ai, "
                "total_poin=EXCLUDED.total_poin, provider=EXCLUDED.provider, model=EXCLUDED.model, "
                "generated_at=CURRENT_TIMESTAMP",
                (row["id"], poin_a, aspek_a["checklist"], aspek_a["total_kriteria"], aspek_a["narasi"],
                 poin_b, aspek_b["checklist"], aspek_b["total_indikator"], aspek_b["narasi"],
                 narasi, poin_a + poin_b, provider, model),
            )
            diproses += 1
    if diproses == 0:
        # lindungi frontend dari loop tak berujung kalau model menjawab tapi id-nya salah semua
        raise HTTPException(502, f"{provider} menjawab, tetapi tidak ada narasi valid untuk batch ini.")
    _ijd_bulk_cache.clear()  # narasi AI baru -> preview & export xlsx kadaluarsa
    return {"diproses": diproses, "sisa": len(pending) - (len(rows) if force else diproses),
            "provider": provider, "model": model, "provinsi": provinsi,
            "next_after_id": rows[-1]["id"] if force and rows else after_id}


_KML_COORD_RE = re.compile(
    r"<(?:LineString|Point|Polygon)\b[^>]*>.*?<coordinates>\s*([^<]+?)\s*</coordinates>",
    re.DOTALL,
)


def _parse_kml_linestrings(kml_text: str) -> Optional[dict]:
    lines = []
    for match in _KML_COORD_RE.finditer(kml_text):
        raw = match.group(1).strip()
        pts = []
        for token in raw.split():
            parts = token.split(",")
            if len(parts) < 2:
                continue
            lng, lat = float(parts[0]), float(parts[1])
            pts.append([lng, lat])
        if len(pts) >= 2:
            lines.append(pts)

    if not lines:
        return None
    if len(lines) == 1:
        return {"type": "LineString", "coordinates": lines[0]}
    return {"type": "MultiLineString", "coordinates": lines}


def _geojson_line_to_shapely(geojson: dict):
    """GeoJSON LineString/MultiLineString -> geometri shapely.
    MultiLineString(list_of_linestrings) segfaults-into-TypeError pada kombinasi
    shapely/numpy ini (shapely.creation.multilinestrings tersedak object array)
    — lewat WKT untuk menghindari jalur kode itu."""
    if geojson["type"] == "MultiLineString":
        lines = [LineString(coords) for coords in geojson["coordinates"] if len(coords) >= 2]
        wkt = "MULTILINESTRING (" + ", ".join(
            "(" + ", ".join(f"{c[0]} {c[1]}" for c in ls.coords) + ")" for ls in lines
        ) + ")"
        return shapely.wkt.loads(wkt)
    return LineString(geojson["coordinates"])


def _parse_usulan_geometry(content: bytes) -> Optional[dict]:
    """Parse unggahan geometri SITIA apa adanya: sebagian 'KML' ternyata
    KMZ (zip berisi doc.kml) atau bahkan shapefile ter-zip — deteksi dari
    header PK lalu tangani ketiganya. Return geojson LineString/
    MultiLineString atau None."""
    if content[:2] != b"PK":
        return _parse_kml_linestrings(content.decode("utf-8", "replace"))
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        return None
    names = zf.namelist()

    kml = next((n for n in names if n.lower().endswith(".kml")), None)
    if kml:
        return _parse_kml_linestrings(zf.read(kml).decode("utf-8", "replace"))

    kmz = next((n for n in names if n.lower().endswith(".kmz")), None)
    if kmz:
        # zip berisi .kmz (arsip di dalam arsip) — rekursi sekali
        return _parse_usulan_geometry(zf.read(kmz))

    shp = next((n for n in names if n.lower().endswith(".shp")), None)
    if shp:
        with TemporaryDirectory() as tmp:
            # ekstraksi rata (basename saja) — hindari path traversal dari
            # nama entri arsip
            for n in names:
                base = Path(n).name
                if not base:
                    continue
                (Path(tmp) / base).write_bytes(zf.read(n))
            try:
                gdf = gpd.read_file(Path(tmp) / Path(shp).name)
                if gdf.crs and gdf.crs.to_epsg() != 4326:
                    gdf = gdf.to_crs(epsg=4326)
            except Exception:
                return None
        lines = []
        for geom in gdf.geometry:
            if geom is None:
                continue
            if geom.geom_type == "LineString":
                lines.append([[c[0], c[1]] for c in geom.coords])
            elif geom.geom_type == "MultiLineString":
                lines.extend([[c[0], c[1]] for c in g.coords] for g in geom.geoms)
        lines = [l for l in lines if len(l) >= 2]
        if not lines:
            return None
        if len(lines) == 1:
            return {"type": "LineString", "coordinates": lines[0]}
        return {"type": "MultiLineString", "coordinates": lines}
    return None


MAPS_DIR = BASE_DIR / "Maps"

# MAP_LAYER_LABELS / _map_layer_label pindah ke map_layer_labels.py (dipakai
# bersama scripts/import_maps_to_postgis.py, lihat import di atas). MAPS_DIR
# sendiri TETAP di sini apa adanya: masih dipakai scripts/spatial_join_*.py
# dan import_*.py yang baca Maps/ langsung dari disk (di luar cakupan
# migrasi PostGIS ini, lihat CLAUDE.md).
_map_layer_geojson_cache: dict = {}


# --- Layer khusus: Maps/BATAS_ADMINISTRASI.gdb (layer ADMINISTRASI_KECAMATAN_AR
# BIG, 7.283 poligon kecamatan definitif). Diimpor via
# scripts/import_batas_administrasi_kecamatan.py, MENGGANTIKAN sumber lama
# (SHP batas kecamatan nasional Dukcapil Des 2019, Maps/BATAS KECAMATAN/,
# 6.810 poligon TANPA kolom provinsi/kabupaten — dulu perlu dicocokkan lewat
# nama ke penduduk_kecamatan plus heuristik ketetanggaan utk kecamatan
# homonim lintas daerah, lihat riwayat git kalau perlu referensi). Sumber
# gdb ini sudah punya kolom provinsi/kabupaten/kecamatan ASLI per poligon,
# jadi disimpan LANGSUNG sbg kolom kabupaten (=provinsi asli) dan layer
# (=kabupaten/kota asli) di map_layers/map_layer_meta — bukan flat+attrs lagi
# — sehingga endpoint /api/maps/* yang GENERIK (maps_provinces/kabupaten/
# layers/layer di bawah) sudah otomatis melayani hierarki provinsi Indonesia
# -> kabupaten/kota -> poligon kecamatan tanpa pencocokan nama runtime sama
# sekali; tidak ada lagi cabang khusus BATAS_KEC_DIRNAME di endpoint2 itu.
# Provinsi bucket teratas dipatok konstan "BATAS KECAMATAN" supaya topbar
# picker tidak berubah bentuk (satu "provinsi" di picker = seluruh Indonesia,
# lalu "kabupaten" di picker = 34 provinsi asli, lalu "layer" = kabupaten/kota).
#
# _batas_kec_shp() (SHP Dukcapil lama, dibaca langsung dari disk) TETAP ada
# apa adanya — dipakai scripts/spatial_join_kecamatan*.py utk mengisi
# usulan_inpres.kode_kecamatan/usulan_kecamatan_dilalui, jalur yang independen
# dari map_layers/PostGIS dan sengaja TIDAK ikut diganti di sini (beda risiko:
# itu memberi makan skor IJD nasional, ganti sumbernya perlu proses verifikasi
# terpisah, bukan cuma tukar layer overlay peta).
BATAS_KEC_DIRNAME = "BATAS KECAMATAN"
_batas_kec_index_cache: Optional[dict] = None


def _batas_kec_shp() -> Optional[Path]:
    d = MAPS_DIR / BATAS_KEC_DIRNAME
    if not d.is_dir():
        return None
    return next(iter(sorted(d.glob("*.shp"))), None)


def _batas_kec_index() -> dict:
    """{provinsi: {kabupaten_kota: [{kecamatan, kode_kecamatan}]}} — dibaca
    langsung dari kolom kabupaten/layer di map_layers (bukan lagi fuzzy-match
    nama runtime, lihat catatan di atas). Dipakai HANYA oleh
    scripts/import_indeks_penanaman_raster.py sekarang (endpoint maps_*
    di bawah sudah generik). Cache sekali per proses — restart server utk
    lihat hasil impor ulang."""
    global _batas_kec_index_cache
    if _batas_kec_index_cache is not None:
        return _batas_kec_index_cache
    with db_cursor() as cur:
        cur.execute(
            "SELECT kabupaten AS provinsi, layer AS kabupaten_kota, "
            "attrs->>'KECAMATAN' AS kecamatan, (attrs->>'KODE_KECAMATAN')::int AS kode_kecamatan "
            "FROM map_layers WHERE provinsi=%s",
            (BATAS_KEC_DIRNAME,),
        )
        rows = cur.fetchall()
    if not rows:
        raise HTTPException(404, "Layer batas kecamatan belum diimpor ke PostGIS (map_layers) "
                                  "— jalankan scripts/import_batas_administrasi_kecamatan.py")
    index: dict = {}
    for r in rows:
        index.setdefault(r["provinsi"], {}).setdefault(r["kabupaten_kota"], []).append(
            {"kecamatan": r["kecamatan"], "kode_kecamatan": r["kode_kecamatan"]}
        )
    _batas_kec_index_cache = index
    return index


def _batas_kec_layer_geojson(prov: str, kab: str) -> dict:
    """FeatureCollection poligon kecamatan definitif satu kabupaten/kota.
    Dipakai HANYA oleh scripts/import_indeks_penanaman_raster.py (endpoint
    maps_layer() di bawah sudah generik lewat map_layers/map_layer_meta
    langsung). Query SQL biasa via ST_AsGeoJSON — BUKAN lagi WKB+shapely
    dengan heuristik pemilihan bagian poligon homonim spt versi lama: sumber
    gdb ini sudah punya kolom provinsi/kabupaten per poligon jadi tidak ada
    lagi homonim lintas daerah yang perlu dipilah."""
    index = _batas_kec_index()
    if prov not in index or kab not in index[prov]:
        raise HTTPException(404, "Provinsi/kabupaten tidak dikenal di layer batas kecamatan")
    with db_cursor() as cur:
        cur.execute(
            """SELECT jsonb_build_object(
                    'type', 'FeatureCollection',
                    'features', COALESCE(jsonb_agg(jsonb_build_object(
                        'type', 'Feature',
                        'geometry', ST_AsGeoJSON(geom)::jsonb,
                        'properties', attrs
                    )), '[]'::jsonb)
                ) AS fc
               FROM map_layers WHERE provinsi=%s AND kabupaten=%s AND layer=%s""",
            (BATAS_KEC_DIRNAME, prov, kab),
        )
        geojson = cur.fetchone()["fc"]
    geojson["label"] = f"Batas Kecamatan — {kab} ({prov})"
    geojson["kecamatan_tanpa_poligon"] = []
    return geojson


# Join atribut poligon kecamatan (identify di overlay BATAS KECAMATAN) ke
# tabel database — pengguna memilih tabel mana yang dilihat di popup.
KECAMATAN_JOIN_TABLES = {
    "kecamatan_data_turunan": "Data turunan (kepadatan, kendaraan)",
    "penduduk_kecamatan": "Master penduduk kecamatan",
    "usulan_inpres": "Usulan Inpres di kecamatan ini",
    "bps_kecamatan_potensi_tematik": "Potensi & Produksi Tematik (Dalam Angka)",
    "kawasan_tematik": "Kawasan Tematik (A3 IJD)",
    "bappenas_lokus_a": "Lokus Aspek A Bappenas",
}
_USULAN_JOIN_COLS = (
    "id", "nama_ruas", "jenis_penanganan", "panjang_ruas_km",
    "alokasi_usulan_pemda", "prioritas", "verifikasi_balai", "verifikasi_kompetensi",
)


@app.get("/api/kecamatan/{kode_kecamatan}/data")
def kecamatan_join_data(kode_kecamatan: int, tabel: str = "kecamatan_data_turunan"):
    if tabel not in KECAMATAN_JOIN_TABLES:
        raise HTTPException(400, "Tabel tidak dikenal untuk join kecamatan")
    kode_kab = kode_kecamatan // 1000
    kode_prov = kode_kab // 100
    with db_cursor() as cur:
        if tabel == "usulan_inpres":
            cur.execute(
                f"SELECT {', '.join(_USULAN_JOIN_COLS)} FROM usulan_inpres "
                "WHERE kode_kecamatan = %s ORDER BY id LIMIT 50",
                (kode_kecamatan,),
            )
        elif tabel == "kawasan_tematik":
            # Sebagian besar baris level KABUPATEN (kode_kecamatan NULL) --
            # WHERE kode_kecamatan=%s polos melewatkan hampir semuanya.
            # Cocokkan juga via kode_kabupaten turunan dari kode_kecamatan.
            cur.execute(
                f'SELECT * FROM "{tabel}" WHERE kode_kecamatan = %s '
                "OR (kode_kabupaten = %s AND kode_kecamatan IS NULL) LIMIT 20",
                (kode_kecamatan, kode_kab),
            )
        elif tabel == "bappenas_lokus_a":
            # Level bisa KABUPATEN, KECAMATAN, atau PROVINSI -- sama alasan
            # spt kawasan_tematik, ditambah fallback level provinsi.
            cur.execute(
                f'SELECT * FROM "{tabel}" WHERE kode_kecamatan = %s '
                "OR (kode_kabupaten = %s AND (level = 'KABUPATEN' OR kode_kecamatan IS NULL)) "
                "OR (kode_provinsi = %s AND level = 'PROVINSI') LIMIT 20",
                (kode_kecamatan, kode_kab, kode_prov),
            )
        else:
            cur.execute(
                f'SELECT * FROM "{tabel}" WHERE kode_kecamatan = %s LIMIT 20',
                (kode_kecamatan,),
            )
        rows = cur.fetchall()
    columns = [c for c in (rows[0].keys() if rows else []) if c not in DATA_TABLE_SKIP_COLS]
    return {
        "tabel": tabel,
        "label": KECAMATAN_JOIN_TABLES[tabel],
        "tabel_tersedia": [{"tabel": k, "label": v} for k, v in KECAMATAN_JOIN_TABLES.items()],
        "columns": columns,
        "rows": [[jsonable_encoder(r[c]) for c in columns] for r in rows],
    }


@app.get("/api/maps/provinces")
def maps_provinces():
    with db_cursor() as cur:
        cur.execute(
            "SELECT provinsi, COUNT(DISTINCT kabupaten) AS kabupaten_count "
            "FROM map_layer_meta GROUP BY provinsi ORDER BY provinsi"
        )
        rows = cur.fetchall()
    return [{"provinsi": r["provinsi"], "kabupaten_count": r["kabupaten_count"]} for r in rows]


@app.get("/api/maps/kabupaten")
def maps_kabupaten(provinsi: str):
    with db_cursor() as cur:
        cur.execute(
            "SELECT kabupaten, COUNT(*) AS layer_count FROM map_layer_meta "
            "WHERE provinsi=%s GROUP BY kabupaten ORDER BY kabupaten",
            (provinsi,),
        )
        rows = cur.fetchall()
    return [
        {
            "kabupaten": r["kabupaten"],
            "layer_count": r["layer_count"],
            **({"label": provinsi} if r["kabupaten"] == "" else {}),
        }
        for r in rows
    ]


@app.get("/api/maps/layers")
def maps_layers(provinsi: str, kabupaten: str = ""):
    with db_cursor() as cur:
        cur.execute(
            "SELECT layer, label, size_mb FROM map_layer_meta "
            "WHERE provinsi=%s AND kabupaten=%s ORDER BY layer",
            (provinsi, kabupaten),
        )
        rows = cur.fetchall()
    return [
        {"layer": r["layer"], "label": r["label"] or _map_layer_label(r["layer"]),
         "size_mb": float(r["size_mb"]) if r["size_mb"] is not None else None}
        for r in rows
    ]


# Layer peta (dulu SHP statis selama proses server jalan, kini PostGIS --
# masih di-cache in-process karena payload besar tak berubah dalam masa
# hidup satu proses; restart server utk lihat hasil impor ulang). Aman
# dikasih Cache-Control lumayan panjang supaya klien intranet tidak perlu
# unduh ulang payload besar (mis. Jalan Nasional ~10MB) tiap buka halaman/
# pindah tab -- cuma sekali per jam per browser, bukan tiap request.
_MAP_LAYER_CACHE_HEADERS = {"Cache-Control": "public, max-age=3600"}


@app.get("/api/maps/layer")
def maps_layer(provinsi: str, layer: str, kabupaten: str = ""):
    key = (provinsi, kabupaten, layer)
    if key in _map_layer_geojson_cache:
        return JSONResponse(content=_map_layer_geojson_cache[key], headers=_MAP_LAYER_CACHE_HEADERS)

    with db_cursor() as cur:
        cur.execute(
            "SELECT label, feature_count FROM map_layer_meta "
            "WHERE provinsi=%s AND kabupaten=%s AND layer=%s",
            (provinsi, kabupaten, layer),
        )
        meta = cur.fetchone()
        if not meta:
            raise HTTPException(404, "Layer tidak ditemukan")

        # Heavy layers (e.g. KONTUR, tens of MB of contour lines) are
        # simplified so the browser doesn't choke on rendering; tolerance is
        # in degrees (~0.00015 deg ~ 15-17m at this latitude), fine for
        # on-screen display.
        simplify = (meta["feature_count"] or 0) > 3000
        geom_expr = "ST_SimplifyPreserveTopology(geom, 0.00015)" if simplify else "geom"
        cur.execute(
            f"""SELECT jsonb_build_object(
                    'type', 'FeatureCollection',
                    'features', COALESCE(jsonb_agg(jsonb_build_object(
                        'type', 'Feature',
                        'geometry', ST_AsGeoJSON({geom_expr})::jsonb,
                        'properties', attrs
                    )), '[]'::jsonb)
                ) AS fc
                FROM map_layers WHERE provinsi=%s AND kabupaten=%s AND layer=%s""",
            (provinsi, kabupaten, layer),
        )
        geojson = cur.fetchone()["fc"]

    geojson["label"] = meta["label"] or _map_layer_label(layer)
    _map_layer_geojson_cache[key] = geojson
    return JSONResponse(content=geojson, headers=_MAP_LAYER_CACHE_HEADERS)


def _fetch_usulan_geometry(usulan_id: int) -> dict:
    with db_cursor() as cur:
        cur.execute(
            "SELECT kml_original_url, geom_geojson FROM usulan_inpres WHERE id = %s",
            (usulan_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Usulan tidak ditemukan")

        if row["geom_geojson"]:
            return json.loads(row["geom_geojson"])

        if not row["kml_original_url"]:
            raise HTTPException(404, "Usulan ini tidak memiliki data geometri (KML)")

        try:
            resp = requests.get(row["kml_original_url"], timeout=15)
            resp.raise_for_status()
        except requests.RequestException as e:
            raise HTTPException(502, f"Gagal mengambil KML dari SITIA Bina Marga: {e}")

        geojson = _parse_usulan_geometry(resp.content)
        if not geojson:
            raise HTTPException(404, "KML tidak berisi geometri jalur yang dapat dibaca")

        cur.execute(
            "UPDATE usulan_inpres SET geom_geojson = %s, geom_fetched_at = NOW() WHERE id = %s",
            (json.dumps(geojson), usulan_id),
        )

    return geojson


@app.get("/api/usulan-inpres/{usulan_id}/geometry")
def usulan_inpres_geometry(usulan_id: int):
    return _fetch_usulan_geometry(usulan_id)


_BPS_API_BASE = "https://webapi.bps.go.id/v1/api/list"


def _dalam_angka_domain_code(kode_wilayah: int, jenis_wilayah: str) -> str:
    """Format domain BPS Web API: provinsi = 2 digit kode + '00', kabupaten/
    kota = 4 digit kode_kabupaten apa adanya."""
    return f"{kode_wilayah:02d}00" if jenis_wilayah == "PROVINSI" else f"{kode_wilayah:04d}"


def _dalam_angka_fresh_url(pub_id: str, domain_code: str) -> Optional[str]:
    """Regenerasi link download.php?f=<token> segar dari pub_id lewat BPS
    Web API -- token yang tersimpan di dalam_angka_publikasi.url_publikasi
    BISA KEDALUWARSA (lihat schema_dalam_angka_publikasi.sql), jadi selalu
    diregenerasi live tiap kali endpoint ini dipanggil, bukan andalkan
    cache. Butuh BPS_API_KEY di .env -- return None kalau key tidak ada
    atau request gagal (caller fallback ke url_publikasi tersimpan)."""
    key = os.getenv("BPS_API_KEY")
    if not key:
        return None
    try:
        r = requests.get(_BPS_API_BASE, params={
            "model": "publication", "lang": "ind", "domain": domain_code,
            "id": pub_id, "key": key,
        }, timeout=15)
        r.raise_for_status()
        data = r.json()
        items = data.get("data", [])
        items = items[1] if len(items) > 1 else []
        # BPS Web API tidak benar-benar memfilter berdasarkan param "id" --
        # ia balikin listing default domain tsb (terbaru dulu), jadi
        # items[0] bisa jadi publikasi lain sama sekali (mis. PDRB, bukan
        # Dalam Angka). Cocokkan eksplisit ke pub_id yang diminta.
        for item in items:
            if str(item.get("pub_id")) == str(pub_id):
                return item.get("pdf")
        return None
    except (requests.RequestException, ValueError, KeyError, IndexError):
        return None


@app.get("/api/usulan-inpres/{usulan_id}/dalam-angka")
def usulan_inpres_dalam_angka(usulan_id: int):
    """Link publikasi resmi BPS "<Kabupaten/Kota> Dalam Angka <tahun>" utk
    kabupaten/kota usulan ini -- diisi scripts/sync_dalam_angka_bps_api.py
    (BELUM cakupan nasional penuh per 24 Jul 2026, cuma provinsi/kab yang
    sudah disinkron -- "tersedia": false kalau belum ada baris utk kab ini,
    bukan error). Dipakai tombol preview PDF di panel detail usulan."""
    with db_cursor() as cur:
        cur.execute(
            "SELECT provinsi, kabupaten_kota FROM usulan_inpres WHERE id = %s",
            (usulan_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Usulan tidak ditemukan")

        cur.execute(
            "SELECT kode_kabupaten FROM wilayah_mapping "
            "WHERE provinsi_sitia = %s AND kabupaten_kota_sitia = %s",
            (row["provinsi"], row["kabupaten_kota"]),
        )
        kab = cur.fetchone()
        if not kab:
            return {"tersedia": False, "keterangan": "Kabupaten/kota usulan ini belum "
                    "terpetakan ke kode BPS (wilayah_mapping)."}
        kode_kab = kab["kode_kabupaten"]

        cur.execute(
            "SELECT tahun, judul, pub_id, url_publikasi FROM dalam_angka_publikasi "
            "WHERE kode_wilayah = %s ORDER BY tahun DESC",
            (kode_kab,),
        )
        pubs = cur.fetchall()

    if not pubs:
        return {"tersedia": False, "kode_kabupaten": kode_kab,
                "keterangan": "Belum ada link Dalam Angka tersinkron utk kabupaten/kota "
                "ini -- jalankan scripts/sync_dalam_angka_bps_api.py."}

    domain_code = _dalam_angka_domain_code(kode_kab, "KABUPATEN_KOTA")
    hasil = []
    for p in pubs:
        url = _dalam_angka_fresh_url(p["pub_id"], domain_code) or p["url_publikasi"]
        hasil.append({"tahun": p["tahun"], "judul": p["judul"], "url": url})

    return {"tersedia": True, "kode_kabupaten": kode_kab, "publikasi": hasil}


@app.get("/api/dalam-angka/list")
def dalam_angka_list(q: str = "", jenis: str = ""):
    """Daftar seluruh publikasi "Dalam Angka" (provinsi + kabupaten/kota)
    yang sudah tersinkron (scripts/sync_dalam_angka_bps_api.py), dikelompokkan
    per wilayah dengan daftar tahun tersedia. Dipakai panel topbar "Dalam
    Angka" (list + pencarian nama wilayah). Mengembalikan url_publikasi
    TERSIMPAN apa adanya (bukan diregenerasi live) supaya daftar ratusan
    wilayah tetap cepat dimuat -- link segar per baris diambil on-demand
    lewat /api/dalam-angka/preview saat tombol Pratinjau diklik."""
    jenis = jenis.upper() if jenis else ""
    if jenis and jenis not in ("PROVINSI", "KABUPATEN_KOTA"):
        raise HTTPException(400, "jenis harus PROVINSI atau KABUPATEN_KOTA")

    sql = ("SELECT kode_wilayah, jenis_wilayah, nama_wilayah, tahun, pub_id, "
           "judul, url_publikasi FROM dalam_angka_publikasi WHERE 1=1")
    params: list = []
    if q:
        sql += " AND nama_wilayah ILIKE %s"
        params.append(f"%{q}%")
    if jenis:
        sql += " AND jenis_wilayah = %s"
        params.append(jenis)
    sql += " ORDER BY nama_wilayah, tahun DESC"

    with db_cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    wilayah: dict = {}
    for r in rows:
        key = (r["jenis_wilayah"], r["kode_wilayah"])
        entry = wilayah.setdefault(key, {
            "kode_wilayah": r["kode_wilayah"],
            "jenis_wilayah": r["jenis_wilayah"],
            "nama_wilayah": r["nama_wilayah"],
            "publikasi": [],
        })
        entry["publikasi"].append({
            "tahun": r["tahun"], "judul": r["judul"], "pub_id": r["pub_id"],
            "url": r["url_publikasi"],
        })

    return sorted(wilayah.values(), key=lambda w: (w["nama_wilayah"], -w["publikasi"][0]["tahun"]))


@app.get("/api/dalam-angka/preview")
def dalam_angka_preview(kode_wilayah: int, jenis_wilayah: str, tahun: int):
    """Regenerasi satu link "Dalam Angka" segar on-demand (dipanggil saat
    tombol Pratinjau/Tab Baru diklik di panel list), bukan untuk seluruh
    daftar sekaligus -- lihat catatan performa di dalam_angka_list()."""
    jenis_wilayah = jenis_wilayah.upper()
    with db_cursor() as cur:
        cur.execute(
            "SELECT pub_id, judul, url_publikasi FROM dalam_angka_publikasi "
            "WHERE kode_wilayah = %s AND jenis_wilayah = %s AND tahun = %s",
            (kode_wilayah, jenis_wilayah, tahun),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Publikasi tidak ditemukan")

    domain_code = _dalam_angka_domain_code(kode_wilayah, jenis_wilayah)
    url = _dalam_angka_fresh_url(row["pub_id"], domain_code) or row["url_publikasi"]
    return {"judul": row["judul"], "url": url}


# User-Agent "browser-like" wajib -- webapi.bps.go.id/download.php ada di
# belakang WAF anti-bot (cookie sesi "TS...", pola khas Imperva/F5 ASM) yang
# menolak (403, atau malah hang/timeout tanpa respons) request dgn UA default
# requests/curl, tapi meloloskan UA browser biasa -- dikonfirmasi manual 24
# Jul 2026 (curl UA default -> 403 instan; UA "Mozilla/5.0" -> 200 + PDF).
_BPS_DOWNLOAD_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"


@app.get("/api/dalam-angka/pdf")
def dalam_angka_pdf(kode_wilayah: int, jenis_wilayah: str, tahun: int):
    """Proxy PDF "Dalam Angka" lewat backend (bukan iframe langsung ke
    webapi.bps.go.id) -- iframe cross-origin ke link BPS kena WAF anti-bot
    tsb secara tidak konsisten (kadang lolos kadang tidak, tergantung
    Referer/fingerprint browser saat itu terhadap WAF-nya, di luar kendali
    kita); server-side fetch dgn _BPS_DOWNLOAD_UA sudah terbukti selalu
    lolos, jadi klien cukup embed URL same-origin ini, tidak pernah bicara
    langsung ke BPS. Lihat static/js/utils.js openPdfPreviewModal()."""
    jenis_wilayah = jenis_wilayah.upper()
    with db_cursor() as cur:
        cur.execute(
            "SELECT pub_id, url_publikasi FROM dalam_angka_publikasi "
            "WHERE kode_wilayah = %s AND jenis_wilayah = %s AND tahun = %s",
            (kode_wilayah, jenis_wilayah, tahun),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Publikasi tidak ditemukan")

    domain_code = _dalam_angka_domain_code(kode_wilayah, jenis_wilayah)
    url = _dalam_angka_fresh_url(row["pub_id"], domain_code) or row["url_publikasi"]
    try:
        r = requests.get(url, headers={"User-Agent": _BPS_DOWNLOAD_UA}, timeout=30, stream=True)
        r.raise_for_status()
    except requests.RequestException as e:
        raise HTTPException(502, f"Gagal mengambil PDF dari BPS: {e}")
    return StreamingResponse(r.iter_content(chunk_size=65536), media_type="application/pdf")


# Tab "Data per Subjek" di panel Dalam Angka -- jelajah data tabel dinamis
# BPS Web API per kategori/subjek/variabel (terpisah dari katalog publikasi
# PDF di atas). Restart server bersihkan cache -- pola sama dgn
# _map_layer_geojson_cache/_npr_kelas_cache. UA browser dipakai konsisten
# di semua panggilan (bukan cuma download.php) -- dikonfirmasi manual 24 Jul
# 2026 endpoint model=list JUGA bisa kena "Perimeter WAF Block" tanpa UA,
# meski _dalam_angka_fresh_url() di atas (model=publication) kebetulan lolos
# tanpa UA -- tidak diandalkan utk endpoint baru ini.
_bps_subjek_cache: dict = {}


def _bps_api_get(model: str, **params) -> Optional[dict]:
    key = os.getenv("BPS_API_KEY")
    if not key:
        return None
    params = {"model": model, "key": key, **params}
    try:
        r = requests.get(_BPS_API_BASE, params=params,
                          headers={"User-Agent": _BPS_DOWNLOAD_UA}, timeout=20)
        r.raise_for_status()
        data = r.json()
    except (requests.RequestException, ValueError):
        return None
    if data.get("status") != "OK":
        return None
    return data


def _bps_api_list_all(model: str, **params) -> list:
    """Kumpulkan semua halaman model=list (subcat/subject/var/th) jadi satu
    list flat -- endpoint ini page-nya kecil (puluhan item), aman diagregasi
    sekali panggil daripada bikin frontend urus paging BPS sendiri."""
    out: list = []
    page = 1
    while True:
        data = _bps_api_get(model, page=page, **params)
        if not data:
            break
        payload = data.get("data", [])
        if len(payload) < 2:
            break
        meta, items = payload[0], payload[1]
        out.extend(items)
        if page >= meta.get("pages", 1):
            break
        page += 1
    return out


def _bps_subjek_cached(cache_key: tuple, builder) -> list:
    if cache_key not in _bps_subjek_cache:
        _bps_subjek_cache[cache_key] = builder()
    return _bps_subjek_cache[cache_key]


_BPS_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _bps_clean_wilayah_label(label: Optional[str]) -> tuple:
    """vervar.label BPS suka dibungkus <b>...</b> utk baris level provinsi
    saat 1 variabel mencampur provinsi & kab/kota sekaligus (indikasi
    hierarki visual dari portal BPS sendiri, mis. var Angka Harapan Hidup,
    546 wilayah) -- dipetakan jadi flag is_provinsi drpd ditampilkan mentah
    (escapeHtml() di frontend bikin tag-nya kelihatan literal sbg teks,
    bukan dirender sbg bold)."""
    if not label:
        return "", False
    is_provinsi = bool(re.search(r"<b>", label, re.IGNORECASE))
    return _BPS_HTML_TAG_RE.sub("", label).strip(), is_provinsi


@app.get("/api/bps-subjek/subcat")
def bps_subjek_subcat():
    """Kategori subjek BPS (Sosial-Kependudukan, Ekonomi-Perdagangan, dst) --
    level teratas navigasi tab "Data per Subjek"."""
    if not os.getenv("BPS_API_KEY"):
        raise HTTPException(503, "BPS_API_KEY belum dikonfigurasi di .env")
    return _bps_subjek_cached(("subcat",),
                               lambda: _bps_api_list_all("subcat", lang="ind", domain="0000"))


@app.get("/api/bps-subjek/subject")
def bps_subjek_subject(subcat: int):
    return _bps_subjek_cached(("subject", subcat),
                               lambda: _bps_api_list_all("subject", lang="ind", domain="0000", subcat=subcat))


@app.get("/api/bps-subjek/var")
def bps_subjek_var(subject: int):
    return _bps_subjek_cached(("var", subject),
                               lambda: _bps_api_list_all("var", lang="ind", domain="0000", subject=subject))


@app.get("/api/bps-subjek/{var_id}/tahun")
def bps_subjek_tahun(var_id: int):
    return _bps_subjek_cached(("th", var_id),
                               lambda: _bps_api_list_all("th", domain="0000", var=var_id))


@app.get("/api/bps-subjek/{var_id}/turvar")
def bps_subjek_turvar(var_id: int):
    """Kategori turunan (turvar) variabel ini, kalau ada -- mis. Perkotaan/
    Perdesaan/Perkotaan+Perdesaan utk variabel kemiskinan P0/P1/P2. List
    kosong berarti variabel ini tak punya sub-kategori sama sekali (frontend
    sembunyikan dropdown-nya, turvar dianggap 0 di /data). HANYA bergantung
    pada var (dikonfirmasi manual: tidak butuh th)."""
    return _bps_subjek_cached(("turvar", var_id),
                               lambda: _bps_api_list_all("turvar", domain="0000", var=var_id))


@app.get("/api/bps-subjek/{var_id}/turth")
def bps_subjek_turth(var_id: int, th: int):
    """Periode turunan (turth) variabel ini utk 1 tahun tertentu, kalau ada --
    mis. Semester 1 (Maret)/Semester 2 (September)/Tahunan utk variabel yg
    disurvei 2x setahun (Susenas). BEDA dari turvar: turth bergantung pada th
    yg dipilih, bisa jadi beda opsi antar tahun -- jadi selalu refetch tiap
    tahun berganti, bukan cuma sekali per var spt turvar."""
    return _bps_subjek_cached(("turth", var_id, th),
                               lambda: _bps_api_list_all("turth", domain="0000", var=var_id, th=th))


@app.get("/api/bps-subjek/{var_id}/data")
def bps_subjek_data(var_id: int, th: int, turvar: int = 0, turth: int = 0):
    """Nilai 1 variabel BPS utk 1 tahun (+ turvar/turth kalau berlaku), seluruh
    wilayah yg tersedia (granularitas ditentukan definisi variabel itu sendiri
    oleh BPS, bukan dipilih di sini). turvar/turth WAJIB dipilih dulu oleh
    caller lewat /turvar /turth kalau list-nya tidak kosong -- dibiarkan 0
    (default) kalau variabelnya memang tak punya dimensi itu. Kunci
    datacontent dikonfirmasi manual = f"{vervar_val}{var_id}{turvar}{th}
    {turth}" (var 128 IKK tanpa turvar/turth: "1100"+"128"+"0"+"125"+"0" =
    "110012801250"; var 503 Kemiskinan P1 dgn turvar=191 Perkotaan+Perdesaan,
    turth=61 Semester 1: "1100"+"503"+"191"+"125"+"61" = "110050319112561" --
    keduanya dites cocok 24 Jul 2026). vervar di-lookup per-key (bukan
    order-zip) krn vervar suka menyelipkan baris agregat "INDONESIA" tanpa
    entri datacontent-nya sendiri (39 vervar vs 38 datacontent utk var 128)."""
    if not os.getenv("BPS_API_KEY"):
        raise HTTPException(503, "BPS_API_KEY belum dikonfigurasi di .env")

    data = _bps_api_get("data", domain="0000", var=var_id, th=th)
    if not data:
        raise HTTPException(502, "Gagal mengambil data dari BPS Web API.")

    vervar = data.get("vervar") or []
    datacontent = data.get("datacontent") or {}
    if not vervar or not datacontent:
        raise HTTPException(502, "Data BPS kosong utk variabel/tahun ini.")

    var_meta = (data.get("var") or [{}])[0]
    tahun_meta = (data.get("tahun") or [{}])[0]
    rows = []
    for v in vervar:
        key = f"{v.get('val')}{var_id}{turvar}{th}{turth}"
        if key not in datacontent:
            continue
        label, is_provinsi = _bps_clean_wilayah_label(v.get("label"))
        rows.append({"kode_wilayah": v.get("val"), "wilayah": label,
                     "is_provinsi": is_provinsi, "nilai": datacontent[key]})
    if not rows:
        raise HTTPException(502, "Tidak ada wilayah yang cocok dgn data BPS -- "
                             "kombinasi turvar/turth mungkin tidak sesuai.")
    rows.sort(key=lambda r: r["wilayah"] or "")

    return {
        "var": var_meta.get("label"), "unit": var_meta.get("unit"),
        "tahun": tahun_meta.get("label"), "last_update": data.get("last_update"),
        "rows": rows,
    }


@app.get("/api/usulan-inpres/{usulan_id}/export/shp")
def usulan_inpres_export_shp(usulan_id: int):
    geojson = _fetch_usulan_geometry(usulan_id)

    with db_cursor() as cur:
        cur.execute(
            "SELECT nama_kegiatan, nama_ruas, kode_ruas, kabupaten_kota, provinsi, "
            "jenis_penanganan, panjang_ruas_km FROM usulan_inpres WHERE id = %s",
            (usulan_id,),
        )
        u = cur.fetchone()
    if not u:
        raise HTTPException(404, "Usulan tidak ditemukan")

    geometry = _geojson_line_to_shapely(geojson)

    gdf = gpd.GeoDataFrame(
        [{
            "nama": (u["nama_kegiatan"] or u["nama_ruas"] or "")[:80],
            "kode_ruas": u["kode_ruas"],
            "kab_kota": u["kabupaten_kota"],
            "provinsi": u["provinsi"],
            "penanganan": u["jenis_penanganan"],
            "panjang_km": u["panjang_ruas_km"],
            "geometry": geometry,
        }],
        geometry="geometry",
        crs="EPSG:4326",
    )

    with TemporaryDirectory() as tmp:
        shp_path = Path(tmp) / "usulan_inpres.shp"
        gdf.to_file(shp_path, driver="ESRI Shapefile", engine="pyogrio")

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in Path(tmp).glob("usulan_inpres.*"):
                zf.write(f, arcname=f.name)
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=usulan_inpres_{usulan_id}_shp.zip"},
        )


@app.post("/api/chat")
def chat(payload: ChatRequest):
    if not payload.messages:
        raise HTTPException(400, "Tidak ada pesan")
    reply, actions = chat_providers._call_chat(payload.messages, payload.context)
    return {"reply": reply, "actions": actions}


class NoCacheStaticFiles(StaticFiles):
    """StaticFiles tanpa Cache-Control membuat browser memakai caching
    heuristik: js/css lama bisa terus dipakai setelah file berubah (tanpa
    build step tidak ada cache-busting hash). no-cache = browser tetap
    menyimpan file tapi selalu revalidasi ETag (304 bila tidak berubah)."""

    def file_response(self, *args, **kwargs):
        resp = super().file_response(*args, **kwargs)
        resp.headers["Cache-Control"] = "no-cache"
        return resp


app.mount("/", NoCacheStaticFiles(directory=str(STATIC_DIR), html=True), name="static")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app:app",
        host=os.getenv("HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", 8000)),
        reload=True,
    )
