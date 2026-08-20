# -*- coding: utf-8 -*-
"""Impor inventaris ALUT (Alat Utama) SAR -- udara, darat, laut --
(docs/New/6. BASARNAS/2. Data Sarana dan Prasarana/, 3 file xlsx) ke
tabel referensi basarnas_alut. Lihat scripts/schema_basarnas_alut.sql
untuk skema, dan docs/kajian_data_baru_docs_new.md §8.2 untuk hasil
telaah datanya. TIDAK terkait usulan_inpres/IJD.

Header berbeda posisi kolom antar file (sheet LAUT tidak punya kolom
spacer kosong di depan seperti UDARA/DARAT) -- dicari otomatis lewat
posisi sel berisi 'NO', bukan index tetap.

Seluruh baris diimpor apa adanya (termasuk baris penanda kategori tanpa
detail unit, dan baris "Tidak Memiliki") -- lihat docstring skema untuk
alasannya. kondisi_kategori dinormalisasi dari kondisi_saat_ini_raw yang
penulisannya bebas di sumber (S/US/Baik/Rusak Ringan/dll.).

Idempotent: DELETE + INSERT ulang seluruh tabel tiap run.

Usage (venv aktif):
    python scripts/import_basarnas_alut.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

from db import db_cursor as pg_cursor  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
BASE_DIR = (
    REPO_ROOT / "docs" / "New" / "6. BASARNAS-20260820T015304Z-1-001"
    / "6. BASARNAS" / "2. Data Sarana dan Prasarana"
)
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_basarnas_alut.sql"

SOURCES = [
    ("UDARA", BASE_DIR / "DATA_ALUT_SAR_UDARA_2025.xlsx", "REKAP DATA UDARA"),
    ("DARAT", BASE_DIR / "REKAP_DATA_ALUT_SAR_DARAT_2025_.xlsx", "REKAP"),
    ("LAUT", BASE_DIR / "REKAP_DATA_ALUT_SAR_LAUT_2025.xlsx", "REKAP"),
]


def _normalize_kondisi(raw):
    if not raw:
        return None
    s = str(raw).strip().upper()
    if not s or s == "\xa0":
        return None
    if s in ("S", "BAIK"):
        return "SIAP"
    if "TERBATAS" in s:
        return "SIAP_TERBATAS"
    if s.startswith("US") or "RUSAK" in s:
        return "TIDAK_SIAP"
    return "LAINNYA"


def _text(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", "").strip()
    return s or None


def _find_header_offset(ws):
    """Cari index kolom 'NO' pada baris header (posisi beda antar file)."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for idx, val in enumerate(row):
            if isinstance(val, str) and val.strip().upper() == "NO":
                return idx
    raise ValueError("kolom header 'NO' tidak ditemukan")


def _load_matra(matra, xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    off = _find_header_offset(ws)
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, kode_daerah, kantor_sar, kendaraan, plat, merk, tempat, tahun, mesin, rangka, kondisi, ket = (
            row[off + i] if off + i < len(row) else None for i in range(12)
        )
        if kode_daerah is None and kantor_sar is None and kendaraan is None:
            continue
        rows.append((
            matra, int(no) if isinstance(no, (int, float)) else None,
            _text(kode_daerah), _text(kantor_sar), _text(kendaraan), _text(plat), _text(merk),
            _text(tempat), _text(tahun), _text(mesin), _text(rangka),
            _text(kondisi), _normalize_kondisi(kondisi), _text(ket),
        ))
    return rows


def main():
    with pg_cursor() as cur:
        cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

    all_rows = []
    for matra, path, sheet in SOURCES:
        if not path.exists():
            print(f"GAGAL: tidak ditemukan {path}")
            continue
        rows = _load_matra(matra, path, sheet)
        print(f"  {matra}: {len(rows)} baris")
        all_rows.extend(rows)

    with pg_cursor() as cur:
        cur.execute("DELETE FROM basarnas_alut")
        cur.executemany(
            """INSERT INTO basarnas_alut
                   (matra, no_urut, kode_daerah, kantor_sar, kendaraan, plat_lambung,
                    merk_type, tempat, tahun, no_mesin, no_rangka,
                    kondisi_saat_ini_raw, kondisi_kategori, keterangan)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            all_rows,
        )

    print(f"\nSelesai: {len(all_rows)} baris ALUT di-refresh ke basarnas_alut.")


if __name__ == "__main__":
    main()
