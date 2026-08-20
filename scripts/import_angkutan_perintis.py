# -*- coding: utf-8 -*-
"""Impor 5 sheet Daftar Angkutan Perintis Darat Tahun 2026
(docs/New/5. DARAT/Daftar Angkutan Perintis Darat Tahun 2026.xlsx) ke
tabel referensi gabungan angkutan_perintis. Lihat
scripts/schema_angkutan_perintis.sql untuk skema, dan
docs/kajian_data_baru_docs_new.md §7.1. TIDAK terkait usulan_inpres/IJD.

Provinsi hanya dicantumkan sumber saat berganti (baris kelanjutan trayek
dalam provinsi yang sama kosong) -- dilacak dengan variabel berjalan per
sheet, sama pola dengan Terminal Tipe A/LRK 2026.

Idempotent: DELETE + INSERT ulang seluruh tabel tiap run.

Usage (venv aktif):
    python scripts/import_angkutan_perintis.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

from db import db_cursor as pg_cursor  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = (
    REPO_ROOT / "docs" / "New" / "5. DARAT-20260820T015302Z-1-001"
    / "5. DARAT" / "Daftar Angkutan Perintis Darat Tahun 2026.xlsx"
)
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_angkutan_perintis.sql"


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _load_barang_perintis():
    # Kolom: NO, Provinsi, Kabupaten/Kota, Jarak/KM, KORIDOR YANG DILAYANI
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)["Angkutan Barang Perintis"]
    prov = None
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, p, kab, jarak, koridor = row[:5]
        if not koridor:
            continue
        if p:
            prov = p
        rows.append(("BARANG_PERINTIS", int(no) if isinstance(no, (int, float)) else None,
                      _s(prov), _s(kab), None, _s(koridor), _num(jarak), "KM", None))
    return rows


def _load_perkotaan_bts():
    # Kolom: NO, Provinsi, WILAYAH, NO KORIDOR, KORIDOR YANG DILAYANI, JARAK PP (KM)
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)["Angkutan Perkotaan BTS"]
    prov, wilayah = None, None
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, p, w, no_koridor, koridor, jarak = row[:6]
        if not koridor:
            continue
        if p:
            prov = p
        if w:
            wilayah = w
        rows.append(("PERKOTAAN_BTS", int(no) if isinstance(no, (int, float)) else None,
                      _s(prov), _s(wilayah), _s(no_koridor), _s(koridor), _num(jarak), "KM PP", None))
    return rows


def _load_penyeberangan_perintis():
    # Kolom: NO, PROVINSI, NAMA LINTAS, JARAK (MIL), TARGET TRIP 2026
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)["Angkutan Penyeberangan Perintis"]
    prov = None
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, p, lintas, jarak, target = row[:5]
        if not lintas:
            continue
        if p:
            prov = p
        rows.append(("PENYEBERANGAN_PERINTIS", int(no) if isinstance(no, (int, float)) else None,
                      _s(prov), None, None, _s(lintas), _num(jarak), "MIL",
                      int(target) if isinstance(target, (int, float)) else None))
    return rows


def _load_kspn():
    # Kolom: NO, PROVINSI/KAWASAN STRATEGIS, NO TRAYEK, TRAYEK YANG DILAYANI, JARAK (KM)
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)["Angkutan KSPN"]
    kawasan = None
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, k, no_trayek, trayek, jarak = row[:5]
        if not trayek:
            continue
        if k:
            kawasan = k
        rows.append(("KSPN", int(no) if isinstance(no, (int, float)) else None,
                      None, _s(kawasan), _s(no_trayek), _s(trayek), _num(jarak), "KM", None))
    return rows


def _load_jalan_perintis():
    # Kolom: NO, PROVINSI, TRAYEK YANG DILAYANI, JARAK (KM) PP
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)["Angkutan Jalan Perintis"]
    prov = None
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, p, trayek, jarak = row[:4]
        if not trayek:
            continue
        if p:
            prov = p
        rows.append(("JALAN_PERINTIS", int(no) if isinstance(no, (int, float)) else None,
                      _s(prov), None, None, _s(trayek), _num(jarak), "KM PP", None))
    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"GAGAL: tidak ditemukan {XLSX_PATH}")
        sys.exit(1)

    with pg_cursor() as cur:
        cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

    print(f"Membaca {XLSX_PATH.name}...")
    all_rows = []
    for label, loader in [
        ("Barang Perintis", _load_barang_perintis),
        ("Perkotaan BTS", _load_perkotaan_bts),
        ("Penyeberangan Perintis", _load_penyeberangan_perintis),
        ("KSPN", _load_kspn),
        ("Jalan Perintis", _load_jalan_perintis),
    ]:
        rows = loader()
        print(f"  {label}: {len(rows)} baris")
        all_rows.extend(rows)

    with pg_cursor() as cur:
        cur.execute("DELETE FROM angkutan_perintis")
        cur.executemany(
            """INSERT INTO angkutan_perintis
                   (jenis, no_urut, provinsi, wilayah, no_koridor_trayek, nama_trayek,
                    jarak, satuan_jarak, target_trip_2026)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            all_rows,
        )

    print(f"\nSelesai: {len(all_rows)} baris di-refresh ke angkutan_perintis.")


if __name__ == "__main__":
    main()
