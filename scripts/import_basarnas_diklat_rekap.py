# -*- coding: utf-8 -*-
"""Impor rekap tahunan kapasitas diklat SAR 2021-2025 -- gabungan 2 file
kecil (docs/New/6. BASARNAS/7. Peserta Pendidikan dan Pelatihan SAR/ &
8. Tenaga Pendidik Pendidikan dan Pelatihan SAR/) ke tabel referensi
basarnas_diklat_rekap. Lihat scripts/schema_basarnas_diklat_rekap.sql
untuk skema, dan docs/kajian_data_baru_docs_new.md §8.6. TIDAK terkait
usulan_inpres/IJD.

Idempotent: UPSERT per tahun.

Usage (venv aktif):
    python scripts/import_basarnas_diklat_rekap.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

from db import db_cursor as pg_cursor  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
BASARNAS_DIR = REPO_ROOT / "docs" / "New" / "6. BASARNAS-20260820T015304Z-1-001" / "6. BASARNAS"
PELATIHAN_PATH = (
    BASARNAS_DIR / "7. Peserta Pendidikan dan Pelatihan SAR"
    / "Rekapitulasi Pelatihan 2021-2025.xlsx"
)
PENDIDIK_PATH = (
    BASARNAS_DIR / "8. Tenaga Pendidik Pendidikan dan Pelatihan SAR"
    / "Rekapitulasi Tenaga Pendidik Pendidikan dan Pelatihan Pencarian dan Pertolongan 2021-2025.xlsx"
)
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_basarnas_diklat_rekap.sql"


def _int(v):
    return int(v) if isinstance(v, (int, float)) else None


def _load_pelatihan():
    wb = openpyxl.load_workbook(PELATIHAN_PATH, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        tahun, kegiatan, peserta = row[1], row[2], row[3]
        if not isinstance(tahun, (int, float)):
            continue
        out[int(tahun)] = (_int(kegiatan), _int(peserta))
    return out


def _load_pendidik():
    wb = openpyxl.load_workbook(PENDIDIK_PATH, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        tahun, pendidik = row[1], row[2]
        if not isinstance(tahun, (int, float)):
            continue
        out[int(tahun)] = _int(pendidik)
    return out


def main():
    if not PELATIHAN_PATH.exists() or not PENDIDIK_PATH.exists():
        print("GAGAL: salah satu file sumber tidak ditemukan")
        sys.exit(1)

    with pg_cursor() as cur:
        cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

    pelatihan = _load_pelatihan()
    pendidik = _load_pendidik()
    tahun_set = sorted(set(pelatihan) | set(pendidik))
    rows = [
        (t, *(pelatihan.get(t, (None, None))), pendidik.get(t))
        for t in tahun_set
    ]
    print(f"  {len(rows)} tahun")

    with pg_cursor() as cur:
        cur.executemany(
            """INSERT INTO basarnas_diklat_rekap (tahun, jumlah_kegiatan, jumlah_peserta, jumlah_tenaga_pendidik)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (tahun) DO UPDATE SET
                   jumlah_kegiatan=EXCLUDED.jumlah_kegiatan, jumlah_peserta=EXCLUDED.jumlah_peserta,
                   jumlah_tenaga_pendidik=EXCLUDED.jumlah_tenaga_pendidik, imported_at=now()""",
            rows,
        )

    print("\nSelesai: basarnas_diklat_rekap di-refresh.")


if __name__ == "__main__":
    main()
