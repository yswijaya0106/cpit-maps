# -*- coding: utf-8 -*-
"""Impor Lalu Lintas Harian Rata-Rata (LHR/AADT) per ruas Jalan Nasional
2024 (docs/New/8. KESELAMATAN/Data Lalu Lintas Harian Rata-Rata (LHR)
Jalan Nasional Tahun 2024.xlsx, sheet "Per Ruas") ke tabel referensi
bps_lhr_ruas_nasional -- lihat scripts/schema_bps_lhr_ruas_nasional.sql
untuk skema & alasan tabel terpisah (bukan merge ke attrs JSONB layer
peta JALAN NASIONAL, walau linkid cocok 1:1/~100% dengan attrs->>'LINKID'
di sana -- diverifikasi 3306/3306 match saat kajian, lihat
docs/kajian_data_baru_docs_new.md §9.3).

Tabel referensi non-spasial murni, ditampilkan lewat "Data" viewer
(DATA_TABLES di app.py), TIDAK terkait usulan_inpres/IJD.

Idempotent: UPSERT per linkid (ON CONFLICT DO UPDATE), aman dijalankan
ulang.

Kolom provinsi/kabupaten/kecamatan (opsional) diambil dari header sheet yang
sama kalau scripts/lhr_spatial_join.py sudah pernah dijalankan thd file ini
(ditulis di baris 1, posisi kolom dicari by nama header, bukan indeks tetap
-- lihat _find_enrichment_cols). Kalau belum, ketiga kolom itu tetap NULL
(atau nilai lama dipertahankan via COALESCE saat re-run, tidak ditimpa NULL).

Usage (venv aktif):
    python scripts/lhr_spatial_join.py       # sekali, isi kolom Provinsi/Kabupaten/Kecamatan di xlsx
    python scripts/import_lhr_ruas_nasional.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

from db import db_cursor as pg_cursor  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = (
    REPO_ROOT / "docs" / "New" / "8. KESELAMATAN-20260820T015309Z-1-001"
    / "8. KESELAMATAN" / "Data Lalu Lintas Harian Rata-Rata (LHR) Jalan Nasional Tahun 2024.xlsx"
)
SCHEMA_PATH = Path(__file__).resolve().parent / "schema_bps_lhr_ruas_nasional.sql"

# Indeks kolom (0-based) pada sheet "Per Ruas", diverifikasi manual thd
# header 3-baris merge (lihat docstring & kajian §9.3):
# 0=kode wilayah (tidak dipakai), 1=Linkid, 2=Linkname, 3=Lintas,
# 4=Panjang SK (KM), 5=Tahun, 6=AADT TOTAL, 7-18=Veh1..Veh8,
# 19=Volume, 20=Capacity, 21=VCR
COL_LINKID, COL_LINKNAME, COL_LINTAS, COL_PANJANG, COL_TAHUN = 1, 2, 3, 4, 5
COL_AADT_TOTAL = 6
COL_VEH = list(range(7, 19))  # Veh1, Veh2, Veh3, Veh4, Veh5a, Veh5b, Veh6a, Veh6b, Veh7a, Veh7b, Veh7c, Veh8
COL_VOLUME, COL_CAPACITY, COL_VCR = 19, 20, 21


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


# Provinsi/Kabupaten/Kecamatan ditulis scripts/lhr_spatial_join.py sbg kolom
# baru di baris header (row 1), posisi TIDAK tetap (tergantung ws.max_column
# saat script itu jalan) -- dicari by nama header, bukan indeks hardcode.
# Kosong (None, None, None) kalau lhr_spatial_join.py belum pernah dijalankan
# thd file ini -- enrichment-nya opsional, import tetap jalan tanpanya.
def _find_enrichment_cols(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = {str(v).strip(): i for i, v in enumerate(header_row) if v}
    return header.get("Provinsi"), header.get("Kabupaten/Kota"), header.get("Kecamatan")


def _load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["Per Ruas"]
    col_provinsi, col_kabupaten, col_kecamatan = _find_enrichment_cols(ws)
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        linkid = row[COL_LINKID]
        if not linkid:
            continue
        veh = [_num(row[c]) for c in COL_VEH]
        rows.append((
            str(linkid).strip(),
            row[COL_LINKNAME],
            row[COL_LINTAS],
            _num(row[COL_PANJANG]),
            int(row[COL_TAHUN]) if row[COL_TAHUN] else None,
            int(_num(row[COL_AADT_TOTAL])) if _num(row[COL_AADT_TOTAL]) is not None else None,
            *[int(v) if v is not None else None for v in veh],
            _num(row[COL_VOLUME]),
            _num(row[COL_CAPACITY]),
            _num(row[COL_VCR]),
            _str(row[col_provinsi]) if col_provinsi is not None and col_provinsi < len(row) else None,
            _str(row[col_kabupaten]) if col_kabupaten is not None and col_kabupaten < len(row) else None,
            _str(row[col_kecamatan]) if col_kecamatan is not None and col_kecamatan < len(row) else None,
        ))
    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"GAGAL: tidak ditemukan {XLSX_PATH}")
        sys.exit(1)

    with pg_cursor() as cur:
        cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

    print(f"Membaca {XLSX_PATH.name}...")
    rows = _load_rows()
    print(f"  {len(rows)} baris ruas")

    with pg_cursor() as cur:
        cur.executemany(
            """INSERT INTO bps_lhr_ruas_nasional
                   (linkid, linkname, lintas, panjang_sk_km, tahun_data, aadt_total,
                    aadt_veh1, aadt_veh2, aadt_veh3, aadt_veh4, aadt_veh5a, aadt_veh5b,
                    aadt_veh6a, aadt_veh6b, aadt_veh7a, aadt_veh7b, aadt_veh7c, aadt_veh8,
                    volume, capacity, vcr, provinsi, kabupaten, kecamatan)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT (linkid) DO UPDATE SET
                   linkname=EXCLUDED.linkname, lintas=EXCLUDED.lintas,
                   panjang_sk_km=EXCLUDED.panjang_sk_km, tahun_data=EXCLUDED.tahun_data,
                   aadt_total=EXCLUDED.aadt_total,
                   aadt_veh1=EXCLUDED.aadt_veh1, aadt_veh2=EXCLUDED.aadt_veh2,
                   aadt_veh3=EXCLUDED.aadt_veh3, aadt_veh4=EXCLUDED.aadt_veh4,
                   aadt_veh5a=EXCLUDED.aadt_veh5a, aadt_veh5b=EXCLUDED.aadt_veh5b,
                   aadt_veh6a=EXCLUDED.aadt_veh6a, aadt_veh6b=EXCLUDED.aadt_veh6b,
                   aadt_veh7a=EXCLUDED.aadt_veh7a, aadt_veh7b=EXCLUDED.aadt_veh7b,
                   aadt_veh7c=EXCLUDED.aadt_veh7c, aadt_veh8=EXCLUDED.aadt_veh8,
                   volume=EXCLUDED.volume, capacity=EXCLUDED.capacity, vcr=EXCLUDED.vcr,
                   provinsi=COALESCE(EXCLUDED.provinsi, bps_lhr_ruas_nasional.provinsi),
                   kabupaten=COALESCE(EXCLUDED.kabupaten, bps_lhr_ruas_nasional.kabupaten),
                   kecamatan=COALESCE(EXCLUDED.kecamatan, bps_lhr_ruas_nasional.kecamatan),
                   imported_at=now()""",
            rows,
        )

    n_enriched = sum(1 for r in rows if r[21] is not None)
    print(f"\nSelesai: {len(rows)} baris LHR di-upsert ke bps_lhr_ruas_nasional "
          f"({n_enriched} punya provinsi/kabupaten/kecamatan dari lhr_spatial_join.py).")


if __name__ == "__main__":
    main()
